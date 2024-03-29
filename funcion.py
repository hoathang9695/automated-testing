import time
import var
import action
import openpyxl
import logging
import checkdata
import retry
from retry import retry







logging.basicConfig(handlers=[logging.FileHandler(filename="C:/Users/Admin/PycharmProjects/pythonProject/emso.log",
                                                 encoding='utf-8', mode='a+')],
                    format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                    datefmt="%F %A %T",
                    level=logging.INFO)



def getRowCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getColumnCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value
def writeData(file,sheetName,rowum,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowum,column=columnno).value = data
    wordbook.save(file)

rows = getRowCount(var.path_baocao,'Sheet1')





class login():
    def thanh_cong_tk_emso(self):
        action.login.login3(self, "truongvck33@gmail.com", "voncamk22")
        writeData(var.path_baocao, "Sheet1", 11, 2, "x")
        writeData(var.path_baocao, "Sheet1", 25, 2, "x")
        action.logout()
    def khong_thanh_cong_tk_emso1(self):
        action.login.login1(self, "truongv22@gmail.com", "Neko101213111999")
        time.sleep(2)
        writeData(var.path_baocao, "Sheet1", 12, 2, "x")
        writeData(var.path_baocao, "Sheet1", 28, 2, "x")
    def khong_thanh_cong_tk_emso2(self):
        time.sleep(2)
        action.login.login2(self, "","")
        writeData(var.path_baocao, "Sheet1", 13, 2, "x")
        writeData(var.path_baocao, "Sheet1", 29, 2, "x")
        time.sleep(1)
    def thanh_cong_tk_google1(self):
        action.login.login4(self, "truongvck33@gmail.com", "voncamk22")
        writeData(var.path_baocao, "Sheet1", 15, 2, "x")
        action.logout()

    def thanh_cong_tk_google2(self):
        action.login.login_google(self, "truongvck22@gmail.com", "voncamk22")
        # action.logout()
    def chon_tk_dang_nhap_gan_day_chua_luu_mk(self):
        action.login.login_chon_tk_dang_nhap_gan_day_chua_luu_mk(self, "voncamk22")
        writeData(var.path_baocao, "Sheet1", 21, 2, "x")
        writeData(var.path_baocao, "Sheet1", 24, 2, "x")
        action.logout()
    def chon_tk_dang_nhap_gan_day_da_luu_mk(self):
        action.login.login_chon_tk_dang_nhap_gan_day_da_luu_mk(self)
        writeData(var.path_baocao, "Sheet1", 20, 2, "x")
        writeData(var.path_baocao, "Sheet1", 23, 2, "x")
        action.logout()
    def chon_tk_dang_nhap_gan_day_nhap_sai_pass(self):
        action.login.login_chon_tk_dang_nhap_gan_day_nhap_sai_pass(self, "truongvccds")
        writeData(var.path_baocao, "Sheet1", 18, 2, "x")
        writeData(var.path_baocao, "Sheet1", 26, 2, "x")
    def chon_tk_dang_nhap_gan_day_khong_nhap_pass(self):
        action.login.login_chon_tk_dang_nhap_gan_day_khong_nhap_pass(self)
        writeData(var.path_baocao, "Sheet1", 19, 2, "x")
        writeData(var.path_baocao, "Sheet1", 27, 2, "x")



# @retry(tries=3, delay=2, backoff=1, jitter=5, )
def thongtincanhan_anhdaidien(self):
    action.anhdaidien.anhdaidien_themmoi(self)
    action.checkdata_be.trangcanhan_thongtincanhan_anhdaidien_themmoi(self)

    action.anhdaidien.anhdaidien_khac(self)
    action.anhdaidien.anhdaidien_tuanhcosan(self)
    action.anhdaidien.anhdaidien_themkhung(self)
    action.checkdata_be.trangcanhan_thongtincanhan_anhdaidien_themkhung(self)

def thongtincanhan_anhbia(self):
    action.anhbia.anhbia_tailen(self)
    action.checkdata_be.trangcanhan_thongtincanhan_anhbia_tailen(self)
    action.anhbia.anhbia_thayanh(self)
    action.anhbia.anhbia_chonanh(self)
    action.anhbia.anhbia_chinhsuavitri(self)



class gioithieu():
    def gioithieu_tongquan(self):
        action.tongquan.tongquan(self)
        action.checkdata_be.trangcanhan_gioithieu_tongquan(self)

        action.tongquan.tongquan_dulieusai(self)

    def congviec_vahocvan(self):
        action.cong_viec_va_hoc_van.congviecvahocvan_congviec(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_congviec(self)

        action.cong_viec_va_hoc_van.congviecvahocvan_daihoc(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_daihoc(self)

        action.cong_viec_va_hoc_van.congviecvahocvan_trunghoc(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_trunghoc(self)

    def xemsukientrongdoi_congviecvahocvan(self):

        action.xemkientrongdoi.congviecvahocvan_congviec(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_xemsukientrongdoi_congviec(self)
        #
        action.xemkientrongdoi.congviecvahocvan_daihoc(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_xemsukientrongdoi_daihoc(self)

        action.xemkientrongdoi.congviecvahocvan_trunghoc(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_xemsukientrongdoi_trunghoc(self)

    def trangcanhan_gioithieu_congviecvahocvan_addthem(self):
        action.cong_viec_va_hoc_van.congviecvahocvan_addthem(self)

    def trangcanhan_gioithieu_thongtincoban(self):
        action.thongtincoban.thongtincoban(self)
        action.checkdata_be.trangcanhan_gioithieu_thongtincoban(self)

        action.thongtincoban.thongtincoban_dulieusai(self)

    def trangcanhan_gioithieu_gd_va_mqh(self):
        action.giadinh_va_cacmoiquanhe.gd_va_mqh_taikhoan1(self)
        action.checkdata_be.trangcanhan_gioithieu_giadinh_va_cacmqh(self)

    def trangcanhan_gioithieu_sukientrongdoi(self):
        action.sukientrongdoi.sukientrongdoi_xem(self)

        action.sukientrongdoi.sukientrongdoi_themmoisukien_check(self)
        action.sukientrongdoi.taosukienrieng(self)

        action.sukientrongdoi.nhacuavadoisong(self)
        action.sukientrongdoi.moiquanhe(self)
        action.sukientrongdoi.work(self)
        action.sukientrongdoi.tuongnho(self)
        action.sukientrongdoi.daumocvathanhtuu(self)
        action.sukientrongdoi.suckhoethechatvatinhthan(self)
        action.sukientrongdoi.sothichhoatdong(self)
        action.sukientrongdoi.dulich(self)
        action.sukientrongdoi.giadinh(self)
        action.sukientrongdoi.hocvan(self)

        action.sukientrongdoi.sukientrongdoi_themmoisukien(self)
        action.checkdata_be.trangcanhan_gioithieu_sukientrongdoi(self)
        action.sukientrongdoi.sukientrongdoi_themmoisukien_chinhsua_xoa(self)

    def trangcanhan_gioithieu_noitungsong(self):
        action.noitungsong.noitungsong(self)


class banbe():
    def banbe_tatcabanbe(self):
        action.banbe.tatcabanbe(self)
        action.banbe.banbe_chinhsuaquyenriengtu(self)



class anh_video():
    def anh_video(self):
        action.anh_video.anh(self)
        action.anh_video.video(self)


class check_thongtin_trangcanhan():
    def check_thongtin_trangcanhan(self):
        action.check_thongtin_trangcanhan.check_thongtin_trangcanhan(self)
    def phandangchuy(self):
        action.check_thongtin_trangcanhan.phandangchuy(self)


class trangchu():
    def taobaiviet_khoangkhac(self):
        action.trangchu.taobaiviet_congkhai(self)
        action.trangchu.taobaiviet_banbe(self)
        action.trangchu.taobaiviet_riengtu(self)
        action.trangchu.taobaimoment(self)
        action.trangchu.camxuc_hoatdong(self)

    def menu(self):
        action.trangchu.menu(self)
        action.trangchu.menu_tao(self)

    def tinnhanmoi(self):
        action.trangchu.tinnhanmoi(self)

    def chat(self):
        action.trangchu.chat(self)

    def chat_xemtatca(self):
        action.trangchu.chat_xemtatca(self)


    def taonhom(self):
        action.trangchu.taonhom(self)


    def thongtindoanchat_nhom(self):
        action.trangchu.thongtindoanchat_nhom(self)


    def caidatcanhan_chuyentaikhoan(self):
        action.trangchu.caidatcanhan_chuyentaikhoan((self))

    def canhan_caidatvaquyenriengtu(self):
        action.trangchu.caidatvaquyenriengtu_chung(self)
        action.trangchu.caidatvaquyenriengtu_baomatvadangnhap(self)
        action.trangchu.caidatvaquyenriengtu_thongtinbantrenemso(self)
        action.trangchu.caidatvaquyenriengtu_trangcanhanvaganthe(self)
        action.trangchu.caidatvaquyenriengtu_baivietcongkhai(self)
        action.trangchu.caidatvaquyenriengtu_chan(self)
        # action.trangchu.caidatvaquyenriengtu_batkiemtien(self)        #lỗi xác minh page từ Tuấn

    def trogiupvahotro(self):
        action.trangchu.trogiupvahotro_sudungemso(self)

    def caidatcanhan_manhinh(self):
        action.trangchu.caidatcanhan_manhinh(self)

    def caidatcanhan_donggopykien(self):
        action.trangchu.caidatcanhan_donggopykien(self)

    def caidatcanhan_dangxuat(self):
        action.trangchu.caidatcanhan_dangxuat(self)

    def trangchu_timkiem(self):
        action.trangchu.trangchu_timkiem(self)


class khoanhkhac():
    def dangtheodoi(self):
        action.khoanhkhac.dangtheodoi(self)
    def tructiep(self):
        action.khoanhkhac.tructiep(self)

    def taokhoanhkhac(self):
        action.khoanhkhac.taokhoanhkhac(self)

    def taikhoan_duocdexuat(self):
        action.khoanhkhac.taikhoan_duocdexuat(self)

    def taikhoan_dangtheodoi(self):
        action.khoanhkhac.taikhoan_dangtheodoi(self)

    def khoanhkhac_timkiem(self):
        action.khoanhkhac.khoanhkhac_timkiem(self)


class watch():
    def timkiem(self):
        action.watch.timkiem(self)

    def trangchu(self):
        action.watch.trangchu(self)

    def chuongtrinh(self):
        action.watch.chuongtrinh(self)

    def videodaluu(self):
        action.watch.videodaluu(self)

    def dangtheodoi(self):
        action.watch.dangtheodoi(self)


class trang():
    def timkiem_trangcuaban(self):
        action.trang.timkiem_trangcuaban(self)
    def khampha(self):
        action.trang.khampha(self)
    def trangdathich(self):
        action.trang.trangdathich(self)
    def trangdathich_dau3cham(self):
        action.login.login4(self, "truongvck333@gmail.com", "voncamk22")
        action.trang.trangdathich_dau3cham(self)
    def loimoi(self):
        action.trang.loimoi(self)
    def taotrangmoi(self):
        action.login.login4(self, "truongvck22@gmail.com", "voncamk22")
        action.trang.taotrangmoi_banhang(self)
        action.trang.taotrangmoi_bankhoahoc(self)
        action.trang.taotrangmoi_trangnoidung(self)

    def trang_gioithieu(self):
        action.trang.trang_gioithieu(self)
        action.trang.trang_gioithieu_dulieusai(self)

    def anh(self):
        action.trang.anh(self)

    def music(self):
        action.trang.music(self)

    def video(self):
        action.trang.video(self)

    def xemthem(self):
        action.trang.xemthem(self)

    def cuahang(self):
        action.trang.cuahang(self)

    def cacnutkhac(self):
        action.trang.cacnutkhac(self)

    def themnuthanhdong(self):
        action.trang.themnuthanhdong(self)

    def anhdaidien_anhbia(self):
        action.trang.anhdaidien_anhbia(self)

    def taobaiviet(self):
        action.trang.taobaiviet(self)

    def hopthu(self):
        action.trang.hopthu(self)

    def thongbao(self):
        action.trang.thongbao(self)

    def chatluongtrang(self):
        action.trang.chatluongtrang(self)

    def baivietdalenlich(self):
        action.trang.baivietdalenlich(self)

    def chinhsuathongtintrang(self):
        action.trang.chinhsuathongtintrang(self)

    def caidat(self):
        action.trang.caidat_chung(self)
        action.trang.caidat_thongbao(self)
        action.trang.caidat_vaitrotrentrang(self)
        action.trang.caidat_chutaikhoantrang(self)
        action.trang.caidat_nguoivatrangkhac(self)  #trắng trang khi chọn 7720
        action.trang.caidat_nhatkyhoatdong(self)
        action.trang.caidat_caidathopthu(self)


class nhom:
    def timkiem(self):
        action.nhom.timkiem(self)
    def bangtincuaban(self):
        action.nhom.bangtincuaban(self)
    def khampha(self):
        action.nhom.khampha(self)
    def moithamgianhom(self):
        action.nhom.moithamgianhom(self)

    def taonhommoi(self):
        action.nhom.taonhommoi(self, "Nhóm công khai", "Test tạo nhóm công khai", var.congkhai)
        action.nhom.taonhommoi(self, "Nhóm riêng tư", "Test tạo nhóm riêng tư", var.riengtu)

    def nhombanquanly(self):
        action.nhom.nhombanquanly(self)

    def nhombanthamgia(self):
        action.nhom.nhombanthamgia(self)

    def thaoluan(self):
        action.nhom.thaoluan(self)

    def thanhvien(self):
        action.nhom.thanhvien(self)

    def filephuongtien(self):
        action.nhom.filephuongtien(self)
        action.nhom.anh(self)
        action.nhom.video(self)
        action.nhom.khoanhkhac(self)
        action.nhom.album(self)

    def yeucaulamthanhvien(self):
        action.nhom.yeucaulamthanhvien(self)

    def baivietdalenlich(self):
        action.nhom.baivietdalenlich(self)

    def nhatkyhoatdong(self):
        action.nhom.nhatkyhoatdong(self)

    def quytacnhom(self):
        action.nhom.quytacnhom(self)

    def noidungthanhvienbaocao(self):
        action.nhom.noidungthanhvienbaocao(self)

    def thongbaokiemduyet(self):
        action.nhom.thongbaokiemduyet(self)

    def cauhoichonthanhvien(self):
        action.nhom.cauhoichonthanhvien(self)

    def caidatnhom(self):
        action.nhom.caidatnhom(self)
        action.nhom.thietlapnhom(self)
        action.nhom.tuychinhnhom(self)
        action.nhom.quanlynoidungthaoluan(self)



class quatrinhmuahang():
    def timkiemsanpham(self):
        action.quatrinhmuahang.timkiemsanpham(self)

    def timkiemsanpham_thanhtoanonline(self):
        action.quatrinhmuahang.timkiemsanpham_thanhtoanonline(self)

    def timkiemsanpham_voucher(self):
        action.quatrinhmuahang.timkiemsanpham_voucher(self)

    def xacnhandon(self):
        action.quatrinhmuahang.xacnhandon(self)

    def chuanbihang(self):
        action.quatrinhmuahang.chuanbihang(self)

    def donvivanchuyendanglayhang(self):
        action.quatrinhmuahang.donvivanchuyendanglayhang(self)

    def donvivanchuyendalayhang(self):
        action.quatrinhmuahang.donvivanchuyendalayhang(self)

    def danggiaohang(self):
        action.quatrinhmuahang.danggiaohang(self)

    def danhanhang(self):
        action.quatrinhmuahang.danhanhang(self)

    def chothanhtoan_dahuy(self):
        action.quatrinhmuahang.timkiemsanpham(self)
        action.quatrinhmuahang.chothanhtoan_dahuy(self)

    def shop_het_hang(self):
        action.quatrinhmuahang.shop_het_hang(self)


    def trahanghoantien_dongy(self):
        action.quatrinhmuahang.timkiemsanpham(self)
        action.quatrinhmuahang.xacnhandon(self)
        action.quatrinhmuahang.chuanbihang(self)
        action.quatrinhmuahang.donvivanchuyendanglayhang(self)
        action.quatrinhmuahang.donvivanchuyendalayhang(self)
        action.quatrinhmuahang.danggiaohang(self)
        action.quatrinhmuahang.trahanghoantien_dongy(self)


    def trahanghoantien_khongdongy_dongy(self):
        action.quatrinhmuahang.timkiemsanpham(self)
        action.quatrinhmuahang.xacnhandon(self)
        action.quatrinhmuahang.chuanbihang(self)
        action.quatrinhmuahang.donvivanchuyendanglayhang(self)
        action.quatrinhmuahang.donvivanchuyendalayhang(self)
        action.quatrinhmuahang.danggiaohang(self)
        action.quatrinhmuahang.trahanghoantien_khongdongy_dongy(self)


    def trahanghoantien_khongdongy_khongdongy(self):
        action.quatrinhmuahang.timkiemsanpham(self)
        action.quatrinhmuahang.xacnhandon(self)
        action.quatrinhmuahang.chuanbihang(self)
        action.quatrinhmuahang.donvivanchuyendanglayhang(self)
        action.quatrinhmuahang.donvivanchuyendalayhang(self)
        action.quatrinhmuahang.danggiaohang(self)
        action.quatrinhmuahang.trahanghoantien_khongdongy_khongdongy(self)





    def trahanghoantien_nguoimuahuy(self):
        action.quatrinhmuahang.timkiemsanpham(self)
        action.quatrinhmuahang.xacnhandon(self)
        action.quatrinhmuahang.chuanbihang(self)
        action.quatrinhmuahang.donvivanchuyendanglayhang(self)
        action.quatrinhmuahang.donvivanchuyendalayhang(self)
        action.quatrinhmuahang.danggiaohang(self)
        action.quatrinhmuahang.trahanghoantien_nguoimuahuy(self)


    def shophuydon(self):
        action.quatrinhmuahang.timkiemsanpham(self)
        action.quatrinhmuahang.shophuydon(self)


class quanlysanpham():
    def themsanphammoi(self):
        action.quanlysanpham.themsanphammoi_guipheduyet(self)
        action.thuongmai.duyetsanpham_tuchoi(self)

        action.quanlysanpham.themsanphammoi_guipheduyet(self)
        action.thuongmai.duyetsanpham_duyet(self)
        #
        action.quanlysanpham.themsanphammoi_luunhap(self)

    def danhsachsanpham(self):
        action.quanlysanpham.danhsachsanpham(self)


class add_dulieuemso():
    def themsanphammoi1(self):
        action.add_dulieuemso.themsanphammoi1(self)

    def themsanphammoi1_khongconghanhcha(self):
        action.add_dulieuemso.themsanphammoi1_khongconghanhcha(self)

    def vietfilesanpham(self):
        action.add_dulieuemso.vietfilesanpham(self)

    # @retry(tries=5, delay=2, backoff=1, jitter=5, )
    def get_data_lazada(self):
        action.add_dulieuemso.get_data_lazada(self)

    def get_data_lazada_theonghanh(self):
        action.add_dulieuemso.get_data_lazada_theonghanh(self)


    def get_image(self):
        action.add_dulieuemso.get_image(self)


class kenhmarketing():
    def magiamgiacuashop(self):
        # action.kenhmarketing.magiamgiacuashop(self)
        action.kenhmarketing.danhsachmagiamgia(self)

    def chuongtrinhkhuyenmai(self):
        action.kenhmarketing.chuongtrinhkhuyenmai_tatca(self)

        action.kenhmarketing.chuongtrinhkhuyenmai_flashsale(self)
        action.kenhmarketing.chuongtrinhkhuyenmai_pheduyetsp_flashsale(self)

        action.kenhmarketing.chuongtrinhkhuyenmai_emsocampaign(self)
        action.kenhmarketing.chuongtrinhkhuyenmai_pheduyetsp_cambpaign(self)

    def flashsalecuashop(self):
        action.kenhmarketing.flashsalecuashop_admin(self)
        action.kenhmarketing.flashsalecuashop_themsanpham(self)
        action.kenhmarketing.flashsalecuashop_danhsachchuongtrinh(self)


    def chuongtrinhcuashop(self):
        action.kenhmarketing.flashsalecuashop_chuongtrinhcuashop(self)



