from urllib import response

import openpyxl
import asyncio
import requests
# from pytz import unicode
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
options = Options()
from seleniumwire.utils import decode as sw_decode
from selenium.common.exceptions import NoSuchElementException
import time
from selenium.webdriver.support.ui import Select
import json
import var
from retry import retry
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
import urllib.request as request
import subprocess
import urllib.error as error

import datetime as dt
from seleniumwire import webdriver
from seleniumwire.utils import decode

chrome_options = Options()
driver = webdriver.Chrome(var.PATH, chrome_options=chrome_options)

import pandas as pd
import codecs
import decode
from selenium.common.exceptions import InvalidSessionIdException
import unittest
import logging
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys


#read wrirte csv
def getRowCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getColumnCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rowum,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowum,column=columnno).value = data
    wordbook.save(file)


rows = getRowCount(var.path_baocao,'Sheet1')


file_name = 'data_emso.json'
with open(file_name, 'r', encoding='utf-8') as f:
    data = json.load(f, strict = False)




logging.basicConfig(handlers=[logging.FileHandler(filename="C:/Users/Admin/PycharmProjects/pythonProject/emso.log",
                                                 encoding='utf-8', mode='a+')],
                    format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                    datefmt="%F %A %T",
                    level=logging.INFO)



logging.debug("đây là debug")
logging.info("đây là info")
logging.warning("đây là warning")
logging.error("đây là error")
logging.critical("đây là critical")

@retry(tries=3, delay=2, backoff=1, jitter=5, )
def logout():
    driver.implicitly_wait(15)
    time.sleep(2)
    driver.set_window_size(1024, 768)
    driver.find_element(By.XPATH, var.logout_chontaikhoan1).click()
    driver.find_element(By.XPATH, var.logout_dangxuat).click()
    writeData(var.path_baocao, "Sheet1", 31, 2, "x")
    time.sleep(2)


class login():
    @retry(tries=3, delay=2, backoff=1, jitter=5, )
    def login1(self, user, password):
        driver.implicitly_wait(15)
        driver.get(var.url)
        driver.maximize_window()
        driver.find_element(By.XPATH, var.login_user).send_keys(user)
        driver.find_element(By.XPATH, var.login_password).send_keys(password)
        driver.find_element(By.XPATH, var.login_submit).click()
        time.sleep(1)
        message_sai_tk_mk = driver.find_element(By.XPATH, var.message_sai_tk_mk1).text
        print(message_sai_tk_mk)
        logging.info("Đăng nhập bằng tài khoản emso-(Đăng nhập không thành công/Chuyển tài khoản không thành công): Nhập sai tài khoản/mật khẩu")
        logging.info("Message: Tài khoản hoặc mật khẩu không đúng, vui lòng kiểm tra lại.")
        logging.info( message_sai_tk_mk == "Tài khoản hoặc mật khẩu không đúng, vui lòng kiểm tra lại.")
        print("tên tài khoản login:", user)
        print("mật khẩu login:", password)
        print("login vào phần mềm emso khong thành công")

    @retry(tries=3, delay=2, backoff=1, jitter=5, )
    def login2(self, user, password):
        driver.implicitly_wait(15)
        driver.get(var.url)
        driver.maximize_window()
        driver.find_element(By.XPATH, var.login_user).send_keys(user)
        driver.find_element(By.XPATH, var.login_password).send_keys(password)
        driver.find_element(By.XPATH, var.login_submit).click()
        time.sleep(1)
        # bỏ trống tài khoản
        message_bo_trong_tk = driver.find_element(By.XPATH, var.message_bo_trong_tk1).text
        print(message_bo_trong_tk)
        logging.info("Đăng nhập bằng tài khoản emso-(Đăng nhập không thành công/Chuyển tài khoản không thành công): Không nhập tài khoản/mật khẩu")
        logging.info("Message Email, Số điện thoại:Email hoặc số điện thoại không được để trống ")
        logging.info(message_bo_trong_tk == "Email hoặc số điện thoại không được để trống")
        # bỏ trống mật khẩu
        message_bo_trong_mk = driver.find_element(By.XPATH, var.message_bo_trong_mk1).text
        print(message_bo_trong_mk)
        logging.info("Đăng nhập bằng tài khoản emso-(Đăng nhập không thành công/Chuyển tài khoản không thành công): Không nhập tài khoản/mật khẩu")
        logging.info("Message Mật khẩu: Mật khẩu không được để trống")
        logging.info(message_bo_trong_mk == "Mật khẩu không được để trống")

        print("tên tài khoản login:", user)
        print("mật khẩu login:", password)
        print("login vào phần mềm emso khong thành công")

    @retry(tries=3, delay=2, backoff=1, jitter=5, )
    def login3(self, user, password):
        driver.implicitly_wait(15)
        driver.get(var.url)
        driver.maximize_window()
        # driver.set_window_size(400, 600)

        driver.find_element(By.XPATH, var.login_user).send_keys(user)
        driver.find_element(By.XPATH, var.login_password).send_keys(password)
        driver.find_element(By.XPATH, var.login_submit).click()
        time.sleep(3.5)
        login_thanh_cong = driver.find_element(By.XPATH, var.login_thanhcong_khoanh_khac).text
        print(login_thanh_cong)
        logging.info("Đăng nhập bằng tài khoản emso/Gmail")
        logging.info( login_thanh_cong == "Khoảnh khắc")
        print("tên tài khoản login:", user)
        print("mật khẩu login:", password)
        print("login vào phần mềm emso thành công")
        driver.find_element(By.XPATH, var.trangcanhan).click()
        time.sleep(1)

    @retry(tries=3, delay=2, backoff=1, jitter=5, )
    def login4(self, user, password):
        driver.implicitly_wait(15)
        driver.get(var.url)
        driver.maximize_window()
        driver.find_element(By.XPATH, var.login_user).send_keys(user)
        driver.find_element(By.XPATH, var.login_password).send_keys(password)
        driver.find_element(By.XPATH, var.login_submit).click()
        time.sleep(3.5)
        login_thanh_cong = driver.find_element(By.XPATH, var.login_thanhcong_khoanh_khac).text
        print(login_thanh_cong)
        logging.info("Đăng nhập bằng tài khoản google - Đăng nhập thành công - tài khoản đã có sẵn và đã được đồng bộ")
        logging.info( login_thanh_cong == "Khoảnh khắc")
        print("tên tài khoản login:", user)
        print("mật khẩu login:", password)
        print("login vào phần mềm emso thành công")


    def login5(self, user, password):
        driver.implicitly_wait(15)
        driver.get(var.url_admin)
        driver.maximize_window()
        driver.find_element(By.XPATH, var.login_admin_user).send_keys(user)
        driver.find_element(By.XPATH, var.login_admin_password).send_keys(password)
        driver.find_element(By.XPATH, var.login_admin_submit).click()
        time.sleep(3.5)
        login_thanh_cong_admin = driver.find_element(By.XPATH, var.login_thanh_cong_admin).text
        print("Đã login vào trang " + login_thanh_cong_admin + "Thành Công")





    @retry(tries=3, delay=2, backoff=1, jitter=5, )
    def login_google(self, user, passsword):
        driver.implicitly_wait(20)
        driver.get(var.url)
        # driver.maximize_window()
        driver.find_element(By.XPATH, var.login_google).click()
        handles = driver.window_handles
        for handle in handles:
            driver.switch_to.window(handle)
            print(driver.title)
        time.sleep(1)
        driver.find_element(By.XPATH, var.login_google_nhap_gmail).send_keys(user)
        driver.find_element(By.XPATH, var.login_google_tiep).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.login_google_nhap_password).send_keys(passsword)
        time.sleep(1)
        driver.find_element(By.XPATH, var.login_google_tiep).click()
        time.sleep(15)
        for handle in handles:
            driver.switch_to.window(handle)
            print(driver.title)
            if driver.title =="Mạng xã hội Emso":
                login_thanh_cong = driver.find_element(By.XPATH, var.login_thanhcong_khoanh_khac).text
                print(login_thanh_cong)
                logging.info("Đăng nhập bằng tài khoản google - Đăng nhập thành công - tài khoản chưa có sẵn và đã được đồng bộ")
                logging.info(login_thanh_cong == "Khoảnh khắc")

                logout()
                writeData(var.path_baocao, "Sheet1", 16, 2, "x")
                print("Current session is {}".format(driver.session_id))
                driver.close()


    @retry(tries=3, delay=2, backoff=1, jitter=5, )
    def login_chon_tk_dang_nhap_gan_day_nhap_sai_pass(self, password):
        driver.implicitly_wait(15)
        driver.get(var.url)
        driver.maximize_window()
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day_password).send_keys(password)
        driver.find_element(By.XPATH, var.login_nho_mat_khau).click()
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day_dang_nhap).click()
        time.sleep(1)
        message_ganday_sai_mk = driver.find_element(By.XPATH, var.message_sai_tk_mk1).text
        print(message_ganday_sai_mk)
        logging.info("Chọn tài khoản đăng nhập gần đây - (Đăng nhập không thành công/Chuyển tài khoan không thành công) - Nhập sai pass")
        logging.info("Message: Tài khoản hoặc mật khẩu không đúng, vui lòng kiểm tra lại.")
        logging.info( message_ganday_sai_mk == "Tài khoản hoặc mật khẩu không đúng, vui lòng kiểm tra lại.")

    @retry(tries=3, delay=2, backoff=1, jitter=5, )
    def login_chon_tk_dang_nhap_gan_day_khong_nhap_pass(self):
        driver.implicitly_wait(15)
        driver.get(var.url)
        driver.maximize_window()
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day_password).clear()
        driver.find_element(By.XPATH, var.login_nho_mat_khau).click()
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day_dang_nhap).click()
        time.sleep(1)
        message_ganday_khong_nhap_mk = driver.find_element(By.XPATH, var.message_ganday_khong_nhap_mk1).text
        print(message_ganday_khong_nhap_mk)
        logging.info("Chọn tài khoản đăng nhập gần đây - (Đăng nhập không thành công/Chuyển tài khoan không thành công) - Không nhập pass")
        logging.info("Message: Mật khẩu không được để trống")
        logging.info( message_ganday_khong_nhap_mk == "Mật khẩu không được để trống")

    @retry(tries=3, delay=2, backoff=1, jitter=5, )
    def login_chon_tk_dang_nhap_gan_day_chua_luu_mk(self, password):
        driver.implicitly_wait(15)
        driver.get(var.url)
        driver.maximize_window()
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day_password).send_keys(password)
        driver.find_element(By.XPATH, var.login_nho_mat_khau).click()
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day_dang_nhap).click()
        time.sleep(3.5)
        login_thanh_cong = driver.find_element(By.XPATH, var.login_thanhcong_khoanh_khac).text
        print(login_thanh_cong)
        logging.info("Chọn tài khoản đăng nhập gần đây - (login thành công/Chuyển tài khoản thành công) - chưa lưu mật khẩu")
        logging.info( login_thanh_cong == "Khoảnh khắc")

    @retry(tries=3, delay=2, backoff=1, jitter=5, )
    def login_chon_tk_dang_nhap_gan_day_da_luu_mk(self):
        driver.implicitly_wait(15)
        driver.get(var.url)
        driver.maximize_window()
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day).click()
        time.sleep(3.5)
        login_thanh_cong = driver.find_element(By.XPATH, var.login_thanhcong_khoanh_khac).text
        print(login_thanh_cong)
        logging.info("Chọn tài khoản đăng nhập gần đây - (login thành công/Chuyển tài khoản thành công) - đã lưu mật khẩu")
        logging.info( login_thanh_cong == "Khoảnh khắc")



class anhdaidien():
    def anhdaidien_themmoi(self):
        driver.implicitly_wait(15)
        #thêm ảnh đại diện mới
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_iconmayanh).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_taianhlen).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_icon_taianhlen).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien1.exe")
        time.sleep(1)
        del driver.requests
        driver.find_element(By.XPATH, var.capnhatanhdaidien_mota).send_keys(data['trangcanhan_anhdaidien_anhbia']['anhdaidien_mota'])
        driver.find_element(By.XPATH, var.capnhatanhdaidien_luu).click()
        writeData(var.path_baocao, "Sheet1", 52, 2, "x")
        time.sleep(7)

    def anhdaidien_khac(self):
        driver.implicitly_wait(15)
        #tải lên ảnh đại diện khác
        driver.find_element(By.XPATH, var.trangcanhan_iconmayanh).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_taianhlen).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_icon_taianhlen).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien3.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_mota).send_keys(data['trangcanhan_anhdaidien_anhbia']['anhdaidienkhac_mota'])
        driver.find_element(By.XPATH, var.capnhatanhdaidien_luu).click()
        writeData(var.path_baocao, "Sheet1", 53, 2, "x")
        time.sleep(7)


    def anhdaidien_tuanhcosan(self):
        driver.implicitly_wait(15)
        # Chọn từ 1 ảnh có sẵn
        # huỷ
        driver.find_element(By.XPATH, var.trangcanhan_iconmayanh).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonanhthu2).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonanhcosan_luu).click()
        time.sleep(7)
        writeData(var.path_baocao, "Sheet1", 54, 2, "x")

    def anhdaidien_themkhung(self):
        driver.implicitly_wait(15)
        #Thêm khung
        driver.find_element(By.XPATH, var.trangcanhan_iconmayanh).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_themkhung).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_dangquauday).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_dangbuonday).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_yeuthuongvn).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_wowvn).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_tuhaovn).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_trungthu).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_rangrovietnam).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_quockhanhvietnam).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_phunuvietnam).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_noel).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_nhagiaovietnam).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_lacquanvietnam).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_cmsn).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_vulanbaohieu).click()
        time.sleep(0.3)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_x).click()
        time.sleep(1)
        #chọn khung yeu thuong vn
        driver.find_element(By.XPATH, var.trangcanhan_iconmayanh).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_themkhung).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_yeuthuongvn).click()
        del driver.requests
        time.sleep(1)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_luu2).click()
        writeData(var.path_baocao, "Sheet1", 55, 2, "x")
        time.sleep(3.5)

class anhbia():
    def anhbia_tailen(self):
        driver.implicitly_wait(15)
        # driver.find_element(By.XPATH, var.trangcanhan).click()
        time.sleep(1.5)
        #tải lên ảnh bìa từ thiết bị
        driver.execute_script("window.scrollBy(0,-100)", "")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_iconchinhsua_anhbia).click()
        driver.find_element(By.XPATH, var.trangcanhan_anhbia_taianhlen).click()
        time.sleep(1)
        del driver.requests
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia1.exe")
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.capnhat)
        driver.execute_script("arguments[0].click();", button)

        writeData(var.path_baocao, "Sheet1", 57, 2, "x")
        time.sleep(2.5)

    def anhbia_thayanh(self):
        driver.implicitly_wait(15)
        #thay ảnh bìa từ thiết bị
        driver.find_element(By.XPATH, var.trangcanhan_iconchinhsua_anhbia).click()
        driver.find_element(By.XPATH, var.trangcanhan_anhbia_taianhlen).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia2.exe")
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.capnhat)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2.5)
        writeData(var.path_baocao, "Sheet1", 58, 2, "x")
    def anhbia_chonanh(self):
        driver.implicitly_wait(15)
        #cập nhật ảnh bìa từ ảnh của bạn
        driver.find_element(By.XPATH, var.trangcanhan_iconchinhsua_anhbia).click()
        driver.find_element(By.XPATH, var.trangcanhan_anhbia_chontuanhcuaban).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_anhbia_chonanhthu2).click()
        time.sleep(2)
        button = driver.find_element(By.XPATH, var.capnhat)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2.5)
        writeData(var.path_baocao, "Sheet1", 59, 2, "x")

    def anhbia_chinhsuavitri(self):
        driver.implicitly_wait(15)
        # chỉnh sửa vị tri ảnh bìa
        driver.find_element(By.XPATH, var.trangcanhan_iconchinhsua_anhbia).click()
        driver.find_element(By.XPATH, var.trangcanhan_anhbia_datlaivitri).click()
        time.sleep(1)
        anh_bia = driver.find_element(By.XPATH, var.trangcanhan_anhbia)
        button_themmoi = driver.find_element(By.XPATH, var.trangcanhan_button_themmoi)
        action1 = ActionChains(driver)
        action1.drag_and_drop(anh_bia, button_themmoi).perform()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.capnhat)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2.5)
        writeData(var.path_baocao, "Sheet1", 60, 2, "x")



class checkdata_be():
    def trangcanhan_thongtincanhan_anhdaidien_themmoi(self):
        driver.implicitly_wait(10)
        time.sleep(2)
        for request in driver.requests:
            if request.url == "https://snapi.emso.asia/api/v1/me":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: có file ảnh đại diện to không?")
                logging.info(res['avatar_media']['url'] != None)
                logging.info((res['avatar_media']['url']))

                logging.info("check back-end: có file ảnh đại diện nhỏ không?")
                logging.info(res['avatar_media']['preview_url'] != None)
                logging.info((res['avatar_media']['preview_url']))

                logging.info("check back-end: Mô tả ảnh đại diện")
                logging.info("Ảnh đại diện - Mô tả - " + data['trangcanhan_anhdaidien_anhbia']['anhdaidien_mota'])
                logging.info(res['avatar_media']['description'] == data['trangcanhan_anhdaidien_anhbia']['anhdaidien_mota'])
                break
            else:
                print("không có  response")
    def trangcanhan_thongtincanhan_anhdaidien_themkhung(self):
        driver.implicitly_wait(10)
        time.sleep(2)
        for request in driver.requests:
            if request.url == "https://snapi.emso.asia/api/v1/me":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)
                logging.info("check back-end: Chọn khung cho avatar")
                logging.info("Khung ảnh đại diện - " + data['trangcanhan_anhdaidien_anhbia']['khung_anhdaidien'])
                logging.info(res['avatar_media']['frame']['name'] == data['trangcanhan_anhdaidien_anhbia']['khung_anhdaidien'])
                logging.info((res['avatar_media']['frame']['url']))
                break
            else:
                print("không có  response")
    def trangcanhan_thongtincanhan_anhbia_tailen(self):
        driver.implicitly_wait(15)
        for request in driver.requests:
            if request.url == "https://snapi.emso.asia/api/v1/me":
                data = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data = data.decode("utf8")
                res = json.loads(data)

                logging.info("check back-end: có file ảnh bìa không?")
                logging.info(res['banner']['show_url'] != None)
                logging.info(res['banner']['show_url'])

                logging.info("check back-end: có file ảnh bìa trên dòng thời gian không?")
                logging.info(res['banner']['url'] != None)
                logging.info(res['banner']['url'])

                logging.info("check back-end: có file ảnh bìa trong danh mục ảnh không?")
                logging.info(res['banner']['preview_url'] != None)
                logging.info(res['banner']['preview_url'])

                break
            else:
                print("không có  response")
    def trangcanhan_gioithieu_tongquan(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        for request in driver.requests:
            if request.url == "https://snapi.emso.asia/api/v1/accounts/111169896815147328/abouts":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Tổng quan")
                logging.info("check back-end: Sống tại - " + data['trangcanhan_gioithieu_tongquan']['songtai'])
                logging.info((res['general_information']['place_live']['title']) == data['trangcanhan_gioithieu_tongquan']['songtai'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Tổng quan")
                logging.info("check back-end: Đến từ - " + data['trangcanhan_gioithieu_tongquan']['dentu'])
                logging.info((res['general_information']['hometown']['title']) == data['trangcanhan_gioithieu_tongquan']['dentu'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Tổng quan")
                logging.info("check back-end: Tình trạng mqh/Ngày bắt đầu - Ly Thân/2022-01-01")
                logging.info("Mối quan hệ - Ly thân")

                #123
                logging.info(res['account_relationship']['relationship_category']['name'])
                logging.info((res['account_relationship']['relationship_category']['name']) == "Ly Thân")
                logging.info("Mối quan hệ - Ngày bắt đầu")
                logging.info((res['account_relationship']['start_date']) == "2022-01-01")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Tổng quan")
                logging.info("check back-end: Số điện thoại - " + data['trangcanhan_gioithieu_tongquan']['sodienthoai'])
                logging.info(res['general_information']['phone_number'])
                logging.info((res['general_information']['phone_number']) == data['trangcanhan_gioithieu_tongquan']['sodienthoai'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Tổng quan")
                logging.info("check back-end: Biệt danh - " + data['trangcanhan_gioithieu_tongquan']['bietdanh'])
                logging.info((res['general_information']['other_name']))
                logging.info((res['general_information']['other_name']) == data['trangcanhan_gioithieu_tongquan']['bietdanh'])

                break
            else:
                print("không có  response")
    def trangcanhan_gioithieu_cvvahocvan_congviec(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        for request in driver.requests:
            if request.url[0:70] == "https://snapi.emso.asia/api/v1/accounts/111169896815147328/life_events":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc")
                logging.info("check back-end: Công ty - " + data['trangcanhan_gioithieu_congviecvahocvan']['congty'])
                logging.info(data['trangcanhan_gioithieu_congviecvahocvan']['congty'])
                logging.info((res['company']) == data['trangcanhan_gioithieu_congviecvahocvan']['congty'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc")
                logging.info("check back-end: Chức vụ - " + data['trangcanhan_gioithieu_congviecvahocvan']['chucvu'])
                logging.info(data['trangcanhan_gioithieu_congviecvahocvan']['chucvu'])
                logging.info((res['position']) == data['trangcanhan_gioithieu_congviecvahocvan']['chucvu'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc")
                logging.info("check back-end: Ngày bắt đầu - 03-10-2021")
                logging.info((res['start_date']))
                logging.info((res['start_date']) == "2021-10-03")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc")
                logging.info("check back-end: Ngày kết thúc - 11-09-2022")
                logging.info((res['end_date']))
                logging.info((res['end_date']) == "2022-09-11T00:00:00.000+07:00")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc")
                logging.info("check back-end: Thành phố/Thị xã - " + data['trangcanhan_gioithieu_congviecvahocvan']['thanhpho'])
                logging.info((res['city']) == data['trangcanhan_gioithieu_congviecvahocvan']['thanhpho'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc")
                logging.info("check back-end: Mô tả - " + data['trangcanhan_gioithieu_congviecvahocvan']['mota'])
                logging.info((res['description']) == data['trangcanhan_gioithieu_congviecvahocvan']['mota'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc")
                logging.info("check back-end: Trạng thái - Công khai")
                logging.info((res['visibility']) == "public")

                break
            else:
                print("không có  response")
    def trangcanhan_gioithieu_cvvahocvan_daihoc(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        for request in driver.requests:
            if request.url[0:70] == "https://snapi.emso.asia/api/v1/accounts/111169896815147328/life_events":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học")
                logging.info("check back-end: Chuyên nghành - " + data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_chuyennghanh1'])
                logging.info((res['concentration1']) == data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_chuyennghanh1'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học")
                logging.info("check back-end: Tên đại học - " + data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_truonghoc'])
                logging.info((res['company']) == data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_truonghoc'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học")
                logging.info("check back-end: Bằng cấp - " + data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_bangcap'])
                logging.info((res['degree']) == data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_bangcap'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học")
                logging.info("check back-end: Ngày bắt đầu - 16-11-2019")
                logging.info((res['start_date']))
                logging.info((res['start_date']) == "2019-11-16")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học")
                logging.info("check back-end: Ngày kết thúc - 08-11-2023")
                logging.info((res['end_date']))
                logging.info((res['end_date']) == "2023-11-08T00:00:00.000+07:00")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học")
                logging.info("check back-end: Mô tả - " + data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_mota'])
                logging.info((res['description']) == data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_mota'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học")
                logging.info("check back-end: Loại trường = Đại học")
                logging.info((res['school_type']) == "UNIVERSITY")

                break
            else:
                print("không có  response")
    def trangcanhan_gioithieu_cvvahocvan_trunghoc(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        for request in driver.requests:
            if request.url[0:70] == "https://snapi.emso.asia/api/v1/accounts/111169896815147328/life_events":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học")
                logging.info("check back-end: Tên trung học - " + data['trangcanhan_gioithieu_congviecvahocvan']['trunghoc_truonghoc'])
                logging.info((res['company']) == data['trangcanhan_gioithieu_congviecvahocvan']['trunghoc_truonghoc'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học")
                logging.info("check back-end: Ngày bắt đầu - 13-10-2017")
                logging.info((res['start_date']))
                logging.info((res['start_date']) == "2017-10-13")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học")
                logging.info("check back-end: Ngày kết thúc - 24-10-2020")
                logging.info((res['end_date']))
                logging.info((res['end_date']) == "2020-10-24T00:00:00.000+07:00")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học")
                logging.info("check back-end: Mô tả - " + data['trangcanhan_gioithieu_congviecvahocvan']['trunghochoc_mota'])
                logging.info((res['description']) == data['trangcanhan_gioithieu_congviecvahocvan']['trunghochoc_mota'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học")
                logging.info("check back-end: Loại trường - Trung học")
                logging.info((res['school_type']) == "HIGH_SCHOOL")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học")
                logging.info("check back-end: Trạng thái - Công khai")
                logging.info((res['visibility']) == "public")
                break
            else:
                print("không có  response")
    def trangcanhan_gioithieu_cvvahocvan_xemsukientrongdoi_congviec(self):
        driver.implicitly_wait(10)
        time.sleep(2)
        for request in driver.requests:
            if request.url[0:39] == "https://snapi.emso.asia/api/v1/statuses":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc - Xem sự kiện trong đời")
                logging.info("check back-end: Công ty - " + data['trangcanhan_gioithieu_congviecvahocvan']['congty'])
                logging.info((res['life_event']['company']) == data['trangcanhan_gioithieu_congviecvahocvan']['congty'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc - Xem sự kiện trong đời")
                logging.info("check back-end: Công ty - Ngày bắt đầu - 03-10-2021")
                logging.info((res['life_event']['start_date']) == "2021-10-03")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc - Xem sự kiện trong đời")
                logging.info("check back-end: Công ty - Chức vụ - " + data['trangcanhan_gioithieu_congviecvahocvan']['chucvu'])
                logging.info((res['life_event']['position']) == data['trangcanhan_gioithieu_congviecvahocvan']['chucvu'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc - Xem sự kiện trong đời")
                logging.info("check back-end: Công ty - Thành phố/Thị xã - " + data['trangcanhan_gioithieu_congviecvahocvan']['thanhpho'])
                logging.info((res['life_event']['city']) == data['trangcanhan_gioithieu_congviecvahocvan']['thanhpho'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc - Xem sự kiện trong đời")
                logging.info("check back-end: Công ty - Mô tả - " + data['trangcanhan_gioithieu_congviecvahocvan']['mota'])
                logging.info((res['life_event']['description']) == data['trangcanhan_gioithieu_congviecvahocvan']['mota'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc - Xem sự kiện trong đời")
                logging.info("check back-end: Công ty - Trạng thái - Công khai")
                logging.info((res['life_event']['visibility']) == "public")

                break
            else:
                print("không có  response")
    def trangcanhan_gioithieu_cvvahocvan_xemsukientrongdoi_daihoc(self):
        driver.implicitly_wait(10)
        time.sleep(2)
        for request in driver.requests:
            if request.url[0:39] == "https://snapi.emso.asia/api/v1/statuses":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học - Xem sự kiện trong đời")
                logging.info("check back-end: Đại học - Tên đại học - " + data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_truonghoc'])
                logging.info((res['life_event']['company']) == data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_truonghoc'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học - Xem sự kiện trong đời")
                logging.info("check back-end: Đại học - Ngày bắt đầu - 16-11-2019")
                logging.info((res['life_event']['start_date']) == "2019-11-16")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học - Xem sự kiện trong đời")
                logging.info("check back-end: Đại học - loại trường - Đại học")
                logging.info((res['life_event']['school_type']) == "UNIVERSITY")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học - Xem sự kiện trong đời")
                logging.info("check back-end: Đại học - Chuyên nghành - " + data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_chuyennghanh1'])
                logging.info((res['life_event']['concentration1']) == data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_chuyennghanh1'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Đại học - Xem sự kiện trong đời")
                logging.info("check back-end: Đại học - Bằng cấp - " + data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_bangcap'])
                logging.info((res['life_event']['degree']) == data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_bangcap'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Công việc - Xem sự kiện trong đời")
                logging.info("check back-end: Đại học - Trạng thái - Công khai")
                logging.info((res['life_event']['visibility']) == "public")
                break
            else:
                print("không có  response")
    def trangcanhan_gioithieu_cvvahocvan_xemsukientrongdoi_trunghoc(self):
        driver.implicitly_wait(10)
        time.sleep(2)
        for request in driver.requests:
            if request.url[0:39] == "https://snapi.emso.asia/api/v1/statuses":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học - Xem sự kiện trong đời")
                logging.info("check back-end: Trung học - Tên Trung học - " + data['trangcanhan_gioithieu_congviecvahocvan']['trunghoc_truonghoc'])
                logging.info((res['life_event']['company']) == data['trangcanhan_gioithieu_congviecvahocvan']['trunghoc_truonghoc'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học - Xem sự kiện trong đời")
                logging.info("check back-end: Trung học - Ngày bắt đầu - 13-10-2017")
                logging.info((res['life_event']['start_date']) == "2017-10-13")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học - Xem sự kiện trong đời")
                logging.info("check back-end: Trung học - loại trường - Trung học")
                logging.info((res['life_event']['school_type']) == "HIGH_SCHOOL")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học - Xem sự kiện trong đời")
                logging.info("check back-end: Trung học - Mô tả - " + data['trangcanhan_gioithieu_congviecvahocvan']['trunghochoc_mota'])
                logging.info((res['life_event']['description']) == data['trangcanhan_gioithieu_congviecvahocvan']['trunghochoc_mota'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Công việc và học vấn - Trung học - Xem sự kiện trong đời")
                logging.info("check back-end: Trung học - Trạng thái - Công khai")
                logging.info((res['life_event']['visibility']) == "public")

                break
            else:
                print("không có  response")
    def trangcanhan_gioithieu_thongtincoban(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        for request in driver.requests:
            if request.url == "https://snapi.emso.asia/api/v1/account_general_infomation":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin liên hệ(Số điện thoại)")
                logging.info("check back-end: Số điện thoại - " + data['trangcanhan_gioithieu_thongtincoban']['sodienthoai'])
                logging.info((res['phone_number']) == data['trangcanhan_gioithieu_thongtincoban']['sodienthoai'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Thông tin cơ bản - Liên kết và xã hội")
                logging.info("check back-end: Web - " + data['trangcanhan_gioithieu_thongtincoban']['web'])
                logging.info((res['account_web_link'][0]['url']) == data['trangcanhan_gioithieu_thongtincoban']['web'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Thông tin cơ bản - Liên kết và xã hội")
                logging.info("check back-end: Liên kết - " + data['trangcanhan_gioithieu_thongtincoban']['lienket'])
                logging.info((res['account_social'][0]['text']) == data['trangcanhan_gioithieu_thongtincoban']['lienket'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
                logging.info("check back-end: Tiểu sử - " + data['trangcanhan_gioithieu_thongtincoban']['tieusu'])
                logging.info((res['description']) == data['trangcanhan_gioithieu_thongtincoban']['tieusu'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
                logging.info("check back-end: Giới tinh - female")
                logging.info((res['gender']) == "female")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
                logging.info("check back-end: Ngày sinh - 3 tháng 10, 2000")
                logging.info("Ngày 3")
                logging.info((res['birth_date']) == 3)

                logging.info("Tháng 10")
                logging.info((res['birth_month']) == 10)

                logging.info("Năm 2000")
                logging.info((res['birth_year']) == 2000)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
                logging.info("check back-end: Biệt danh - " + data['trangcanhan_gioithieu_thongtincoban']['bietdanh'])
                logging.info((res['other_name']) == data['trangcanhan_gioithieu_thongtincoban']['bietdanh'])

                break
            else:
                print("không có  response")
    def trangcanhan_gioithieu_giadinh_va_cacmqh(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        for request in driver.requests:
            if request.url == "https://snapi.emso.asia/api/v1/accounts/111169896815147328/abouts":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Gia đình và các mối quan hệ - Mối quan hệ")
                logging.info("check back-end: Mối quan hệ name - " + data['trangcanhan_gd_va_mqh']['name'])
                logging.info((res['account_relationship']['partner']['display_name']) == data['trangcanhan_gd_va_mqh']['name'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Gia đình và các mối quan hệ - Mối quan hệ")
                logging.info("check back-end: Tình trạng mqh/Ngày bắt đầu - Hẹn hò/2022-01-01")
                logging.info("Mối quan hệ - Hẹn hò")
                logging.info(res['account_relationship']['relationship_category']['name'])
                logging.info((res['account_relationship']['relationship_category']['name']) == "Hẹn hò")
                logging.info("Mối quan hệ - Ngày bắt đầu")
                logging.info((res['account_relationship']['start_date']) == "2022-01-01")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Gia đình và các mối quan hệ - Gia đình")
                logging.info("check back-end: Gia đình name - " + data['trangcanhan_gd_va_mqh']['name1'])
                logging.info((res['family_members'][0]['partner']['display_name']) == data['trangcanhan_gd_va_mqh']['name1'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Gia đình và các mối quan hệ - Gia đình")
                logging.info("check back-end: Mối quan hệ - Gia đình")
                logging.info((res['family_members'][0]['family_relationship_category']['name']) =="Gia đình")
                break
            else:
                print("không có  response")
    def trangcanhan_gioithieu_sukientrongdoi(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        for request in driver.requests:
            if request.url == "https://snapi.emso.asia/api/v1/accounts/111169896815147328/statuses?exclude_replies=true&limit=3":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Sự kiện trong đời - Thêm mới sự kiện lên dòng thời gian")
                logging.info("check back-end: Mô tả - " + data['trangcanhan_sukientrongdoi']['dulich_mota'])
                logging.info((res[0]['content']) == data['trangcanhan_sukientrongdoi']['dulich_mota'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Sự kiện trong đời - Thêm mới sự kiện lên dòng thời gian")
                logging.info("check back-end: Tiêu đề - " + data['trangcanhan_sukientrongdoi']['dulich_tieude'])
                logging.info((res[0]['life_event']['name']) == data['trangcanhan_sukientrongdoi']['dulich_tieude'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Sự kiện trong đời - Thêm mới sự kiện lên dòng thời gian")
                logging.info("check back-end: Địa điểm - " + data['trangcanhan_sukientrongdoi']['dulich_diachi'])
                logging.info((res[0]['life_event']['place']['title']) == data['trangcanhan_sukientrongdoi']['dulich_diachi'])

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Sự kiện trong đời - Thêm mới sự kiện lên dòng thời gian")
                logging.info("check back-end: Thời gian - 2023-09-01")
                logging.info((res[0]['life_event']['start_date']) == "2023-09-01")

                logging.info("check back-end: Trang cá nhân - Giới thiệu - Sự kiện trong đời - Thêm mới sự kiện lên dòng thời gian")
                logging.info("check back-end:Có ảnh sự kiện trên dòng thời gian không? ")
                logging.info(res[0]['media_attachments'][0]['url'])
                logging.info((res[0]['media_attachments'][0]['url']) != None)
                logging.info(res[0]['media_attachments'][0]['preview_url'])
                logging.info((res[0]['media_attachments'][0]['preview_url']) != None)

                break
            else:
                print("không có  response")



class tongquan():
    def tongquan(self):
        driver.implicitly_wait(15)
        # driver.find_element(By.XPATH, var.trangcanhan).click()
        time.sleep(1.5)
        # Trang cá nhân - Giới thiệu tổng quan
        #Sống tại
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_chinhsua).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_input).send_keys(data['trangcanhan_gioithieu_tongquan']['songtai'])

        wait = WebDriverWait(driver, 10)
        chon_hanoi = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_songtai_hanoi)))

        try:
            chon_hanoi.click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.execute_script("window.scrollBy(0,700)", "")
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_chinhsua).click()
            time.sleep(0.5)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_input).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_x).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_input).send_keys(data['trangcanhan_gioithieu_tongquan']['songtai'])
            chon_hanoi.click()

        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_luu).click()
        time.sleep(1.5)
        writeData(var.path_baocao, "Sheet1", 62, 2, "x")
        check_trangcanhan_gioithieu_songtai = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai1).text
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Sống tại")
        logging.info("check font-end: Sống tại - " + data['trangcanhan_gioithieu_tongquan']['songtai'])
        logging.info(check_trangcanhan_gioithieu_songtai == data['trangcanhan_gioithieu_tongquan']['songtai'])

        # Đến từ
        driver.execute_script("window.scrollBy(0,500)", "")
        # time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_chinhsua).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_input).send_keys(data['trangcanhan_gioithieu_tongquan']['dentu'])
        wait = WebDriverWait(driver, 10)
        dentu_langson = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_dentu_langson)))

        try:
            dentu_langson.click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.execute_script("window.scrollBy(0,700)", "")
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_chinhsua).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_input).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_x).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_input).send_keys(data['trangcanhan_gioithieu_tongquan']['dentu'])
            dentu_langson.click()

        # time.sleep(1)
        # driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_langson).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_luu).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 63, 2, "x")
        check_gioithieu_dentu = driver.find_element(By.XPATH, var.trangcanhan_check_gioithieu_dentu1).text
        print(check_gioithieu_dentu)
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Đến từ")
        logging.info("check font-end: Đến từ - " + data['trangcanhan_gioithieu_tongquan']['dentu'])
        logging.info(check_gioithieu_dentu == data['trangcanhan_gioithieu_tongquan']['dentu'])

        # Mối quan hệ
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_chinhsua).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_icon_chonmqh).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_lythan).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_nam).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_nam_2022).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_thang).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_thang_1).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_ngay).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_ngay_1).click()

        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_luu).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 64, 2, "x")
        check_gioithieu_mqh = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh1).text
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Mối quan hệ")
        logging.info("check font-end: Ly Thân(Đang chờ) started 2022-01-01")
        logging.info(check_gioithieu_mqh == "Ly Thân(Đang chờ) started 2022-01-01")



        # Số điện thoại
        try:
            check_gioithieu_mqh = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh1).text
            logging.info("Trang cá nhân - Giới thiệu tổng quan - Mối quan hệ")
            logging.info("check font-end: Ly Thân(Đang chờ) started 2022-01-01")
            logging.info(check_gioithieu_mqh == "Ly Thân(Đang chờ) started 2022-01-01")
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai).click()   #mqh_loi
        except:
            driver.find_element(By.XPATH, var.trangcanhan_gioithieuhuy).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_chinhsua).click()
        time.sleep(1)
        sdt1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_input)
        sdt1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_input).send_keys(data['trangcanhan_gioithieu_tongquan']['sodienthoai'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_luu).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 65, 2, "x")
        check_gioithieu_sodienthoai = driver.find_element(By.XPATH, var.trangcanhan_check_gioithieu_sodienthoai1).text
        print(check_gioithieu_sodienthoai)
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Số điện thoại")
        logging.info("check font-end: Số điện thoại - " + data['trangcanhan_gioithieu_tongquan']['sodienthoai'])
        logging.info(check_gioithieu_sodienthoai == data['trangcanhan_gioithieu_tongquan']['sodienthoai'])

        # Biệt danh
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_chinhsua).click()
        time.sleep(1)
        bietdanh1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_input)
        bietdanh1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_input).send_keys(data['trangcanhan_gioithieu_tongquan']['bietdanh'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_luu).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 66, 2, "x")
        check_gioithieu_bietdanh = driver.find_element(By.XPATH, var.trangcanhan_check_gioithieu_bietdanh1).text
        print(check_gioithieu_bietdanh)
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Biệt danh")
        logging.info("check font-end: Biệt danh - " + data['trangcanhan_gioithieu_tongquan']['bietdanh'])
        logging.info(check_gioithieu_bietdanh == data['trangcanhan_gioithieu_tongquan']['bietdanh'])
        time.sleep(1)
        del driver.requests
        driver.refresh()
        time.sleep(1)

    def tongquan_dulieusai(self):
        driver.implicitly_wait(15)
        # driver.find_element(By.XPATH, var.trangcanhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        # Trang cá nhân - Giới thiệu tổng quan
        # Sống tại
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_chinhsua).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_input).send_keys(data['trangcanhan_gioithieu_tongquan']['songtai_dulieusai'])
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_luu).click()
        time.sleep(1)

        check_gioithieu_songtai_dulieusai = driver.find_element(By.XPATH, var.trangcanhan_check_gioithieu_songtai_dulieusai1).text
        print(check_gioithieu_songtai_dulieusai)
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Sống tại - nhập địa điểm không tồn tại")
        logging.info("check font-end: Chúng tôi không thể nhận dạng vị tri bạn đã chỉ định.")
        logging.info(check_gioithieu_songtai_dulieusai == "Chúng tôi không thể nhận dạng vị trí bạn đã chỉ định.")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_ok).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_huy).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 67, 2, "x")
        writeData(var.path_baocao, "Sheet1", 67, 3, "đã pass: Thành phố hiện tại")
        # Đến từ
        driver.execute_script("window.scrollBy(0,500)", "")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_chinhsua).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_input).send_keys(data['trangcanhan_gioithieu_tongquan']['dentu_dulieusai'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_luu).click()
        time.sleep(2)
        check_gioithieu_dentu_dulieusai = driver.find_element(By.XPATH, var.trangcanhan_check_gioithieu_dentu_dulieusai1).text
        print(check_gioithieu_dentu_dulieusai)
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Đến từ - nhập địa điểm không tồn tại")
        logging.info("check font-end: Chúng tôi không thể nhận dạng vị tri bạn đã chỉ định.")
        logging.info(check_gioithieu_dentu_dulieusai == "Chúng tôi không thể nhận dạng vị trí bạn đã chỉ định.")
        time.sleep(1)

        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_ok).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_huy).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 67, 2, "x")
        writeData(var.path_baocao, "Sheet1", 67, 3, "đã pass: Thành phố hiện tại, quê quán")

        #mối quan hệ
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_chinhsua).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_icon_chonmqh).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_goa).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh_luu).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 68, 2, "x")
        driver.execute_script("window.scrollBy(0,500)", "")
        # check_gioithieu_mqh1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh4).text         #mqh_loi
        check_gioithieu_mqh1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh2).text
        print(check_gioithieu_mqh1)
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Mối quan hệ")
        logging.info("check font-end: Góa(Đang chờ) started 2022-01-01")
        logging.info(check_gioithieu_mqh1)
        logging.info(check_gioithieu_mqh1 == "Góa(Đang chờ) started 2022-01-01")
        time.sleep(1)

        # try:
        #     check_gioithieu_mqh1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_mqh2).text
        #     print(check_gioithieu_mqh1)
        #     logging.info("Trang cá nhân - Giới thiệu tổng quan - Mối quan hệ")
        #     logging.info("check font-end: Góa started 2022-01-01")
        #     logging.info(check_gioithieu_mqh1 == "Góa started 2022-01-01")
        #     time.sleep(1)
        #     driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai).click()
        # except:
        #     driver.find_element(By.XPATH, var.trangcanhan_gioithieuhuy).click()
        # # driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai).click()
        #     driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_chinhsua2).click()
        # driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_chinhsua).click()


        # Số điện thoại1
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_chinhsua).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_input).click()
        time.sleep(1)
        sdt2 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_input)
        sdt2.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_input).send_keys(data['trangcanhan_gioithieu_tongquan']['sodienthoai_dai'])
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_luu).click()
        time.sleep(1)
        check_gioithieu_sodienthoai_dai = driver.find_element(By.XPATH, var.trangcanhan_check_gioithieu_sodienthoai_dai1).text
        print(check_gioithieu_sodienthoai_dai)
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Số điện thoại dài quá 10 ký tự")
        logging.info("check font-end: Số điện thoại độ dài không đúng (phải là 10 ký tự)")
        logging.info(check_gioithieu_sodienthoai_dai == "Số điện thoại độ dài không đúng (phải là 10 ký tự)")
        time.sleep(1)

        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_ok).click()
        writeData(var.path_baocao, "Sheet1", 70, 2, "x")
        writeData(var.path_baocao, "Sheet1", 70, 3, "đã pass: SĐT quá dài")
        sdt2.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_input).send_keys(data['trangcanhan_gioithieu_tongquan']['sodienthoai_ngan'])
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_luu).click()
        time.sleep(1)
        check_gioithieu_sodienthoai_ngan = driver.find_element(By.XPATH, var.trangcanhan_check_gioithieu_sodienthoai_ngan1).text
        print(check_gioithieu_sodienthoai_ngan)
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Số điện thoại it hơn 10 ký tự")
        logging.info("check font-end: Số điện thoại độ dài không đúng (phải là 10 ký tự)")
        logging.info(check_gioithieu_sodienthoai_ngan == "Số điện thoại độ dài không đúng (phải là 10 ký tự)")
        time.sleep(1)

        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_ok).click()
        writeData(var.path_baocao, "Sheet1", 70, 2, "x")
        writeData(var.path_baocao, "Sheet1", 70, 3, "đã pass: SĐT quá dài, SĐT quá ngắn")
        sdt2.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_input).send_keys(data['trangcanhan_gioithieu_tongquan']['sodienthoai_trung'])
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_luu).click()
        time.sleep(1)
        check_gioithieu_sodienthoai_trungso = driver.find_element(By.XPATH, var.trangcanhan_check_gioithieu_sodienthoai_trungso1).text
        print(check_gioithieu_sodienthoai_trungso)
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Số điện thoại bị trùng")
        logging.info("check font-end: Số điện thoại đã có")
        logging.info(check_gioithieu_sodienthoai_trungso == "Số điện thoại đã có")
        time.sleep(1)

        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_ok).click()
        writeData(var.path_baocao, "Sheet1", 70, 2, "x")
        writeData(var.path_baocao, "Sheet1", 70, 3, "đã pass: SĐT quá dài, SĐT quá ngắn, trùng SĐT")
        sdt2.send_keys(Keys.CONTROL, "a")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_input).send_keys(data['trangcanhan_gioithieu_tongquan']['sodienthoai2'])
        writeData(var.path_baocao, "Sheet1", 69, 2, "x")
        writeData(var.path_baocao, "Sheet1", 69, 3, "đã pass: số điện thoại")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sodienthoai_luu).click()
        time.sleep(1)

        # Biệt danh1
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_chinhsua).click()
        bietdanh2 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_input)
        bietdanh2.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_input).send_keys(data['trangcanhan_gioithieu_tongquan']['bietdanh_dai'])
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_luu).click()
        time.sleep(1)
        check_gioithieu_bietdanh_quadai = driver.find_element(By.XPATH, var.trangcanhan_check_gioithieu_bietdanh_quadai1).text
        print(check_gioithieu_bietdanh_quadai)
        logging.info("Trang cá nhân - Giới thiệu tổng quan - Biệt danh quá dài, quá 20 ký tự")
        logging.info("check font-end: Tên biệt danh không vượt quá 20 ký tự.Vui lòng điền lại thông tin!")
        logging.info(check_gioithieu_bietdanh_quadai == "Tên biệt danh không vượt quá 20 ký tự.Vui lòng điền lại thông tin!")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_ok).click()
        writeData(var.path_baocao, "Sheet1", 70, 2, "x")
        writeData(var.path_baocao, "Sheet1", 70, 3, "đã pass: SĐT quá dài, SĐT quá ngắn, trùng SĐT, biệt danh quá dài")
        time.sleep(1)
        bietdanh2.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_input).send_keys(data['trangcanhan_gioithieu_tongquan']['bietdanh2'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_bietdanh_luu).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 69, 2, "x")
        writeData(var.path_baocao, "Sheet1", 69, 3, "đã pass: số điện thoại, biệt danh")



class cong_viec_va_hoc_van():
    def congviecvahocvan_congviec(self):
        driver.implicitly_wait(15)
        # driver.find_element(By.XPATH, var.trangcanhan).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_vahocvan).click()
        #Công việc và học vấn
        #Công việc-công ty
        driver.execute_script("window.scrollBy(0,700)", "")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_iconcongviec).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_chinhsuacongviec).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_congty_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_congty_input_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_congty_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['congty'])
        wait = WebDriverWait(driver, 10)
        gioithieu_congty_quantrada = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_congty_input_quantrada)))

        try:
            gioithieu_congty_quantrada.click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.execute_script("window.scrollBy(0,700)", "")
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_iconcongviec).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_chinhsuacongviec).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_congty_input).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_congty_input_x).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_congty_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['congty'])
            gioithieu_congty_quantrada.click()

        #Công việc-chức vụ
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_chucvu_input).click()
        chucvu1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_chucvu_input)
        chucvu1.send_keys(Keys.CONTROL, "a")
        # time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_chucvu_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['chucvu'])
        #Công việc-thành phố, thị xã
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_tp_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_tp_input_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_tp_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['thanhpho'])

        wait = WebDriverWait(driver, 10)
        gioithieu_congty_tp_hanoi = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_cv_tp_hanoi)))
        gioithieu_congty_tp_hanoi.click()
        time.sleep(0.5)
        #Công việc-mô tả
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_mota_input).click()
        mota1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_mota_input)
        mota1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_mota_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['mota'])
        time.sleep(2)

        # Công việc-ngày bắt đầu, ngày kết thúc
        #Ngày bắt đầu
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchonnam_2021).click()
        time.sleep(1)
        driver.implicitly_wait(0.1)
        count = 0
        while (count < 13):
            count = count + 1
            try:
                driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchuyenthang).click()
                if driver.find_element(By.XPATH, var.check_lich_thang).text == "October 2021":
                    driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_chonngay_3).click()
                    time.sleep(1)
                    break
            except:
                pass

        #Ngày kết thúc
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngayketthuc_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngayketthuc_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngayketthuc_iconchonnam_2022).click()
        time.sleep(1)
        count = 0
        while (count < 13):
            count = count + 1
            try:
                driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchuyenthang).click()
                if driver.find_element(By.XPATH, var.check_lich_thang).text == "September 2022":
                    driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngayketthuc_chonngay_11).click()
                    time.sleep(1)
                    break
            except:
                pass
        del driver.requests
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_luu).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 71, 2, "x")


    def congviecvahocvan_daihoc(self):
        driver.implicitly_wait(15)
        # Đại học - trường học
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_icondaihoc).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_chinhsuadaihoc).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_truonghoc'])
        time.sleep(1.5)
        wait = WebDriverWait(driver, 10)
        gioithieu_daihoc_truonghoc_dhbachkhoa = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_dhbachkhoa)))

        try:
            gioithieu_daihoc_truonghoc_dhbachkhoa.click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.execute_script("window.scrollBy(0,700)", "")
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_icondaihoc).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_chinhsuadaihoc).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input_x).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_truonghoc'])
            gioithieu_daihoc_truonghoc_dhbachkhoa.click()
        # driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_datotnghiep).click()

        # #Đại học - khoang thời gian
        #Ngày bắt đầu
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngaybatdau_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngaybatdau_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngaybatdau_iconchonnam_2019).click()
        time.sleep(1)
        driver.implicitly_wait(0.1)
        count = 0
        while (count < 13):
            count = count + 1
            try:
                driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchuyenthang).click()
                if driver.find_element(By.XPATH, var.check_lich_thang).text == "November 2019":
                    driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngaybatdau_chonngay_16).click()
                    time.sleep(1)
                    break
            except:
                pass

        #Ngày kết thúc
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngayketthuc_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngayketthuc_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngayketthuc_iconchonnam_2023).click()
        time.sleep(1)
        count = 0
        while (count < 13):
            count = count + 1
            try:
                driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchuyenthang).click()
                if driver.find_element(By.XPATH, var.check_lich_thang).text == "November 2023":
                    driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngayketthuc_chonngay_8).click()
                    time.sleep(1)
                    break
            except:
                pass

        #Đại học - mô tả
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_mota_input).click()
        mota2 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_mota_input)
        mota2.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_mota_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_mota'])
        #Đại học - chuyên nghành 1
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_chuyennghanh1_input).click()
        chuyennghanh1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_chuyennghanh1_input)
        chuyennghanh1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_chuyennghanh1_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_chuyennghanh1'])
        time.sleep(0.5)
        # Đại học - đã học tại
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_dahoctai_caodang).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_dahoctai_daihoc).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_bangcap_input).click()
        bangcap1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_bangcap_input)
        bangcap1.send_keys(Keys.CONTROL, "a")
        del driver.requests
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_bangcap_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['daihoc_bangcap'])
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_luu).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 72, 2, "x")

    def congviecvahocvan_trunghoc(self):
        driver.implicitly_wait(15)
        # Trung học - trường học
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_icontrunghoc).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_chinhsuatrunghoc).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghochoc_truonghoc_input_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['trunghoc_truonghoc'])

        wait = WebDriverWait(driver, 10)
        gioithieu_trunghoc_truonghoc_thptdinhlap = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_thpt_dinhlap)))

        try:
            gioithieu_trunghoc_truonghoc_thptdinhlap.click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.execute_script("window.scrollBy(0,700)", "")
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_icontrunghoc).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_chinhsuatrunghoc).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_input).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghochoc_truonghoc_input_x).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['trunghoc_truonghoc'])
            gioithieu_trunghoc_truonghoc_thptdinhlap.click()
        # Trung học - mô tả
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghochoc_mota_input).click()
        mota_trunghoc = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghochoc_mota_input)
        mota_trunghoc.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghochoc_mota_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan']['trunghochoc_mota'])
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_datotnghiep).click()

        # Trung học - khoảng thời gian
        #Ngày bắt đầu
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngaybatdau_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngaybatdau_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngaybatdau_iconchonnam_2017).click()
        time.sleep(1)
        driver.implicitly_wait(0.1)
        count = 0
        while (count < 13):
            count = count + 1
            try:
                driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchuyenthang).click()
                if driver.find_element(By.XPATH, var.check_lich_thang).text == "October 2017":
                    driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngaybatdau_chonngay_13).click()
                    time.sleep(1)
                    break
            except:
                pass

        #Ngày kết thúc
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngayketthuc_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngayketthuc_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngayketthuc_iconchonnam_2020).click()
        time.sleep(1)
        count = 0
        while (count < 13):
            count = count + 1
            try:
                driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchuyenthang).click()
                if driver.find_element(By.XPATH, var.check_lich_thang).text == "October 2020":
                    driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngayketthuc_chonngay_24).click()
                    time.sleep(1)
                    break
            except:
                pass
        del driver.requests
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_luu).click()
        time.sleep(1)
        driver.refresh()
        time.sleep(2)
        writeData(var.path_baocao, "Sheet1", 73, 2, "x")

    def congviecvahocvan_addthem(self):
        driver.implicitly_wait(15)
        # driver.find_element(By.XPATH, var.trangcanhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_vahocvan).click()
        driver.execute_script("window.scrollBy(0,700)", "")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_addthem_congviec).click()
        #Công việc và học vấn
        #Công việc-công ty
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_congty_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_congty_input_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_congty_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['congty'])
        try:
            wait = WebDriverWait(driver, 10)
            gioithieu_congty_input_quancf = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_congty_input_quancf)))
            gioithieu_congty_input_quancf.click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_vahocvan).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_addthem_congviec).click()
            wait = WebDriverWait(driver, 10)
            gioithieu_congty_input_quancf = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_congty_input_quancf)))
            gioithieu_congty_input_quancf.click()

        #Công việc-chức vụ
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_chucvu_input).click()
        chucvu1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_chucvu_input)
        chucvu1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_chucvu_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['chucvu'])
        #Công việc-thành phố, thị xã
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_tp_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_tp_input_x).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_tp_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['thanhpho'])

        wait = WebDriverWait(driver, 10)
        gioithieu_cv_tp_langson = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_cv_tp_langson)))
        gioithieu_cv_tp_langson.click()

        #Công việc-mô tả
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_mota_input).click()
        mota1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_mota_input)
        mota1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_mota_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['mota'])
        time.sleep(0.5)

        # Công việc-ngày bắt đầu, ngày kết thúc
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_toidanglamviecoday).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_iconchonnam_2021).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngaybatdau_chonngay_5).click()

        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngayketthuc_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngayketthuc_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngayketthuc_iconchonnam_2022).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_ngayketthuc_chonngay_16).click()

        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_luu).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cvvahocvan_iconxoa).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cvvahocvan_xoa).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cvvahocvan_xoa1).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 74, 2, "x")

        # Đại học - trường học
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_addtruongdaihoc).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['daihoc_truonghoc1'])
        try:
            wait = WebDriverWait(driver, 10)
            gioithieu_daihoc_truonghoc_dhkingkong = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_dhkingkong)))
            gioithieu_daihoc_truonghoc_dhkingkong.click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_addtruongdaihoc).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input_x).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['daihoc_truonghoc1'])
            wait = WebDriverWait(driver, 10)
            gioithieu_daihoc_truonghoc_dhkingkong = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_dhkingkong)))
            gioithieu_daihoc_truonghoc_dhkingkong.click()


        #
        # time.sleep(2.5)
        # driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_dhkingkong).click()
        # time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_truonghoc_datotnghiep).click()

        # Đại học - khoang thời gian
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngaybatdau_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngaybatdau_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngaybatdau_iconchonnam_2019).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngaybatdau_chonngay_9).click()

        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngayketthuc_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngayketthuc_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngayketthuc_iconchonnam_2023).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dh_ngayketthuc_chonngay_18).click()
        time.sleep(1)



        time.sleep(0.5)
        # Đại học - mô tả
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_mota_input).click()
        mota2 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_mota_input)
        mota2.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_mota_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['daihoc_mota1'])
        # Đại học - chuyên nghành 1
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_chuyennghanh1_input).click()
        chuyennghanh1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_chuyennghanh1_input)
        chuyennghanh1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_chuyennghanh1_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['daihoc_chuyennghanh11'])
        time.sleep(0.5)
        # Đại học - đã học tại
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_dahoctai_caodang).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_dahoctai_daihoc).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_bangcap_input).click()
        bangcap1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_bangcap_input)
        bangcap1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_bangcap_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['daihoc_bangcap1'])
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_luu).click()
        time.sleep(2)
        writeData(var.path_baocao, "Sheet1", 75, 2, "x")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_iconxoa).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_xoa).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_xoa1).click()
        time.sleep(1)

    # Trung học - trường học
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_addthemtrunghoc).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghochoc_truonghoc_input_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['trunghoc_truonghoc1'])
        try:
            wait = WebDriverWait(driver, 10)
            gioithieu_trunghoc_truonghoc_thpt_phubinh = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_thpt_phubinh)))
            gioithieu_trunghoc_truonghoc_thpt_phubinh.click()
        except:
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_addthemtrunghoc).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_input).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghochoc_truonghoc_input_x).click()
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['trunghoc_truonghoc1'])
            wait = WebDriverWait(driver, 10)
            gioithieu_trunghoc_truonghoc_thpt_phubinh = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_thpt_phubinh)))
            gioithieu_trunghoc_truonghoc_thpt_phubinh.click()
        #
        # time.sleep(2.5)
        # driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_truonghoc_thpt_phubinh).click()
        # time.sleep(1.5)
        # Trung học - mô tả
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghochoc_mota_input).click()
        mota_trunghoc = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghochoc_mota_input)
        mota_trunghoc.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghochoc_mota_input).send_keys(data['trangcanhan_gioithieu_congviecvahocvan_addthem']['trunghochoc_mota1'])
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_datotnghiep).click()

        # Trung học - khoảng thời gian
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngaybatdau_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngaybatdau_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngaybatdau_iconchonnam_2017).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngaybatdau_chonngay_16).click()

        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngayketthuc_iconchon).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngayketthuc_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngayketthuc_iconchonnam_2020).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_ngayketthuc_chonngay_22).click()
        time.sleep(1)

        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_luu).click()
        time.sleep(2)
        writeData(var.path_baocao, "Sheet1", 76, 2, "x")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_iconxoa).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_xoa).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_xoa1).click()
        time.sleep(1)
        driver.refresh()
        time.sleep(2)




class xemkientrongdoi():
    def congviecvahocvan_congviec(self):
        driver.implicitly_wait(15)
        # driver.find_element(By.XPATH, var.trangcanhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_cv_vahocvan).click()
        #Công việc
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_iconcongviec).click()
        del driver.requests
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_xemsukientrongdoi).click()
        time.sleep(2)
        check_congviec_congty = driver.find_element(By.XPATH, var.trangcanhan_xemsukientrongdoi_congviec_congty).text
        print(check_congviec_congty)
        logging.info("Trang cá nhân - giới thiệu - công việc và học vấn - công việc - công ty - xem sự kiện trong đời ")
        logging.info("check font-end: Bắt đầu công việc mới tại Quán trà đá")
        logging.info(check_congviec_congty == "Bắt đầu công việc mới tại Quán trà đá")

        check_congviec_ngaythangchucvu = driver.find_element(By.XPATH, var.trangcanhan_xemsukientrongdoi_congviec_ngaythangchucvu).text
        print(check_congviec_ngaythangchucvu)
        logging.info("Trang cá nhân - giới thiệu - công việc và học vấn - công việc - xem sự kiện trong đời")
        logging.info("check font-end: Ngày bắt đầu làm việc, chức vụ - 03 tháng 10, 2021 - Nhân Viên")
        logging.info(check_congviec_ngaythangchucvu == "03 tháng 10, 2021 - Nhân Viên")

        check_congviec_thanhpho = driver.find_element(By.XPATH, var.trangcanhan_xemsukientrongdoi_congviec_thanhpho).text
        print(check_congviec_thanhpho)
        logging.info("Trang cá nhân - giới thiệu - công việc và học vấn - công việc - xem sự kiện trong đời")
        logging.info("check font-end: Thành phố làm việc - " + data['trangcanhan_gioithieu_congviecvahocvan']['thanhpho'])
        logging.info(check_congviec_thanhpho == data['trangcanhan_gioithieu_congviecvahocvan']['thanhpho'])
        writeData(var.path_baocao, "Sheet1", 77, 2, "x")

    def congviecvahocvan_daihoc(self):
        driver.implicitly_wait(15)
        driver.back()
        time.sleep(2)
        #Đại học
        driver.execute_script("window.scrollBy(0,700)", "")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_icondaihoc).click()
        del driver.requests
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_daihoc_xemsukientrongdoi).click()
        time.sleep(2)
        check_daihoc_tentruong = driver.find_element(By.XPATH, var.trangcanhan_xemsukientrongdoi_daihoc_tentruong).text
        print(check_daihoc_tentruong)
        logging.info("Trang cá nhân - giới thiệu - công việc và học vấn - Đại học - tên trường - xem sự kiện trong đời ")
        logging.info("check font-end: Bắt đầu học tại Đại học bách khoa hà nội")
        logging.info(check_daihoc_tentruong == "Bắt đầu học tại Đại học bách khoa hà nội")

        check_daihoc_ngaythang = driver.find_element(By.XPATH,var.trangcanhan_xemsukientrongdoi_check_daihoc_ngaythang).text
        print(check_daihoc_ngaythang)
        logging.info("Trang cá nhân - giới thiệu - công việc và học vấn - đại học - thời gian bắt đầu - xem sự kiện trong đời")
        logging.info("check font-end: Ngày nhập học: 16 tháng 11, 2019")
        logging.info(check_daihoc_ngaythang == "16 tháng 11, 2019")

        check_daihoc_dahoctai_chuyennghanh = driver.find_element(By.XPATH, var.trangcanhan_xemsukientrongdoi_daihoc_dahoctai_chuyennghanh).text
        print(check_daihoc_dahoctai_chuyennghanh)
        logging.info("Trang cá nhân - giới thiệu - công việc và học vấn - đại học- đã học tại, chuyên nghành - xem sự kiện trong đời")
        logging.info("check font-end: đã học tại, chuyên nghành - Đại học - Kỹ thuật máy tinh")
        logging.info(check_daihoc_dahoctai_chuyennghanh)
        logging.info(check_daihoc_dahoctai_chuyennghanh == "Đại học - Kỹ thuật máy tính")
        writeData(var.path_baocao, "Sheet1", 78, 2, "x")

    def congviecvahocvan_trunghoc(self):
        driver.implicitly_wait(15)
        driver.back()
        time.sleep(2)
        #Trung học
        driver.execute_script("window.scrollBy(0,700)", "")
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_icontrunghoc).click()
        del driver.requests
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_trunghoc_xemsukientrongdoi).click()
        time.sleep(2)
        check_trunghoc_tentruong = driver.find_element(By.XPATH, var.trangcanhan_xemsukientrongdoi_trunghoc_tentruong).text
        print(check_trunghoc_tentruong)
        logging.info("Trang cá nhân - giới thiệu - công việc và học vấn - Trung học - tên trường - xem sự kiện trong đời ")
        logging.info("check font-end: Bắt đầu học tại Trường THPT Đình Lập")
        logging.info(check_trunghoc_tentruong == "Bắt đầu học tại Trường THPT Đình Lập")

        check_trunghoc_ngaythang = driver.find_element(By.XPATH,var.trangcanhan_xemsukientrongdoi_check_trunghoc_ngaythang).text
        print(check_trunghoc_ngaythang)
        logging.info("Trang cá nhân - giới thiệu - công việc và học vấn - trung học - thời gian bắt đầu - xem sự kiện trong đời")
        logging.info("check font-end: Ngày nhập học: 13 tháng 10, 2017")
        logging.info(check_trunghoc_ngaythang == "13 tháng 10, 2017")

        check_trunghoc_captruong = driver.find_element(By.XPATH, var.trangcanhan_xemsukientrongdoi_trunghoc_captruong).text
        print(check_trunghoc_captruong)
        logging.info("Trang cá nhân - giới thiệu - công việc và học vấn - trung học- cấp trường - xem sự kiện trong đời")
        logging.info("check font-end: Trường trung học")
        logging.info(check_trunghoc_captruong == "Trường trung học")
        writeData(var.path_baocao, "Sheet1", 79, 2, "x")
        driver.back()
        time.sleep(2)


class thongtincoban():
    def thongtincoban(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,500)", "")
        driver.find_element(By.XPATH, var.thongtincoban).click()

        #thông tin liên hệ
        driver.find_element(By.XPATH, var.thongtincoban_icon_sdt).click()
        driver.find_element(By.XPATH, var.thongtincoban_sdt_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_sdt_input).click()
        thongtincoban_sdt1 = driver.find_element(By.XPATH, var.thongtincoban_sdt_input)
        thongtincoban_sdt1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_sdt_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['sodienthoai'])
        driver.find_element(By.XPATH, var.thongtincoban_sdt_luu).click()
        time.sleep(1)
        thongtincoban_sodienthoai = driver.find_element(By.XPATH, var.thongtincoban_sdt_fe).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin liên hệ(số điện thoại)")
        logging.info("check font-end: Số điện thoại - " + data['trangcanhan_gioithieu_thongtincoban']['sodienthoai'])
        logging.info(thongtincoban_sodienthoai == data['trangcanhan_gioithieu_thongtincoban']['sodienthoai'])

        # Các trang web và liên kết xã hội
        #web
        driver.execute_script("window.scrollBy(0,700)", "")
        time.sleep(1)
        driver.find_element(By.XPATH, var.thongtincoban_icon_web).click()
        driver.find_element(By.XPATH, var.thongtincoban_web_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_web_input).click()
        thongtincoban_web1 = driver.find_element(By.XPATH, var.thongtincoban_web_input)
        thongtincoban_web1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_web_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['web'])
        driver.find_element(By.XPATH, var.thongtincoban_web_luu).click()
        time.sleep(1)
        thongtincoban_web = driver.find_element(By.XPATH, var.thongtincoban_web_fe).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Các trang web và liên kết xã hội")
        logging.info("check font-end: Web - " + data['trangcanhan_gioithieu_thongtincoban']['web'])
        logging.info(thongtincoban_web == data['trangcanhan_gioithieu_thongtincoban']['web'])

        #liên kết xã hội
        driver.find_element(By.XPATH, var.thongtincoban_icon_lienket).click()
        driver.find_element(By.XPATH, var.thongtincoban_lienket_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_lienket_input).click()
        thongtincoban_lienket1 = driver.find_element(By.XPATH, var.thongtincoban_lienket_input)
        thongtincoban_lienket1.send_keys(Keys.CONTROL, "a")
        del driver.requests
        driver.find_element(By.XPATH, var.thongtincoban_lienket_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['lienket'])
        driver.find_element(By.XPATH, var.thongtincoban_lienket_luu).click()
        time.sleep(1)
        thongtincoban_lienket = driver.find_element(By.XPATH, var.thongtincoban_lienket_fe).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Các trang web và liên kết xã hội")
        print(thongtincoban_lienket)
        logging.info(thongtincoban_lienket)
        logging.info("check font-end: Liên kết - " + data['trangcanhan_gioithieu_thongtincoban']['lienket'])
        logging.info(thongtincoban_lienket == data['trangcanhan_gioithieu_thongtincoban']['lienket'])

        #thông tin cơ bản
        #tiểu sử
        driver.find_element(By.XPATH, var.thongtincoban_icon_tieusu).click()
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_input).click()
        thongtincoban_tieusu1 = driver.find_element(By.XPATH, var.thongtincoban_tieusu_input)
        thongtincoban_tieusu1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['tieusu'])
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_luu).click()
        time.sleep(1)
        thongtincoban_tieusu = driver.find_element(By.XPATH, var.thongtincoban_tieusu_fe).text
        print(thongtincoban_tieusu)
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
        logging.info("check font-end: Tiểu sử - " + data['trangcanhan_gioithieu_thongtincoban']['tieusu'])
        logging.info(thongtincoban_tieusu == data['trangcanhan_gioithieu_thongtincoban']['tieusu'])

        #Giới tinh
        driver.find_element(By.XPATH, var.thongtincoban_icon_gioitinh).click()
        driver.find_element(By.XPATH, var.thongtincoban_gioitinh_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_gioitinh_iconchongioitinh).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.thongtincoban_gioitinh_nu).click()
        driver.find_element(By.XPATH, var.thongtincoban_gioitinh_luu).click()
        time.sleep(1)
        thongtincoban_gioitinh = driver.find_element(By.XPATH, var.thongtincoban_gioitinh_fe).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
        logging.info("check font-end: Giới tinh - Nữ")
        logging.info(thongtincoban_gioitinh == "Nữ")

        #Ngày sinh
        driver.find_element(By.XPATH, var.thongtincoban_icon_ngaysinh).click()
        driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_chinhsua).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_luu).click()
        time.sleep(1)
        thongtincoban_ngaysinh = driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_fe).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
        logging.info("check font-end: Ngày sinh - 3 tháng 10, 2000")
        logging.info(thongtincoban_ngaysinh == "3 tháng 10, 2000")

        #biệt danh
        driver.find_element(By.XPATH, var.thongtincoban_icon_bietdanh).click()
        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_input).click()
        thongtincoban_bietdanh1 = driver.find_element(By.XPATH, var.thongtincoban_bietdanh_input)
        thongtincoban_bietdanh1.send_keys(Keys.CONTROL, "a")
        del driver.requests
        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['bietdanh'])
        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_luu).click()
        time.sleep(1)
        thongtincoban_bietdanh = driver.find_element(By.XPATH, var.thongtincoban_bietdanh_fe).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
        logging.info("check font-end: Biệt danh - " + data['trangcanhan_gioithieu_thongtincoban']['bietdanh'])
        logging.info(thongtincoban_bietdanh == data['trangcanhan_gioithieu_thongtincoban']['bietdanh'])

    def thongtincoban_dulieusai(self):
        driver.implicitly_wait(15)
        driver.execute_script("window.scrollBy(0,-200)", "")
        #thông tin liên hệ
        #số điện thoại
        driver.find_element(By.XPATH, var.thongtincoban_icon_sdt).click()
        driver.find_element(By.XPATH, var.thongtincoban_sdt_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_sdt_input).click()
        thongtincoban_sdt1 = driver.find_element(By.XPATH, var.thongtincoban_sdt_input)
        thongtincoban_sdt1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_sdt_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['sodienthoai_dai'])
        driver.find_element(By.XPATH, var.thongtincoban_sdt_luu).click()
        time.sleep(1)
        check_thongtincoban_sodienthoai_dai = driver.find_element(By.XPATH,var.trangcanhan_check_thongtincoban_sodienthoai_dai1).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin liên hệ - Số điện thoại dài quá 10 ký tự")
        logging.info("check font-end: Số điện thoại độ dài không đúng (phải là 10 ký tự)")
        logging.info(check_thongtincoban_sodienthoai_dai == "Số điện thoại độ dài không đúng (phải là 10 ký tự)")
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_thongtincoban_sodienthoai_ok).click()

        thongtincoban_sdt1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_sdt_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['sodienthoai_ngan'])
        driver.find_element(By.XPATH, var.thongtincoban_sdt_luu).click()
        time.sleep(1)
        check_thongtincoban_sodienthoai_ngan = driver.find_element(By.XPATH,var.trangcanhan_check_thongtincoban_sodienthoai_ngan1).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin liên hệ - Số điện thoại it hơn 10 ký tự")
        logging.info("check font-end: Số điện thoại độ dài không đúng (phải là 10 ký tự)")
        logging.info(check_thongtincoban_sodienthoai_ngan == "Số điện thoại độ dài không đúng (phải là 10 ký tự)")
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_thongtincoban_sodienthoai_ok).click()

        thongtincoban_sdt1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_sdt_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['sodienthoai_trung'])
        driver.find_element(By.XPATH, var.thongtincoban_sdt_luu).click()
        time.sleep(1)
        check_thongtincoban_sodienthoai_trungso = driver.find_element(By.XPATH,var.trangcanhan_check_thongtincoban_sodienthoai_trungso1).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin liên hệ - Số điện thoại bị trùng")
        logging.info("check font-end: Số điện thoại đã có")
        print(check_thongtincoban_sodienthoai_trungso)
        logging.info(check_thongtincoban_sodienthoai_trungso == "Số điện thoại đã có")
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_thongtincoban_sodienthoai_ok).click()

        thongtincoban_sdt1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_sdt_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['sodienthoai2'])
        driver.find_element(By.XPATH, var.thongtincoban_sdt_luu).click()
        time.sleep(1)

        # Các trang web và liên kết xã hội
        #web
        driver.execute_script("window.scrollBy(0,700)", "")
        time.sleep(1)
        driver.find_element(By.XPATH, var.thongtincoban_icon_web).click()
        driver.find_element(By.XPATH, var.thongtincoban_web_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_web_input).click()
        thongtincoban_web1 = driver.find_element(By.XPATH, var.thongtincoban_web_input)
        thongtincoban_web1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_web_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['web_dulieusai'])
        driver.find_element(By.XPATH, var.thongtincoban_web_luu).click()
        time.sleep(1)
        thongtincoban_web_dulieusai = driver.find_element(By.XPATH, var.thongtincoban_web_fe_dulieusai).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Các trang web và liên kết xã hội")
        logging.info("check font-end: Web(nhập sai định dạng web) - message: URL You provided invalid URL")
        logging.info(thongtincoban_web_dulieusai == "URL You provided invalid URL")
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.thongtincoban_web_ok).click()
        thongtincoban_web1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_web_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['web1'])
        driver.find_element(By.XPATH, var.thongtincoban_web_luu).click()
        time.sleep(3)

        #liên kết xã hội
        try:
            driver.find_element(By.XPATH, var.thongtincoban_icon_lienket).click()
            driver.find_element(By.XPATH, var.thongtincoban_lienket_chinhsua).click()
            time.sleep(2)
            driver.find_element(By.XPATH, var.thongtincoban_lienket_input).click()
            thongtincoban_lienket1 = driver.find_element(By.XPATH, var.thongtincoban_lienket_input)
            thongtincoban_lienket1.send_keys(Keys.CONTROL, "a")
            driver.find_element(By.XPATH, var.thongtincoban_lienket_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['lienket_dulieusai'])
            driver.find_element(By.XPATH, var.xong).click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
            driver.find_element(By.XPATH, var.thongtincoban).click()
            driver.find_element(By.XPATH, var.thongtincoban_icon_lienket).click()
            driver.find_element(By.XPATH, var.thongtincoban_lienket_chinhsua).click()
            driver.find_element(By.XPATH, var.thongtincoban_lienket_input).click()
            thongtincoban_lienket1 = driver.find_element(By.XPATH, var.thongtincoban_lienket_input)
            thongtincoban_lienket1.send_keys(Keys.CONTROL, "a")
            driver.find_element(By.XPATH, var.thongtincoban_lienket_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['lienket_dulieusai'])
            driver.find_element(By.XPATH, var.xong).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.thongtincoban_icon_lienket).click()
        driver.find_element(By.XPATH, var.thongtincoban_lienket_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_lienket_input).click()
        thongtincoban_lienket1 = driver.find_element(By.XPATH, var.thongtincoban_lienket_input)
        thongtincoban_lienket1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_lienket_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['lienket1'])
        driver.find_element(By.XPATH, var.thongtincoban_lienket_luu).click()
        time.sleep(1)


        #thông tin cơ bản
        #tiểu sử
        driver.find_element(By.XPATH, var.thongtincoban_icon_tieusu).click()
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_input).click()
        thongtincoban_tieusu1 = driver.find_element(By.XPATH, var.thongtincoban_tieusu_input)
        thongtincoban_tieusu1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['tieusu_dai'])
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_luu).click()
        time.sleep(1)

        thongtincoban_tieusu_dulieusai = driver.find_element(By.XPATH, var.thongtincoban_tieusu_fe_dulieusai).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Các trang web và liên kết xã hội")
        logging.info("check font-end: Tiểu sử(Nhập quá 100 ký tự) - message: Tiểu sử chỉ được phép nhập tối đa 100 ký tự!")
        logging.info(thongtincoban_tieusu_dulieusai == "Tiểu sử chỉ được phép nhập tối đa 100 ký tự!")
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_huy).click()
        driver.find_element(By.XPATH, var.thongtincoban_icon_tieusu).click()
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_input).click()
        thongtincoban_tieusu1 = driver.find_element(By.XPATH, var.thongtincoban_tieusu_input)
        thongtincoban_tieusu1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['tieusu1'])
        driver.find_element(By.XPATH, var.thongtincoban_tieusu_luu).click()
        time.sleep(1)

        #Giới tinh
        driver.find_element(By.XPATH, var.thongtincoban_icon_gioitinh).click()
        driver.find_element(By.XPATH, var.thongtincoban_gioitinh_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_gioitinh_iconchongioitinh).click()
        driver.find_element(By.XPATH, var.thongtincoban_gioitinh_nam).click()
        driver.find_element(By.XPATH, var.thongtincoban_gioitinh_luu).click()
        time.sleep(1)
        thongtincoban_gioitinh = driver.find_element(By.XPATH, var.thongtincoban_gioitinh_fe).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
        logging.info("check font-end: Giới tinh - Nam")
        logging.info(thongtincoban_gioitinh == "Nam")

        #Ngày sinh
        driver.find_element(By.XPATH, var.thongtincoban_icon_ngaysinh).click()
        driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_iconchonnam).click()
        driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_duoi14t_2022).click()
        driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_luu).click()
        time.sleep(1)
        thongtincoban_ngaysinh_message = driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_fe_chuadu14t).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
        logging.info("check font-end: Ngày sinh - message: Tuổi của bạn không đủ 14 tuổi.Vui lòng điền lại thông tin!")
        logging.info(thongtincoban_ngaysinh_message == "Tuổi của bạn không đủ 14 tuổi.Vui lòng điền lại thông tin!")
        driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_ok).click()
        time.sleep(0.5)

        driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_iconchonnam).click()
        driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_tren14t_2000).click()
        driver.find_element(By.XPATH, var.thongtincoban_ngaysinh_luu).click()
        time.sleep(1)

        #biệt danh
        driver.find_element(By.XPATH, var.thongtincoban_icon_bietdanh).click()
        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_chinhsua).click()
        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_input).click()
        thongtincoban_bietdanh1 = driver.find_element(By.XPATH, var.thongtincoban_bietdanh_input)
        thongtincoban_bietdanh1.send_keys(Keys.CONTROL, "a")

        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['bietdanh_dai'])
        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_luu).click()
        time.sleep(1)
        thongtincoban_bietdanh_dai = driver.find_element(By.XPATH, var.thongtincoban_bietdanhdai_fe).text
        logging.info("Trang cá nhân - Giới thiệu - Thông tin cơ bản - Thông tin cơ bản")
        logging.info("check font-end: Biệt danh(dài quá 100 ký tự) - message: Tiểu sử chỉ được phép nhập tối đa 100 ký tự!")
        logging.info(thongtincoban_bietdanh_dai)
        logging.info(thongtincoban_bietdanh_dai == "Tên biệt danh không vượt quá 20 ký tự.Vui lòng điền lại thông tin!")
        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_ok).click()
        time.sleep(0.5)

        thongtincoban_bietdanh1 = driver.find_element(By.XPATH, var.thongtincoban_bietdanh_input)
        thongtincoban_bietdanh1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_input).send_keys(data['trangcanhan_gioithieu_thongtincoban']['bietdanh1'])
        driver.find_element(By.XPATH, var.thongtincoban_bietdanh_luu).click()
        time.sleep(1)
        driver.refresh()
        time.sleep(2)


class giadinh_va_cacmoiquanhe():
    def gd_va_mqh_taikhoan1(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        time.sleep(1)
        #gia đình và các mối quan hệ
        #Mối quan hệ
        driver.find_element(By.XPATH, var.giadinhvamqh).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_icon_mqh).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_chinhsua).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_icon_chonmqh).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_henho).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.giadinhvamqh_chonnguoithan_input).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_chonnguoithan_input_x).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_chonnguoithan_input).send_keys(data['trangcanhan_gd_va_mqh']['name'])

        wait = WebDriverWait(driver, 10)
        giadinhvamqh_ngocmai1 = wait.until(EC.element_to_be_clickable((By.XPATH, var.giadinhvamqh_ngocmai)))
        giadinhvamqh_ngocmai1.click()
        driver.find_element(By.XPATH, var.giadinhvamqh_luu).click()
        time.sleep(1)

        # mqh_ngocmai = driver.find_element(By.XPATH, var.giadinhvamqh_mqh_ngocmai1).text
        # logging.info("Trang cá nhân - Giới thiệu - Gia đình và các mối quan hệ - Mối quan hệ")
        # logging.info("check font-end: Người thân - " + data['trangcanhan_gd_va_mqh']['name'])
        # logging.info(mqh_ngocmai == data['trangcanhan_gd_va_mqh']['name'])
        #
        # # mqh_trangthai_thoigian = driver.find_element(By.XPATH, var.giadinhvamqh_mqh_trangthai_thoigian4).text   #mqh_loi
        # mqh_trangthai_thoigian = driver.find_element(By.XPATH, var.giadinhvamqh_mqh_trangthai_thoigian1).text
        # logging.info("Trang cá nhân - Giới thiệu - Gia đình và các mối quan hệ - Gia đình")
        # logging.info("check font-end: Gia đình - Trạng thái/Thời gian - Hẹn hò started 2022-01-01" )
        # logging.info(mqh_trangthai_thoigian == "Hẹn hò started 2022-01-01")

        #Gia đình
        try:
            mqh_ngocmai = driver.find_element(By.XPATH, var.giadinhvamqh_mqh_ngocmai1).text
            logging.info("Trang cá nhân - Giới thiệu - Gia đình và các mối quan hệ - Mối quan hệ")
            logging.info("check font-end: Người thân - " + data['trangcanhan_gd_va_mqh']['name'])
            logging.info(mqh_ngocmai == data['trangcanhan_gd_va_mqh']['name'])

            # mqh_trangthai_thoigian = driver.find_element(By.XPATH, var.giadinhvamqh_mqh_trangthai_thoigian4).text   #loi
            mqh_trangthai_thoigian = driver.find_element(By.XPATH, var.giadinhvamqh_mqh_trangthai_thoigian1).text
            logging.info("Trang cá nhân - Giới thiệu - Gia đình và các mối quan hệ - Gia đình")
            logging.info("check font-end: Gia đình - Trạng thái/Thời gian - Hẹn hò(Đang chờ) started 2022-01-01")
            logging.info(mqh_trangthai_thoigian)
            logging.info(mqh_trangthai_thoigian == "Hẹn hò(Đang chờ) started 2022-01-01")
            driver.find_element(By.XPATH, var.giadinhvamqh_icon_nguoithan).click()
            driver.find_element(By.XPATH, var.giadinhvamqh_sdt_chinhsua).click()
        except:
            driver.find_element(By.XPATH, var.trangcanhan_gioithieuhuy).click()
            driver.find_element(By.XPATH, var.giadinhvamqh_icon_nguoithan).click()
        # driver.find_element(By.XPATH, var.giadinhvamqh_icon_nguoithan).click()
            driver.find_element(By.XPATH, var.giadinhvamqh_sdt_chinhsua).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_nguoithan_input).click()
        # driver.find_element(By.XPATH, var.giadinhvamqh_nguoithan_input).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_nguoithan_input_x).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_nguoithan_input).send_keys(data['trangcanhan_gd_va_mqh']['name1'])

        wait = WebDriverWait(driver, 10)
        giadinhvamqh_manhcuong1 = wait.until(EC.element_to_be_clickable((By.XPATH, var.giadinhvamqh_manhcuong)))
        giadinhvamqh_manhcuong1.click()
        driver.find_element(By.XPATH, var.giadinhvamqh_nguoithan_chongiadinh).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_giadinh).click()
        driver.find_element(By.XPATH, var.giadinhvamqh_giadinh_luu).click()
        time.sleep(1)

        gd_manhcuong = driver.find_element(By.XPATH, var.giadinhvamqh_gd_manhcuong1).text
        logging.info("Trang cá nhân - Giới thiệu - Gia đình và các mối quan hệ - Mối quan hệ")
        logging.info("check font-end: Gia đình - " + data['trangcanhan_gd_va_mqh']['name1'])
        logging.info(gd_manhcuong == data['trangcanhan_gd_va_mqh']['name1'])

        gd_tinhtrang = driver.find_element(By.XPATH, var.giadinhvamqh_mqh_gd_tinhtrang1).text
        logging.info("Trang cá nhân - Giới thiệu - Gia đình và các mối quan hệ - Gia đình")
        logging.info("check font-end: Tình trạng - Họ hàng(Đang chờ)" )
        logging.info(gd_tinhtrang == "Họ hàng(Đang chờ)")
        del driver.requests
        driver.refresh()
        time.sleep(2)


class sukientrongdoi():
    def sukientrongdoi_xem(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sukientrongdoi).click()
        driver.execute_script("window.scrollBy(0,400)", "")
        driver.find_element(By.XPATH, var.sukientrongdoi_trunghoc_icon).click()
        driver.find_element(By.XPATH, var.sukientrongdoi_trunghoc_xem).click()
        time.sleep(3)
        driver.back()
        time.sleep(2)
        driver.find_element(By.XPATH, var.sukientrongdoi_daihoc_icon).click()
        driver.find_element(By.XPATH, var.sukientrongdoi_daihoc_xem).click()
        time.sleep(3)
        driver.back()
        time.sleep(2)
        driver.find_element(By.XPATH, var.sukientrongdoi_congviec_icon).click()
        driver.find_element(By.XPATH, var.sukientrongdoi_congviec_xem).click()
        time.sleep(3)
        driver.back()
        time.sleep(2)
    def sukientrongdoi_themmoisukien_check(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sukientrongdoi).click()
        driver.execute_script("window.scrollBy(0,400)", "")
        driver.find_element(By.XPATH, var.sukientrongdoi_iconthemmoi).click()
    def mau1(self, tieude):
        # Tạo sự kiện riêng
        driver.find_element(By.XPATH, var.taosukienrieng_tailenfile).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien1.exe")
        time.sleep(1)
        sukientrongdoi_clear = driver.find_element(By.XPATH, var.taosukienrieng_input)
        sukientrongdoi_clear.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.taosukienrieng_input).send_keys(tieude)
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukientrongdoi_x).click()
        time.sleep(1)
    def mau2(self, tieude, diachi):
        driver.find_element(By.XPATH, var.nhacuavadoisong_tailenfile).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien1.exe")
        time.sleep(1)
        sukientrongdoi_clear1 = driver.find_element(By.XPATH, var.nhacuavadoisong_tieude_input)
        sukientrongdoi_clear1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.nhacuavadoisong_tieude_input).send_keys(tieude)
        driver.find_element(By.XPATH, var.nhacuavadoisong_chondiachi_input).send_keys(diachi)
        wait = WebDriverWait(driver, 10)
        nhacuavadoisong_chondiachi = wait.until(EC.element_to_be_clickable((By.XPATH, var.nhacuavadoisong_chondiachi_thainguyen)))
        nhacuavadoisong_chondiachi.click()
        driver.find_element(By.XPATH, var.nhacuavadoisong_thoigian).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukientrongdoi_x).click()
        time.sleep(1)
    def taosukienrieng(self):
        try:
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.sukientrongdoi).click()
            driver.execute_script("window.scrollBy(0,400)", "")
            driver.find_element(By.XPATH, var.sukientrongdoi_iconthemmoi).click()
            wait = WebDriverWait(driver, 10)
            taosukienrieng = wait.until(EC.element_to_be_clickable((By.XPATH, var.taosukienrieng1)))
            taosukienrieng.click()
            sukientrongdoi.mau1(self, data['trangcanhan_sukientrongdoi']['taosukienrieng'])
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.sukientrongdoi).click()
            driver.execute_script("window.scrollBy(0,400)", "")
            driver.find_element(By.XPATH, var.sukientrongdoi_iconthemmoi).click()
            wait = WebDriverWait(driver, 10)
            taosukienrieng = wait.until(EC.element_to_be_clickable((By.XPATH, var.taosukienrieng1)))
            taosukienrieng.click()
            sukientrongdoi.mau1(self, data['trangcanhan_sukientrongdoi']['taosukienrieng'])
    def nhacuavadoisong(self):
        driver.find_element(By.XPATH, var.nhacuavadoisong).click()
        driver.find_element(By.XPATH, var.nhacuavadoisong_nhaptieude).click()
        sukientrongdoi.mau2(self, data['trangcanhan_sukientrongdoi']['nhacuavadoisong_quequan'], data['trangcanhan_sukientrongdoi']['nhacuavadoisong_diachi'])
        driver.find_element(By.XPATH, var.taosukientrongdoi_x).click()
    def moiquanhe(self):
        driver.find_element(By.XPATH, var.moiquanhe).click()
        driver.find_element(By.XPATH, var.moiquanhe_nhaptieude).click()
        sukientrongdoi.mau2(self, data['trangcanhan_sukientrongdoi']['moiquanhe_tieude'], data['trangcanhan_sukientrongdoi']['moiquanhe_diachi'])
        driver.find_element(By.XPATH, var.taosukientrongdoi_x).click()
    def work(self):
        driver.find_element(By.XPATH, var.work).click()
        driver.find_element(By.XPATH, var.work_nhaptieude).click()
        sukientrongdoi.mau2(self, data['trangcanhan_sukientrongdoi']['work_tieude'], data['trangcanhan_sukientrongdoi']['work_diachi'])
        driver.find_element(By.XPATH, var.taosukientrongdoi_x).click()
    def tuongnho(self):
        driver.find_element(By.XPATH, var.tuongnho).click()
        sukientrongdoi.mau1(self, data['trangcanhan_sukientrongdoi']['tuongnho_tieude'])
    def daumocvathanhtuu(self):
        driver.find_element(By.XPATH, var.daumocvathanhtuu).click()
        sukientrongdoi.mau1(self, data['trangcanhan_sukientrongdoi']['daumocvathanhtuu_tieude'])
    def suckhoethechatvatinhthan(self):
        driver.find_element(By.XPATH, var.suckhoethechatvatinhthan).click()
        sukientrongdoi.mau1(self, data['trangcanhan_sukientrongdoi']['suckhoethechatvatinhthan_tieude'])
    def sothichhoatdong(self):
        driver.find_element(By.XPATH, var.sothichhoatdong).click()
        sukientrongdoi.mau1(self, data['trangcanhan_sukientrongdoi']['sothichhoatdong_tieude'])
    def dulich(self):
        driver.find_element(By.XPATH, var.dulich).click()
        sukientrongdoi.mau2(self, data['trangcanhan_sukientrongdoi']['dulich_tieude'], data['trangcanhan_sukientrongdoi']['dulich_diachi'])
    def giadinh(self):
        driver.find_element(By.XPATH, var.giadinh).click()
        driver.find_element(By.XPATH, var.work_nhaptieude).click()
        sukientrongdoi.mau2(self, data['trangcanhan_sukientrongdoi']['giadinh_tieude'], data['trangcanhan_sukientrongdoi']['giadinh_diachi'])
        driver.find_element(By.XPATH, var.taosukientrongdoi_x).click()
    def hocvan(self):
        driver.find_element(By.XPATH, var.hocvan).click()
        driver.find_element(By.XPATH, var.work_nhaptieude).click()
        sukientrongdoi.mau2(self, data['trangcanhan_sukientrongdoi']['hocvan_tieude'], data['trangcanhan_sukientrongdoi']['hocvan_diachi'])
        driver.find_element(By.XPATH, var.taosukientrongdoi_x).click()
        driver.find_element(By.XPATH, var.taosukientrongdoi_x).click()
    def sukientrongdoi_themmoisukien(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sukientrongdoi).click()
        driver.execute_script("window.scrollBy(0,400)", "")
        driver.find_element(By.XPATH, var.sukientrongdoi_iconthemmoi).click()
        wait = WebDriverWait(driver, 10)
        dulich_add = wait.until(EC.element_to_be_clickable((By.XPATH, var.dulich)))
        dulich_add.click()
        driver.find_element(By.XPATH, var.nhacuavadoisong_tailenfile).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien1.exe")
        time.sleep(1)
        sukientrongdoi_clear1 = driver.find_element(By.XPATH, var.nhacuavadoisong_tieude_input)
        sukientrongdoi_clear1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.nhacuavadoisong_tieude_input).send_keys(data['trangcanhan_sukientrongdoi']['dulich_tieude'])
        driver.find_element(By.XPATH, var.nhacuavadoisong_chondiachi_input).send_keys(data['trangcanhan_sukientrongdoi']['dulich_diachi'])
        wait = WebDriverWait(driver, 10)
        nhacuavadoisong_chondiachi = wait.until(EC.element_to_be_clickable((By.XPATH, var.nhacuavadoisong_chondiachi_thainguyen)))
        nhacuavadoisong_chondiachi.click()
        driver.find_element(By.XPATH, var.nhacuavadoisong_thoigian).click()
        # driver.find_element(By.XPATH, var.nhacuavadoisong_thoigian_nam).click()
        # driver.find_element(By.XPATH, var.nhacuavadoisong_thoigian_nam_2022).click()
        driver.find_element(By.XPATH, var.nhacuavadoisong_thoigian_thang).click()
        driver.find_element(By.XPATH, var.nhacuavadoisong_thoigian_thang_9).click()
        driver.find_element(By.XPATH, var.nhacuavadoisong_thoigian_ngay).click()
        driver.find_element(By.XPATH, var.nhacuavadoisong_thoigian_ngay_1a).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukientrongdoi_luu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukientrongdoi_mota).send_keys(data['trangcanhan_sukientrongdoi']['dulich_mota'])
        driver.find_element(By.XPATH, var.taosukientrongdoi_dang).click()
        time.sleep(10)
        driver.find_element(By.XPATH, var.trangcanhan).click()
        time.sleep(2)
        del driver.requests
        driver.refresh()
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,1100)", "")
        dongthoigian_sukien_mota = driver.find_element(By.XPATH, var.dongthoigian_sukienthemmoi1).text
        logging.info("Trang cá nhân - Giới thiệu - Thêm mới sự kiện trong đời lên trang cá nhân(du lịch)")
        logging.info("check font-end: Dòng thời gian - Mô tả - " + data['trangcanhan_sukientrongdoi']['dulich_mota'])
        logging.info(dongthoigian_sukien_mota == data['trangcanhan_sukientrongdoi']['dulich_mota'])

        dongthoigian_sukiendiadiem = driver.find_element(By.XPATH, var.dongthoigian_sukien_diadiem1).text
        logging.info("Trang cá nhân - Giới thiệu - Thêm mới sự kiện trong đời lên trang cá nhân(du lịch)")
        logging.info("check font-end: Dòng thời gian - Sự kiện, địa điểm - Du lịch, Trường test tại Thái Nguyên")
        logging.info(dongthoigian_sukiendiadiem == "Du lịch, Trường test tại Thái Nguyên")

        dongthoigian_sukien_thoigian = driver.find_element(By.XPATH, var.dongthoigian_sukien_thoigian1).text
        logging.info("Trang cá nhân - Giới thiệu - Thêm mới sự kiện trong đời lên trang cá nhân(du lịch)")
        logging.info("check font-end: Dòng thời gian - Thời gian - 01 tháng 09, 2024")
        logging.info(dongthoigian_sukien_thoigian == "01 tháng 09, 2024")

        dongthoigian_sukien_mota = driver.find_element(By.XPATH, var.dongthoigian_sukien_mota).text
        logging.info("Trang cá nhân - Giới thiệu - Thêm mới sự kiện trong đời lên trang cá nhân(du lịch)")
        logging.info("check font-end: Dòng thời gian - Mô tả - " + data['trangcanhan_sukientrongdoi']['dulich_mota'])
        logging.info(dongthoigian_sukien_mota == data['trangcanhan_sukientrongdoi']['dulich_mota'])

        #chỉnh sửa dữ liệu thêm mới sự kiên
        dongthoigian_sukien_diadiem = driver.find_element(By.XPATH, var.dongthoigian_sukien_diadiem).text
        print(dongthoigian_sukien_diadiem)
        logging.info("Trang cá nhân - Giới thiệu - Thêm mới sự kiện trong đời lên trang cá nhân(du lịch)")
        logging.info("check font-end: Dòng thời gian - Địa điểm - " + data['trangcanhan_sukientrongdoi']['dulich_diachi'])
        logging.info(dongthoigian_sukien_diadiem == data['trangcanhan_sukientrongdoi']['dulich_diachi'])
        time.sleep(1)
    def sukientrongdoi_themmoisukien_chinhsua_xoa(self):
        driver.implicitly_wait(15)
        driver.find_element(By.XPATH, var.icon_chinhsuabaiviet).click()
        driver.find_element(By.XPATH, var.chinhsuabaiviet).click()
        driver.find_element(By.XPATH, var.chinhsuabaiviet_mota_input).send_keys("123")
        driver.find_element(By.XPATH, var.chinhsuabaiviet_luu).click()
        driver.find_element(By.XPATH, var.icon_chinhsuabaiviet_message_luu).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.icon_chinhsuabaiviet).click()
        driver.find_element(By.XPATH, var.icon_chinhsuabaiviet_xoabaiviet).click()
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(2)
        driver.refresh()
        time.sleep(2)


class noitungsong():
    def noitungsong(self):
        driver.implicitly_wait(15)
        # driver.find_element(By.XPATH, var.trangcanhan).click()
        time.sleep(3)
        # Trang cá nhân - Giới thiệu - Nơi từng sống
        #Sống tại
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_noitungsong).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_chinhsua).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_input).send_keys(data['trangcanhan_noitungsong']['noitungsong_songtai'])

        wait = WebDriverWait(driver, 10)
        chon_hungyen = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_songtai_hungyen)))
        chon_hungyen.click()

        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai_luu).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 62, 2, "x")
        check_trangcanhan_gioithieu_songtai = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai2).text
        logging.info("Trang cá nhân - Giới thiệu  - Nơi từng sống - Sống tại")
        logging.info("check font-end: Sống tại - " + data['trangcanhan_noitungsong']['noitungsong_songtai'])
        logging.info(check_trangcanhan_gioithieu_songtai == data['trangcanhan_noitungsong']['noitungsong_songtai'])

        # Đến từ
        driver.execute_script("window.scrollBy(0,500)", "")
        # time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_chinhsua).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_input).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_x).click()
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_input).send_keys(data['trangcanhan_noitungsong']['noitungsong_dentu'])

        wait = WebDriverWait(driver, 10)
        dentu_thainguyen = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_dentu_thainguyen)))
        dentu_thainguyen.click()

        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu_luu).click()
        time.sleep(1)
        writeData(var.path_baocao, "Sheet1", 63, 2, "x")
        check_gioithieu_dentu = driver.find_element(By.XPATH, var.trangcanhan_check_gioithieu_dentu2).text
        print(check_gioithieu_dentu)
        logging.info("Trang cá nhân - Giới thiệu - Nơi từng sống - Đến từ")
        logging.info("check font-end: Đến từ - " + data['trangcanhan_noitungsong']['noitungsong_dentu'])
        logging.info(check_gioithieu_dentu == data['trangcanhan_noitungsong']['noitungsong_dentu'])
        del driver.requests
        driver.refresh()
        time.sleep(2)



class banbe():
    def tatcabanbe(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.trangcanhan_banbe)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        # driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.find_element(By.XPATH, var.trangcanhan_banbe_ngocmai).click()
        time.sleep(3)
        check_ngocmai =driver.find_element(By.XPATH, var.trangcanhan_banbe_ngocmai1).text
        logging.info("Trang cá nhân - Bạn bè  - Tất cả bạn bè - Xem trang cá nhân của bạn bè")
        logging.info("check font-end: Trang cá nhân của - " + data['trangcanhan_banbe']['trangcanhan_banbe_ngocmai'])
        logging.info(check_ngocmai)
        logging.info(check_ngocmai == data['trangcanhan_banbe']['trangcanhan_banbe_ngocmai'])
        driver.back()
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,700)", "")
        #bo theo dõi
        driver.find_element(By.XPATH, var.trangcanhan_banbe_icon_manhcuong).click()
        driver.find_element(By.XPATH, var.trangcanhan_banbe_botheodoi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(4)
        #theodoi
        driver.find_element(By.XPATH, var.trangcanhan_banbe_icon_manhcuong).click()
        driver.find_element(By.XPATH, var.theodoi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(2)
        driver.refresh()
        #huỷ kết bạn
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_banbe_icon_manhcuong).click()
        driver.find_element(By.XPATH, var.huyketban).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(2)
        huyketban_manhcuong =driver.find_element(By.XPATH, var.thembanbe).text
        print(huyketban_manhcuong)
        logging.info("Trang cá nhân - Bạn bè  - Tất cả bạn bè - Huỷ kết bạn")
        logging.info("check font-end: Huỷ kết bạn thành công")
        logging.info(huyketban_manhcuong == "Thêm bạn bè")
        time.sleep(3)

        #thêm bạn be - huỷ
        # wait = WebDriverWait(driver, 10)
        # element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_banbe_thembanbe_manhcuong)))
        # element.click()
        driver.find_element(By.XPATH, var.thembanbe).click()
        time.sleep(1)
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_banbe_loimoi_huy).click()
        time.sleep(2.5)
        driver.get("https://cmc-fe.emso.vn/user/truongvck33")
        time.sleep(2.5)
        check_banbe_tranquangtruong =driver.find_element(By.XPATH, var.check_banbe_tranquangtruong_tinhtrang).text
        check_banbe_name = driver.find_element(By.XPATH, var.check_banbe_tranquangtruong2).text
        print(check_banbe_tranquangtruong)
        logging.info("Xoá lời mời kết bạn - check trang cá nhân người vừa kết bạn")
        logging.info("check font-end: tên Người gửi lời mời kết bạn: " + check_banbe_name)
        logging.info("check font-end: tình trạng - Thêm bạn bè")
        logging.info(check_banbe_tranquangtruong == "Thêm bạn bè")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_banbe_guiloimoi).click()
        time.sleep(2)

        # thêm bạn bè - đồng ý
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        driver.find_element(By.XPATH, var.trangcanhan_banbe_loimoi_chapnhan).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan).click()
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.find_element(By.XPATH, var.trangcanhan_banbe_icon_manhcuong).click()
        driver.find_element(By.XPATH, var.huyketban).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(5)
        driver.find_element(By.XPATH, var.thembanbe).click()
        time.sleep(2)
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_banbe_loimoi_chapnhan).click()
        time.sleep(2.5)
        driver.get("https://cmc-fe.emso.vn/user/truongvck33")
        time.sleep(2)
        check_banbe_tranquangtruong = driver.find_element(By.XPATH, var.check_banbe_tranquangtruong1).text
        check_banbe_name = driver.find_element(By.XPATH, var.check_banbe_tranquangtruong2).text
        print(check_banbe_tranquangtruong)
        logging.info("Chấp nhận kết bạn - check trang cá nhân người vừa kết bạn")
        logging.info("check font-end: tên Người gửi lời mời kết bạn: " + check_banbe_name)
        logging.info("check font-end: tình trạng - Bạn bè")
        logging.info(check_banbe_tranquangtruong == "Bạn bè")
        time.sleep(1)
        login.login3(self, "truongvck33@gmail.com", "voncamk22")

    def banbe_chinhsuaquyenriengtu(self):
        driver.implicitly_wait(15)

        #Danh sách bạn be
        #quyền riengtu
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.find_element(By.XPATH, var.trangcanhan_banbe_iconchinhsuaquyen).click()
        driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu_icon_dsbb1).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu_dsbb_riengtu1).click()
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,700)", "")
        time.sleep(1)
        check_banbe_riengtu1 = driver.find_element(By.XPATH, var.huenguyen).is_displayed()
        logging.info("Trang cá nhân - bạn bè - chỉnh sửa quyền riêng tư ")
        logging.info("check font-end: Danh sách bạn bè khi chỉnh quyền thành riêng tư có hiển thị không?")
        logging.info(check_banbe_riengtu1)
        #Người lạ
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        driver.get("https://cmc-fe.emso.vn/user/truongvck33")
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.execute_script("window.scrollBy(0,510)", "")
        time.sleep(1)
        check_nguoila_riengtu = driver.find_element(By.XPATH, var.check_nguoila_riengtu3).text
        logging.info("Trang cá nhân - bạn bè - chỉnh sửa quyền riêng tư - riêng tư")
        logging.info("check font-end: login bằng tài khoản khác(chưa kết bạn) và check  - Không có bạn bè nào")
        logging.info(check_nguoila_riengtu)
        logging.info(check_nguoila_riengtu == "Không có bạn bè nào")
        time.sleep(1.5)
        #Bạn bè
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        driver.get("https://cmc-fe.emso.vn/user/truongvck33")
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.execute_script("window.scrollBy(0,510)", "")
        time.sleep(1)
        check_banbe_riengtu = driver.find_element(By.XPATH, var.check_banbe_riengtu3).text
        logging.info("Trang cá nhân - bạn bè - chỉnh sửa quyền riêng tư - Rêng tư")
        logging.info("check font-end: login bằng tài khoản khác(chưa kết bạn) và check  - Không có bạn bè nào")
        logging.info(check_banbe_riengtu == "Không có bạn bè nào")
        time.sleep(1.5)



        #quyền bạn bè
        login.login3(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.find_element(By.XPATH, var.trangcanhan_banbe_iconchinhsuaquyen).click()
        driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu_icon_dsbb1).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu_dsbb_banbe2).click()
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,700)", "")
        time.sleep(1)
        check_banbe_riengtu = driver.find_element(By.XPATH, var.huenguyen).is_displayed()
        time.sleep(1)
        logging.info("Trang cá nhân - bạn bè - chỉnh sửa quyền riêng tư")
        logging.info("check font-end: Danh sách bạn bè khi chỉnh quyền thành Bạn bè có hiển thị không")
        logging.info(check_banbe_riengtu)
        time.sleep(0.5)
        #người lạ
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        driver.get("https://cmc-fe.emso.vn/user/truongvck33")
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.execute_script("window.scrollBy(0,510)", "")
        time.sleep(1)
        check_banbe_riengtu = driver.find_element(By.XPATH, var.huenguyen).text
        logging.info("Trang cá nhân - bạn bè - chỉnh sửa quyền riêng tư - bạn bè")
        print(check_banbe_riengtu)
        logging.info("check font-end: login bằng tài khoản khác(chưa kết bạn) và check  - chỉ hiện thị bạn bè chung")
        logging.info(check_banbe_riengtu == "hue nguyen")
        time.sleep(1.5)
        #Bạn bè
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        driver.get("https://cmc-fe.emso.vn/user/truongvck33")
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.execute_script("window.scrollBy(0,510)", "")
        time.sleep(1)
        check_banbe_quyenbanbe = driver.find_element(By.XPATH, var.trangcanhan_banbe_ngocmai).text
        logging.info("Trang cá nhân - bạn bè - chỉnh sửa quyền riêng tư - Bạn bè")
        logging.info("check font-end: login bằng tài khoản khác(bạn bè) và check  - có bạn bè tên Ngọc Mai")
        logging.info(check_banbe_quyenbanbe[0:8] == "Ngọc Mai")
        time.sleep(1.5)



        # quyền công khai
        login.login3(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.find_element(By.XPATH, var.trangcanhan_banbe_iconchinhsuaquyen).click()
        driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu_icon_dsbb1).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu_dsbb_congkhai3).click()

        # driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu_icon_dsbb).click()
        # driver.find_element(By.XPATH, var.trangcanhan_chinhsuaquyenriengtu_dsbb_congkhai).click()
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,700)", "")
        time.sleep(1)
        check_banbe_quyenbanbe_huenguyen = driver.find_element(By.XPATH, var.huenguyen)
        logging.info("Trang cá nhân - bạn bè - chỉnh sửa quyền riêng tư")
        logging.info(check_banbe_quyenbanbe_huenguyen.text)
        logging.info("check font-end: Danh sách bạn bè khi chỉnh quyền thành Công khai có hiển thị không")
        logging.info(check_banbe_quyenbanbe_huenguyen.is_displayed())
        time.sleep(0.5)

        # người lạ
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        driver.get("https://cmc-fe.emso.vn/user/truongvck33")
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.execute_script("window.scrollBy(0,510)", "")
        time.sleep(1)
        check_nguoila_congkhai = driver.find_element(By.XPATH, var.huenguyen).text
        logging.info("Trang cá nhân - bạn bè - chỉnh sửa quyền riêng tư - Công khai")
        logging.info("check font-end: login bằng tài khoản khác(chưa kết bạn) và check  - xem được bạn bè")
        logging.info(check_nguoila_congkhai == "hue nguyen")
        time.sleep(1.5)

        # Bạn bè
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        driver.get("https://cmc-fe.emso.vn/user/truongvck33")
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.trangcanhan_banbe).click()
        driver.execute_script("window.scrollBy(0,510)", "")
        time.sleep(1)
        check_banbe_quyenbanbe = driver.find_element(By.XPATH, var.check_banbe_quyenbanbe).text
        print(check_banbe_quyenbanbe)
        logging.info("Trang cá nhân - bạn bè - chỉnh sửa quyền riêng tư - Công khai")
        logging.info("check font-end: login bằng tài khoản khác(bạn bè) và check  - xem được bạn bè")
        logging.info(check_banbe_quyenbanbe == "hue nguyen\n1 bạn chung")
        time.sleep(1.5)


class anh_video:
    def anh(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_anh).click()
        driver.execute_script("window.scrollBy(0,700)", "")
        #ảnh
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_anh_xemanh).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_thich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_binhluan).send_keys(data['trangcanhan_anh_video']['trangcanhan_anh_binhluan'])
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_binhluan).submit()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_anh_xemanh_x).click()

        #album
        driver.find_element(By.XPATH, var.trangcanhan_anh_album).click()
        driver.find_element(By.XPATH, var.trangcanhan_anh_alum_taomoi).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_ten).send_keys(data['trangcanhan_anh_video']['trangcanhan_anh_anh'])
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_mota).send_keys(data['trangcanhan_anh_video']['trangcanhan_anh_mota'])
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_tailen).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia1.exe")
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_anhtailen_chuthich).send_keys(data['trangcanhan_anh_video']['chuthich'])
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_dang).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_iconxoa).click()
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_xoa).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_xoa).click()
        time.sleep(1)

    def video(self):
        #video
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.execute_script("window.scrollBy(0,300)", "")
        driver.find_element(By.XPATH, var.trangcanhan_video).click()
        driver.find_element(By.XPATH, var.trangcanhan_video_xem).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_thich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_binhluan).send_keys(data['trangcanhan_anh_video']['trangcanhan_video_binhluan'])
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_binhluan).submit()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_video_xemvideo_x).click()

        #khoanh khắc
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac).click()
        driver.find_element(By.XPATH, var.trangcanhan_video_xem).click()
        time.sleep(3)
        # driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_like).click()
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_thich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_binhluan).send_keys(data['trangcanhan_anh_video']['trangcanhan_video_binhluan1'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_binhluan).submit()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_x).click()
        del driver.requests
        time.sleep(2)


class check_thongtin_trangcanhan():
    def check_thongtin_trangcanhan(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        #giới thiệu(tổng quan, cong việc, nơi sống...)
        # driver.refresh()
        # time.sleep(3)
        try:
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
            tongquan_songtai1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai1).text
            tongquan_dentu1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu1).text
        except:
            driver.refresh()
            time.sleep(3)
            driver.find_element(By.XPATH, var.trangcanhan_gioithieu).click()
            tongquan_songtai1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_songtai1).text
            tongquan_dentu1 = driver.find_element(By.XPATH, var.trangcanhan_gioithieu_dentu1).text

        tongquan_tinhtrang_mqh1 = driver.find_element(By.XPATH, var.tongquan_tinhtrang_mqh1).text
        tongquan_nguoithan_mqh1 = driver.find_element(By.XPATH, var.tongquan_nguoithan_mqh1).text
        tongquan_sodienthoai1 = driver.find_element(By.XPATH, var.tongquan_sodienthoai1).text
        tongquan_bietdanh1 = driver.find_element(By.XPATH, var.tongquan_bietdanh1).text
        time.sleep(0.5)

        #công việc và học vấn
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_congviecvahocvan).click()
        cvvahocvan_cv_chucvu1 = driver.find_element(By.XPATH, var.cvvahocvan_cv_chucvu1).text
        cvvahocvan_cv_congty1 = driver.find_element(By.XPATH, var.cvvahocvan_cv_chucvu1).text
        cvvahocvan_daihoc_truonghoc1 = driver.find_element(By.XPATH, var.cvvahocvan_daihoc_truonghoc1).text
        cvvahocvan_trunghoc_truonghoc1 = driver.find_element(By.XPATH, var.cvvahocvan_trunghoc_truonghoc1).text
        time.sleep(0.5)

        # nơi từng sống
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_noitungsong).click()
        noitungsong_songtai1 = driver.find_element(By.XPATH, var.noitungsong_songtai1).text
        noitungsong_dentu1 = driver.find_element(By.XPATH, var.noitungsong_dentu1).text
        time.sleep(0.5)

        # thông tin cơ bản
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_thongtincoban).click()
        thongtincoban_sodienthoai1 = driver.find_element(By.XPATH, var.thongtincoban_sodienthoai1).text
        thongtincoban_web1 = driver.find_element(By.XPATH, var.thongtincoban_web1).text
        thongtincoban_lienket1 = driver.find_element(By.XPATH, var.thongtincoban_lienket1).text
        thongtincoban_tieusu1 = driver.find_element(By.XPATH, var.thongtincoban_tieusu1).text
        thongtincoban_bietdanh1 = driver.find_element(By.XPATH, var.thongtincoban_bietdanh1).text
        time.sleep(0.5)

        # Gia đình và các mối quan hệ
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_gdvacacmqh).click()
        tongquan_gdvacacmqh_tingtrang_mqh1 = driver.find_element(By.XPATH, var.tongquan_gdvacacmqh_tingtrang_mqh1).text
        tongquan_gdvacacmqh_nguoithan_mqh1 = driver.find_element(By.XPATH, var.tongquan_gdvacacmqh_nguoithan_mqh1).text
        time.sleep(0.5)

        # Sự kiện trong đời
        driver.find_element(By.XPATH, var.trangcanhan_gioithieu_sukientrongdoi).click()
        tongquan_sukientrongdoi_trunghoc_tentruong1 = driver.find_element(By.XPATH, var.tongquan_sukientrongdoi_trunghoc_tentruong1).text
        tongquan_sukientrongdoi_daihoc_tentruong1 = driver.find_element(By.XPATH, var.tongquan_sukientrongdoi_daihoc_tentruong1).text
        tongquan_sukientrongdoi_congviec_congty1 = driver.find_element(By.XPATH, var.tongquan_sukientrongdoi_congviec_congty1).text
        time.sleep(0.5)

        # GIỚI THIỆU(trang cá nhân)
        driver.find_element(By.XPATH, var.trangcanhan_baiviet).click()
        bietdanh2 = driver.find_element(By.XPATH, var.bietdanh2).text   #
        tieusu2 = driver.find_element(By.XPATH, var.tieusu2).text   #Mãi chả xong zzz
        congviec_chucvu2 = driver.find_element(By.XPATH, var.congviec_chucvu2).text  #Nhân Viên
        congviec_congty2 = driver.find_element(By.XPATH, var.congviec_chucvu2).text     #Cafe+
        daihoc_truonghoc2 = driver.find_element(By.XPATH, var.daihoc_truonghoc2).text       #Đại học bách khoa hà nội
        trunghoc_truonghoc2 = driver.find_element(By.XPATH, var.trunghoc_truonghoc2).text       #Trường THPT Phú Bình ( Phú Bình High School )
        songtai2 = driver.find_element(By.XPATH, var.songtai2).text     #Lạng Sơn
        dentu2 = driver.find_element(By.XPATH, var.dentu2).text     #Hưng Yên
        moiquanhe_tinhtrang2 = driver.find_element(By.XPATH, var.moiquanhe_tinhtrang2).text     #Hẹn hò
        moiquanhe_nguoithan2 = driver.find_element(By.XPATH, var.moiquanhe_nguoithan2).text     #Ngọc Mai
        web2 = driver.find_element(By.XPATH, var.web2).text     #https://plusplus.vn/
        lienket2 = driver.find_element(By.XPATH, var.lienket2).text     #https://pypi.org/
        sothich2 = driver.find_element(By.XPATH, var.sothich2).text
        time.sleep(0.5)
        # del driver.requests

        # Chỉnh sửa trang cá nhân
        driver.find_element(By.XPATH, var.trangcanhan_chinhsuatrangcanhan).click()
        time.sleep(1)
        tieusu3 = driver.find_element(By.XPATH, var.tieusu3).text   #Mãi chả xong zzz
        congviec_chucvu3 = driver.find_element(By.XPATH, var.congviec_chucvu3).text  #Nhân Viên
        congviec_congty3 = driver.find_element(By.XPATH, var.congviec_chucvu3).text     #Cafe+
        daihoc_truonghoc3 = driver.find_element(By.XPATH, var.daihoc_truonghoc3).text       #Đại học bách khoa hà nội
        trunghoc_truonghoc3 = driver.find_element(By.XPATH, var.trunghoc_truonghoc3).text       #Trường THPT Phú Bình ( Phú Bình High School )
        songtai3 = driver.find_element(By.XPATH, var.songtai3).text     #Lạng Sơn
        dentu3 = driver.find_element(By.XPATH, var.dentu3).text     #Hưng Yên
        moiquanhe_tinhtrang3 = driver.find_element(By.XPATH, var.moiquanhe_tinhtrang3).text     #Hẹn hò
        moiquanhe_nguoithan3 = driver.find_element(By.XPATH, var.moiquanhe_tinhtrang3).text     #Ngọc Mai
        sodienthoai3 = driver.find_element(By.XPATH, var.sodienthoai3).text     #Số diẹn thoại
        bietdanh3 = driver.find_element(By.XPATH, var.bietdanh3).text     #Bit danh
        web3 = driver.find_element(By.XPATH, var.web3).text     #https://plusplus.vn/
        lienket3 = driver.find_element(By.XPATH, var.lienket3).text     #https://pypi.org/
        sothich3 = driver.find_element(By.XPATH, var.sothich3).text
        driver.refresh()
        time.sleep(2)
        for request in driver.requests:
            if request.url == "https://snapi.emso.asia/api/v1/accounts/111169896815147328/abouts":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)

                #tiểu sử
                logging.info("check back-end, font-end trường: Trang cá nhân - Tiểu sử ")
                logging.info("respone: " + res['general_information']['description'])
                logging.info("Giới thiêu - thông tin cơ bản: " + thongtincoban_tieusu1)
                logging.info("Giới thiệu - Trang cá nhân: " + tieusu2)
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + tieusu3)
                logging.info(res['general_information']['description'] == thongtincoban_tieusu1 == tieusu2 == tieusu3)

                #sống tại
                logging.info("check back-end, font-end trường: Trang cá nhân - Sống tai ")
                logging.info("respone: " + res['general_information']['place_live']['title'])
                logging.info("Giới thiêu - Tổng quan: " + tongquan_songtai1)
                logging.info("Giới thiêu - Nơi từng sống: " + noitungsong_songtai1)
                logging.info("Giới thiệu - Trang cá nhân: " + songtai2[9::])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + songtai3[9::])
                logging.info(res['general_information']['place_live']['title'] == tongquan_songtai1 == noitungsong_songtai1 == songtai2[9::] == songtai3[9::])

                #đến từ
                logging.info("check back-end, font-end trường: Trang cá nhân - Đến từ ")
                logging.info("respone: " + res['general_information']['hometown']['title'])
                logging.info("Giới thiêu - Tổng quan: " + tongquan_dentu1)
                logging.info("Giới thiêu - Nơi từng sống: " + noitungsong_dentu1)
                logging.info("Giới thiệu - Trang cá nhân: " + dentu2[7::])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + dentu3[7::])
                logging.info(res['general_information']['hometown']['title'] == tongquan_dentu1 == noitungsong_dentu1 == dentu2[7::] == dentu3[7::])

                #Mối quan hệ
                #tình trạng
                logging.info("check back-end, font-end trường: Trang cá nhân - Mối quan hệ - Tình trạng ")
                logging.info("respone: " + res['account_relationship']['relationship_category']['name'])
                logging.info("Giới thiêu - Tổng quan: " + tongquan_tinhtrang_mqh1[0:6])
                logging.info("Giới thiêu - Gia đình và các mối quan hệ: " + tongquan_gdvacacmqh_tingtrang_mqh1[0:6])
                logging.info("Giới thiệu - Trang cá nhân: " + moiquanhe_tinhtrang2[0:6])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + moiquanhe_tinhtrang3[0:6])
                logging.info(res['account_relationship']['relationship_category']['name'] == tongquan_tinhtrang_mqh1[0:6] == tongquan_gdvacacmqh_tingtrang_mqh1[0:6] == moiquanhe_tinhtrang2[0:6] == moiquanhe_tinhtrang3[0:6])

                #Người thân
                logging.info("check back-end, font-end trường: Trang cá nhân - Mối quan hệ - Người thân ")
                logging.info("respone: " + res['account_relationship']['partner']['display_name'])
                logging.info("Giới thiêu - Tổng quan: " + tongquan_nguoithan_mqh1)
                logging.info("Giới thiêu - Gia đình và các mối quan hệ: " + tongquan_gdvacacmqh_nguoithan_mqh1)
                logging.info("Giới thiệu - Trang cá nhân: " + moiquanhe_nguoithan2[17:25])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + moiquanhe_nguoithan3[16:24])
                logging.info(res['account_relationship']['partner']['display_name'] == tongquan_nguoithan_mqh1 == tongquan_gdvacacmqh_nguoithan_mqh1 == moiquanhe_nguoithan2[17:25] == moiquanhe_nguoithan3[16:24])

                #Số điện thoại
                logging.info("check back-end, font-end trường: Trang cá nhân - Số điện thoai ")
                logging.info("respone: " + res['general_information']['phone_number'])
                logging.info("Giới thiêu - Tổng quan: " + tongquan_sodienthoai1)
                logging.info("Giới thiệu - Thông tin cơ bản: " + thongtincoban_sodienthoai1)
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + sodienthoai3)
                logging.info(res['general_information']['phone_number'] == tongquan_sodienthoai1 == thongtincoban_sodienthoai1 == sodienthoai3)

                #biêt danh
                logging.info("check back-end, font-end trường: Trang cá nhân - Biệt danh ")
                logging.info("respone: " + res['general_information']['other_name'])
                logging.info("Giới thiêu - Tổng quan: " + tongquan_bietdanh1)
                logging.info("Giới thiêu - Thông tin cơ bản: " + thongtincoban_bietdanh1)
                logging.info("Giới thiệu - Trang cá nhân: " + bietdanh2[1:13])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + bietdanh3)
                logging.info(res['general_information']['other_name'] == tongquan_bietdanh1 == thongtincoban_bietdanh1 == bietdanh2[1:13] == bietdanh3)

                #web
                logging.info("check back-end, font-end trường: Trang cá nhân - Web ")
                logging.info("respone: " + res['general_information']['account_web_link'][0]['url'])
                logging.info("Giới thiêu - thông tin cơ bản: " + thongtincoban_web1)
                logging.info("Giới thiệu - Trang cá nhân: " + web2[::])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + web3[::])
                logging.info(res['general_information']['account_web_link'][0]['url'] == web2[::] == thongtincoban_web1 == web3[::])

                #lien kết
                logging.info("check back-end, font-end trường: Trang cá nhân - Liên kết ")
                logging.info("respone: " + res['general_information']['account_social'][0]['text'])
                logging.info("Giới thiêu - thông tin cơ bản: " + thongtincoban_lienket1)
                logging.info("Giới thiệu - Trang cá nhân: " + lienket2[::])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + lienket3[::])
                logging.info(res['general_information']['account_social'][0]['text'] == thongtincoban_lienket1 == lienket2[::] == lienket3[::])

                #Sở thích
                logging.info("check back-end, font-end trường: Trang cá nhân - Sở thích ")
                logging.info("respone: " + res['hobbies'][0]['text'])
                logging.info("Giới thiệu - Trang cá nhân: " + sothich2)
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + sothich3)
                logging.info(res['hobbies'][0]['text'] == sothich2 == sothich3)

                break
            else:
                pass
                # print("không có  response api abouts")
        time.sleep(0.5)
        for request in driver.requests:
            if request.url[0:70] == "https://snapi.emso.asia/api/v1/accounts/111169896815147328/life_events":
                data1 = sw_decode(request.response.body, request.response.headers.get('Content-Encoding', 'identity'))
                data1 = data1.decode("utf8")
                res = json.loads(data1)
                # print("r1")
                # print(res[0]['life_event'])
                # print("r2")
                # print(res[1]['life_event'])
                # print("r3")
                # print(res[2]['life_event'])
                # print("r4")
                # print(res[3]['life_event'])
                # print("r5")
                # print(res[4]['life_event'])

                #công viêc - công ty
                logging.info("check back-end, font-end trường: Trang cá nhân - Công ty ")
                logging.info("respone: " + res[0]['life_event']['company'])
                logging.info("Giới thiêu - công việc và học vấn: " + cvvahocvan_cv_congty1[12::])
                logging.info("Giới thiêu - Sự kiên trong đời: " + tongquan_sukientrongdoi_congviec_congty1[4::])
                logging.info("Giới thiệu - Trang cá nhân: " + congviec_congty2[14::])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + congviec_congty3[14::])
                logging.info(res[0]['life_event']['company'] == cvvahocvan_cv_congty1[12::] == tongquan_sukientrongdoi_congviec_congty1[4::] == congviec_congty2[14::] == congviec_congty3[14::])

                #công viêc - chức vu
                logging.info("check back-end, font-end trường: Trang cá nhân - Chức vụ ")
                logging.info("respone: " + res[0]['life_event']['position'])
                logging.info("Giới thiêu - công việc và học vấn: " + cvvahocvan_cv_chucvu1[0:9])
                logging.info("Giới thiệu - Trang cá nhân: " + congviec_chucvu2[0:9])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + congviec_chucvu3[0:9])
                logging.info(res[0]['life_event']['position'] == cvvahocvan_cv_chucvu1[0:9] == congviec_chucvu2[0:9] == congviec_chucvu3[0:9])

                #đai hoc
                logging.info("check back-end, font-end trường: Trang cá nhân - Trường đại học ")
                logging.info("respone: " + res[1]['life_event']['company'])
                logging.info("Giới thiêu - công việc và học vấn: " + cvvahocvan_daihoc_truonghoc1[4::])
                logging.info("Giới thiêu - Sự kiên trong đời: " + tongquan_sukientrongdoi_daihoc_tentruong1[4::])
                logging.info("Giới thiệu - Trang cá nhân: " + daihoc_truonghoc2[13::])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + daihoc_truonghoc3[13::])
                logging.info(res[1]['life_event']['company'] == cvvahocvan_daihoc_truonghoc1[4::] == tongquan_sukientrongdoi_daihoc_tentruong1[4::] == daihoc_truonghoc2[13::] == daihoc_truonghoc3[13::])

                #trung hoc
                logging.info("check back-end, font-end trường: Trang cá nhân - Trường trung học ")
                logging.info("respone: " + res[2]['life_event']['company'])
                logging.info("Giới thiêu - công việc và học vấn: " + cvvahocvan_trunghoc_truonghoc1)
                logging.info("Giới thiêu - Sự kiện trong đời: " + tongquan_sukientrongdoi_trunghoc_tentruong1[4::])
                logging.info("Giới thiệu - Trang cá nhân: " + trunghoc_truonghoc2[11::])
                logging.info("Giới thiệu - Chỉnh sửa trang cá nhân: " + trunghoc_truonghoc3[11::])
                logging.info(res[2]['life_event']['company'] == cvvahocvan_trunghoc_truonghoc1 == tongquan_sukientrongdoi_trunghoc_tentruong1[4::] == trunghoc_truonghoc2[11::] == trunghoc_truonghoc3[11::])

                break
            else:
                pass
                # print("không có  response api life events")

    def phandangchuy(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.baiviet).click()
        cuon = driver.find_element(By.XPATH, var.chinhsuaphandangchuy)
        driver.execute_script("arguments[0].scrollIntoView();", cuon)

        button = driver.find_element(By.XPATH, var.chinhsuaphandangchuy)
        driver.execute_script("arguments[0].click();", button)
        # driver.find_element(By.XPATH, var.chinhsuaphandangchuy).click()
        driver.find_element(By.XPATH, var.chinhsuaphandangchuy_themmoi).click()
        driver.find_element(By.XPATH, var.chinhsuaphandangchuy_chonanh1).click()
        driver.find_element(By.XPATH, var.chinhsuaphandangchuy_tiep).click()
        time.sleep(1)
        #đặt tên
        driver.find_element(By.XPATH, var.chinhsuaphandangchuy_input).click()
        xoa = driver.find_element(By.XPATH, var.chinhsuaphandangchuy_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuaphandangchuy_input).send_keys("Test mục đángchuy")
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)
        #Xóa mục đang chú ý
        button = driver.find_element(By.XPATH, var.chinhsuaphandangchuy)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.chinhsuaphandangchuy_chonmuc1).click()
        driver.find_element(By.XPATH, var.chinhsuaphandangchuy_xoabosuutap).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1.5)



class trangchu():
    def taobaiviet(self, trangthai, mota, camxuc, hoatdong, tinhtrang):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.taobaiviet_anhvideo).click()
        driver.find_element(By.XPATH, var.taobaiviet_trangthai_icon).click()
        # driver.find_element(By.XPATH, var.taobaiviet_trangthai_riengtu).click()
        driver.find_element(By.XPATH, trangthai).click()
        driver.find_element(By.XPATH, var.taobaiviet_mota).send_keys(mota)
        driver.find_element(By.XPATH, var.taobaiviet_tailenanhvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia1.exe")
        time.sleep(2)
        driver.find_element(By.XPATH, var.chinhsua).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsua1).click()
        driver.find_element(By.XPATH, var.chinhsua_cat).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsua_xoay).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsua_congcuchenvanvan).click()
        driver.find_element(By.XPATH, var.chinhsua_congcuchenvanvan_themvanban).click()
        driver.find_element(By.XPATH, var.chinhsua_nhapvanvan).send_keys(data['trangchu_taobaiviet']['chinhsua_nhapvanvan'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsua_luu).click()
        # driver.find_element(By.XPATH, var.huy).click()
        # button = driver.find_element(By.XPATH, var.themanhvideo1)
        # driver.execute_script("arguments[0].click();", button)
        time.sleep(10)
        driver.find_element(By.XPATH, var.themanhvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia1.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsua).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsua_chuthich1).send_keys(data['trangchu_taobaiviet']['chinhsua_chuthich1'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsua_chuthich2).click()
        driver.find_element(By.XPATH, var.chinhsua_chuthich2).send_keys(data['trangchu_taobaiviet']['chinhsua_chuthich2'])
        driver.find_element(By.XPATH, var.chinhsua_xong).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.taobaiviet_camxuchoatdong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taobaiviet_camxuchoatdong_camxuc).click()
        time.sleep(1)
        driver.find_element(By.XPATH, camxuc).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taobaiviet_camxuchoatdong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taobaiviet_hoatdong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, hoatdong).click()
        # driver.find_element(By.XPATH, var.hoatdong_x).click()
        button = driver.find_element(By.XPATH, var.hoatdong_x)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.taobaiviet_ganthenguoikhac).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.taobaiviet_ngocmai).click()
        driver.find_element(By.XPATH, var.xong).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(1)
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.messgae_taobaiviet1)))
        print(element.text)
        logging.info(element.text)
        check_taobaiviet = driver.find_element(By.XPATH, var.check_taobaiviet2).text
        logging.info("Trang chủ - Tạo bài viết ")
        logging.info("Tạo bài viết - Mô tả:  Đây là bai viết "+ tinhtrang)
        logging.info("check font-end: Bài viết đang để quyền " + tinhtrang)
        logging.info(check_taobaiviet == "Đây là bai viết "+ tinhtrang)
        time.sleep(1)

    def taobaimoment(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        driver.find_element(By.XPATH, var.khoanhkhac).click()
        driver.find_element(By.XPATH, var.khoanhkhac_nhapnoidung).send_keys(data['trangchu_taobaiviet']['khoanhkhac_noidung'])
        driver.find_element(By.XPATH, var.khoanhkhac_tailenvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/quangaygiongbao.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.khoanhkhac_dangbai).click()
        time.sleep(23)

    def camxuc_hoatdong(self):
        driver.implicitly_wait(15)
        time.sleep(2)
        driver.find_element(By.XPATH, var.camxuc_hoatdong).click()
        driver.find_element(By.XPATH, var.camxuc).click()
        driver.find_element(By.XPATH, var.camxuc_xucdong).click()   #mất icon xúc động
        time.sleep(1)
        check_taobaiviet_camxuc = driver.find_element(By.XPATH,var.check_taobaiviet_camxuc1).text
        logging.info("Trang chủ - Cảm xúc/Hoạt động")
        logging.info("check font-end: đang cảm thấy xúc động ")
        logging.info(check_taobaiviet_camxuc)
        logging.info(check_taobaiviet_camxuc =="đang  cảm thấy xúc động ")
        driver.find_element(By.XPATH, var.x).click()

        driver.find_element(By.XPATH, var.camxuc_hoatdong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.camxuc_hoatdong_hoatdong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.hoatdong_dangxem).click()
        logging.info("Trang chủ - Cảm xúc/Hoạt động")
        logging.info("check font-end: Options của hoạt động")
        logging.info("Chức năng chưa hoạt động")


        driver.find_element(By.XPATH, var.x).click()
        time.sleep(1)
        check_taobaiviet_hoatdong = driver.find_element(By.XPATH,var.check_taobaiviet_hoatdong1).text
        logging.info("Trang chủ - Cảm xúc/Hoạt động")
        logging.info("check font-end: đang xem")
        logging.info(check_taobaiviet_hoatdong)
        logging.info(check_taobaiviet_hoatdong == "đang xem")
        driver.find_element(By.XPATH, var.x).click()

    def taobaiviet_congkhai(self):
        trangchu.taobaiviet(self, var.taobaiviet_trangthai_congkhai, "Đây là bai viết công khai", var.taobaiviet_camxuc_tuyet, var.taobaiviet_hoatdong_dangxem, "công khai" )

    def taobaiviet_banbe(self):
        trangchu.taobaiviet(self, var.taobaiviet_trangthai_banbe, "Đây là bai viết bạn bè", var.taobaiviet_camxuc_yeu, var.taobaiviet_hoatdong_dangnghive, "bạn bè" )

    def taobaiviet_riengtu(self):
        trangchu.taobaiviet(self, var.taobaiviet_trangthai_riengtu, "Đây là bai viết riêng tư", var.taobaiviet_camxuc_vuive2, var.taobaiviet_hoatdong_dangtim, "riêng tư" )

    def menu(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).clear()
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).send_keys(data['trangchu_menu']['timkiem'])
        driver.find_element(By.XPATH, var.trangchu_menu_banbe).click()
        time.sleep(2)
        check_menu_name1 = driver.find_element(By.XPATH,var.check_menu_name1).text
        logging.info("Trang chủ - Menu - Bạn bè")
        logging.info("check font-end: Tiêu đề: " + check_menu_name1)
        logging.info(check_menu_name1 == "Bạn bè")
        driver.refresh()
        time.sleep(2)

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).clear()
        driver.find_element(By.XPATH, var.trangchu_menu_nhom).click()
        check_menu_name2 = driver.find_element(By.XPATH,var.check_menu_name2).text
        logging.info("Trang chủ - Menu - Nhóm")
        logging.info("check font-end: Tiêu đề: " + check_menu_name2)
        logging.info(check_menu_name2 == "Nhóm")



        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_trang).click()
        check_menu_name3 = driver.find_element(By.XPATH,var.check_menu_name3).text
        logging.info("Trang chủ - Menu - Trang")
        logging.info("check font-end: Tiêu đề: " + check_menu_name3)
        logging.info(check_menu_name3 == "Trang")


        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_sukien).click()
        check_menu_name4 = driver.find_element(By.XPATH,var.check_menu_name4).text
        logging.info("Trang chủ - Menu - Sự kiện")
        logging.info("check font-end: Tiêu đề: " + check_menu_name4)
        logging.info(check_menu_name4 == "Sự kiện")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_goivon).click()
        check_menu_name5 = driver.find_element(By.XPATH,var.check_menu_name5).text
        logging.info("Trang chủ - Menu - Gọi vốn cộng đồng")
        logging.info("check font-end: Tiêu đề: " + check_menu_name5)
        logging.info(check_menu_name5 == "Gọi vốn cộng đồng")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_tuyendung).click()
        check_menu_name6 = driver.find_element(By.XPATH,var.check_menu_name6).text
        logging.info("Trang chủ - Menu - Tuyển dụng")
        logging.info("check font-end: Tiêu đề: " + check_menu_name6)
        logging.info(check_menu_name6 == "Tuyển dụng")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_khoahoc).click()
        check_menu_name7 = driver.find_element(By.XPATH,var.check_menu_name7).text
        logging.info("Trang chủ - Menu - Khoá học")
        logging.info("check font-end: Tiêu đề: " + check_menu_name7)
        logging.info(check_menu_name7 == "Khóa học")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_hienmau).click()
        check_menu_name8 = driver.find_element(By.XPATH,var.check_menu_name8).text
        logging.info("Trang chủ - Menu - Hiến máu")
        logging.info("check font-end: Tiêu đề: " + check_menu_name8)
        logging.info(check_menu_name8 == "Hiến máu")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).send_keys("Không gian âm nhạc")
        driver.find_element(By.XPATH, var.trangchu_menu_khonggianamnhac).click()
        check_menu_name9 = driver.find_element(By.XPATH,var.check_menu_name9).text
        logging.info("Trang chủ - Menu - Không gian âm nhạc")
        logging.info("check font-end: Tiêu đề: " + check_menu_name9)
        logging.info(check_menu_name9 == "Không gian Âm nhạc")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        time.sleep(1.5)
        xoa = driver.find_element(By.XPATH, var.trangchu_timkiem_menu)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).send_keys("Trò chuyện")
        driver.find_element(By.XPATH, var.trangchu_menu_trochuyen).click()
        check_menu_name10 = driver.find_element(By.XPATH,var.check_menu_name10).text
        logging.info("Trang chủ - Menu - Trò chuyện")
        logging.info("check font-end: Tiêu đề: " + check_menu_name10)
        logging.info(check_menu_name10 == "Chat")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trangchu_timkiem_menu)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).send_keys("Khoảnh khắc")
        driver.find_element(By.XPATH, var.trangchu_menu_khoanhkhac).click()
        check_menu_name11 = driver.find_element(By.XPATH,var.check_menu_name11).text
        logging.info("Trang chủ - Menu - Khoảnh khắc")
        logging.info("check font-end: Tiêu đề: " + check_menu_name11)
        logging.info(check_menu_name11 == "Khoảnh khắc")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trangchu_timkiem_menu)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).send_keys("Watch")
        driver.find_element(By.XPATH, var.trangchu_menu_watch).click()
        check_menu_name12 = driver.find_element(By.XPATH,var.check_menu_name12).text
        logging.info("Trang chủ - Menu - Watch")
        logging.info("check font-end: Tiêu đề: " + check_menu_name12)
        logging.info(check_menu_name12 == "Watch")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trangchu_timkiem_menu)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).send_keys("Không gian thương mại")
        driver.find_element(By.XPATH, var.trangchu_menu_khonggianthuongmai).click()
        check_menu_name13 = driver.find_element(By.XPATH,var.check_menu_name13).get_attribute("aria-label")
        logging.info("Trang chủ - Menu - Không gian thương mại")
        logging.info("check font-end: Tiêu đề: " + driver.title)
        logging.info(check_menu_name13)
        logging.info(check_menu_name13 == "Shop của tôi")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trangchu_timkiem_menu)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).send_keys("Kỷ niệm")
        driver.find_element(By.XPATH, var.trangchu_menu_kyniem).click()
        check_menu_name14 = driver.find_element(By.XPATH,var.check_menu_name14).text
        logging.info(check_menu_name14)
        logging.info("Trang chủ - Menu - Kỷ niệm")
        logging.info("check font-end: Tiêu đề: " + check_menu_name14)
        logging.info(check_menu_name14 == "Kỷ niệm")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trangchu_timkiem_menu)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).send_keys("Đã lưu")
        driver.find_element(By.XPATH, var.trangchu_menu_daluu).click()
        check_menu_name15 = driver.find_element(By.XPATH,var.check_menu_name15).text
        print(check_menu_name15)
        logging.info("Trang chủ - Menu - Đã lưu")
        logging.info("check font-end: Tiêu đề: " + check_menu_name15)
        logging.info(check_menu_name15 == "Đã lưu")

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trangchu_timkiem_menu)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangchu_timkiem_menu).send_keys("Đơn đặt hàng và thanh toán")
        driver.find_element(By.XPATH, var.trangchu_menu_dondathangvathanhtoan).click()
        driver.find_element(By.XPATH, var.trangchu_menu_dondathangvathanhtoan_huy).click()
        check_menu_name16 = driver.find_element(By.XPATH,var.check_menu_name16).text
        logging.info("Trang chủ - Menu - Đơn đặt hàng và thanh toán")
        logging.info("check font-end: Tiêu đề: " + check_menu_name16)
        logging.info(check_menu_name16 == "Đơn đặt hàng và thanh toán")
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(1)
        driver.refresh()
        time.sleep(2)

    def menu_tao(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_tao_post).click()
        check_taobaiviet = driver.find_element(By.XPATH, var.check_taobaiviet1).text
        print(check_taobaiviet)
        logging.info("Trang chủ - Tạo bài viết - Post")
        logging.info("check font-end: chuyển tới POPUP: " + check_taobaiviet)
        logging.info(check_taobaiviet == "Tạo bài viết")
        driver.find_element(By.XPATH, var.taobaiviet_x).click()

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_tao_khoanhkhac).click()
        check_taobaimoment = driver.find_element(By.XPATH, var.check_taobaimoment1).text
        print(check_taobaimoment)
        logging.info("Trang chủ - Tạo bài viết - Khoảnh khắc")
        logging.info("check font-end: chuyển tới POPUP: " + check_taobaimoment)
        logging.info(check_taobaimoment == "Tạo bài khoảnh khắc")
        driver.find_element(By.XPATH, var.check_taobaimoment_x).click()

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_tao_trang).click()
        check_taotrang = driver.find_element(By.XPATH, var.check_taotrang1).text
        print(check_taotrang)
        logging.info("Trang chủ - Tạo bài viết - Trang")
        logging.info("check font-end: chuyển tới POPUP: " + check_taotrang)
        logging.info(check_taotrang == "Tạo Trang")
        driver.find_element(By.XPATH, var.check_taotrang_x).click()

        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_tao_nhom).click()
        check_taonhom = driver.find_element(By.XPATH, var.check_taonhom1).text
        print(check_taonhom)
        logging.info("Trang chủ - Tạo bài viết - Nhóm")
        logging.info("check font-end: chuyển tới POPUP: " + check_taonhom)
        logging.info(check_taonhom == "Tạo nhóm")
        driver.find_element(By.XPATH, var.check_taonhom_x).click()


        driver.find_element(By.XPATH, var.trangchu_menu).click()
        driver.find_element(By.XPATH, var.trangchu_menu_tao_sukien).click()
        check_taosukien = driver.find_element(By.XPATH, var.check_taosukien1).text
        print(check_taosukien)
        logging.info("Trang chủ - Tạo bài viết - Sự kiện")
        logging.info("check font-end: chuyển tới POPUP: " + check_taonhom)
        logging.info(check_taosukien == "Tạo sự kiện")
        driver.find_element(By.XPATH, var.check_taosukien_x).click()
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(2)

    def tinnhanmoi(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_den_input).send_keys(data['trangchu_tinnhan']['tinnhanmoi_den'])
        time.sleep(3)
        try:
            driver.find_element(By.XPATH, var.tinnhanmoi_ngocmai).click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trangchu_tinnhanmoi).click()
            driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_den_input).send_keys(data['trangchu_tinnhan']['tinnhanmoi_den'])
            time.sleep(3)
            driver.find_element(By.XPATH, var.tinnhanmoi_ngocmai).click()

        #Ghi âm
        # driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icondaucong).click()        ko aloww dc ghi am
        # driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icondaucong_guiamthanh).click()
        # driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icondaucong_guiamthanh_x).click()
        # driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icondaucong).click()
        # driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icondaucong_guiamthanh).click()
        # time.sleep(2)
        # driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icondaucong_guiamthanh_ghiam).click()
        # time.sleep(1)
        #
        # # chrome_opt = Options()
        # # chrome_opt.add_argument("--incognito")
        # # chrome_opt.add_argument("--enable-infobars")
        # # chrome_opt.add_argument("start-maximized")
        # # chrome_opt.add_argument("--enable-extensions")
        # #
        # # chrome_opt.add_argument("use-fake-device-for-media-stream")  # this argument for accepting permissions
        # # chrome_opt.add_argument("use-fake-ui-for-media-stream")
        #
        # # options.add_argument(r'--user-data-dir=C:\Users\smart\AppData\Local\Google\Chrome\User Data\Default.')
        # # options.set_preference("media.navigator.permission.disabled", True)
        # # options.add_argument('use-fake-device-for-media-stream')
        # # options.add_option("prefs", {
        # #     "profile.default_content_setting_values.media_stream_mic": 1
        # # })
        # # chrome_options.add_argument("--use-fake-ui-for-media-stream")
        # # chrome_options.add_experimental_option('prefs', {'profile.default_content_setting_values.media_stream_mic': 1})
        # time.sleep(2)
        # driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icondaucong_guiamthanh_gui).click()
        #

        #Đính kèm file
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icontailenfile).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien1.exe")
        time.sleep(1)

        #Gif
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icongif).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icongif_input).send_keys(data['trangchu_tinnhan']['tinnhanmoi_meo'])
        time.sleep(5)

        try:
            driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icongif_anh1a).click()
        except:
            driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icongif_anh1).click()
            # driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icongif_anh4).click()
        #Nhãn dán
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan_2).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan_3).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan_4).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan_1).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan_1_chonnhandan).click()

        #input
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_input).send_keys(data['trangchu_tinnhan']['tinnhanmoi_input'])
        #Biểu tượng cảm xúc
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page2).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page3).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page4).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page5).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page6).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page7).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page8).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page9).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_lichsu).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_chon).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_input).send_keys(Keys.ENTER)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc).click()
        #Like
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_like).click()
        # button = driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_like)
        # driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_input).send_keys(Keys.ENTER)
        time.sleep(10)

    def chat(self):
        driver.implicitly_wait(5)
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/user/truongvck33")
        time.sleep(3)
        driver.find_element(By.XPATH, var.trangchu_chat).click()
        time.sleep(1)
        try:
            driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trangchu_chat).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        time.sleep(1)
        check_trangchu_chat_tuychon = driver.find_element(By.XPATH,var.check_trangchu_chat_tuychon1).text
        print(check_trangchu_chat_tuychon)
        logging.info("Trang chủ - Chat - Tuỳ chọn")
        logging.info("check font-end: Cài đặt chat")
        logging.info(check_trangchu_chat_tuychon == "Cài đặt Chat")

        #Âm thanh cuộc gọi đến
        amthanhgoiden_tat = driver.find_element(By.XPATH, var.trangchu_chat_tuychon_amthanhgoiden_tat).text
        logging.info("Trang chủ - Chat - Tuỳ chọn - Âm thanh cuộc gọi đến")
        logging.info("check font-end: Âm thanh cuộc gọi đến: đang tắt")
        logging.info(amthanhgoiden_tat == "Âm thanh cuộc gọi đến: đang tắt")
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon_amthanhgoiden_tat).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        time.sleep(1)
        amthanhgoiden_bat = driver.find_element(By.XPATH, var.trangchu_chat_tuychon_amthanhgoiden_bat).text
        logging.info("Trang chủ - Chat - Tuỳ chọn - Âm thanh cuộc gọi đến")
        logging.info("check font-end: Âm thanh cuộc gọi đến: đang bật")
        logging.info(amthanhgoiden_bat == "Âm thanh cuộc gọi đến: đang bật")
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon_amthanhgoiden_bat).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        time.sleep(1)

        #Âm thanh tin nhắn
        amthanhgoiden_tat = driver.find_element(By.XPATH, var.trangchu_chat_tuychon_amthanhtinnhan_tat).text
        logging.info("Trang chủ - Chat - Tuỳ chọn - Âm thanh tin nhắn")
        logging.info("check font-end: Âm thanh tin nhắn: đang tắt")
        logging.info(amthanhgoiden_tat == "Âm thanh tin nhắn: đang tắt")
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon_amthanhtinnhan_tat).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        time.sleep(1)
        amthanhgoiden_bat = driver.find_element(By.XPATH, var.trangchu_chat_tuychon_amthanhtinnhan_bat).text
        logging.info("Trang chủ - Chat - Tuỳ chọn - Âm thanh tin nhắn")
        logging.info("check font-end: Âm thanh tin nhắn: đang bật")
        logging.info(amthanhgoiden_bat == "Âm thanh tin nhắn: đang bật")
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon_amthanhtinnhan_bat).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        time.sleep(1)
        try:
            tudongmotinnhan_tat = driver.find_element(By.XPATH, var.trangchu_chat_tuychon_tudongmotinnhan_tat).text
        except:
            driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
            tudongmotinnhan_tat = driver.find_element(By.XPATH, var.trangchu_chat_tuychon_tudongmotinnhan_tat).text
        logging.info("Trang chủ - Chat - Tuỳ chọn - Tự động mở tin nhắn")
        logging.info("check font-end: Tự động mở tin nhắn mới: đang tắt")
        logging.info(tudongmotinnhan_tat == "Tự động mở tin nhắn mới: đang tắt")
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon_tudongmotinnhan_tat).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        time.sleep(1)
        tudongmotinnhan_bat = driver.find_element(By.XPATH, var.trangchu_chat_tuychon_tudongmotinnhan_bat).text
        logging.info("Trang chủ - Chat - Tuỳ chọn - Tự động mở tin nhắn")
        logging.info("check font-end: Tự động mở tin nhắn mới: đang bật")
        logging.info(tudongmotinnhan_bat == "Tự động mở tin nhắn mới: đang bật")
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon_tudongmotinnhan_bat).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        time.sleep(1)

        trangthaihoatdong_tat = driver.find_element(By.XPATH, var.trangchu_chat_tuychon_trangthaihoatdong_tat).text  #bật/tắt load lại trang mới đổi
        logging.info("Trang chủ - Chat - Tuỳ chọn - Trạng thái hoạt động")
        logging.info("check font-end: Trạng thái hoạt động: đang tắt")
        logging.info(trangthaihoatdong_tat == "Trạng thái hoạt động: đang tắt")
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon_trangthaihoatdong_tat).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        time.sleep(1)
        trangthaihoatdong_bat = driver.find_element(By.XPATH, var.trangchu_chat_tuychon_trangthaihoatdong_bat).text
        logging.info("Trang chủ - Chat - Tuỳ chọn - Trạng thái hoạt động")
        logging.info("check font-end: Trạng thái hoạt động: đang bật")
        logging.info(trangthaihoatdong_bat == "Trạng thái hoạt động: đang bật")
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon_trangthaihoatdong_bat).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        time.sleep(1)

        driver.find_element(By.XPATH, var.trangchu_chat_tuychon_tinnhancho).click()
        time.sleep(1)
        check_tinnhandangcho = driver.find_element(By.XPATH, var.check_tinnhandangcho1).text
        check_tinnhancho = driver.find_element(By.XPATH, var.check_tinnhancho1).text
        logging.info("Trang chủ - Chat - Tuỳ chọn - Tin nhắn chờ")
        logging.info("check font-end: Tin nhắn đang chờ")
        logging.info(check_tinnhandangcho == "Tin nhắn đang chờ")
        logging.info("check font-end: Người gửi tin nhắn - Ni Mặc")
        logging.info(check_tinnhancho == "Ni Mặc")


        driver.find_element(By.XPATH, var.trangchu_chat).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu_chat).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        time.sleep(1)

        # driver.find_element(By.XPATH, var.trangchu_chat_tuychon_caidatguitinnhan).click() #Đã bỏ chức năng này
        # time.sleep(1)
        # logging.info("Trang chủ - Chat - Tuỳ chọn - Cài đặt gửi tin nhắn")
        # logging.info("Chức năng chưa hoạt động")
        # driver.find_element(By.XPATH, var.trangchu_chat_tuychon).click()
        # time.sleep(1)

        driver.find_element(By.XPATH, var.trangchu_chat_tuychon_caidatchan).click()
        logging.info("Trang chủ - Chat - Tuỳ chọn - Cài đặt chặn")
        logging.info("Chức năng chưa hoạt động")
        time.sleep(2)



    def chat_xemtatca(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_chat).click()
        driver.find_element(By.XPATH, var.trangchu_emsochat).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.emsochat_goi).click()
        driver.find_element(By.XPATH, var.emsochat_video).click()

        #gửi text
        driver.find_element(By.XPATH, var.emsochat_input).send_keys(data['trangchu_tinnhan']['input1'])
        driver.find_element(By.XPATH, var.emsochat_input).click()
        driver.find_element(By.XPATH, var.emsochat_input).send_keys(Keys.ENTER)
        time.sleep(2)

        #xoá, gỡ bỏ
        chat_hover = driver.find_element(By.XPATH, var.chat_hover1)
        actions = ActionChains(driver)
        try:
            actions.move_to_element(chat_hover).perform()
        except:
            actions.move_to_element(chat_hover).perform()
        time.sleep(1)
        xemthem = driver.find_element(By.XPATH, var.emsochat_input_iconxemthem)
        try:
            actions.move_to_element(xemthem).click().perform()
        except:
            actions.move_to_element(xemthem).click().perform()
        time.sleep(1)
        driver.find_element(By.XPATH, var.emsochat_iconxemthem_xoagobo).click()
        time.sleep(1.5)
        check_xoatinnhan = driver.find_element(By.XPATH,var.check_xoatinnhan1).text
        print(check_xoatinnhan)
        logging.info("Trang cá nhân - Chat - Xoá tin nhắn")
        logging.info("check font-end: message - Đã xóa tin nhắn thành công.")
        logging.info(check_xoatinnhan == "Đã xóa tin nhắn thành công.")
        driver.find_element(By.XPATH, var.emsochat_iconxemthem_message_x).click()

        #Chỉnh sửa text
        driver.find_element(By.XPATH, var.emsochat_input).send_keys(data['trangchu_tinnhan']['input'])
        driver.find_element(By.XPATH, var.emsochat_input).click()
        driver.find_element(By.XPATH, var.emsochat_input).send_keys(Keys.ENTER)
        time.sleep(1)
        actions = ActionChains(driver)
        chat_hover = driver.find_element(By.XPATH, var.chat_hover1)
        try:
            actions.move_to_element(chat_hover).perform()
        except:
            chat_hover = driver.find_element(By.XPATH, var.chat_hover1)
            actions.move_to_element(chat_hover).perform()
        time.sleep(1)
        xemthem = driver.find_element(By.XPATH, var.emsochat_input_iconxemthem)
        try:
            actions.move_to_element(xemthem).click().perform()
        except:
            actions.move_to_element(xemthem).click().perform()
        driver.find_element(By.XPATH, var.emsochat_iconxemthem_chinhsua).click()
        time.sleep(0.5)
        input_chat = driver.find_element(By.XPATH, var.emsochat_input)
        input_chat.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.emsochat_input).send_keys(data['trangchu_tinnhan']['input2'])
        driver.find_element(By.XPATH, var.emsochat_input).send_keys(Keys.ENTER)
        time.sleep(0.5)
        check_suatinnhan = driver.find_element(By.XPATH,var.check_suatinnhan1).text
        print(check_suatinnhan)
        logging.info("Trang cá nhân - Chat - Sửa tin nhắn")
        logging.info("check font-end: Sửa tin nhắn - alo alo 123")
        logging.info(check_suatinnhan == "alo alo 123")

        # Chuyển tiếp
        chat_hover = driver.find_element(By.XPATH, var.chat_hover1)
        try:
            actions.move_to_element(chat_hover).perform()
        except:
            actions.move_to_element(chat_hover).perform()
        xemthem = driver.find_element(By.XPATH, var.emsochat_input_iconxemthem)
        try:
            actions.move_to_element(xemthem).click().perform()
        except:
            actions.move_to_element(xemthem).click().perform()
        driver.find_element(By.XPATH, var.emsochat_iconxemthem_chuyentiep).click()
        time.sleep(1)
        check_chyentiep = driver.find_element(By.XPATH,var.check_chyentiep1).text
        print(check_chyentiep)
        logging.info("Trang cá nhân - Chat - Chuyển tiếp tin nhắn")
        logging.info("check font-end: đang ở popup - Chuyển tiếp ")
        logging.info(check_chyentiep == "Chuyển tiếp")
        driver.find_element(By.XPATH, var.emsochat_chuyentiep_timkiem).click()
        driver.find_element(By.XPATH, var.emsochat_chuyentiep_timkiem).send_keys(data['trangchu_tinnhan']['timkiem_huenguyen'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.emsochat_chuyentiep_timkiem_gui).click()
        time.sleep(2)
        check_chyentiep_dagui = driver.find_element(By.XPATH,var.check_chyentiep_dagui1).text
        print(check_chyentiep_dagui)
        logging.info("Trang cá nhân - Chat - Chuyển tiếp tin nhắn")
        logging.info("check font-end: tình trạng - Đã gửi ")
        logging.info(check_chyentiep_dagui == "Đã gửi")
        driver.find_element(By.XPATH, var.x).click()
        time.sleep(0.5)

        # # # Gim
        try:
            actions.move_to_element(chat_hover).perform()    #lỗi không ghim đuoc
            actions.move_to_element(xemthem).click().perform()
        except:
            actions.move_to_element(chat_hover).perform()
            actions.move_to_element(xemthem).click().perform()
        time.sleep(1)
        check_gimtinnhan = driver.find_element(By.XPATH,var.emsochat_iconxemthem_gim).text
        print(check_gimtinnhan)
        logging.info("Trang cá nhân - Chat - Ghim tin nhắn")
        logging.info("check font-end: Ghim")
        logging.info(check_gimtinnhan)
        logging.info(check_gimtinnhan == "Ghim")
        driver.find_element(By.XPATH, var.emsochat_iconxemthem_gim).click()
        time.sleep(1)
        # Bỏ gim
        try:
            actions.move_to_element(chat_hover).perform()
            actions.move_to_element(xemthem).click().perform()
        except:
            actions.move_to_element(chat_hover).perform()
            actions.move_to_element(xemthem).click().perform()
        time.sleep(1)
        check_bogimtinnhan = driver.find_element(By.XPATH,var.emsochat_iconxemthem_bogim).text
        print(check_bogimtinnhan)
        logging.info("Trang cá nhân - Chat - Bỏ Ghim tin nhắn")
        logging.info("check font-end: Bỏ ghim ")
        logging.info(check_bogimtinnhan == "Bỏ ghim")

        driver.find_element(By.XPATH, var.emsochat_iconxemthem_bogim).click()
        #Gim
        chat_hover4 = driver.find_element(By.XPATH, var.chat_hover4)
        actions.move_to_element(chat_hover4).perform()
        time.sleep(1)
        xemthem = driver.find_element(By.XPATH, var.emsochat_input_iconxemthem)
        actions = ActionChains(driver)
        actions.move_to_element(xemthem).click().perform()
        time.sleep(1)
        driver.find_element(By.XPATH, var.emsochat_iconxemthem_gim).click()
        time.sleep(1)
        actions = ActionChains(driver)
        #Trả lời
        try:
            actions.move_to_element(chat_hover).perform()
            traloi = driver.find_element(By.XPATH, var.emsochat_input_icontraloi)
        except:
            actions.move_to_element(chat_hover).perform()
            traloi = driver.find_element(By.XPATH, var.emsochat_input_icontraloi)
        try:
            actions.move_to_element(traloi).click().perform()
        except:
            actions.move_to_element(traloi).click().perform()
        driver.find_element(By.XPATH, var.emsochat_traloi).send_keys(data['trangchu_tinnhan']['traloi'])
        driver.find_element(By.XPATH, var.emsochat_traloi).send_keys(Keys.ENTER)
        time.sleep(1.5)
        check_traloitinnhan = driver.find_element(By.XPATH,var.check_traloitinnhan1).text
        print(check_traloitinnhan)
        logging.info("Trang cá nhân - Chat - Trả lời tin nhắn")
        logging.info("check font-end: Trả lời tin nhắn  - Bạn đã trả lời chính mình")
        logging.info(check_traloitinnhan == "Bạn đã trả lời chính mình")
        time.sleep(2)

        #Bày tỏ cảm xúc
        actions = ActionChains(driver)
        chat_hover3 = driver.find_element(By.XPATH, var.chat_hover3)        #Không trả lời tin nhắn được nên ko bày tỏ cảm xsuc được
        actions.move_to_element(chat_hover3).perform()
        print("r1")
        time.sleep(1)
        baytocamxuc = driver.find_element(By.XPATH, var.emsochat_input_iconbaytocamxuc)
        actions.move_to_element(baytocamxuc).perform()
        time.sleep(2)
        driver.find_element(By.XPATH, var.icon_yeuthich).click()
        time.sleep(1)

        #Icon ! dấu thăng hộp thoại
        #Trạng thái hoạt động
        driver.find_element(By.XPATH, var.emsochat_caidathopthoai).click()
        check_chat_tatthongbao = driver.find_element(By.XPATH,var.trangthaihoatdong_tatthongbao).text
        print(check_chat_tatthongbao)
        logging.info("Chat - tuỳ chon - Trạng thái hoạt động")
        logging.info("check font-end: Trạng thái hoạt động - Tắt thông báo")
        logging.info(check_chat_tatthongbao == "Tắt thông báo")
        time.sleep(1)
        driver.find_element(By.XPATH, var.emsochat_icontatthongbao).click()
        time.sleep(3)
        check_chat_batthongbao = driver.find_element(By.XPATH,var.trangthaihoatdong_batthongbao).text
        print(check_chat_batthongbao)
        logging.info("Chat - tuỳ chon - Trạng thái hoạt động")
        logging.info("check font-end: Trạng thái hoạt động - Bật thông báo")
        logging.info(check_chat_batthongbao == "Bật thông báo")
        # driver.find_element(By.XPATH, var.trangthaihoatdong_batthongbao).click()      #bật thông báo

        #Tìm kiếm
        driver.find_element(By.XPATH, var.emsochat_icontimkiem).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.emsochat_timkiem_input).click()
        driver.find_element(By.XPATH, var.emsochat_timkiem_input).send_keys(data['trangchu_tinnhan']['timkiem'])
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.emsochat_timkiem_input).click()
        driver.find_element(By.XPATH, var.emsochat_timkiem_input).send_keys(Keys.ENTER)
        time.sleep(1)
        check_chat_timkiem = driver.find_element(By.XPATH,var.check_chat_timkiem1).text
        logging.info("Trang chủ - Chat - tuỳ chon - tìm kiếm")
        logging.info("check font-end: Có kết quả hay không")
        logging.info(check_chat_timkiem)
        logging.info(check_chat_timkiem != "0 kết quả")
        time.sleep(1)
        driver.find_element(By.XPATH, var.emsochat_timkiem_dong).click()

        #thông tin về đoạn chat
        driver.find_element(By.XPATH, var.emsochat_thongtinvedoanchat).click()
        driver.find_element(By.XPATH, var.emsochat_xemtinnhandaghim).click()
        time.sleep(1)
        check_tinnhandaghim = driver.find_element(By.XPATH,var.check_tinnhandaghim1).text
        print(check_tinnhandaghim)
        logging.info("Trang chủ - Chat - tuỳ chon - tin nhắn đã ghim")
        logging.info("check font-end: đang ở popup - Tin nhắn đã ghim")
        logging.info(check_tinnhandaghim == "Tin nhắn đã ghim")
        driver.find_element(By.XPATH, var.emsochat_xemtinnhandaghim_x).click()

        # # Tuỳ chỉnh đoạn chat
        driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat).click()        #lại lỗi r
        # #màu hồng tìm
        # driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat_doichude).click()
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat_doichude_mauhongtim).click()
        # driver.find_element(By.XPATH, var.luu).click()
        # time.sleep(1)
        # driver.refresh()
        # time.sleep(2)
        # driver.find_element(By.XPATH, var.emsochat_caidathopthoai).click()
        # time.sleep(1)
        # try:
        #     check_mauchude = driver.find_element(By.XPATH,var.check_mauchude1).text
        #     print(check_mauchude)
        #     logging.info("Trang chủ - Chat - tuỳ chon - Màu chủ đề")
        #     logging.info("check font-end: Bạn đã đổi chủ đề thành Tuổi thơ.")
        #     logging.info(check_mauchude == "Bạn đã đổi chủ đề thành Tuổi thơ.")
        # except:
        #     driver.refresh()
        #     time.sleep(2)
        #     driver.find_element(By.XPATH, var.emsochat_caidathopthoai).click()
        #     check_mauchude = driver.find_element(By.XPATH,var.check_mauchude1).text
        #     print(check_mauchude)
        #     logging.info("Trang chủ - Chat - tuỳ chon - Màu chủ đề")
        #     logging.info("check font-end: Bạn đã đổi chủ đề thành Tuổi thơ.")
        #     logging.info(check_mauchude == "Bạn đã đổi chủ đề thành Tuổi thơ.")
        # # màu mặc định
        # driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat).click()
        # driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat_doichude).click()
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat_doichude_maumacdinh).click()
        # driver.find_element(By.XPATH, var.luu).click()
        # try:
        #     check_mauchude2 = driver.find_element(By.XPATH,var.check_mauchude2).text
        #     print(check_mauchude2)
        #     logging.info("Trang chủ - Chat - tuỳ chon - Màu chủ đề")
        #     logging.info(check_mauchude2)
        #     logging.info("check font-end: Bạn đã đổi chủ đề thành Mặc định.")
        #     logging.info(check_mauchude2 == "Bạn đã đổi chủ đề thành Mặc định.")
        # except:
        #     driver.refresh()
        #     time.sleep(2)
        #     check_mauchude2 = driver.find_element(By.XPATH,var.check_mauchude2).text
        #     print(check_mauchude2)
        #     logging.info("Trang chủ - Chat - tuỳ chon - Màu chủ đề")
        #     logging.info(check_mauchude2)
        #     logging.info("check font-end: Bạn đã đổi chủ đề thành Mặc định.")
        #     logging.info(check_mauchude2 == "Bạn đã đổi chủ đề thành Mặc định.")
        #
        # # Biểu tượng cảm xúc
        # driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat).click()    #lại lỗi r
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat_thaydoibieutuongcamxuc).click()
        # time.sleep(2.5)
        # driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat_thaydoibieutuongcamxuc_iconthu6).click()
        # time.sleep(1)
        # check_bieutuongcamxuc1 = driver.find_element(By.XPATH,var.check_bieutuongcamxucthaydoi).text
        # print(check_bieutuongcamxuc1)
        # logging.info("Trang chủ - Chat - tuỳ chon - Cảm xúc nhanh")
        # logging.info("check font-end: Bạn đã đặt cảm xúc nhanh thành 😱.")
        # logging.info(check_bieutuongcamxuc1 == "Bạn đã đặt cảm xúc nhanh thành 😱.")
        #
        # time.sleep(2)
        # driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat_thaydoibieutuongcamxuc).click()
        # time.sleep(1.5)
        # driver.find_element(By.XPATH, var.emsochat_tuychinhdoanchat_thaydoibieutuongcamxuc_macdinh).click()
        # time.sleep(1)
        # check_bieutuongcamxuc2 = driver.find_element(By.XPATH,var.check_bieutuongcamxucmacdinh).text
        # print(check_bieutuongcamxuc2)
        # logging.info("Trang chủ - Chat - tuỳ chon - Cảm xúc nhanh")
        # logging.info("check font-end: Bạn đã gỡ biểu tượng cảm xúc nhanh.")
        # logging.info(check_bieutuongcamxuc2)
        # logging.info(check_bieutuongcamxuc2 == "Bạn đã gỡ biểu tượng cảm xúc nhanh.")
        # time.sleep(2)

        #File phương tiện và file liên kết
        driver.find_element(By.XPATH, var.emsochat_file).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.emsochat_file_file_phuongtien).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.emsochat_file_file_phuongtien_anh).click()
        time.sleep(1.5)
        try:
            check_filephuongtien_anh = driver.find_element(By.XPATH, var.check_filephuongtien_anh1).is_displayed()
            print(check_filephuongtien_anh)
            logging.info("Trang chủ - Chat - tuỳ chon - File phương tiện")
            logging.info("check font-end: Có ảnh trong file phương tiện hay không")
            logging.info(check_filephuongtien_anh)
        except NoSuchElementException:
            logging.info("Trang chủ - Chat - tuỳ chon - File phương tiện")
            logging.info("check font-end: Có ảnh trong file phương tiện hay không")
            logging.info("False")

        driver.find_element(By.XPATH, var.emsochat_file_file_phuongtien_video).click()
        time.sleep(1.5)
        try:
            check_filephuongtien_video = driver.find_element(By.XPATH, var.check_filephuongtien_video1).is_displayed()
            print(check_filephuongtien_video)
            logging.info("Trang chủ - Chat - tuỳ chon - File phương tiện")
            logging.info("check font-end: Có video trong file phương tiện hay không")
            logging.info(check_filephuongtien_video)
        except NoSuchElementException:
            logging.info("Trang chủ - Chat - tuỳ chon - File phương tiện")
            logging.info("check font-end: Có video trong file phương tiện hay không")
            logging.info("False")

        driver.find_element(By.XPATH, var.emsochat_file_file_phuongtien_file).click()
        time.sleep(1)
        try:
            check_filephuongtien_file = driver.find_element(By.XPATH, var.check_filephuongtien_file1).is_displayed()
            print(check_filephuongtien_file)
            logging.info("Trang chủ - Chat - tuỳ chon - File phương tiện")
            logging.info("check font-end: Có file trong file phương tiện hay không")
            logging.info(check_filephuongtien_file)
        except NoSuchElementException:
            logging.info("Trang chủ - Chat - tuỳ chon - File phương tiện")
            logging.info("check font-end: Có file trong file phương tiện hay không")
            logging.info("False")

        driver.find_element(By.XPATH, var.emsochat_file_file_phuongtien_lienket).click()
        time.sleep(1)
        try:
            check_filephuongtien_lienket = driver.find_element(By.XPATH, var.check_filephuongtien_lienket1).is_displayed()
            print(check_filephuongtien_lienket)
            logging.info("Trang chủ - Chat - tuỳ chon - File phương tiện")
            logging.info("check font-end: Có liên kết trong file phương tiện hay không")
            logging.info(check_filephuongtien_lienket)
        except NoSuchElementException:
            logging.info("Trang chủ - Chat - tuỳ chon - File phương tiện")
            logging.info("check font-end: Có liên kết trong file phương tiện hay không")
            logging.info("False")
        driver.find_element(By.XPATH, var.emsochat_file_file_phuongtien_back).click()
        time.sleep(2)

        # Quyền riêng tư & hỗ trợ
        driver.find_element(By.XPATH, var.emsochat_quuyenriengtuhotro).click()
        #Tắt thông báo
        driver.find_element(By.XPATH, var.emsochat_quuyenriengtuhotro_tatthongbao).click()
        check_batthongbao_tuychon = driver.find_element(By.XPATH,var.check_batthongbao_tuychon1).text
        check_batthongbao_quyenriengtu = driver.find_element(By.XPATH, var.check_batthongbao_quyenriengtu1).text
        print(check_batthongbao_tuychon)
        print(check_batthongbao_quyenriengtu)
        logging.info("Trang chủ - Chat - tuỳ chon - Thông báo")
        logging.info("check font-end: Thông báo(tuỳ chọn) - " + check_batthongbao_tuychon)
        logging.info("check font-end: Thông báo quyền riêng tư - " + check_batthongbao_quyenriengtu)
        logging.info(check_batthongbao_quyenriengtu)
        logging.info(check_batthongbao_tuychon)
        logging.info(check_batthongbao_tuychon == check_batthongbao_quyenriengtu)
        time.sleep(1)

        #Bật thông báo
        driver.find_element(By.XPATH, var.emsochat_quuyenriengtuhotro_batthongbao).click()
        time.sleep(2)
        check_tatthongbao_tuychon = driver.find_element(By.XPATH,var.check_tatthongthongbao_tuychon1).text
        check_tatthongbao_quyenriengtu = driver.find_element(By.XPATH, var.check_tatthongbao_quyenriengtu1).text
        print(check_tatthongbao_tuychon)
        print(check_tatthongbao_quyenriengtu)
        logging.info("Trang chủ - Chat - tuỳ chon - Thông báo")
        logging.info("check font-end: Thông báo(tuỳ chọn) - " + check_batthongbao_tuychon)
        logging.info("check font-end: Thông báo quyền riêng tư - " + check_batthongbao_quyenriengtu)
        logging.info(check_tatthongbao_tuychon == check_tatthongbao_quyenriengtu)

        #Chặn
        driver.find_element(By.XPATH, var.emsochat_quuyenriengtuhotro_chan).click()
        check_chan_ngocmai = driver.find_element(By.XPATH, var.check_chan_ngocmai1).text
        print(check_chan_ngocmai)
        logging.info("Trang chủ - Chat - tuỳ chon - Chặn")
        logging.info("check font-end: popup - Chặn Ngọc Mai")
        logging.info(check_chan_ngocmai)
        logging.info(check_chan_ngocmai == "Chặn Ngọc Mai")
        driver.find_element(By.XPATH, var.emsochat_quuyenriengtuhotro_chan_xacnhan).click()
        time.sleep(1)
        check_chan_ngocmai_dachan = driver.find_element(By.XPATH, var.check_chan_ngocmai_dachan).text
        print(check_chan_ngocmai_dachan)
        logging.info("Trang chủ - Chat - tuỳ chon - Chặn")
        logging.info("check font-end: Bạn đã chặn Ngọc Mai")
        logging.info(check_chan_ngocmai_dachan == "Bạn đã chặn Ngọc Mai")

        #Bỏ chặn
        # driver.find_element(By.XPATH, var.emsochat_quuyenriengtuhotro_bochan1).click()
        driver.find_element(By.XPATH, var.emsochat_quuyenriengtuhotro_bochan2).click()
        driver.find_element(By.XPATH, var.emsochat_quuyenriengtuhotro_chan_xacnhan).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(2)


    def taonhom(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/messages/111225894003999190111340222342420143")
        driver.find_element(By.XPATH, var.trangchu_emsochat).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.emsochat_icontaotinnhanmoi).click()

        #Chọn Ngọc Mai lần 1 đê xoá
        driver.find_element(By.XPATH, var.tinnhanmoi_den).click()       #Chọn người nhưng ko tắt ds chọn
        driver.find_element(By.XPATH, var.tinnhanmoi_den).send_keys(data['trangchu_tinnhan']['den'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.tinnhanmoi_den_ngocmai)))
        element.click()
        time.sleep(1)
        #Chọn hue nguyen lần 1 đê xoá
        driver.find_element(By.XPATH, var.tinnhanmoi_den).click()
        driver.find_element(By.XPATH, var.tinnhanmoi_den).send_keys(data['trangchu_tinnhan']['den1'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.tinnhanmoi_den_huenguyen)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.emsochat_icontaotinnhanmoi_x).click()
        time.sleep(1)

        #Chọn Ngọc Mai lần 2
        driver.find_element(By.XPATH, var.emsochat_icontaotinnhanmoi).click()
        driver.find_element(By.XPATH, var.tinnhanmoi_den).click()
        driver.find_element(By.XPATH, var.tinnhanmoi_den).send_keys(data['trangchu_tinnhan']['den'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.tinnhanmoi_den_ngocmai)))
        element.click()
        time.sleep(1)
        #Chọn hue nguyen lần 2
        driver.find_element(By.XPATH, var.tinnhanmoi_den).click()
        driver.find_element(By.XPATH, var.tinnhanmoi_den).send_keys(data['trangchu_tinnhan']['den1'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.tinnhanmoi_den_huenguyen)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.tinnhanmoi_input).send_keys(data['trangchu_tinnhan']['tinnhanmoi_input'])
        driver.find_element(By.XPATH, var.tinnhanmoi_input).send_keys(Keys.ENTER)
        time.sleep(2)
        check_tinnhanmoi_taonhom = driver.find_element(By.XPATH, var.check_tinnhanmoi_taonhom1).text
        logging.info("Chat - icon tạo tin nhắn mới - Tạo nhóm")
        logging.info("check font-end: đã tạo nhóm -  Mai, nguyen")
        logging.info(check_tinnhanmoi_taonhom == "Mai, nguyen")

        #Tùy chọn nhóm
        actions = ActionChains(driver)
        hover_hopchat = driver.find_element(By.XPATH, var.hover_hopchat1)
        actions.move_to_element(hover_hopchat).perform()
        time.sleep(1)
        driver.find_element(By.XPATH, var.tuychonnhom).click()
        #Đánh dáu là chưa đọc

        driver.find_element(By.XPATH, var.tuychonnhom_danhdaulachuadoc).click()
        time.sleep(1)
        element1 = driver.find_element(By.XPATH, var.check_color_mautennhom)
        color = element1.value_of_css_property("color")
        logging.info("Chat - Nhóm - Cài đặt nhóm")
        logging.info("check font-end: Đánh dấu là chưa đọc")
        logging.info(color)
        logging.info(color == "rgba(53, 120, 229, 1)")
        print(color)

        # Đánh dáu là đã đọc
        actions.move_to_element(hover_hopchat).perform()
        driver.find_element(By.XPATH, var.tuychonnhom).click()
        driver.find_element(By.XPATH, var.tuychonnhom_danhdauladadoc).click()
        time.sleep(1)
        element1 = driver.find_element(By.XPATH, var.check_color_mautennhom)
        color = element1.value_of_css_property("color")
        logging.info("Chat - Nhóm - Cài đặt nhóm")
        logging.info("check font-end: Đánh dấu là đã đọc")
        logging.info(color)
        logging.info(color == "rgba(101, 103, 107, 1)")
        print(color)

        #Ghim
        actions.move_to_element(hover_hopchat).perform()
        driver.find_element(By.XPATH, var.tuychonnhom).click()
        driver.find_element(By.XPATH, var.tuychonnhom_ghim).click()
        time.sleep(1)
        try:
            check_nhomcaidat_ghim = driver.find_element(By.XPATH, var.check_nhomcaidat_ghim1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Ghim - Có icon Ghim")
            logging.info(check_nhomcaidat_ghim)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Ghim - Có icon Ghim")
            logging.info("False")


        #Bỏ ghim
        hover_hopchat2 = driver.find_element(By.XPATH, var.hover_hopchat2)
        actions.move_to_element(hover_hopchat2).perform()
        time.sleep(1)
        driver.find_element(By.XPATH, var.tuychonnhom2).click()
        driver.find_element(By.XPATH, var.tuychonnhom_boghim).click()
        time.sleep(2)

        driver.implicitly_wait(3)
        try:      #Ko cần mở comment này
            check_nhomcaidat_ghim = driver.find_element(By.XPATH, var.check_nhomcaidat_ghim1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Bỏ Ghim - Không Có icon Ghim")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end:Bỏ Ghim - Không Có icon Ghim")
            logging.info("True")
        driver.implicitly_wait(15)


        #Tắt thông báo
        actions = ActionChains(driver)
        hover_hopchat = driver.find_element(By.XPATH, var.hover_hopchat1)
        actions.move_to_element(hover_hopchat).perform()
        time.sleep(1)
        actions.move_to_element(hover_hopchat).perform()
        driver.find_element(By.XPATH, var.tuychonnhom).click()
        driver.find_element(By.XPATH, var.tuychonnhom_tatthongbao).click()
        time.sleep(1)
        try:
            check_nhomcaidat_tatthongbao = driver.find_element(By.XPATH, var.check_nhomcaidat_tatthongbao1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Tắt thông báo - có icon tắt thông báo")
            logging.info(check_nhomcaidat_tatthongbao)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Tắt thông báo - có icon tắt thông báo")
            logging.info("False")

        #Bật thông báo
        actions.move_to_element(hover_hopchat).perform()
        driver.find_element(By.XPATH, var.tuychonnhom).click()
        check_nhomcaidat_batthongbao = driver.find_element(By.XPATH, var.check_nhomcaidat_batthongbao1).text
        logging.info("Chat - Nhóm - Cài đặt nhóm")
        logging.info("check font-end: Bật thông báo")
        logging.info(check_nhomcaidat_batthongbao == "Bật thông báo")
        driver.find_element(By.XPATH, var.tuychonnhom_batthongbao).click()
        time.sleep(2)


        #Xoá đoạn chat
        actions = ActionChains(driver)
        actions.move_to_element(hover_hopchat).perform()
        driver.find_element(By.XPATH, var.tuychonnhom).click()
        button = driver.find_element(By.XPATH, var.tuychonnhom_xoadoanchat)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        check_nhomcaidat_xoadoanchat = driver.find_element(By.XPATH, var.check_nhomcaidat_xoadoanchat1).text
        logging.info("Chat - Nhóm - Cài đặt nhóm")
        logging.info("check font-end: Xoá đoạn chat message - Bạn không thể hoàn tác sau khi xoá bản sao của cuộc trò chuyện này.")
        logging.info(check_nhomcaidat_xoadoanchat == "Bạn không thể hoàn tác sau khi xoá bản sao của cuộc trò chuyện này.")
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(3)

        # #Rời khỏi nhóm
        driver.find_element(By.XPATH, var.emsochat_icontaotinnhanmoi).click()
        driver.find_element(By.XPATH, var.tinnhanmoi_den).click()
        driver.find_element(By.XPATH, var.tinnhanmoi_den).send_keys(data['trangchu_tinnhan']['den'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.tinnhanmoi_den_ngocmai)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.tinnhanmoi_den).click()
        driver.find_element(By.XPATH, var.tinnhanmoi_den).send_keys(data['trangchu_tinnhan']['den1'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.tinnhanmoi_den_huenguyen)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.tinnhanmoi_input).send_keys(data['trangchu_tinnhan']['tinnhanmoi_input'])
        driver.find_element(By.XPATH, var.tinnhanmoi_input).send_keys(Keys.ENTER)
        time.sleep(2)

        #Message rời khỏi nhóm
        actions = ActionChains(driver)
        hover_hopchat = driver.find_element(By.XPATH, var.hover_hopchat1)     #Tự động chọn qtv khi rời nhóm
        actions.move_to_element(hover_hopchat).perform()
        driver.find_element(By.XPATH, var.tuychonnhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.tuychonnhom_roikhoinhom1).click()
        time.sleep(1)
        check_popup_roikhoinhom = driver.find_element(By.XPATH, var.check_popup_roikhoinhom1).text
        logging.info("Chat - Nhóm - Cài đặt nhóm")
        logging.info("check font-end: Popup Rời khỏi nhóm - Rời đi mà không chọn Quản trị viên ?")
        logging.info(check_popup_roikhoinhom)
        logging.info(check_popup_roikhoinhom == "Rời đi mà không chọn Quản trị viên ?")

        check_messagepopup_roikhoinhom = driver.find_element(By.XPATH, var.check_message_roikhoinhom1).text
        logging.info("Chat - Nhóm - Cài đặt nhóm")
        logging.info("check font-end: Message POPUP - Bạn có thể chọn quản trị viên mới trong số những người có tên ở phần Thành viên. Nếu bạn rời khỏi nhóm mà không chọn ai thay thế, thành viên kỳ cựu nhất của nhóm sẽ trở thành quản trị viên.")
        logging.info(check_messagepopup_roikhoinhom == "Bạn có thể chọn quản trị viên mới trong số những người có tên ở phần Thành viên. Nếu bạn rời khỏi nhóm mà không chọn ai thay thế, thành viên kỳ cựu nhất của nhóm sẽ trở thành quản trị viên.")
        driver.find_element(By.XPATH, var.huy).click()
        time.sleep(1)
        actions.move_to_element(hover_hopchat).perform()
        driver.find_element(By.XPATH, var.tuychonnhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.tuychonnhom_roikhoinhom1).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.tuychonnhom_roikhoinhom1_ok).click()
        time.sleep(1)
        # check_message_roikhoinhom_thatbai = driver.find_element(By.XPATH, var.check_message_roikhoinhom_thatbai1).text
        # logging.info("Chat - Nhóm - Cài đặt nhóm")
        # logging.info("check font-end: Message Rời khỏi nhóm thất bại - Vui lòng chọn một quản trị viên mới trước khi rời khỏi nhóm.")
        # logging.info(check_message_roikhoinhom_thatbai == "Vui lòng chọn một quản trị viên mới trước khi rời khỏi nhóm.")


    def thongtindoanchat_nhom(self):
        driver.implicitly_wait(3)
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_chat).click()
        driver.find_element(By.XPATH, var.trangchu_emsochat).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.emsochat_icontaotinnhanmoi).click()
        # Chọn hue nguyen
        driver.find_element(By.XPATH, var.tinnhanmoi_den).click()
        driver.find_element(By.XPATH, var.tinnhanmoi_den).send_keys(data['trangchu_tinnhan']['den1'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.tinnhanmoi_den_huenguyen)))
        element.click()
        time.sleep(1)
        #Chọn Ngọc Mai
        driver.find_element(By.XPATH, var.tinnhanmoi_den).click()
        driver.find_element(By.XPATH, var.tinnhanmoi_den).send_keys(data['trangchu_tinnhan']['den'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.tinnhanmoi_den_ngocmai)))
        element.click()
        time.sleep(1)

        driver.find_element(By.XPATH, var.tinnhanmoi_input).send_keys(data['trangchu_tinnhan']['tinnhanmoi_input'])
        driver.find_element(By.XPATH, var.tinnhanmoi_input).send_keys(Keys.ENTER)
        time.sleep(2)

        driver.find_element(By.XPATH, var.emsochat_caidathopthoai).click()
        time.sleep(1)
        #Thay ảnh
        driver.find_element(By.XPATH, var.nhom_caidathopthoai_tuychinhdoanchat).click()
        driver.find_element(By.XPATH, var.nhom_caidathopthoai_tuychinhdoanchat_thaydoidanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhmes1.exe")
        time.sleep(1)
        try:
            check_nhom_thayanh = driver.find_element(By.XPATH, var.check_nhom_thayanh1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Thay ảnh - Thông báo - Bạn đã đổi ảnh đại diện nhóm.")
            logging.info(check_nhom_thayanh)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Thay ảnh - Thông báo - Bạn đã đổi ảnh đại diện nhóm.")
            logging.info("False")

        #Tên cuộc trò chuyện
        driver.find_element(By.XPATH, var.nhom_caidathopthoai_tuychinhdoanchat_tencuoctrochuyen).click()
        xoa = driver.find_element(By.XPATH, var.tencuoctrochuyen_input)
        xoa.send_keys(Keys.CONTROL, "a")
        #Tên nhóm có chưa ký tự đặc biệt
        driver.find_element(By.XPATH, var.tencuoctrochuyen_input).send_keys(data['trangchu_tinnhan']['dattennhom'])
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        check_nhom_doiten_sai = driver.find_element(By.XPATH, var.check_nhom_doiten_sai1).text
        logging.info("Chat - Nhóm - Cài đặt nhóm")
        logging.info("check font-end: Nhập ký tự đặc biệt - Tên nhóm không được chứa kí tự đặc biệt.")
        logging.info(check_nhom_doiten_sai == "Tên nhóm không được chứa kí tự đặc biệt.")

        xoa = driver.find_element(By.XPATH, var.tencuoctrochuyen_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.tencuoctrochuyen_input).send_keys(data['trangchu_tinnhan']['dattennhom1'])
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        try:
            check_nhom_doitennhom = driver.find_element(By.XPATH, var.check_nhom_doitennhom1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Đổi tên nhóm - Bạn đã đặt tên nhóm là Bực mình.")
            logging.info(check_nhom_doitennhom)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Đổi tên nhóm - Bạn đã đặt tên nhóm là Bực mình.")
            logging.info("False")

        #Tuỳ chọn nhóm
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_caidathopthoai_tuychinhdoanchat_tuychonnhom).click()
        #Chỉ quản trị viên thêm thành viên mới
        button = driver.find_element(By.XPATH, var.tuychonnhom_chiquantrithemthanhvien)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        try:
            check_tuychonnhom_themthanhvienmoi_tat = driver.find_element(By.XPATH, var.check_tuychonnhom_themthanhvienmoi_tat1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Chỉ quản trị viên thêm thành viên mới - Bạn đã tắt tính năng phê duyệt thành viên và bất cứ ai đều có thể thêm người mới vào nhóm.")
            logging.info(check_tuychonnhom_themthanhvienmoi_tat)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Chỉ quản trị viên thêm thành viên mới - Bạn đã tắt tính năng phê duyệt thành viên và bất cứ ai đều có thể thêm người mới vào nhóm.")
            logging.info("False")

        button = driver.find_element(By.XPATH, var.tuychonnhom_chiquantrithemthanhvien)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        check_tuychonnhom_themthanhvienmoi_bat = driver.find_element(By.XPATH, var.check_tuychonnhom_themthanhvienmoi_bat1).text
        logging.info("Chat - Nhóm - Cài đặt nhóm")
        logging.info("check font-end: Chỉ quản trị viên thêm thành viên mới - Bạn đã bật tính năng phê duyệt thành viên và chỉ có quản trị viên có thể thêm người mới vào nhóm.")
        logging.info(check_tuychonnhom_themthanhvienmoi_bat)
        logging.info(check_tuychonnhom_themthanhvienmoi_bat == "Bạn đã bật tính năng phê duyệt thành viên và chỉ có quản trị viên có thể thêm người mới vào nhóm.")
        time.sleep(1)


        #Chỉ quản trị viên có thể nhắn tin
        button = driver.find_element(By.XPATH, var.tuychonnhomchiquanchinhantin)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        try:
            check_tuychonnhom_quantricothenhan_bat = driver.find_element(By.XPATH, var.check_tuychonnhom_quantricothenhan_bat1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Chỉ quản trị viên có thể nhắn tin - Bạn chỉ cho phép quản trị viên gửi tin nhắn.")
            logging.info(check_tuychonnhom_quantricothenhan_bat)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Chỉ quản trị viên có thể nhắn tin - Bạn chỉ cho phép quản trị viên gửi tin nhắn.")
            logging.info("False")
        time.sleep(1)

        button = driver.find_element(By.XPATH, var.tuychonnhomchiquanchinhantin)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        try:
            check_tuychonnhom_quantricothenhan_tat = driver.find_element(By.XPATH, var.check_tuychonnhom_quantricothenhan_tat1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Chỉ quản trị viên có thể nhắn tin - Bạn cho phép tất cả mọi người gửi tin nhắn.")
            logging.info(check_tuychonnhom_quantricothenhan_tat)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Chỉ quản trị viên có thể nhắn tin - Bạn cho phép tất cả mọi người gửi tin nhắn.")
            logging.info("False")
        time.sleep(1)

        #Thành viên trong đoạn chat
        driver.find_element(By.XPATH, var.nhom_caidathopthoai_tuychinhdoanchat_thanhvientrongdoanchat).click()
        #Thêm người
        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_themnguoi).click()
        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_themnguoi_input).send_keys(data['trangchu_tinnhan']['thanhvien_themnguoi'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.vuonglam)))
        element.click()
        driver.find_element(By.XPATH, var.them).click()
        time.sleep(1)
        try:
            check_thanhvientrongdoanchat_themnguoi = driver.find_element(By.XPATH, var.check_thanhvientrongdoanchat_themnguoi1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Thông báo thêm người - Bạn đã thêm Vương Lâm vào nhóm.")
            logging.info(check_thanhvientrongdoanchat_themnguoi)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Thông báo thêm người - Bạn đã thêm Vương Lâm vào nhóm.")
            logging.info("False")
        time.sleep(1)

        driver.execute_script("window.scrollBy(0,700)", "")

        #Xoá người
        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham_xoa)
        driver.execute_script("arguments[0].click();", button)
        # driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham_xoa).click()
        time.sleep(1)
        try:
            check_thanhvientrongdoanchat_xoanguoi = driver.find_element(By.XPATH, var.check_thanhvientrongdoanchat_xoanguoi1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Thông báo xoá người - Bạn đã xóa Vương Lâm khỏi nhóm.")
            logging.info(check_thanhvientrongdoanchat_xoanguoi)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Thông báo xoá người - Bạn đã xóa Vương Lâm khỏi nhóm.")
            logging.info("False")
        time.sleep(1)


        #Chỉ định làm quản trị viên
        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham1).click()
        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham1_chidinhlamquantrivien).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham1_chidinhlamquantrivien1).click()
        time.sleep(1)
        try:
            check_thanhvientrongdoanchat_chidinhlamqtv = driver.find_element(By.XPATH, var.check_thanhvientrongdoanchat_chidinhlamqtv1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Thông báo chỉ định làm quản trị viên - Bạn đã thêm Ngọc Mai làm quản trị viên.")
            logging.info(check_thanhvientrongdoanchat_chidinhlamqtv)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Thông báo chỉ định làm quản trị viên - Bạn đã thêm Ngọc Mai làm quản trị viên.")
            logging.info("False")
        time.sleep(1)

        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham1).click()
        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham1_xoaquantrivien).click()
        time.sleep(1)
        try:
            check_thanhvientrongdoanchat_xoachidinhlamqtv = driver.find_element(By.XPATH, var.check_thanhvientrongdoanchat_xoachidinhlamqtv1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Thông báo xoá chỉ định làm quản trị viên - Bạn đã gỡ quyền quản trị viên của Ngọc Mai.")
            logging.info(check_thanhvientrongdoanchat_xoachidinhlamqtv)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Thông báo xoá chỉ định làm quản trị viên - Bạn đã gỡ quyền quản trị viên của Ngọc Mai.")
            logging.info("False")
        time.sleep(1)
        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham1_chidinhlamquantrivien).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham1_chidinhlamquantrivien1).click()
        time.sleep(1)
        #Tham gia bằng link nhóm
        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham1_thamgiabanglinknhom).click()
        check_nhom_thamgiabanglink = driver.find_element(By.XPATH, var.check_nhom_thamgiabanglink1).text
        logging.info("Chat - Nhóm - Cài đặt nhóm")
        logging.info("check font-end: Popup Tham gia bằng link - Link tham gia nhóm")
        logging.info(check_nhom_thamgiabanglink == "Link tham gia nhóm")

        driver.find_element(By.XPATH, var.thanhvientrongdoanchat_icon3cham1_thamgiabanglinknhom_iconcoppy).click()
        time.sleep(1)
        try:
            check_nhom_thamgiabanglink_icon = driver.find_element(By.XPATH, var.check_nhom_thamgiabanglink_icon1).is_displayed()
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Popup Tham gia bằng link - Link tham gia nhóm - đã coppy")
            logging.info(check_nhom_thamgiabanglink_icon)
        except NoSuchElementException:
            logging.info("Chat - Nhóm - Cài đặt nhóm")
            logging.info("check font-end: Popup Tham gia bằng link - Link tham gia nhóm - đã coppy")
            logging.info("False")
        time.sleep(1)
        driver.implicitly_wait(15)
        driver.back()
        time.sleep(0.5)
        driver.forward()
        time.sleep(2)

        #Đổi tên nhóm
        # driver.find_element(By.XPATH, var.emsochat_caidathopthoai).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_caidathopthoai_tuychinhdoanchat).click()
        driver.find_element(By.XPATH, var.nhom_caidathopthoai_tuychinhdoanchat_tencuoctrochuyen).click()
        xoa = driver.find_element(By.XPATH, var.tencuoctrochuyen_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.tencuoctrochuyen_input).send_keys(data['trangchu_tinnhan']['dattennhom3'])
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)

        #Quyền riêng tư & hỗ trợ
        driver.find_element(By.XPATH, var.nhom_caidathopthoai_tuychinhdoanchat_quyenriengtuvahotro).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quyenriengtuvahotro_roikhoinhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quyenriengtuvahotro_roikhoinhom_ok).click()
        time.sleep(2)






    def caidatcanhan_chuyentaikhoan(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(2)
        driver.find_element(By.XPATH, var.icon_caidatcanhan).click()
        driver.find_element(By.XPATH, var.caidatcanhan_chuyentaikhoankhac).click()
        driver.find_element(By.XPATH, var.caidatcanhan_chuyentaikhoankhac_tranquangtruong).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day_password).send_keys("voncamk22")
        driver.find_element(By.XPATH, var.login_nho_mat_khau).click()
        driver.find_element(By.XPATH, var.login_chon_tk_dang_nhap_gan_day_dang_nhap).click()
        time.sleep(3)
        check_chuyentaikhoan = driver.find_element(By.XPATH,var.check_chuyentaikhoan1).text
        print(check_chuyentaikhoan)
        logging.info("Trang chủ -Tài khoản - Chuyển tài khoản")
        logging.info("check font-end: Đã chuyển qua tài khoản - Trần Quang Trường")
        logging.info(check_chuyentaikhoan == "Trần Quang Trường")

    def caidatvaquyenriengtu_chung(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.icon_caidatcanhan).click()
        driver.find_element(By.XPATH, var.caidatcanhan_caidatvaquyenriengtu).click()
        time.sleep(1.5)
        check_caidatvaquyenriengtu_chung = driver.find_element(By.XPATH,var.check_caidatvaquyenriengtu_chung1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư")
        logging.info("check font-end: Cài đặt")
        logging.info(check_caidatvaquyenriengtu_chung == "Cài đặt")

        #CHUNG
        #Tên
        driver.find_element(By.XPATH, var.chung_ten_chinhsua).click()
        check_chung_ten = driver.find_element(By.XPATH,var.check_chung_ten1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Chung")
        logging.info("check font-end: Tên - Chỉnh sửa")
        logging.info(check_chung_ten == "Họ và tên")
        driver.find_element(By.XPATH, var.chung_ten_huy).click()
        time.sleep(0.5)
        #Tên người dùng
        driver.find_element(By.XPATH, var.chung_tennguoidung_chinhsua).click()
        check_chung_tennguoidung = driver.find_element(By.XPATH,var.check_chung_tennguoidung1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Chung")
        logging.info("check font-end: Tên người dùng - Chỉnh sửa")
        logging.info(check_chung_tennguoidung == "Tên người dùng:")
        driver.find_element(By.XPATH, var.chung_ten_huy).click()
        time.sleep(0.5)
        #Số điện thoại
        driver.find_element(By.XPATH, var.chung_sodienthoai_chinhsua).click()
        check_chung_sodienthoai = driver.find_element(By.XPATH,var.check_chung_sodienthoai1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Chung")
        logging.info("check font-end: Số điện thoại - Chỉnh sửa")
        logging.info(check_chung_sodienthoai == "Số điện thoại:")
        driver.find_element(By.XPATH, var.chung_ten_huy).click()
        time.sleep(0.5)
        #Xác nhận danh tính
        driver.find_element(By.XPATH, var.chung_xacnhandanhtinh_chinhsua).click()
        check_chung_xacnhandanhtinh = driver.find_element(By.XPATH,var.check_chung_xacnhandanhtinh1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Chung")
        logging.info("check font-end: Xác nhận danh tính - Xem")
        logging.info(check_chung_xacnhandanhtinh == "Bật kiếm tiền")
        driver.back()
        time.sleep(0.5)
        #Liên hệ
        driver.find_element(By.XPATH, var.chung_lienhe_chinhsua).click()
        time.sleep(1)
        check_chung_lienhe = driver.find_element(By.XPATH,var.tinhangdangphattrien).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Chung")
        logging.info("check font-end: Liên hệ - Chỉnh sửa")
        logging.info("Chức năng chưa hoạt động")
        logging.info(check_chung_lienhe)
        logging.info("check font-end: message - " + check_chung_lienhe)
        logging.info(check_chung_lienhe == "Tính năng đang phát triển.Vui lòng thử lại sau!!")
        driver.find_element(By.XPATH, var.ok).click()
        time.sleep(0.5)

    def caidatvaquyenriengtu_baomatvadangnhap(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        #BẢO MẬT VÀ ĐĂNG NHẬP
        driver.find_element(By.XPATH, var.baomatvadangnhap).click()
        #Đổi mật khẩu
        driver.find_element(By.XPATH, var.baomatvadangnhap_doimatkhau).click()
        check_baomatvadangnhap_doimatkhau = driver.find_element(By.XPATH,var.check_baomatvadangnhap_doimatkhau1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Bảo mật và đăng nhập")
        logging.info("check font-end: Đổi mật khẩu - Nhập mật khẩu cũ")
        logging.info(check_baomatvadangnhap_doimatkhau == "Nhập mật khẩu cũ")
        driver.find_element(By.XPATH, var.chung_ten_huy).click()
        time.sleep(0.5)
        #Lưu thông tin đăng nhập
        driver.find_element(By.XPATH, var.baomatvadangnhap_luuthongtindangnhap).click()
        time.sleep(1)
        check_baomatvadangnhap_luuthongtindangnhap = driver.find_element(By.XPATH,var.tinhangdangphattrien).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Bảo mật và đăng nhập")
        logging.info("check font-end: Lưu thông tin đăng nhập của bạn - Chỉnh sửa")
        logging.info("Chức năng chưa hoạt động")
        logging.info("check font-end: message - " + check_baomatvadangnhap_luuthongtindangnhap)
        logging.info(check_baomatvadangnhap_luuthongtindangnhap)
        logging.info(check_baomatvadangnhap_luuthongtindangnhap == "Tính năng đang phát triển.Vui lòng thử lại sau!!")
        driver.find_element(By.XPATH, var.ok).click()
        time.sleep(0.5)
        #Dùng tính năng xác thực 2 yếu tố
        driver.find_element(By.XPATH, var.baomatvadangnhap_xacthuc2yeuto).click()
        time.sleep(1)
        check_baomatvadangnhap_xacthuc = driver.find_element(By.XPATH,var.tinhangdangphattrien).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Bảo mật và đăng nhập")
        logging.info("check font-end: Dùng tính năng xác thực 2 yếu tố - Chỉnh sửa")
        logging.info("Chức năng chưa hoạt động")
        logging.info("check font-end: message - " + check_baomatvadangnhap_xacthuc)
        logging.info(check_baomatvadangnhap_xacthuc)
        logging.info(check_baomatvadangnhap_xacthuc == "Tính năng đang phát triển.Vui lòng thử lại sau!!")
        driver.find_element(By.XPATH, var.ok).click()
        time.sleep(0.5)
        #Đăng nhập hợp lệ
        driver.find_element(By.XPATH, var.baomatvadangnhap_dangnhaphople).click()
        time.sleep(1)
        check_baomatvadangnhap_dangnhaphople = driver.find_element(By.XPATH,var.tinhangdangphattrien).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Bảo mật và đăng nhập")
        logging.info("check font-end: Đăng nhập hợp lệ - Xem")
        logging.info("Chức năng chưa hoạt động")
        logging.info("check font-end: message - " + check_baomatvadangnhap_dangnhaphople)
        logging.info(check_baomatvadangnhap_dangnhaphople)
        logging.info(check_baomatvadangnhap_dangnhaphople == "Tính năng đang phát triển.Vui lòng thử lại sau!!")
        driver.find_element(By.XPATH, var.ok).click()
        time.sleep(0.5)
        #Nhận cảnh báo về những lần đăng nhập lạ
        driver.find_element(By.XPATH, var.baomatvadangnhap_nhancanhbao).click()
        time.sleep(1)
        check_baomatvadangnhap_dangnhapla = driver.find_element(By.XPATH,var.tinhangdangphattrien).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Bảo mật và đăng nhập")
        logging.info("check font-end: Nhận cảnh báo về những lần đăng nhập lạ - Chỉnh sửa")
        logging.info("Chức năng chưa hoạt động")
        logging.info("check font-end: message - " + check_baomatvadangnhap_dangnhapla)
        logging.info(check_baomatvadangnhap_dangnhapla)
        logging.info(check_baomatvadangnhap_dangnhapla == "Tính năng đang phát triển.Vui lòng thử lại sau!!")
        driver.find_element(By.XPATH, var.ok).click()
        time.sleep(0.5)
        #Mã hóa email thông báo
        driver.find_element(By.XPATH, var.baomatvadangnhap_mahoaemail).click()
        time.sleep(1)
        check_baomatvadangnhap_emailthongbao = driver.find_element(By.XPATH,var.tinhangdangphattrien).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Bảo mật và đăng nhập")
        logging.info("check font-end: Mã hóa email thông báo - Chỉnh sửa")
        logging.info("Chức năng chưa hoạt động")
        logging.info("check font-end: message - " + check_baomatvadangnhap_emailthongbao)
        logging.info(check_baomatvadangnhap_emailthongbao)
        logging.info(check_baomatvadangnhap_emailthongbao == "Tính năng đang phát triển.Vui lòng thử lại sau!!")
        driver.find_element(By.XPATH, var.ok).click()
        time.sleep(0.5)
        #Xem email gần đây từ emso
        driver.find_element(By.XPATH, var.baomatvadangnhap_emailganday).click()
        time.sleep(1)
        check_baomatvadangnhap_xememailganday = driver.find_element(By.XPATH,var.tinhangdangphattrien).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Bảo mật và đăng nhập")
        logging.info("check font-end: Xem email gần đây từ emso - Xem")
        logging.info("Chức năng chưa hoạt động")
        logging.info("check font-end: message - " + check_baomatvadangnhap_xememailganday)
        logging.info(check_baomatvadangnhap_xememailganday)
        logging.info(check_baomatvadangnhap_xememailganday == "Tính năng đang phát triển.Vui lòng thử lại sau!!")
        driver.find_element(By.XPATH, var.ok).click()
        time.sleep(0.5)

    def caidatvaquyenriengtu_thongtinbantrenemso(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        #THÔNG TIN CỦA BẠN TRÊN EMSO
        driver.find_element(By.XPATH, var.thongtincuabantrenemso).click()
        #Truy cập thông tin của bạn
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_truycapthongtincuaban).click()
        time.sleep(1)
        check_thongtincuaban_truycapthongtin = driver.find_element(By.XPATH,var.tinhangdangphattrien).text
        print(check_thongtincuaban_truycapthongtin)
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bạn trên emso")
        logging.info("check font-end: Truy cập thông tin của bạn - Xem")
        logging.info("Chức năng chưa hoạt động")
        logging.info("check font-end: message - " + check_thongtincuaban_truycapthongtin)
        logging.info(check_thongtincuaban_truycapthongtin)
        logging.info(check_thongtincuaban_truycapthongtin == "Tính năng đang phát triển.Vui lòng thử lại sau!!")
        driver.find_element(By.XPATH, var.ok).click()
        time.sleep(1)
        #Nhật ký hoạt động
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_nhatkyhoatdong).click()
        time.sleep(1)
        # driver.find_element(By.XPATH, var.baivietcuaban_lichsuxembaiviet).click()     #ko load dc trang
        # time.sleep(3)
        # check_baivietcuaban_lichsuxembaiviet = driver.find_element(By.XPATH,var.check_baivietcuaban_lichsuxembaiviet1).text
        # logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Lịch sử xem bài viết")
        # logging.info("check font-end:Có thông báo hay không ")
        # logging.info(check_baivietcuaban_lichsuxembaiviet != None)
        # driver.back()
        driver.find_element(By.XPATH, var.baivietcuaban_lichsutimkiem).click()        #ko load dc trang
        time.sleep(3)
        check_baivietcuaban_lichsutimkiem = driver.find_element(By.XPATH,var.check_baivietcuaban_lichsutimkiem1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Lịch sử tìm kiếm")
        logging.info("check font-end:Có thông báo hay không ")
        logging.info(check_baivietcuaban_lichsutimkiem != None)
        driver.back()
        driver.find_element(By.XPATH, var.baivietcuaban_nhombandatimkiem).click()       #ko load dc trang
        time.sleep(3)
        check_baivietcuaban_nhombandatimkiem = driver.find_element(By.XPATH,var.check_baivietcuaban_nhombandatimkiem1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Nhóm bạn đã tìm kiếm")
        logging.info("check font-end:Có thông báo hay không ")
        logging.info(check_baivietcuaban_nhombandatimkiem != None)
        driver.back()
        driver.find_element(By.XPATH, var.baivietcuaban_trangluotthichsothich).click()
        time.sleep(3)
        check_baivietcuaban_trangluotthichsothich = driver.find_element(By.XPATH,var.check_baivietcuaban_trangluotthichsothich1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Trang, lượt thích Trang và sở thích")
        logging.info("check font-end:Có thông báo hay không ")
        logging.info(check_baivietcuaban_trangluotthichsothich != None)
        driver.back()
        driver.find_element(By.XPATH, var.baivietcuaban_binhluan).click()
        time.sleep(3)
        check_baivietcuaban_binhluan = driver.find_element(By.XPATH,var.check_baivietcuaban_binhluan1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Bình luận")
        logging.info("check font-end:Có thông báo hay không ")
        logging.info(check_baivietcuaban_binhluan != None)
        driver.back()
        driver.find_element(By.XPATH, var.baivietcuaban_luotthichvacamxuc).click()
        time.sleep(3)
        check_baivietcuaban_luotthichvacamxuc = driver.find_element(By.XPATH,var.check_baivietcuaban_luotthichvacamxuc1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Lượt thích và cảm xúc")
        logging.info("check font-end:Có thông báo hay không ")
        logging.info(check_baivietcuaban_luotthichvacamxuc != None)
        driver.back()
        driver.find_element(By.XPATH, var.baivietcuaban_baivietluotcheckinanhvavideo).click()
        time.sleep(3)
        check_baivietcuaban_baivietluotcheckinanhvavideo = driver.find_element(By.XPATH,var.check_baivietcuaban_baivietluotcheckinanhvavideo1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Bài viết, lượt checkin, ảnh và video của bạn")
        logging.info("check font-end:Có thông báo hay không ")
        logging.info(check_baivietcuaban_baivietluotcheckinanhvavideo != None)
        driver.back()
        driver.find_element(By.XPATH, var.baivietcuaban_baivietcuabantrendongthoigiannguoikhac).click()
        time.sleep(2)
        check_baivietcuaban_baivietcuabantrendongthoigiannguoikhac = driver.find_element(By.XPATH,var.check_baivietcuaban_baivietcuabantrendongthoigiannguoikhac1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Bài viết của bạn trên dòng thời gian của người khác")
        logging.info("check font-end:Có thông báo hay không ")
        logging.info(check_baivietcuaban_baivietcuabantrendongthoigiannguoikhac != None)
        driver.back()
        driver.find_element(By.XPATH, var.baivietcuaban_baivietnguoikhacdanglenfeedcuaban).click()
        time.sleep(2)
        check_baivietcuaban_baivietnguoikhacdanglenfeedcuaban = driver.find_element(By.XPATH,var.check_baivietcuaban_baivietnguoikhacdanglenfeedcuaban1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Bài viết người khác đăng lên bảng feed của bạn")
        logging.info("check font-end:Có thông báo hay không ")
        logging.info(check_baivietcuaban_baivietnguoikhacdanglenfeedcuaban != None)
        driver.back()
        # driver.find_element(By.XPATH, var.baivietcuaban_livestream).click()       #Không có dữ liệu để test
        # time.sleep(2)
        # check_baivietcuaban_livestream = driver.find_element(By.XPATH,var.check_baivietcuaban_livestream1).text
        # logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Livestream")
        # logging.info("Chức năng chưa hoạt động")
        # logging.info("check font-end:Có thông báo hay không ")
        # logging.info(check_baivietcuaban_livestream != "Không có dữ liệu!")
        # driver.back()
        driver.find_element(By.XPATH, var.baivietcuaban_baivietvabinhluangantheban).click()     #ko load duoc trang
        time.sleep(2)
        check_baivietcuaban_baivietvabinhluangantheban = driver.find_element(By.XPATH,var.check_baivietcuaban_baivietvabinhluangantheban1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Bài viết và bình luận có gắn thẻ bạn")
        logging.info("Chức năng chưa hoạt động")
        logging.info("check font-end:Có thông báo hay không ")
        logging.info(check_baivietcuaban_baivietvabinhluangantheban != "Không có dữ liệu!")
        driver.back()
        driver.find_element(By.XPATH, var.baivietcuaban_anhvavideocogantheban).click()
        time.sleep(2)
        check_baivietcuaban_anhvavideocogantheban = driver.find_element(By.XPATH,var.check_baivietcuaban_anhvavideocogantheban1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bài viết của bạn - Ảnh và video có gắn thẻ bạn")
        logging.info("Chức năng chưa hoạt động")
        logging.info("check font-end:Có thông báo hay không ")
        logging.info(check_baivietcuaban_anhvavideocogantheban != "Không có dữ liệu!")
        driver.back()
        #Hoạt động có gắn thẻ bạn
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_hoatdongcogantheban).click()
        time.sleep(2)
        driver.implicitly_wait(3)
        try:
            check_hoatdongcogantheban = driver.find_element(By.XPATH,var.check_hoatdongcogantheban1).is_displayed()
            logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Hoạt động có gắn thẻ bạn")
            logging.info("check font-end:Có thông báo hay không ")
            logging.info(check_hoatdongcogantheban)
        except NoSuchElementException:
            logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Hoạt động có gắn thẻ bạn")
            logging.info("check font-end:Có thông báo hay không ")
            logging.info("False")
        driver.implicitly_wait(15)

        #Bộ sưu tập và mục đã lưu
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_bosuutapvamucdaluu).click()
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_bosuutapvamucdaluu_mucbandaluu).click()
        check_mucbandaluu = driver.find_element(By.XPATH,var.check_mucbandaluu1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bộ sưu tập và mục đã lưu")
        logging.info("check font-end:Mục bạn đã lưu - Có thông báo hay không ")
        logging.info(check_mucbandaluu != "Không có dữ liệu!")
        time.sleep(1)

        driver.find_element(By.XPATH, var.thongtincuabantrenemso_bosuutapvamucdaluu_bosuutap).click()
        check_bosuutap = driver.find_element(By.XPATH,var.check_bosuutap1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Bộ sưu tập và mục đã lưu")
        logging.info("check font-end:Bộ sưu tập Có thông báo hay không ")
        logging.info(check_bosuutap != "Không có dữ liệu!")
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_bosuutapvamucdaluu).click()
        time.sleep(1)

        #Trang
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_trang).click()
        time.sleep(2)
        check_trang = driver.find_element(By.XPATH,var.check_trang1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Trang")
        logging.info("check font-end:Trang Có thông báo hay không ")
        logging.info(check_trang != "Không có dữ liệu!")
        time.sleep(1)

        #các chức năng khác
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_nhomsukienthuocphim).click()
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Nhóm, sự kiện và thước phim")
        logging.info("Chức năng chưa hoạt động")
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_thongtintrentrangcanhan).click()
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Thông tin trên trang cá nhân")
        logging.info("Chức năng chưa hoạt động")
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_quanheketnoi).click()
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Quan hệ kết nối")
        logging.info("Chức năng chưa hoạt động")
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_hoatdongdaghilaivahoatdongkhac).click()
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Hoạt động đã ghi lại và hoạt động khác")
        logging.info("Chức năng chưa hoạt động")
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_kholuutru).click()
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Kho lưu trữ")
        logging.info("Chức năng chưa hoạt động")
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_thungrac).click()
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Thùng rác")
        logging.info("Chức năng chưa hoạt động")
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_lichsuhoatdong).click()
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Lịch sử hoạt động")
        logging.info("Chức năng chưa hoạt động")
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_xemlaidongthoigianvanhthe).click()
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Nhật ký hoạt động - Xem lại dòng thời gian và ảnh thẻ")
        logging.info("Chức năng chưa hoạt động")
        time.sleep(1)
        driver.get("https://cmc-fe.emso.vn/settings/information")
        time.sleep(2)
        #Vô hiẹu hoá tài khoản
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_vohieuhoataikhoan).click()
        time.sleep(0.5)
        check_vohieuquataikhoan = driver.find_element(By.XPATH,var.check_vohieuquataikhoan1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Thông tin của bản trên Emso - Vô hiệu hóa tài khoản")
        logging.info("check font-end: popup - Vô hiệu hóa tài khoản Emso")
        logging.info(check_vohieuquataikhoan == "Vô hiệu hóa tài khoản Emso")
        driver.find_element(By.XPATH, var.huy).click()
        time.sleep(1)
        #Quản lý thông tin của bạn
        driver.find_element(By.XPATH, var.thongtincuabantrenemso_quanlythongtincuaban).click()
        time.sleep(0.5)
        driver.execute_script("window.scrollBy(0,1000)", "")
        time.sleep(2)
        driver.back()
        time.sleep(1)

    def caidatvaquyenriengtu_trangcanhanvaganthe(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        #TRANG CÁ NHÂN VÀ GẮN THẺ
        #Trang cá nhân
        driver.find_element(By.XPATH, var.trangcanhanvaganthe).click()
        #Ai có thể đăng lên trang cá nhân của bạn?
        driver.find_element(By.XPATH, var.trangcanhanvaganthe_1).click()
        driver.find_element(By.XPATH, var.trangcanhan_x).click()
        time.sleep(0.5)
        #Ai có thể xem những gì người khác đăng lên trang cá nhân của bạn?
        driver.find_element(By.XPATH, var.trangcanhanvaganthe_2).click()
        driver.find_element(By.XPATH, var.trangcanhan_x).click()
        time.sleep(0.5)
        #Ẩn bình luận chứa một số từ nhất định khỏi trang của bạn
        driver.find_element(By.XPATH, var.trangcanhanvaganthe_3).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhanvaganthe_3).click()
        time.sleep(0.5)
        #Găn thẻ
        # Ai có thể xem bài viết có gắn thẻ bạn trên trang cá nhân của bạn?
        driver.find_element(By.XPATH, var.trangcanhanvaganthe_4).click()
        driver.find_element(By.XPATH, var.trangcanhan_x).click()
        time.sleep(0.5)
        # Khi bạn được gắn thẻ trong một bài viết, bạn muốn thêm ai vào đối tượng của bài viết nếu họ chưa thể nhìn thấy bài viết?
        driver.find_element(By.XPATH, var.trangcanhanvaganthe_5).click()
        driver.find_element(By.XPATH, var.trangcanhan_x).click()
        time.sleep(0.5)
        #Xem lại
        #Xét duyệt bài viết có gắn thẻ bạn trước khi bài viết đó xuất hiện trên trang cá nhân của bạn
        driver.find_element(By.XPATH, var.trangcanhanvaganthe_6).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhanvaganthe_6).click()
        time.sleep(0.5)
        #Xem lại thẻ mọi người thêm vào bài viết của bạn trước khi thẻ xuất hiện trên EMSO?
        driver.find_element(By.XPATH, var.trangcanhanvaganthe_7).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trangcanhanvaganthe_7).click()
        time.sleep(0.5)

    def caidatvaquyenriengtu_baivietcongkhai(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        #BÀI VIẾT CÔNG KHAI
        driver.find_element(By.XPATH, var.baivietcongkhai).click()
        #Ai có thể theo dõi tôi
        driver.find_element(By.XPATH, var.baivietcongkhai_aicothetheodoitoi).click()
        driver.find_element(By.XPATH, var.baivietcongkhai_x).click()
        time.sleep(1)
        #Bình luận về bài viết công khai
        driver.find_element(By.XPATH, var.baivietcongkhai_binhluanvebaivietcongkhai).click()
        driver.find_element(By.XPATH, var.baivietcongkhai_x).click()
        driver.execute_script("window.scrollBy(0,900)", "")
        time.sleep(1)
        #Thông báo về bài viết công khai
        driver.find_element(By.XPATH, var.baivietcongkhai_thongbaovebaivietcongkhai).click()
        driver.find_element(By.XPATH, var.baivietcongkhai_x).click()
        time.sleep(1)
        #Thông báo công khai trên trang cá nhân
        driver.find_element(By.XPATH, var.baivietcongkhai_thongbaocongkhaitrentrangcanhan).click()
        driver.find_element(By.XPATH, var.baivietcongkhai_x).click()
        time.sleep(1)
        #Bản xem trước ngoài Emso
        driver.find_element(By.XPATH, var.baivietcongkhai_banxemtruocngoaiemsso).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.baivietcongkhai_banxemtruocngoaiemsso).click()
        time.sleep(1)
        #Hiển thị bình luận phù hợp nhất trước tiên
        driver.find_element(By.XPATH, var.baivietcongkhai_hienthiphuhoptruoctien).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.baivietcongkhai_hienthiphuhoptruoctien).click()
        time.sleep(1)
        #Tên người dùng
        driver.find_element(By.XPATH, var.baivietcongkhai_tennguoidung).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.huy).click()
        time.sleep(1)

    def caidatvaquyenriengtu_chan(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        #CHẶN
        driver.find_element(By.XPATH, var.chan).click()
        #Chặn ngưoi dùng
        driver.find_element(By.XPATH, var.chan_channguoidung).click()
        time.sleep(0.5)
        check_channguoidung = driver.find_element(By.XPATH,var.check_channguoidung1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Chặn - Chặn người dùng")
        logging.info("check font-end: Chặn người dùng - chức năng có hoạt động không ")
        logging.info(check_channguoidung)
        logging.info(check_channguoidung == "Bạn đã chặn 1 người")
        driver.find_element(By.XPATH, var.chan_themvaodanhsach).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.x).click()
        time.sleep(1)
        #Chặn tin nhắn
        driver.find_element(By.XPATH, var.chan_tinhan).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.themvaodanhsachchan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chantinnhan_nhaptenmotnguoiban).send_keys("hue nguyen")
        time.sleep(2)
        driver.find_element(By.XPATH, var.themvaodanhsachchan_chan).click()
        time.sleep(1)
        # driver.find_element(By.XPATH, var.themvaodanhsachchan_chan).click()
        # time.sleep(1)
        driver.find_element(By.XPATH, var.chan_tinhan).click()
        time.sleep(0.5)
        check_chantinnhan = driver.find_element(By.XPATH,var.check_chantinnhan1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Chặn - Chặn người dùng")
        logging.info("check font-end: Chặn tin nhắn - chức năng có hoạt động không ")
        logging.info(check_chantinnhan)
        logging.info(check_chantinnhan == "Bạn đã chặn 1 người")
        # driver.find_element(By.XPATH, var.chan_themvaodanhsach).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xemdanhsachchan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.themvaodanhsachchan_bochan).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.icon_x3).click()
        time.sleep(1)
        #Chặn trang
        driver.find_element(By.XPATH, var.chan_trang).click()
        time.sleep(0.5)
        check_chantrang = driver.find_element(By.XPATH,var.check_chantrang1).text
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Chặn - Chặn người dùng")
        logging.info("check font-end: Chặn trang - chức năng có hoạt động không ")
        logging.info(check_chantrang)
        logging.info(check_chantrang == "Bạn đã chặn 1 trang")
        driver.find_element(By.XPATH, var.chan_themvaodanhsach).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.x).click()
        time.sleep(1)

    def caidatvaquyenriengtu_batkiemtien(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        #BẬT KIẾM TIỀN
        driver.find_element(By.XPATH, var.batkiemtien).click()
        # Trần Quang Trường
        driver.find_element(By.XPATH, var.batkiemtien_timkiem).send_keys("Trần Quang Trường")
        time.sleep(2)
        check_batkiemtien_nutbidissable = driver.find_element(By.XPATH,var.check_batkiemtien_nutbidissable).is_enabled() == False
        print(check_batkiemtien_nutbidissable)
        logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Bật kiếm tiền")
        logging.info("check font-end: Tài khoản đã xác minh - Tài khoản đã được phê duyệt - nút bị dissable")
        logging.info(check_batkiemtien_nutbidissable)
        time.sleep(1)

        # driver.find_element(By.XPATH, var.batkiemtien_xacminhdanhtinh).click()
        # time.sleep(1)
        # check_xacminhdanhtinh = driver.find_element(By.XPATH,var.check_xacminhdanhtinh1).text
        # logging.info("Trang chủ - tài khoản - Cài đặt và quyền riêng tư - Bật kiếm tiền")
        # logging.info("check font-end: popup - Xác minh danh tính")
        # logging.info(check_xacminhdanhtinh == "Xác minh danh tính")
        # driver.find_element(By.XPATH, var.xacminhdanhtinh_cccdmattruoc).click()
        # time.sleep(1)
        # subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia1.exe")
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.xacminhdanhtinh_cccdmatsau).click()
        # time.sleep(1)
        # subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia2.exe")
        # driver.find_element(By.XPATH, var.xacminhdanhtinh_video).click()
        # time.sleep(1)
        # subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/quangaygiongbao.exe")
        # time.sleep(3)
        # driver.find_element(By.XPATH, var.x).click()
        # time.sleep(2)

        #Trường test bản tin
        driver.find_element(By.XPATH, var.batkiemtien_timkiem).click()
        imput = driver.find_element(By.XPATH, var.batkiemtien_timkiem)
        imput.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.batkiemtien_timkiem).send_keys("Trường test bản tin")
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.batkiemtien_truongtestbantin).click()
        driver.find_element(By.XPATH, var.batkiemtien_xacminhdanhtinh).click()
        driver.find_element(By.XPATH, var.xacminhdanhtinh_GPKDKD).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia1.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacminhdanhtinh_masothue).send_keys(data['trangchu_banthan']['masothue'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacminhdanhtinh_tenchudoanhnghiep).send_keys(data['trangchu_banthan']['tenchudoanhnghiep'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.x).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(2)

    def trogiupvahotro_sudungemso(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.icon_caidatcanhan).click()
        driver.find_element(By.XPATH, var.caidatcanhan_trogiupvahotro).click()
        time.sleep(2)

        #Sử dụng emso
        driver.find_element(By.XPATH, var.sudungemso).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan1).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan1).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan2).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan2).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan3).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan3).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan4).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan4).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan5).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan5).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan6).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_taotaikhoan6).click()

        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_xacnhantaikhoan1).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_xacnhantaikhoan1).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_xacnhantaikhoan2).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_xacnhantaikhoan2).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_xacnhantaikhoan3).click()
        driver.find_element(By.XPATH, var.sudungemso_cachtaotaikhoan_xacnhantaikhoan3).click()

        #Trang cá nhân của bạn
        #Thêm và chỉnh sửa thông tin trên trang cá nhân của bạn
        driver.find_element(By.XPATH, var.trangcanhancuaban).click()
        driver.find_element(By.XPATH, var.trangcanhancuaban_themvachinhsuatrentrangcanhan).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_thongtincoban1).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_thongtincoban1).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_thongtincoban2).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_thongtincoban2).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_thongtincoban3).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_thongtincoban3).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_thongtincoban4).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_thongtincoban4).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_thongtincoban5).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_thongtincoban5).click()

        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren1).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren1).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren2).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren2).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren3).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren3).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren4).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren4).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren5).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren5).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren6).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren6).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren7).click()
        driver.find_element(By.XPATH, var.themvachinhsuatrentrangcanhan_gioithieutren7).click()

        #Ảnh đại diện và ảnh bìa của bạn
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_1).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_1).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_2).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_2).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_3).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_3).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_4).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_4).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_5).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_5).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_6).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_6).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_7).click()
        driver.find_element(By.XPATH, var.anhdaidienvanhbiacuaban_7).click()

        #Chia sẻ và quản lý bài viết trên trang cá nhân
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_1).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_1).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_2).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_2).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_3).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_3).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_4).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_4).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_5).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_5).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_6).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_6).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_7).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_7).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_8).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_8).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_9).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_9).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_10).click()
        driver.find_element(By.XPATH, var.chiasevaquanlybaiviet_10).click()

        #Khắc phục sự cố
        driver.find_element(By.XPATH, var.khacphucsuco).click()
        driver.find_element(By.XPATH, var.khacphucsuco_1).click()
        driver.find_element(By.XPATH, var.khacphucsuco_1).click()
        driver.find_element(By.XPATH, var.khacphucsuco_2).click()
        driver.find_element(By.XPATH, var.khacphucsuco_2).click()
        driver.find_element(By.XPATH, var.khacphucsuco_3).click()
        driver.find_element(By.XPATH, var.khacphucsuco_3).click()
        driver.find_element(By.XPATH, var.khacphucsuco_4).click()
        driver.find_element(By.XPATH, var.khacphucsuco_4).click()
        driver.find_element(By.XPATH, var.khacphucsuco_5).click()
        driver.find_element(By.XPATH, var.khacphucsuco_5).click()
        driver.find_element(By.XPATH, var.khacphucsuco_6).click()
        driver.find_element(By.XPATH, var.khacphucsuco_6).click()
        driver.find_element(By.XPATH, var.khacphucsuco_7).click()
        driver.find_element(By.XPATH, var.khacphucsuco_7).click()
        driver.find_element(By.XPATH, var.khacphucsuco_8).click()
        driver.find_element(By.XPATH, var.khacphucsuco_8).click()
        driver.find_element(By.XPATH, var.khacphucsuco_9).click()
        driver.find_element(By.XPATH, var.khacphucsuco_9).click()
        driver.find_element(By.XPATH, var.khacphucsuco_10).click()
        driver.find_element(By.XPATH, var.khacphucsuco_10).click()

        #Thêm bạn bè
        driver.find_element(By.XPATH, var.sudungemsso_thembanbe).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sudungemsso_thembanbe_1).click()
        driver.find_element(By.XPATH, var.sudungemsso_thembanbe_2).click()
        driver.find_element(By.XPATH, var.sudungemsso_thembanbe_3).click()
        driver.find_element(By.XPATH, var.sudungemsso_thembanbe_4).click()

        #Trang chủ của bạn
        driver.find_element(By.XPATH, var.sudungemsso_trangchucuaban).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sudungemsso_trangchucuaban_1).click()
        driver.find_element(By.XPATH, var.sudungemsso_trangchucuaban_2).click()
        driver.find_element(By.XPATH, var.sudungemsso_trangchucuaban_3).click()
        driver.find_element(By.XPATH, var.sudungemsso_trangchucuaban_4).click()
        driver.find_element(By.XPATH, var.sudungemsso_trangchucuaban_5).click()

        #Nhắn tin
        driver.find_element(By.XPATH, var.sudungemsso_nhantin).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sudungemsso_nhantin_1).click()
        driver.find_element(By.XPATH, var.sudungemsso_nhantin_2).click()
        driver.find_element(By.XPATH, var.sudungemsso_nhantin_3).click()
        driver.find_element(By.XPATH, var.sudungemsso_nhantin_4).click()

        #Ảnh và video
        driver.find_element(By.XPATH, var.sudungemsso_anhvavideo).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sudungemsso_anhvavideo_1).click()
        driver.find_element(By.XPATH, var.sudungemsso_anhvavideo_2).click()
        driver.find_element(By.XPATH, var.sudungemsso_anhvavideo_3).click()
        driver.find_element(By.XPATH, var.sudungemsso_anhvavideo_4).click()
        driver.find_element(By.XPATH, var.sudungemsso_anhvavideo_5).click()
        driver.find_element(By.XPATH, var.sudungemsso_anhvavideo_6).click()

        #Video trên watch
        driver.find_element(By.XPATH, var.sudungemsso_videotrenwatch).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sudungemsso_videotrenwatch_1).click()
        driver.find_element(By.XPATH, var.sudungemsso_videotrenwatch_2).click()

        #VTrang
        cuon = driver.find_element(By.XPATH, var.sudungemsso_trang_9)
        driver.execute_script("arguments[0].scrollIntoView();", cuon)

        driver.find_element(By.XPATH, var.sudungemsso_trang).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sudungemsso_trang_1).click()
        driver.find_element(By.XPATH, var.sudungemsso_trang_2).click()
        driver.find_element(By.XPATH, var.sudungemsso_trang_3).click()
        driver.find_element(By.XPATH, var.sudungemsso_trang_4).click()
        driver.find_element(By.XPATH, var.sudungemsso_trang_5).click()
        driver.find_element(By.XPATH, var.sudungemsso_trang_6).click()
        driver.find_element(By.XPATH, var.sudungemsso_trang_7).click()
        driver.find_element(By.XPATH, var.sudungemsso_trang_8).click()
        driver.find_element(By.XPATH, var.sudungemsso_trang_9).click()

        #VNhóm
        driver.find_element(By.XPATH, var.sudungemsso_nhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sudungemsso_nhom_1).click()
        driver.find_element(By.XPATH, var.sudungemsso_nhom_2).click()
        driver.find_element(By.XPATH, var.sudungemsso_nhom_3).click()
        driver.find_element(By.XPATH, var.sudungemsso_nhom_4).click()
        driver.find_element(By.XPATH, var.sudungemsso_nhom_5).click()
        driver.find_element(By.XPATH, var.sudungemsso_nhom_6).click()

        #Sự kiện
        driver.find_element(By.XPATH, var.sudungemsso_sukien).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sudungemsso_sukien_1).click()
        driver.find_element(By.XPATH, var.sudungemsso_sukien_2).click()
        driver.find_element(By.XPATH, var.sudungemsso_sukien_3).click()

        #Úng dụng Emso trên di động
        driver.find_element(By.XPATH, var.sudungemsso_emsotrendidong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.sudungemsso_emsotrendidong_1).click()
        driver.find_element(By.XPATH, var.sudungemsso_emsotrendidong_2).click()

        #Trợ năng
        driver.find_element(By.XPATH, var.sudungemsso_tronang).click()
        time.sleep(1)

        #QUẢN LÝ TÀI KHOẢN
        driver.find_element(By.XPATH, var.quanlytaikhoan).click()
        time.sleep(1)
        #đăng nhập và mật khẩu
        driver.find_element(By.XPATH, var.quanlytaikhoan_dangnhapvamatkhau).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_dangnhapvamatkhau_1).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_dangnhapvamatkhau_2).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_dangnhapvamatkhau_3).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_dangnhapvamatkhau_4).click()

        #cài đặt tài khoản
        driver.find_element(By.XPATH, var.quanlytaikhoan_caidattaikhoan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlytaikhoan_caidattaikhoan_1).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_caidattaikhoan_2).click()

        #Thông báo
        driver.find_element(By.XPATH, var.quanlytaikhoan_thongbao).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlytaikhoan_thongbao_1).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_thongbao_2).click()

        #Vô hiệu hoá hoặc xoá tài khoản của bạn
        driver.find_element(By.XPATH, var.quanlytaikhoan_vohieuhoahoacxoataikhoan).click()
        time.sleep(1)

        #QUYỀN RIÊNG TƯ AN TOÀN VÀ BẢO MẬT
        driver.find_element(By.XPATH, var.quyenriengtuantoanbaomat).click()
        time.sleep(1)
        #Quyền riêng tư của bạn
        driver.find_element(By.XPATH, var.quanlytaikhoan_quyenriengtucuaban).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlytaikhoan_quyenriengtucuaban_1).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_quyenriengtucuaban_2).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_quyenriengtucuaban_3).click()

        #Giữ an toàn
        driver.find_element(By.XPATH, var.quanlytaikhoan_giuantoan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlytaikhoan_giuantoan_1).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_giuantoan_2).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_giuantoan_3).click()

        #Duy trì tính bảo mật cho tài khoản của bạn
        driver.find_element(By.XPATH, var.quanlytaikhoan_duytribaomat).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlytaikhoan_duytribaomat_1).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_duytribaomat_2).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_duytribaomat_3).click()

        #An toàn khi mua sắm
        driver.find_element(By.XPATH, var.quanlytaikhoan_antoankhimuasam).click()
        time.sleep(1)
        logging.info("Trợ giúp và hỗ trợ - Quyền riêng tư, an toàn và bảo mật - An toàn khi mua sắm")
        logging.info("Chức năng chưa hoạt động")
        driver.find_element(By.XPATH, var.quanlytaikhoan_antoankhimuasam_1).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_antoankhimuasam_2).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_antoankhimuasam_3).click()
        driver.find_element(By.XPATH, var.quanlytaikhoan_antoankhimuasam_4).click()
        time.sleep(1)

        #CHÍNH SÁCH VÀ BÁO CÁO
        driver.find_element(By.XPATH, var.chinhsachvabaocao).click()
        time.sleep(1)
        #Báo cáo lạm dụng
        driver.find_element(By.XPATH, var.chinhsachvabaocao_baocaolamdung).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsachvabaocao_baocaolamdung_1).click()
        driver.find_element(By.XPATH, var.chinhsachvabaocao_baocaolamdung_2).click()
        #Báo cáo sự cố với Emso
        driver.find_element(By.XPATH, var.chinhsachvabaocao_baocaosucovoiemso).click()
        #Báo cáo vi phạm quyền riêng tư
        driver.find_element(By.XPATH, var.chinhsachvabaocao_baocaoviphamquyenriengtu).click()
        #Tài khoản hack và tài khoản giả
        driver.find_element(By.XPATH, var.chinhsachvabaocao_taikhoanhackvataikhoangia).click()
        #Quyền sở hữu trí tuệ
        driver.find_element(By.XPATH, var.chinhsachvabaocao_quyensohuutritue).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsachvabaocao_quyensohuutritue_1).click()
        driver.find_element(By.XPATH, var.chinhsachvabaocao_quyensohuutritue_2).click()
        time.sleep(1)

        #ĐIỀU KHOAN VÀ CHÍNH SÁCH
        driver.find_element(By.XPATH, var.dieukhoanvachinhsach).click()
        #Điều khoản dịch vụ
        driver.find_element(By.XPATH, var.dieukhoanvachinhsach_dieukhoandichvu).click()
        time.sleep(0.5)
        #Chính sách riêng tư và tiêu chuẩn bảo mật
        driver.find_element(By.XPATH, var.dieukhoanvachinhsach_chinhsachriengtuvatieuchuanbaomat).click()
        time.sleep(0.5)
        #Tiêu chuẩn cộng đồng
        driver.find_element(By.XPATH, var.dieukhoanvachinhsach_tieuchuancongdong).click()
        time.sleep(0.5)
        #Chính sách quảng cáo
        driver.find_element(By.XPATH, var.dieukhoanvachinhsach_chinhsachquangcao).click()
        time.sleep(0.5)
        #Báo cáo khiếu nại
        driver.find_element(By.XPATH, var.dieukhoanvachinhsach_baobaokhieunai).click()
        time.sleep(1)

        #ĐIỀU KHOẢN VÀ CHÍNH SÁCH RIÊNG
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng).click()
        #Tuyển dụng
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_tuyendung).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_tuyendung_1).click()
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_tuyendung_2).click()
        time.sleep(0.5)
        #Gọi vốn cộng đồng
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_goivoncongdong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_goivoncongdong_1).click()
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_goivoncongdong_2).click()
        time.sleep(0.5)
        #Không gian âm nhạc
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_khonggianamnhac).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_khonggianamnhac_1).click()
        time.sleep(0.5)
        #Không gian học tập
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_khonggianhoctap).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_khonggianhoctap_1).click()
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_khonggianhoctap_2).click()
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_khonggianhoctap_3).click()
        time.sleep(0.5)
        #Không gian thương mại
        driver.find_element(By.XPATH, var.dieukhoanvachinhsachrieng_khonggianthuongmai).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(2)

    def caidatcanhan_manhinh(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.icon_caidatcanhan).click()
        driver.find_element(By.XPATH, var.caidatcanhan_manhinh).click()
        time.sleep(1)
        check_chedomanhinh = driver.find_element(By.XPATH,var.check_chedomanhinh1).text
        print(check_chedomanhinh)
        logging.info("Trang chủ - Tài Khoản - Màn hình")
        logging.info("check font-end: popup - Chế độ màn hình")
        logging.info(check_chedomanhinh == "Chế độ màn hình")
        driver.find_element(By.XPATH, var.caidatcanhan_manhinh_bat).click()
        time.sleep(2)
        check_bat = driver.find_element(By.XPATH, var.caidatcanhan_manhinh_bat)
        print("Có đang được bật không", check_bat.is_enabled())
        print("Có đang được chọn không", check_bat.is_selected())
        check_bat2 = driver.find_element(By.XPATH, var.caidatcanhan_manhinh_bat).is_enabled()
        logging.info("Màn hình tối có được chọn không?")
        logging.info(check_bat2)
        check_bat1 = driver.find_element(By.XPATH, var.caidatcanhan_manhinh_bat).is_selected()
        logging.info("Màn hình tối có được bật không?")
        logging.info(check_bat1)

        driver.find_element(By.XPATH, var.caidatcanhan_manhinhtudong).click()
        time.sleep(1)
        check_tudong1 = driver.find_element(By.XPATH, var.caidatcanhan_manhinhtudong).is_selected()
        logging.info("Màn hình tự động có được chọn không?")
        logging.info(check_tudong1)
        check_tudong2 = driver.find_element(By.XPATH, var.caidatcanhan_manhinhtudong).is_enabled()
        logging.info("Màn hình tự động có được bật không?")
        logging.info(check_tudong2)

        driver.find_element(By.XPATH, var.caidatcanhan_manhinh_tat).click()
        time.sleep(1)

        check_tat1 = driver.find_element(By.XPATH, var.caidatcanhan_manhinh_tat).is_selected()
        logging.info("Màn hình tối(tắt) có được chọn không?")
        logging.info(check_tat1)
        check_tat2 = driver.find_element(By.XPATH, var.caidatcanhan_manhinh_tat).is_enabled()
        logging.info("Màn hình tối(tắt) có được bật không?")
        logging.info(check_tat2)

        driver.find_element(By.XPATH, var.caidatcanhan_manhinh_iconquaylai).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.icon_caidatcanhan).click()
        time.sleep(1)

    def caidatcanhan_donggopykien(self):
        driver.implicitly_wait(15)
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.icon_caidatcanhan).click()
        driver.find_element(By.XPATH, var.caidatcanhan_donggopykien).click()
        time.sleep(1)
        #Chung tay cải thiện emso
        driver.find_element(By.XPATH, var.donggopykien_chungtaycaithienemso).click()
        driver.find_element(By.XPATH, var.chungtaycaithienemso_chonsanpham).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chungtaycaithienemso_chonsanpham_choigame).click()
        driver.find_element(By.XPATH, var.chungtaycaithienemso_chitiet).click()
        driver.find_element(By.XPATH, var.chungtaycaithienemso_chitiet).send_keys(data['trangchu_donggopykien']['chitiet'])
        time.sleep(1)
        # driver.find_element(By.XPATH, var.chungtaycaithienemso_tailenfile1).click()
        button = driver.find_element(By.XPATH, var.chungtaycaithienemso_tailenfile1)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia2.exe")
        time.sleep(1)
        check_gui2 = driver.find_element(By.XPATH, var.donggopykien_chungtaycaithienemso_buttongui).is_enabled()
        logging.info("Trang chủ - Tài khoản - Đóng góp ý kiến - Chung tay cải thiện EMSO")
        logging.info("Nút GỬI đang ở trạng thái bật không?")
        logging.info(check_gui2)
        time.sleep(2)
        driver.find_element(By.XPATH, var.gui).click()
        # driver.find_element(By.XPATH, var.chungtaycaithienemso_huy).click()
        time.sleep(1)
        check_donggopykien_chungtay = driver.find_element(By.XPATH, var.check_donggopykien_chungtay).text
        logging.info("Trang chủ - Tài khoản - Đóng góp ý kiến - Chung tay cải thiện EMSO")
        logging.info("Check font-end: Messgae - Gửi đóng góp thành công.")
        logging.info(check_donggopykien_chungtay == "Gửi đóng góp thành công.")
        time.sleep(1)

        #Đã xảy ra lỗi
        driver.find_element(By.XPATH, var.caidatcanhan_donggopykien).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.donggopykien_daxayraloi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chungtaycaithienemso_chonsanpha1).click()
        # driver.find_element(By.XPATH, var.chungtaycaithienemso_chonsanpham_chatdongdong).click()
        button = driver.find_element(By.XPATH, var.chungtaycaithienemso_chonsanpham_chatdongdong)
        driver.execute_script("arguments[0].click();", button)

        # driver.find_element(By.XPATH, var.chungtaycaithienemso_chitiet).click()
        button = driver.find_element(By.XPATH, var.chungtaycaithienemso_chitiet)
        driver.execute_script("arguments[0].click();", button)

        driver.find_element(By.XPATH, var.chungtaycaithienemso_chitiet).clear()
        driver.find_element(By.XPATH, var.chungtaycaithienemso_chitiet).send_keys(data['trangchu_donggopykien']['chitiet_daxayraloi'])
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.chungtaycaithienemso_tailenfile1)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia2.exe")
        time.sleep(2)
        check_gui2 = driver.find_element(By.XPATH, var.donggopykien_chungtaycaithienemso_buttongui).is_enabled()
        logging.info("Trang chủ - Tài khoản - Đóng góp ý kiến - Đã xảy ra lỗi")
        logging.info("Nút GỬI đang ở trạng thái bật không?")
        logging.info(check_gui2)
        time.sleep(1)
        # driver.find_element(By.XPATH, var.chungtaycaithienemso_huy).click()
        # driver.find_element(By.XPATH, var.caidatcanhan_donggopykien_x).click()
        driver.find_element(By.XPATH, var.gui).click()
        time.sleep(1)
        check_donggopykien_daxayraloi = driver.find_element(By.XPATH, var.check_donggopykien_chungtay1).text
        logging.info("Trang chủ - Tài khoản - Đóng góp ý kiến - Đã xảy ra lỗi")
        logging.info("Check font-end: Messgae - Gửi báo lỗi thành công")
        logging.info(check_donggopykien_daxayraloi == "Gửi báo lỗi thành công")
        driver.find_element(By.XPATH, var.icon_caidatcanhan).click()


    def caidatcanhan_dangxuat(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.icon_caidatcanhan).click()
        driver.find_element(By.XPATH, var.caidatcanhan_dangxuat).click()
        time.sleep(1)
        check_dangxuat = driver.find_element(By.XPATH, var.check_dangxuat).text
        print(check_dangxuat)
        logging.info("Trang chủ - Tài khoản - Đăng xuất")
        logging.info(check_dangxuat == "Đăng nhập")
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(2.5)

    def trangchu_timkiem(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu).click()
        driver.find_element(By.XPATH, var.trangchu_timkiem).click()
        driver.find_element(By.XPATH, var.trangchu_timkiem1).send_keys("Phương")
        driver.find_element(By.XPATH, var.trangchu_timkiem_nam).click()
        time.sleep(3)
        check_ketquatimkiem = driver.find_element(By.XPATH, var.check_ketquatimkiem1).text

        #tất cả
        check_timkiem_tatca = driver.find_element(By.XPATH,var.check_timkiem_tatca1).text
        logging.info("Trang chủ - Tìm kiếm - Tất cả")
        logging.info("check font-end: Có kết quả tìm kiếm bộ lọc tất cả")
        logging.info(check_timkiem_tatca !=None)

        #Mọi người
        driver.find_element(By.XPATH, var.trangchu_timkiem_moinguoi).click()
        time.sleep(2)
        check_timkiem_moinguoi = driver.find_element(By.XPATH,var.check_timkiem_tatca1).text
        logging.info("Trang chủ - Tìm kiếm - Mọi người")
        logging.info("check font-end: Có kết quả tìm kiếm bộ lọc mọi người")
        logging.info(check_timkiem_moinguoi)
        logging.info(check_timkiem_moinguoi == "Vân Phương")

        #Bài viết
        driver.find_element(By.XPATH, var.timkiem_input).click()
        xoa = driver.find_element(By.XPATH, var.timkiem_input1)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.timkiem_input1).send_keys("nam")
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.timkiem_input1).send_keys(Keys.ENTER)
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_timkiem_baiviet).click()
        time.sleep(2)
        check_timkiem_baiviet = driver.find_element(By.XPATH,var.check_timkiem_baiviet1).text
        logging.info("Trang chủ - Tìm kiếm - Bài viết")
        logging.info("check font-end: Có kết quả tìm kiếm bộ lọc bài viết")
        logging.info(check_timkiem_baiviet)
        logging.info(check_timkiem_baiviet == "Réponse à @inh.tr hôm nay về Việt Nam rồi 🥳❤️")

        #Nhóm
        driver.find_element(By.XPATH, var.trangchu_timkiem_nhom).click()
        time.sleep(2)
        check_timkiem_nhom = driver.find_element(By.XPATH,var.check_timkiem_nhom1).text
        logging.info("Trang chủ - Tìm kiếm - Nhóm")
        logging.info("check font-end: Có kết quả tìm kiếm bộ lọc Nhóm")
        logging.info(check_timkiem_nhom == "nam test")

        #Trang
        driver.find_element(By.XPATH, var.trangchu_timkiem_trang).click()
        time.sleep(2)
        check_timkiem_trang = driver.find_element(By.XPATH,var.check_timkiem_trang1).text
        logging.info("Trang chủ - Tìm kiếm - Trang")
        logging.info("check font-end: Có kết quả tìm kiếm bộ lọc Trang")
        logging.info(check_timkiem_trang)
        logging.info(check_timkiem_trang == "Nhã Nam")

        #Sự kiện
        driver.find_element(By.XPATH, var.trangchu_timkiem_sukien).click()
        time.sleep(2)
        check_timkiem_sukien = driver.find_element(By.XPATH,var.check_timkiem_sukien1).text
        logging.info("Trang chủ - Tìm kiếm - Sự kiện")
        logging.info("check font-end: Có kết quả tìm kiếm bộ lọc Sự kiện")
        logging.info(check_timkiem_sukien == "Offline Fan Harry Potter 20 năm")

        #Gọi vốn
        driver.find_element(By.XPATH, var.trangchu_timkiem_goivon).click()
        time.sleep(2)
        check_timkiem_goivon = driver.find_element(By.XPATH,var.check_timkiem_goivon1).text
        logging.info("Trang chủ - Tìm kiếm - Gọi vốn")
        logging.info("check font-end: Có kết quả tìm kiếm bộ lọc Gọi vốn")
        logging.info(check_timkiem_goivon == "nam test tìm kiếm")

        # #Tuyển dụng
        driver.find_element(By.XPATH, var.trangchu_timkiem_tuyendung).click()
        time.sleep(2)
        check_timkiem_tuyendung = driver.find_element(By.XPATH,var.check_timkiem_tuyendung1).text
        logging.info("Trang chủ - Tìm kiếm - Tuyển dụng")
        logging.info("check font-end: Có kết quả tìm kiếm bộ lọc Tuyển dụng")
        logging.info(check_timkiem_tuyendung == "tuyển nam")

        #Khoá học
        driver.find_element(By.XPATH, var.trangchu_timkiem_khoahoc).click()
        time.sleep(2)
        check_timkiem_khoahoc = driver.find_element(By.XPATH,var.check_timkiem_khoahoc1).text
        logging.info("Trang chủ - Tìm kiếm - Khoá học")
        logging.info("check font-end: Có kết quả tìm kiếm bộ lọc Khoá học")
        logging.info(check_timkiem_khoahoc == "học nam")
        #
        # #San phẩm
        driver.find_element(By.XPATH, var.trangchu_timkiem_sanpham).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.timkiem_input).click()
        # driver.find_element(By.XPATH, var.timkiem_input).clear()
        xoa = driver.find_element(By.XPATH, var.timkiem_input1)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.timkiem_input1).send_keys("Áo polo nam")
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.timkiem_input1).send_keys(Keys.ENTER)
        time.sleep(2)

        check_timkiem_sanpham = driver.find_element(By.XPATH,var.check_timkiem_sanpham1).text
        logging.info("Trang chủ - Tìm kiếm - Sản phẩm")
        logging.info("check font-end: Có kết quả tìm kiếm bộ lọc Sản phẩm")
        logging.info(check_timkiem_sanpham)
        logging.info(check_timkiem_sanpham == "Áo Polo Nam")
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(2)



    # def youtobe(self):
    #     driver.implicitly_wait(15)
    #     time.sleep(1.5)
    #     driver.get("https://www.youtube.com/")
    #     time.sleep(2)
    #     driver.find_element(By.XPATH, "//*[@placeholder='Search']").send_keys("khoảnh khắc thu vị")
    #     driver.find_element(By.XPATH, "//*[@placeholder='Search']").submit()
    #     time.sleep(2)
    #     button = driver.find_element(By.XPATH, "//*[text()='Filters']")
    #     driver.execute_script("arguments[0].click();", button)
    #     time.sleep(1)
    #     # driver.find_element(By.XPATH, "//*[text()='Filters']").click()
    #     driver.find_element(By.XPATH, "//*[text()='Under 4 minutes']").click()
    #     time.sleep(3)
    #     danhsachvideo = driver.find_elements(By.XPATH, "//*[@class='yt-simple-endpoint inline-block style-scope ytd-thumbnail']")
    #     r = 0
    #     while r<1002:
    #         for video in danhsachvideo:
    #             r += 1
    #             link = video.get_attribute('href')
    #             writeData(var.linkvideo, "Sheet1", r, 1, link)
    #             print(link)
    #             driver.execute_script("window.scrollBy(0,300)", "")
    #             time.sleep(0.3)
    #         if  r ==1000:
    #             break



class khoanhkhac():
    def dangtheodoi(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        # driver.find_element(By.XPATH, var.icon_khoanhkhac).click()
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #đang theo dõi
        driver.find_element(By.XPATH, var.dangtheodoi).click()
        time.sleep(3)
        try:
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        except:
            driver.refresh()
            time.sleep(1)
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.like).click()
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_thich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_binhluan).send_keys(data['khoanhkhac']['binhluan'])
        driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_binhluan).submit()
        time.sleep(1)

        #chia se ngay
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase).click()
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiasengay).click()
        time.sleep(2)
        check_message_chiasengay = driver.find_element(By.XPATH,var.check_message_chiase).text
        print(check_message_chiasengay)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ ngay")
        logging.info("check font-end: Message - Chia sẻ bài viết thành công")
        logging.info(check_message_chiasengay == "Chia sẻ bài viết thành công")
        time.sleep(5)
        driver.get("https://cmc-fe.emso.vn/user/truongvck333")
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,1100)", "")
        check_chiasengay_time = driver.find_element(By.XPATH,var.check_chiasengay1).text
        print(check_chiasengay_time)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ ngay")
        logging.info(check_chiasengay_time)
        logging.info("check font-end: Message - Check Thời gian trên trang cá nhân - Vài giây trước")
        logging.info(check_chiasengay_time == "Vài giây trước ")

        check_chiasengay_tieude = driver.find_element(By.XPATH,var.check_chiasengay_tieude1).text
        print(check_chiasengay_tieude)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ ngay")
        logging.info(check_chiasengay_tieude)
        logging.info("check font-end: Message - Check Tiêu đề trên trang cá nhân - Test khoảng khắc")
        logging.info(check_chiasengay_tieude == "Test khoảng khắc")
        time.sleep(1)

        #Tuỳ chọn bài viết
        #Ghim
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_gimbaiviet).click()
        time.sleep(1)
        check_tuychonbaiviet_chinhsuabaiviet_ghim = driver.find_element(By.XPATH,var.check_tuychonbaiviet_chinhsuabaiviet_ghim1).text
        print(check_tuychonbaiviet_chinhsuabaiviet_ghim)
        logging.info("Trang cá nhân - Tuỳ chọn bài viết - Chỉnh sửa bài viết")
        logging.info("check font-end: Message Ghim - Đã ghim bài viết của bạn ")
        logging.info(check_tuychonbaiviet_chinhsuabaiviet_ghim == "Đã ghim bài viết của bạn")

        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon1).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_bogimbaiviet).click()
        time.sleep(1)
        check_tuychonbaiviet_chinhsuabaiviet_boghim = driver.find_element(By.XPATH,var.check_tuychonbaiviet_chinhsuabaiviet_boghim1).text
        print(check_tuychonbaiviet_chinhsuabaiviet_boghim)
        logging.info("Trang cá nhân - Tuỳ chọn bài viết - Chỉnh sửa bài viết")
        logging.info("check font-end: Message Ghim - Đã bỏ ghim bài viết của bạn")
        logging.info(check_tuychonbaiviet_chinhsuabaiviet_boghim == "Đã bỏ ghim bài viết của bạn")
        time.sleep(1)

        #lưu bài viết
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_luubaiviet).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_luubaiviet_no1).click()
        driver.find_element(By.XPATH, var.xong).click()
        time.sleep(1)
        check_tuychonbaiviet_luubaiviet = driver.find_element(By.XPATH,var.check_tuychonbaiviet_luubaiviet1).text
        print(check_tuychonbaiviet_luubaiviet)
        logging.info("Trang cá nhân - Tuỳ chọn bài viết - Lưu bài viết")
        logging.info("check font-end: Message - Đã lưu vào no1")
        logging.info(check_tuychonbaiviet_luubaiviet == "Đã lưu vào no1")

        #Chỉnh sửa bài viết
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_chinhsuabaiviet).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_chinhsuabaiviet_input).send_keys(data['khoanhkhac']['noidung2'])
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        check_tuychonbaiviet_chinhsuabaiviet = driver.find_element(By.XPATH,var.check_tuychonbaiviet_chinhsuabaiviet1).text
        print(check_tuychonbaiviet_chinhsuabaiviet)
        logging.info("Trang cá nhân - Tuỳ chọn bài viết - Chỉnh sửa bài viết")
        logging.info("check font-end: Message - Chỉnh sửa bài viết thành công")
        logging.info(check_tuychonbaiviet_chinhsuabaiviet == "Chỉnh sửa bài viết thành công")

        #Ai có thể bình luận về bài viết này?
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_aicothebinhluan).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_aicothebinhluan_banbe).click()
        driver.find_element(By.XPATH, var.xong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_aicothebinhluan).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_aicothebinhluan_tcnvtbnd).click()
        driver.find_element(By.XPATH, var.xong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_aicothebinhluan).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_aicothebinhluan_moinguoi).click()
        driver.find_element(By.XPATH, var.xong).click()
        time.sleep(1)

        #Bật thông báo về bài viết này
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_batthongbao).click()
        check_tuychonbaiviet_batthongbao = driver.find_element(By.XPATH,var.check_tuychonbaiviet_batthongbao1).text
        print(check_tuychonbaiviet_batthongbao)
        logging.info("Trang cá nhân - Tuỳ chọn bài viết - Bật thông báo về bài viết này")
        logging.info("check font-end: Message - Đã bật thông báo")
        logging.info(check_tuychonbaiviet_batthongbao == "Đã bật thông báo")
        time.sleep(1)
        #Tắt thông báo về bài viết này
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.check_tuychonbaiviet_tatthongbao11).click()
        check_tuychonbaiviet_tatthongbao = driver.find_element(By.XPATH,var.check_tuychonbaiviet_tatthongbao1).text
        print(check_tuychonbaiviet_tatthongbao)
        logging.info("Trang cá nhân - Tuỳ chọn bài viết - Tắt thông báo về bài viết này")
        logging.info("check font-end: Message - Đã tắt thông báo")
        logging.info(check_tuychonbaiviet_tatthongbao == "Đã tắt thông báo")
        #Nhúng
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_nhung).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_nhung_saochep).click()
        driver.find_element(By.XPATH, var.x).click()
        time.sleep(1)

        #Chỉnh sửa ngày
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_chinhsuangay).click()
        xoa = driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_chinhsuangay_2022)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_chinhsuangay_2022).send_keys("27-10-2022")
        driver.find_element(By.XPATH, var.xong).click()
        time.sleep(1)

        check_tuychonbaiviet_chinhsuangay = driver.find_element(By.XPATH,var.check_tuychonbaiviet_chinhsuangay1).text
        print(check_tuychonbaiviet_chinhsuangay)
        logging.info("Trang cá nhân - Tuỳ chọn bài viết - Chỉnh sửa ngày")
        logging.info("check font-end: Message - Chỉnh sửa ngày thành công.")
        logging.info(check_tuychonbaiviet_chinhsuangay == "Chỉnh sửa ngày thành công.")

        #Xoá bài viết
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_xoabaiviet).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(1)
        check_tuychonbaiviet_xoa = driver.find_element(By.XPATH,var.check_tuychonbaiviet_xoa1).text
        print(check_tuychonbaiviet_xoa)
        logging.info("Trang cá nhân - Tuỳ chọn bài viết - Tắt thông báo về bài viết này")
        logging.info("check font-end: Message - Bài viết của bạn đã bị xóa.")
        logging.info(check_tuychonbaiviet_xoa == "Bài viết của bạn đã bị xóa.")
        time.sleep(1.5)

        # chia sẻ lên bảng tin
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        button = driver.find_element(By.XPATH, var.dangtheodoi)
        driver.execute_script("arguments[0].click();", button)
        # driver.find_element(By.XPATH, var.dangtheodoi).click()

        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.dangtheodoi).click()
        time.sleep(3)
        try:
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        except:
            driver.refresh()
            time.sleep(1)
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        time.sleep(3)
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase).click()
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselenbangtin).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselenbangtin_mota).send_keys(data['khoanhkhac']['binhluan'])
        time.sleep(1)
        try:
            driver.find_element(By.XPATH, var.dang).click()
            time.sleep(3)
            check_message_chiaselenbbangtin = driver.find_element(By.XPATH,var.check_message_chiase).text
            print(check_message_chiaselenbbangtin)
            logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên bảng tin")
            logging.info("check font-end: Message - Chia sẻ bài viết thành công")
            logging.info(check_message_chiaselenbbangtin == "Chia sẻ bài viết thành công")
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase).click()
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselenbangtin).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselenbangtin_mota).send_keys(data['khoanhkhac']['binhluan'])
            time.sleep(1)
            driver.find_element(By.XPATH, var.dang).click()
            time.sleep(3)
            check_message_chiaselenbbangtin = driver.find_element(By.XPATH,var.check_message_chiase).text
            print(check_message_chiaselenbbangtin)
            logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên bảng tin")
            logging.info("check font-end: Message - Chia sẻ bài viết thành công")
            logging.info(check_message_chiaselenbbangtin == "Chia sẻ bài viết thành công")

        time.sleep(5)
        driver.get("https://cmc-fe.emso.vn/user/truongvck333")
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,1100)", "")
        check_chiasengay_time = driver.find_element(By.XPATH,var.check_chiasengay1).text
        print(check_chiasengay_time)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên bảng tin")
        logging.info("check font-end: Message - Check Thời gian trên trang cá nhân - Vài giây trước")
        logging.info(check_chiasengay_time == "Vài giây trước ")

        check_chiasengay_tieude = driver.find_element(By.XPATH,var.check_chiasengay_tieude1).text
        print(check_chiasengay_tieude)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên bảng tin")
        logging.info("check font-end: Message - Check Tiêu đề trên trang cá nhân - Test khoảng khắc")
        logging.info(check_chiasengay_tieude == "Test khoảng khắc")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_chinhsuabaiviet).click()
        driver.find_element(By.XPATH, var.trangcanhan_baiviet_tuychon_chinhsuabaiviet_input).send_keys(data['khoanhkhac']['noidung3'])
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)

        # #Gửi bằng message
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.roikhoi).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.dangtheodoi).click()
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.dangtheodoi).click()
        time.sleep(3)
        try:
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        except:
            driver.refresh()
            time.sleep(1)
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        time.sleep(3)
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase).click()
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_guibangmessage).click()
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_guibangmessage_noidung).send_keys(data['khoanhkhac']['noidung1'])
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_guibangmessage_timkiemtqt).send_keys(data['khoanhkhac']['tranquangtruong'])
        driver.find_element(By.XPATH, var.gui).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.x).click()
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/messages/111340222342420143111504657986364400")
        time.sleep(2)
        try:
            check_chiase_guibangmessage = driver.find_element(By.XPATH,var.check_chiase_guibangmessage2).text
            logging.info(check_chiase_guibangmessage)
            logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Gửi bằng message")
            logging.info("check font-end: chat - có gửi link khoảnh khắc")
            logging.info(check_chiase_guibangmessage == "Mạng xã hội Emso - Mạng xã hội vì người Việt")
        except:
            driver.get("https://cmc-fe.emso.vn/messages/111340222342420143111504657986364400")
            time.sleep(2)
            check_chiase_guibangmessage = driver.find_element(By.XPATH,var.check_chiase_guibangmessage2).text
            logging.info(check_chiase_guibangmessage)
            logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Gửi bằng message")
            logging.info("check font-end: chat - có gửi link khoảnh khắc")
            logging.info(check_chiase_guibangmessage == "Mạng xã hội Emso - Mạng xã hội vì người Việt")

        # #Chia sẻ lên cộng đồng
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.dangtheodoi).click()
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.dangtheodoi).click()
        time.sleep(3)
        try:
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        except:
            driver.refresh()
            time.sleep(1)
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        time.sleep(3)
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase).click()
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselencongdong).click()
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselencongdong_timkiem).send_keys(data['khoanhkhac']['chiaselennhom'])
        driver.find_element(By.XPATH, var.chiase_iconchonnhom).click()
        time.sleep(0.5)
        try:
            driver.find_element(By.XPATH, var.chiase_iconchonnhom_mota).send_keys(data['khoanhkhac']['mota1'])
            driver.find_element(By.XPATH, var.dang).click()
            time.sleep(2)
            check_message_chiaselencongdong = driver.find_element(By.XPATH,var.check_message_chiaselencongdong1).text
            print(check_message_chiaselencongdong)
            logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên cộng đồng")
            logging.info("check font-end: Message - Chia sẻ bài viết thành công")
            logging.info(check_message_chiaselencongdong == "Chia sẻ bài viết thành công")
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase).click()
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselencongdong).click()
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselencongdong_timkiem).send_keys(data['khoanhkhac']['chiaselennhom'])
            driver.find_element(By.XPATH, var.chiase_iconchonnhom).click()
            time.sleep(0.5)
            driver.find_element(By.XPATH, var.chiase_iconchonnhom_mota).send_keys(data['khoanhkhac']['mota1'])
            driver.find_element(By.XPATH, var.dang).click()
            time.sleep(2)
            check_message_chiaselencongdong = driver.find_element(By.XPATH,var.check_message_chiaselencongdong1).text
            print(check_message_chiaselencongdong)
            logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên cộng đồng")
            logging.info("check font-end: Message - Chia sẻ bài viết thành công")
            logging.info(check_message_chiaselencongdong == "Chia sẻ bài viết thành công")
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,500)", "")
        check_chiaselencongdong_time = driver.find_element(By.XPATH,var.check_chiaselencongdong_time1).text
        print(check_chiaselencongdong_time)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên cộng đồng")
        logging.info("check font-end: Message - Check Thời gian trong nhóm - Vài giây trước")
        logging.info(check_chiaselencongdong_time == "Vài giây trước ")

        check_chiaselencongdong_tieude = driver.find_element(By.XPATH,var.check_chiaselencongdong_tieude1).text
        print(check_chiaselencongdong_tieude)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên cộng đồng")
        logging.info("check font-end: Message - Check Tiêu đề trong nhóm - nutri boost")
        logging.info(check_chiaselencongdong_tieude == "nutri boost")
        time.sleep(2)

        # #Chia sẻ lên trang
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.dangtheodoi).click()
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.dangtheodoi).click()
        time.sleep(3)
        try:
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        except:
            driver.refresh()
            time.sleep(1)
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        time.sleep(3)
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase).click()
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselentrang).click()
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselentrang_timkiem).send_keys("Trường test bản tin")
        time.sleep(1)
        # driver.find_element(By.XPATH, var.chiase_iconchontrang).click()
        # time.sleep(0.5)
        driver.find_element(By.XPATH, var.chiase_iconchontrang_truongtestbantin).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chiase_iconchontrang_mota).send_keys(data['khoanhkhac']['mota2'])
        try:
            driver.find_element(By.XPATH, var.dang).click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase).click()
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselentrang).click()
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselentrang_timkiem).send_keys(data['khoanhkhac']['chiaselennhom'])
            driver.find_element(By.XPATH, var.chiase_iconchontrang).click()
            time.sleep(0.5)
            driver.find_element(By.XPATH, var.chiase_iconchontrang_mota).send_keys(data['khoanhkhac']['mota2'])
            driver.find_element(By.XPATH, var.dang).click()
        time.sleep(2)
        check_chiase_chiaselentrang = driver.find_element(By.XPATH, var.check_chiase_chiaselentrang1).text
        print(check_chiase_chiaselentrang)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên trang")
        logging.info("check font-end: Message - Chia sẻ bài viết thành công")
        logging.info(check_chiase_chiaselentrang == "Chia sẻ bài viết thành công")
        driver.get("https://cmc-fe.emso.vn/page/108277159419223992")
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,1000)", "")

        check_chiaselentrang_time = driver.find_element(By.XPATH,var.check_chiaselentrang_time1).text
        print(check_chiaselentrang_time)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên cộng đồng")
        logging.info("check font-end: Message - Check Thời gian trên trang - Vài giây trước")
        logging.info(check_chiaselencongdong_time == "Vài giây trước ")

        check_chiaselentrang_tieude = driver.find_element(By.XPATH,var.check_chiaselentrang_tieude1).text
        print(check_chiaselentrang_tieude)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên cộng đồng")
        logging.info("check font-end: Message - Check Tiêu đề trên trang - nutri boost")
        logging.info(check_chiaselentrang_tieude == "nutri caffee")
        time.sleep(2)

        #Chia sẻ lên trang cá nhân của bạn bè
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.dangtheodoi).click()
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.dangtheodoi).click()
        time.sleep(3)
        try:
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        except:
            driver.refresh()
            time.sleep(1)
            button = driver.find_element(By.XPATH, var.dangtheodoi_xemvideo)
            driver.execute_script("arguments[0].click();", button)
        time.sleep(3)
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase).click()
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselentrangcanhanbanbe).click()
        driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselentrang_timkiem).send_keys(data['khoanhkhac']['chiaselentrangcanhanbanbe'])
        driver.find_element(By.XPATH, var.chiase_iconchonbanbe).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.chiase_iconchontrangcanhanbanbe_mota).send_keys(data['khoanhkhac']['mota3'])
        try:
            driver.find_element(By.XPATH, var.dang).click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase).click()
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselentrangcanhanbanbe).click()
            driver.find_element(By.XPATH, var.khoanhkhac_xemvideo_chiase_chiaselentrang_timkiem).send_keys(data['khoanhkhac']['chiaselentrangcanhanbanbe'])
            driver.find_element(By.XPATH, var.chiase_iconchonbanbe).click()
            time.sleep(0.5)
            driver.find_element(By.XPATH, var.chiase_iconchontrangcanhanbanbe_mota).send_keys(data['khoanhkhac']['mota3'])
            driver.find_element(By.XPATH, var.dang).click()
        time.sleep(2)
        check_chiase_chiaselentrangcanhanbanbe = driver.find_element(By.XPATH, var.check_chiase_chiaselentrangcanhanbanbe1).text
        print(check_chiase_chiaselentrangcanhanbanbe)
        logging.info("Khoảnh khắc - Đang theo dõi - Chia sẻ - Chia sẻ lên trang cá nhân bạn bè")
        logging.info("check font-end: Message - Chia sẻ bài viết thành công")
        logging.info(check_chiase_chiaselentrangcanhanbanbe == "Chia sẻ bài viết thành công")
        driver.get("https://cmc-fe.emso.vn/user/truongvck22")
        time.sleep(2)

    def tructiep(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #trực tiếp
        driver.find_element(By.XPATH, var.tructiep).click()
        time.sleep(3)
        check_tructiep_videocuanguoiphattructiepdatheodoi = driver.find_element(By.XPATH,var.check_tructiep_videocuanguoiphattructiepdatheodoi1).text
        print(check_tructiep_videocuanguoiphattructiepdatheodoi)
        logging.info("Khoảnh khắc - Trực tiếp -  Video của người phát trực tiếp đã theo dõi")
        logging.info("check font-end:  Video của người phát trực tiếp đã theo dõi")
        logging.info(check_tructiep_videocuanguoiphattructiepdatheodoi == "Video của người phát trực tiếp đã theo dõi")

    def taokhoanhkhac(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Tạo khoảnh khắc
        driver.find_element(By.XPATH, var.taokhoanhkhac).click()
        time.sleep(3)
        try:
            driver.find_element(By.XPATH, var.khoanhkhac_nhapnoidung).send_keys(data['khoanhkhac']['taokhoanhkhac'])
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.taokhoanhkhac).click()
            time.sleep(3)
            driver.find_element(By.XPATH, var.khoanhkhac_nhapnoidung).send_keys(data['khoanhkhac']['taokhoanhkhac'])
        driver.find_element(By.XPATH, var.khoanhkhac_tailenvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/quangaygiongbao.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.khoanhkhac_dangbai).click()
        time.sleep(15)
        check_taokhoanhkhac_tieude = driver.find_element(By.XPATH,var.check_taokhoanhkhac_tieude1).text
        print(check_taokhoanhkhac_tieude)
        logging.info("Khoảnh khắc - Tạo khoảnh khắc")
        logging.info("check font-end: Tiêu đề - Test tạo khoảnh khắc")
        logging.info(check_taokhoanhkhac_tieude == "Test tạo khoảnh khắc")

    def taikhoan_duocdexuat(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Tài khoản được đề xuất
        ten_taikhoandexuat1 = driver.find_element(By.XPATH, var.taikhoandexuat1).text
        driver.find_element(By.XPATH, var.taikhoandexuat1).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.taikhoanduocdexuat_xemtongquanvideo)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        check_khoanhkhac_trangcanhan_taikhoandexuat1 = driver.find_element(By.XPATH,var.check_khoanhkhac_trangcanhan_taikhoandexuat1).text
        print(check_khoanhkhac_trangcanhan_taikhoandexuat1)
        logging.info("Khoảnh khắc - Tài khoản được đề xuất")
        logging.info("check font-end: Xem trang cá nhân - "+ ten_taikhoandexuat1)
        logging.info(check_khoanhkhac_trangcanhan_taikhoandexuat1)
        logging.info(check_khoanhkhac_trangcanhan_taikhoandexuat1[0:18] == ten_taikhoandexuat1)
        driver.back()
        time.sleep(2)
        button = driver.find_element(By.XPATH, var.taikhoanduocdexuat_xemvideo)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(3)
        driver.find_element(By.XPATH, var.taikhoandexuat_xemvideo_iconnextxuongduoi).click()
        time.sleep(3)
        check_taikhoandexuat_xemvideo_iconnextxuongduoi = driver.find_element(By.XPATH,var.taikhoandexuat_xemvideo_iconnextxuongduoi1).text
        print(check_taikhoandexuat_xemvideo_iconnextxuongduoi)
        logging.info("Khoảnh khắc - Tài khoản được đề xuất - Xem video")
        logging.info("check font-end: Xem video khi chọn mũi tên xuồng")
        logging.info(check_taikhoandexuat_xemvideo_iconnextxuongduoi)
        logging.info(check_taikhoandexuat_xemvideo_iconnextxuongduoi != None)
        time.sleep(1)
        driver.find_element(By.XPATH, var.taikhoandexuat_xemvideo_iconnextlentren).click()
        time.sleep(3)
        check_taikhoandexuat_xemvideo_iconnextlentren = driver.find_element(By.XPATH,var.check_taikhoandexuat_xemvideo_iconnextlentren1).text
        print(check_taikhoandexuat_xemvideo_iconnextlentren)
        logging.info("Khoảnh khắc - Tài khoản được đề xuất - Xem video")
        logging.info("check font-end: Xem video khi chọn mũi tên lên")
        logging.info(check_taikhoandexuat_xemvideo_iconnextlentren)
        logging.info(check_taikhoandexuat_xemvideo_iconnextlentren !=check_taikhoandexuat_xemvideo_iconnextxuongduoi)
        time.sleep(1)

    def taikhoan_dangtheodoi(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Tài khoản đang theo dõi
        driver.find_element(By.XPATH, var.taikhoandangtheodoi_xemthem).click()
        taikhoandangtheodoi_ten2 = driver.find_element(By.XPATH, var.taikhoandangtheodoi2).text
        driver.find_element(By.XPATH, var.taikhoandangtheodoi2).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.taikhoandangtheodoi_tongquanvideo)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        check_khoanhkhac_trangcanhan_taikhoantheodoi2 = driver.find_element(By.XPATH,var.check_khoanhkhac_trangcanhan_taikhoantheodoi2).text
        print(check_khoanhkhac_trangcanhan_taikhoantheodoi2)
        logging.info("Khoảnh khắc - Tài khoản đang theo dõi")
        logging.info("check font-end: Xem trang cá nhân - "+ taikhoandangtheodoi_ten2)
        logging.info(check_khoanhkhac_trangcanhan_taikhoantheodoi2 == taikhoandangtheodoi_ten2)
        driver.back()
        time.sleep(2)
        button = driver.find_element(By.XPATH, var.taikhoandangtheodoi_xemvideo)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(3)
        driver.find_element(By.XPATH, var.taikhoandexuat_xemvideo_iconnextxuongduoi).click()
        time.sleep(3)
        check_taikhoandangtheodoi_xemvideo_iconnextxuongduoi = driver.find_element(By.XPATH,var.check_taikhoandangtheodoi_xemvideo_iconnextxuongduoi1).text
        print(check_taikhoandangtheodoi_xemvideo_iconnextxuongduoi)
        logging.info("Khoảnh khắc - Tài khoản đang theo dõi - Xem video")
        logging.info("check font-end: Xem video khi chọn mũi tên xuồng")
        logging.info(check_taikhoandangtheodoi_xemvideo_iconnextxuongduoi)
        logging.info(check_taikhoandangtheodoi_xemvideo_iconnextxuongduoi != None)
        time.sleep(1)
        driver.find_element(By.XPATH, var.taikhoandexuat_xemvideo_iconnextlentren).click()
        time.sleep(3)
        check_taikhoandangtheodoi_xemvideo_iconnextlentren = driver.find_element(By.XPATH,var.check_taikhoandangtheodoi_xemvideo_iconnextxuongduoi1).text
        print(check_taikhoandangtheodoi_xemvideo_iconnextlentren)
        logging.info("Khoảnh khắc - Tài khoản đang theo dõi - Xem video")
        logging.info("check font-end: Xem video khi chọn mũi tên lên")
        logging.info(check_taikhoandangtheodoi_xemvideo_iconnextlentren)
        logging.info(check_taikhoandangtheodoi_xemvideo_iconnextlentren != check_taikhoandangtheodoi_xemvideo_iconnextxuongduoi)
        time.sleep(2)

    def khoanhkhac_timkiem(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_khoanhkhac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Tìm kiếm video
        driver.find_element(By.XPATH, var.khoanhkhac_input).send_keys(data['khoanhkhac']['timkiem'])
        driver.find_element(By.XPATH, var.khoanhkhac_input).send_keys(Keys.ENTER)
        time.sleep(4)
        check_timikiem_goiy = driver.find_element(By.XPATH,var.check_timikiem_goiy1).text
        print(check_timikiem_goiy)
        logging.info("Khoảnh khắc - Tìm kiếm - Gợi ý")
        logging.info("check font-end: Gợi ý 1 - Chùa Vanh")
        logging.info(check_timikiem_goiy == "Chùa Vanh")

        #Top
        check_timikiem_top = driver.find_element(By.XPATH,var.check_timikiem_top1).text
        print(check_timikiem_top)
        logging.info("Khoảnh khắc - Tìm kiếm - Top")
        logging.info("check font-end: Top - Chùa Vanh")
        logging.info(check_timikiem_goiy == "Chùa Vanh")
        #Tài khoản
        driver.find_element(By.XPATH, var.khoanhkhac_timkiem_taikhoan).click()
        time.sleep(1)
        check_timikiem_taikhoan = driver.find_element(By.XPATH,var.check_timikiem_taikhoan1).text
        print(check_timikiem_taikhoan)
        logging.info("Khoảnh khắc - Tìm kiếm - Tài khoản")
        logging.info("check font-end: Top - Chùa Vanh")
        logging.info(check_timikiem_taikhoan == "Chùa Vanh")

        #Video
        driver.find_element(By.XPATH, var.timkiem_video).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.khoanhkhac_timkiem_video).click()
        time.sleep(6)
        check_timikiem_video = driver.find_element(By.XPATH,var.check_timikiem_video1).text
        print(check_timikiem_video)
        logging.info("Khoảnh khắc - Tìm kiếm - Video")
        logging.info("check font-end: Mô tả - Hi vọng ông anh vẫn ổn 🤭 @Kiều Oanh #kha7studio #vsl")
        logging.info(check_timikiem_video == "Hi vọng ông anh vẫn ổn 🤭 @Kiều Oanh #kha7studio #vsl")
        time.sleep(1)
        driver.find_element(By.XPATH, var.esc).click()
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(2)



class watch():

    def timkiem(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_watch)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Tìm kiếm video
        driver.find_element(By.XPATH, var.khoanhkhac_input).send_keys(data['watch']['timkiem'])
        driver.find_element(By.XPATH, var.khoanhkhac_input).send_keys(Keys.ENTER)
        time.sleep(5)
        check_watch_timikiem_goiy = driver.find_element(By.XPATH,var.check_watch_timikiem_goiy1).text
        print(check_watch_timikiem_goiy)
        logging.info("Watch - Tìm kiếm - Gợi ý")
        logging.info("check font-end: Gợi ý 1 - Nhạc đã cuốn người còn cuốn hơn #Lacongaiphaixinh")
        logging.info(check_watch_timikiem_goiy)
        logging.info(check_watch_timikiem_goiy == "Nhạc đã cuốn người còn cuốn hơn #Lacongaiphaixinh")

        #Kết quả tìm kiếm
        check_watch_timikiem_ketqua = driver.find_element(By.XPATH,var.check_watch_timikiem_ketqua1).text
        print(check_watch_timikiem_ketqua)
        logging.info("Watch - Tìm kiếm - Gợi ý")
        logging.info("check font-end: Gợi ý 1 - Nhạc đã cuốn người còn cuốn hơn#Lacongaiphaixinh")
        logging.info(check_watch_timikiem_ketqua)
        logging.info(check_watch_timikiem_ketqua == "Nhạc đã cuốn người còn cuốn hơn\n\n#Lacongaiphaixinh")
        button = driver.find_element(By.XPATH, var.watch_xemvideo)
        try:
            driver.execute_script("arguments[0].click();", button)
            time.sleep(0.3)
            # driver.execute_script("arguments[0].click();", button)
        except:
            pass
        time.sleep(3)
        driver.find_element(By.XPATH, var.watch_xemvideo_like).click()
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan).send_keys(data['watch']['mota'])
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan).send_keys(Keys.ENTER)
        time.sleep(2)
        # driver.find_element(By.XPATH, var.watch_xemvideo_x).click()
        button = driver.find_element(By.XPATH, var.watch_xemvideo_x)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)



    def trangchu(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_watch)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Xem video
        button = driver.find_element(By.XPATH, var.watch_trangchu_xemvideo)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(4)

        #like video
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.watch_xemvideo_like)))
        element.click()

        # chọn cách tương tác
        button = driver.find_element(By.XPATH, var.xemvideo_iconchoncachtuongtac)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.xemvideo_iconchoncachtuongtac_trang).click()
        #Bình luận 1
        button = driver.find_element(By.XPATH, var.watch_xemvideo_binhluan_button)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan).send_keys(data['watch']['mota1'])
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan).submit()
        wait = WebDriverWait(driver, 10)
        time.sleep(3)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_like)))
        time.sleep(3)
        element.click()

        check_binhluan_dalike = driver.find_element(By.XPATH,var.check_binhluan_dalike).get_attribute('style')
        print(check_binhluan_dalike)
        logging.info("Watch - Trang chủ - Bình luận - Like")
        logging.info("check font-end: Đã like bình luận")
        logging.info(check_binhluan_dalike == "color: rgb(113, 101, 224);")

        check_choncachtuongtac_trang = driver.find_element(By.XPATH,var.check_choncachtuongtac_trang1).text
        print(check_choncachtuongtac_trang)
        logging.info("Watch - Trang chủ - Chọn cách tương tác")
        logging.info("check font-end: Trang - Trường test bản tin")
        logging.info("Trang - " + check_choncachtuongtac_trang)
        logging.info(check_choncachtuongtac_trang == "Trường test bản tin")

        check_watch_binhluan = driver.find_element(By.XPATH,var.check_watch_binhluan1).text
        print(check_watch_binhluan)
        logging.info("Watch - Trang chủ - Chọn cách tương tác")
        logging.info("check font-end: comment - Test watch trang chủ")
        logging.info(check_watch_binhluan == "Test watch trang chủ")

        #phản hồi
        driver.find_element(By.XPATH, var.binhluan_phanhoi).click()
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi_input)))
        element.send_keys(data['watch']['mota2'])

        #phản hồi - icon cảm xúc
        driver.find_element(By.XPATH, var.binhluan_phanhoi_inputiconcamxuc).click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_inputiconcamxuc_chon).click()
        time.sleep(1)
        #phản hồi - ảnh
        driver.find_element(By.XPATH, var.binhluan_phanhoi_tailen).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien1.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.binhluan_phanhoi_tailen_x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.binhluan_phanhoi_tailen).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien1.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input).submit()
        time.sleep(1.5)

        check_phanhoi_comment_camxuc = driver.find_element(By.XPATH,var.check_phanhoi_comment_camxuc1).text
        print(check_phanhoi_comment_camxuc)
        logging.info("Watch - Trang chủ - Bình luận - Phản hồi")
        logging.info("check font-end: phản hồi comment và cảm xuc - Test watch phản hồi😃")
        logging.info(check_phanhoi_comment_camxuc)
        logging.info(check_phanhoi_comment_camxuc == "Test watch phản hồi😃")

        check_phanhoi_anh = driver.find_element(By.XPATH,var.check_phanhoi_anh1).is_displayed()
        print(check_phanhoi_anh)
        logging.info("Watch - Trang chủ - Bình luận - Phản hồi")
        logging.info("check font-end: phản hồi ảnh - tải ảnh lên thành công")
        logging.info(check_phanhoi_anh)


        wait = WebDriverWait(driver, 10)      #Lỗi trắng trang khi  chưa load xong ảnh mà lại like
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi_like)))
        element.click()

        #phản hồi - gif
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi)))
        element.click()

        driver.find_element(By.XPATH, var.binhluan_phanhoi_gif).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.binhluan_phanhoi_gif_timkiem).send_keys(data['watch']['gif_timkiem'])
        time.sleep(5)
        try:
            wait = WebDriverWait(driver, 15)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi_gif_chon)))
            element.click()
        except:
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi_gif_chon2)))
            element.click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input).send_keys(data['watch']['gif'])
        driver.find_element(By.XPATH, var.binhluan_phanhoi_inputiconcamxuc).click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_inputiconcamxuc_chon).click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input).submit()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_inputiconcamxuc).click()

        check_phanhoi_commentgif = driver.find_element(By.XPATH,var.check_phanhoi_commentgif1).text
        print(check_phanhoi_commentgif)
        logging.info("Watch - Trang chủ - Bình luận - Phản hồi")
        logging.info("check font-end: phản hồi comment và cảm xuc khi chọn Gif - Test giff phản hồi nè😃")
        logging.info(check_phanhoi_commentgif)
        logging.info(check_phanhoi_commentgif == "Test giff phản hồi nè😃")

        check_phanhoi_gif = driver.find_element(By.XPATH,var.check_phanhoi_gif1).is_displayed()
        print(check_phanhoi_gif)
        logging.info("Watch - Trang chủ - Bình luận - Phản hồi")
        logging.info("check font-end: Gif - tải gif lên thành công")
        logging.info(check_phanhoi_gif)


        # phản hồi -nhãn dán
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi1)))
        element.click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_nhandan).click()
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi_nhandan_chon)))
        element.click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input).send_keys(data['watch']['nhandan'])
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input).send_keys(Keys.ENTER)

        check_phanhoi_commentnhandan = driver.find_element(By.XPATH,var.check_phanhoi_commentnhandan1).text
        print(check_phanhoi_commentnhandan)
        logging.info("Watch - Trang chủ - Bình luận - Phản hồi")
        logging.info("check font-end: phản hồi comment khi chọn nhãn dán - Test nhãn dán phản hồi nè")
        logging.info(check_phanhoi_commentnhandan)
        logging.info(check_phanhoi_commentnhandan == "Test nhãn dán phản hồi nè")

        check_phanhoi_nhandan = driver.find_element(By.XPATH,var.check_phanhoi_nhandan1).is_displayed()
        print(check_phanhoi_nhandan)
        logging.info("Watch - Trang chủ - Bình luận - Phản hồi")
        logging.info("check font-end: nhãn dán - tải nhãn dán lên thành công")
        logging.info(check_phanhoi_nhandan)


        # #phản hồi - sửa
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi1)))
        element.click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input).send_keys(data['watch']['phanhoi_sua'])
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input).submit()
        wait = WebDriverWait(driver, 10)
        time.sleep(1)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi_hover)))
        element.click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_icontuychon).click()
        time.sleep(1)

        driver.implicitly_wait(0.1)
        count = 0
        n = 0
        while (count < 25):
            n = int(n)
            n = n + 1
            count = count + 1
            n = str(n)
            binhluan_phanhoi_sua2 = "/html/body/div[" + n + "]/div[3]/ul/li[1]/p"
            print(binhluan_phanhoi_sua2)
            try:
                # button = driver.find_element(By.XPATH, var.binhluan_phanhoi_sua)
                # driver.execute_script("arguments[0].click();", button)
                driver.find_element(By.XPATH, binhluan_phanhoi_sua2).click()
                if driver.find_element(By.XPATH, var.binhluan_phanhoi_input1).is_displayed():
                    break
            except:
                pass
        driver.implicitly_wait(15)

        xoa = driver.find_element(By.XPATH, var.binhluan_phanhoi_input1)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input1).send_keys(data['watch']['phanhoi_sua1'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input1).send_keys(Keys.ENTER)
        time.sleep(2)
        check_watch_phanhoi_suacomment = driver.find_element(By.XPATH,var.check_watch_phanhoi_suacomment1).text
        print(check_watch_phanhoi_suacomment)
        logging.info("Watch - Trang chủ - Bình luận - Phản hồi")
        logging.info("check font-end: Sửa comment - Mua nước uống nheee")
        logging.info(check_watch_phanhoi_suacomment == "Mua nước uống nheee")
        time.sleep(5)

        #phản hồi - xoá           #Lỗi không xoá được comment
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi1)))
        element.click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input).send_keys(data['watch']['phanhoi_sua'])
        driver.find_element(By.XPATH, var.binhluan_phanhoi_input).submit()
        time.sleep(5)
        driver.find_element(By.XPATH, var.binhluan_phanhoi_hover1).click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_icontuychon1).click()
        time.sleep(1)

        driver.implicitly_wait(0.1)
        count = 0
        n = 0
        while (count < 25):
            n = int(n)
            n = n + 1
            count = count + 1
            n = str(n)
            binhluan_phanhoi_xoa1 = "/html/body/div[" + n + "]/div[3]/ul/li[2]/p"
            print(binhluan_phanhoi_xoa1)
            try:
                button = driver.find_element(By.XPATH, binhluan_phanhoi_xoa1)
                driver.execute_script("arguments[0].click();", button)
                if driver.find_element(By.XPATH, var.binhluan_xoa).is_displayed():
                    break
            except:
                pass
        driver.implicitly_wait(15)
        button = driver.find_element(By.XPATH, var.binhluan_xoa)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(4)

        # Bình luận 2
        driver.refresh()
        time.sleep(2)
        button = driver.find_element(By.XPATH, var.xemvideo_iconchoncachtuongtac1)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.xemvideo_iconchoncachtuongtac_ngocmai).click()

        #bình luận - tải lên ảnh
        driver.find_element(By.XPATH, var.binhluan_tailenanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia2.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan).click()
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan).send_keys(data['watch']['mota4'])
        driver.find_element(By.XPATH, var.binhluan_iconcamxuc).click()
        driver.find_element(By.XPATH, var.binhluan_phanhoi_inputiconcamxuc_chon9).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan).submit()
        time.sleep(1)
        driver.find_element(By.XPATH, var.binhluan_iconcamxuc).click()
        time.sleep(1.5)
        check_comment_camxuc_cuaanh = driver.find_element(By.XPATH,var.check_comment_camxuc_cuaanh1).text
        print(check_comment_camxuc_cuaanh)
        logging.info("Watch - Trang chủ - Bình luận - Ảnh")
        logging.info("check font-end: Comment và cảm xúc của ảnh - Test watch bình luận tải lên ảnh😍")
        logging.info(check_comment_camxuc_cuaanh)
        logging.info(check_comment_camxuc_cuaanh == "Test watch bình luận tải lên ảnh😍")

        check_binhluan_anh = driver.find_element(By.XPATH,var.check_binhluan_anh1).is_displayed()
        print(check_binhluan_anh)
        logging.info("Watch - Trang chủ - Bình luận ")
        logging.info("check font-end: Tải ảnh lên thành công")
        logging.info(check_binhluan_anh)

        #bình luận - gif
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.watch_binhluan_gif)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.binhluan_gif_timkiem).send_keys(data['watch']['gif_timkiem'])
        try:
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_gif_timkiem1)))
            element.click()
        except:
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_gif_timkiem6)))
            element.click()
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan).send_keys(data['watch']['gif1'])
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan_iconcamxuc).click()
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan_iconcamxuc_chon13).click()
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan).submit()
        time.sleep(1)
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan_iconcamxuc).click()
        time.sleep(1.5)

        check_binhluan_commentgif = driver.find_element(By.XPATH,var.check_binhluan_commentgif1).text
        print(check_binhluan_commentgif)
        logging.info("Watch - Trang chủ - Bình luận -gif")
        logging.info("check font-end: Bình luận - comment và cảm xuc khi chọn Gif - Test giff bnh luận  nè😋")
        logging.info(check_binhluan_commentgif)
        logging.info(check_binhluan_commentgif == "Test giff bnh luận  nè😋")

        check_binhluan_gif = driver.find_element(By.XPATH,var.check_binhluan_gif1).is_displayed()
        print(check_binhluan_gif)
        logging.info("Watch - Trang chủ - Bình luận - Phản hồi")
        logging.info("check font-end: Gif - tải gif lên thành công")
        logging.info(check_binhluan_gif)


        #bình luận - nhãn dán
        driver.find_element(By.XPATH, var.watch_binhluan_nhandan).click()
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.binhluan_phanhoi_nhandan_chon)))
        element.click()
        driver.find_element(By.XPATH, var.binhluan_nhandan_input).send_keys(data['watch']['nhandan1'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan_iconcamxuc).click()
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan_iconcamxuc_chon12).click()
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan).submit()
        time.sleep(1)
        driver.find_element(By.XPATH, var.watch_xemvideo_binhluan_iconcamxuc).click()
        time.sleep(1.5)
        check_binhluan_commentnhandan = driver.find_element(By.XPATH,var.check_binhluan_commentnhandan1).text
        print(check_binhluan_commentnhandan)
        logging.info("Watch - Trang chủ - Bình luận - Nhãn dán")
        logging.info("check font-end: Bình luận comment khi chọn nhãn dán - Test nhãn dán bình luận nè😚")
        logging.info(check_binhluan_commentnhandan)
        logging.info(check_binhluan_commentnhandan == "Test nhãn dán bình luận nè😚")

        check_binhluan_nhandan = driver.find_element(By.XPATH,var.check_binhluan_nhandan1).is_displayed()
        print(check_binhluan_nhandan)
        logging.info("Watch - Trang chủ - Bình luận - Nhãn dán")
        logging.info("check font-end: nhãn dán - tải nhãn dán lên thành công")
        logging.info(check_binhluan_nhandan)
        time.sleep(1)
        driver.get("https://cmc-fe.emso.vn/")
        time.sleep(3)




    def chuongtrinh(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_watch)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Chương trình
        driver.find_element(By.XPATH, var.watch_chuongtrinh).click()
        check_watch_chuongtrinh_nguyentac1 = driver.find_element(By.XPATH,var.check_watch_chuongtrinh_nguyentac1).is_displayed()
        check_watch_chuongtrinh_nguyentac2 = driver.find_element(By.XPATH,var.check_watch_chuongtrinh_nguyentac2).is_displayed()
        logging.info("Watch - Chương trình - Nguyên tác trên EMSO")
        logging.info("check font-end: Có nguyên tác 1 không?")
        logging.info(check_watch_chuongtrinh_nguyentac1 != None)

        logging.info("check font-end: Có nguyên tác 2 không?")
        logging.info(check_watch_chuongtrinh_nguyentac2 != None)
        time.sleep(1)
        driver.find_element(By.XPATH, var.watch_chuongtrinhdangtheodoi).click()
        check_chuongtrinh_dangtheodoi_mota = driver.find_element(By.XPATH,var.check_chuongtrinh_dangtheodoi_mota1).text
        print(check_chuongtrinh_dangtheodoi_mota)
        logging.info(check_chuongtrinh_dangtheodoi_mota)
        logging.info("Watch - Chương trình - Chương trình bạn bè đang theo dõi")
        logging.info("check font-end: mô tả - Nhìn bả tàn tạ thấy thương🤣🤣")
        logging.info(check_chuongtrinh_dangtheodoi_mota == "Nhìn bả tàn tạ thấy thương🤣🤣")

        check_chuongtrinh_dangtheodoi_trangthai = driver.find_element(By.XPATH,var.check_chuongtrinh_dangtheodoi_trangthai1).text
        print(check_chuongtrinh_dangtheodoi_trangthai)
        logging.info("Watch - Chương trình - Chương trình bạn bè đang theo dõi")
        logging.info("check font-end: Trạng thái - Đang theo dõi")
        logging.info(check_chuongtrinh_dangtheodoi_trangthai)
        logging.info(check_chuongtrinh_dangtheodoi_trangthai == "Đang theo dõi")
        time.sleep(2)

        #Chương trình - xem videp
        driver.find_element(By.XPATH, var.watch_chuongtrinhdangtheodoi_xemvideo1).click()
        time.sleep(3)
        check_chuongtrinh_dangtheodoi_video1 = driver.find_element(By.XPATH,var.check_chuongtrinh_dangtheodoi_video1).text
        print(check_chuongtrinh_dangtheodoi_video1)
        logging.info("Watch - Chương trình - Chương trình bạn bè đang theo dõi")
        logging.info("check font-end: mô tả video 1 - nooo1")
        logging.info(check_chuongtrinh_dangtheodoi_video1 == "nooo1")
        driver.find_element(By.XPATH, var.watch_chuongtrinhdangtheodoi_xemvideo1_x).click()
        time.sleep(1)

        driver.find_element(By.XPATH, var.watch_chuongtrinhdangtheodoi_xemvideo2).click()
        time.sleep(3)
        check_chuongtrinh_dangtheodoi_video2 = driver.find_element(By.XPATH,var.check_chuongtrinh_dangtheodoi_video2).text
        print(check_chuongtrinh_dangtheodoi_video2)
        logging.info("Watch - Chương trình - Chương trình bạn bè đang theo dõi")
        logging.info("check font-end: mô tả video 2 - aaww")
        logging.info(check_chuongtrinh_dangtheodoi_video2 == "aaww")
        driver.find_element(By.XPATH, var.watch_chuongtrinhdangtheodoi_xemvideo1_x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(2)



    def videodaluu(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_watch)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Video đã lưu
        driver.find_element(By.XPATH, var.watch_videodaluu).click()
        time.sleep(2)
        check_watch_videodaluu_video1_trangthai = driver.find_element(By.XPATH,var.check_watch_videodaluu_video1).text
        print(check_watch_videodaluu_video1_trangthai)
        logging.info("Watch - Video đã lưu - video1 - Trạng thái ")
        logging.info("check font-end: Trạng thái - Bỏ lưu")
        logging.info(check_watch_videodaluu_video1_trangthai)
        logging.info(check_watch_videodaluu_video1_trangthai == "Bỏ lưu")

        #Xem video
        #Lưu bài viết
        driver.find_element(By.XPATH, var.watch_videodaluu_xemvideo1).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.videodaluu_xemvideo1_icondau3cham).click()
        check_watch_videodaluu_video1_xemvideo_trangthai = driver.find_element(By.XPATH,var.check_watch_videodaluu_video1_xemvideo_trangthai1).text
        print(check_watch_videodaluu_video1_xemvideo_trangthai)
        logging.info("Watch - Video đã lưu - Video1 - Xem - Trạng thái video1")
        logging.info("check font-end: Trạng thái - Bỏ lưu bài viết Xóa bài viết khỏi danh sách mục đã lưu")
        logging.info(check_watch_videodaluu_video1_xemvideo_trangthai)
        logging.info(check_watch_videodaluu_video1_xemvideo_trangthai == "Bỏ lưu bài viết\nXóa bài viết khỏi danh sách mục đã lưu")

        #Báo cáo
        driver.find_element(By.XPATH, var.watch_videodaluu_xemvideo1_baocao).click()
        check_xemvideo_baocao = driver.find_element(By.XPATH,var.check_xemvideo_baocao1).text
        print(check_xemvideo_baocao)
        logging.info("Watch - Video đã lưu - Video1 - Xem - Báo cáo")
        logging.info("check font-end: Popup - Báo cáo")
        logging.info(check_xemvideo_baocao)
        logging.info(check_xemvideo_baocao == "Báo cáo")

        driver.find_element(By.XPATH, var.baocao_noidungomghiec).click()
        driver.find_element(By.XPATH, var.baocao_quaylai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocao_khungbo).click()
        driver.find_element(By.XPATH, var.baocao_quaylai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocao_ngontugaythughet).click()
        driver.find_element(By.XPATH, var.baocao_quaylai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocao_banhangtraiphep).click()
        driver.find_element(By.XPATH, var.baocao_quaylai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocao_spam).click()
        driver.find_element(By.XPATH, var.baocao_quaylai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocao_tutuhoacgaythuongtich).click()
        driver.find_element(By.XPATH, var.baocao_quaylai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocao_quayroi).click()
        driver.find_element(By.XPATH, var.baocao_quaylai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocao_baoluc).click()
        driver.find_element(By.XPATH, var.baocao_quaylai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocao_anhkhoathan).click()
        driver.find_element(By.XPATH, var.baocao_quaylai).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.baocao_x).click()
        time.sleep(1)

        #Tắt thông báo về bài viết này
        driver.find_element(By.XPATH, var.videodaluu_xemvideo1_icondau3cham).click()
        driver.find_element(By.XPATH, var.icondau3cham_tatthongbaovebaivietnay).click()
        time.sleep(1)
        check_luuvideo_tatthongbaobaiviet = driver.find_element(By.XPATH,var.check_luuvideo_tatthongbaobaiviet1).text
        print(check_luuvideo_tatthongbaobaiviet)
        logging.info("Watch - Video đã lưu - Video1 - Xem - Tắt thông báo bài viết")
        logging.info("check font-end: Message - Đã tắt thông báo")
        logging.info(check_luuvideo_tatthongbaobaiviet)
        logging.info(check_luuvideo_tatthongbaobaiviet == "Đã tắt thông báo")

        #Bật thông báo về bài viết này
        driver.find_element(By.XPATH, var.videodaluu_xemvideo1_icondau3cham).click()
        driver.find_element(By.XPATH, var.icondau3cham_batthongbaovebaivietnay).click()
        time.sleep(1)
        check_luuvideo_batthongbaobaiviet = driver.find_element(By.XPATH,var.check_luuvideo_batthongbaobaiviet1).text
        print(check_luuvideo_batthongbaobaiviet)
        logging.info("Watch - Video đã lưu - Video1 - Xem - Bật thông báo bài viết")
        logging.info("check font-end: Message - Đã bật thông báo")
        logging.info(check_luuvideo_batthongbaobaiviet)
        logging.info(check_luuvideo_batthongbaobaiviet == "Đã bật thông báo")
        time.sleep(1)

        #Nhúng
        driver.find_element(By.XPATH, var.videodaluu_xemvideo1_icondau3cham).click()
        driver.find_element(By.XPATH, var.icondau3cham_nhung).click()
        time.sleep(0.5)
        check_luuvideo_nhung = driver.find_element(By.XPATH,var.check_luuvideo_nhung1).text
        print(check_luuvideo_nhung)
        logging.info("Watch - Video đã lưu - Video1 - Xem - Nhúng")
        logging.info("check font-end: Popup - Nhúng bài viết")
        logging.info(check_luuvideo_nhung)
        logging.info(check_luuvideo_nhung == "Nhúng bài viết")
        time.sleep(1)

        check_luuvideo_nhung_motavideo = driver.find_element(By.XPATH,var.check_luuvideo_nhung_motavideo1).text
        print(check_luuvideo_nhung_motavideo)
        logging.info("Watch - Video đã lưu - Video1 - Xem - Nhúng - Mô tả video")
        logging.info("check font-end: Mô tả video nhúng - Vợ báo quá báo,chồng có ý tốt mà lại bị hiểu nhầm cuối cùng đẩy chồng vào t.ù #phimhay")
        logging.info(check_luuvideo_nhung_motavideo)
        logging.info(check_luuvideo_nhung_motavideo == "Vợ báo quá báo,chồng có ý tốt mà lại bị hiểu nhầm cuối cùng đẩy chồng vào t.ù #phimhay")
        driver.find_element(By.XPATH, var.x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.videodaluu_xem_x).click()
        time.sleep(2)
        try:
            check_luuvideo_title = driver.find_element(By.XPATH,var.check_luuvideo_title).text
            logging.info("Watch - Video đã lưu - xem video - x")
            logging.info("check font-end: Có trở lại trang Video đã lưu không")
            logging.info(check_luuvideo_title == "Video đã lưu")
            time.sleep(2)
        except NoSuchElementException:
            logging.info("Watch - Video đã lưu - xem video - x")
            logging.info("check font-end: Có trở lại trang Video đã lưu không")
            logging.info("False")
            time.sleep(2)



    def dangtheodoi(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_watch)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Dang theo dõi - Trường test
        driver.find_element(By.XPATH, var.watch_dangtheodoi_truongtest).click()
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,-900)", "")
        button = driver.find_element(By.XPATH, var.watch_truongtest_choncachtuongtac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_ngocmai).click()
        time.sleep(1)
        #Trạng thái
        check_dangtheodoi_trangthai = driver.find_element(By.XPATH,var.check_dangtheodoi_trangthai1).text
        print(check_dangtheodoi_trangthai)
        logging.info("Watch - Đang theo dõi - Trường test bản tin")
        logging.info("check font-end: Trạng thái - Đã thích")
        logging.info(check_dangtheodoi_trangthai == "  Đã thích")

        #Phổ biến nhất
        try:
            check_dangtheodoi_phobiennhat = driver.find_element(By.XPATH,var.check_dangtheodoi_phobiennhat1).is_displayed()
            print(check_dangtheodoi_phobiennhat)
            logging.info("Watch - Đang theo dõi - Trường test bản tin")
            logging.info("check font-end: Phổ biến nhất - Có video hiển thị hay không")
            logging.info(check_dangtheodoi_phobiennhat)
        except  NoSuchElementException:
            logging.info("Watch - Đang theo dõi - Trường test bản tin")
            logging.info("check font-end: Phổ biến nhất - Có video hiển thị hay không")
            logging.info("False")

        #Bài đăng
        try:
            check_dangtheodoi_baidang = driver.find_element(By.XPATH,var.check_dangtheodoi_baidang1).is_displayed()
            print(check_dangtheodoi_baidang)
            logging.info("Watch - Đang theo dõi - Trường test bản tin")
            logging.info("check font-end: Bài đăng - Có video hiển thị hay không")
            logging.info(check_dangtheodoi_baidang)
        except  NoSuchElementException:
            logging.info("Watch - Đang theo dõi - Trường test bản tin")
            logging.info("check font-end: Bài đăng - Có video hiển thị hay không")
            logging.info("False")

        #Khoảnh khăc
        driver.find_element(By.XPATH, var.watch_dangtheodoi_khoanhkhac).click()
        time.sleep(2)
        try:
            check_dangtheodoi_khoanhkhac = driver.find_element(By.XPATH,var.check_dangtheodoi_khoanhkhac1).is_displayed()
            print(check_dangtheodoi_khoanhkhac)
            logging.info("Watch - Đang theo dõi - Trường test bản tin")
            logging.info("check font-end: Khoảnh khắc - Có video hiển thị hay không")
            logging.info(check_dangtheodoi_khoanhkhac)
        except  NoSuchElementException:
            logging.info("Watch - Đang theo dõi - Trường test bản tin")
            logging.info("check font-end: Khoảnh khắc - Có video hiển thị hay không")
            logging.info("False")

        #Danh sách phát
        driver.find_element(By.XPATH, var.watch_dangtheodoi_danhsachphat).click()
        time.sleep(2)
        try:
            check_dangtheodoi_danhsachphat = driver.find_element(By.XPATH,var.check_dangtheodoi_danhsachphat1).is_displayed()
            print(check_dangtheodoi_danhsachphat)
            logging.info("Watch - Đang theo dõi - Trường test bản tin")
            logging.info("check font-end: Danh sách phát - Có video hiển thị hay không")
            logging.info(check_dangtheodoi_danhsachphat)
        except  NoSuchElementException:
            logging.info("Watch - Đang theo dõi - Trường test bản tin")
            logging.info("check font-end: Danh sách phát - Có video hiển thị hay không")
            logging.info("False")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu).click()
        time.sleep(2)



class trang():
    def timkiem_trangcuaban(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Tìm kiếm trang page
        driver.find_element(By.XPATH, var.trang_timkiem).send_keys(data['trang']['timkiem'])
        driver.find_element(By.XPATH, var.trang_timkiem).send_keys(Keys.ENTER)
        time.sleep(3)
        trang_timkiem_trang = driver.find_element(By.XPATH,var.trang_timkiem_trang1).text
        print(trang_timkiem_trang)
        logging.info("Trang - Tìm kiếm")
        logging.info("check font-end: Tìm kiếm - Trường test bản tin")
        logging.info(trang_timkiem_trang == "Trường test bản tin")
        driver.back()
        driver.back()
        time.sleep(1)
        #Trang của bạn
        #Bình thuận
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangchu1).click()
        time.sleep(1)
        check_trang_trangcuaban_binhthuan = driver.find_element(By.XPATH,var.trang_trangcuaban_binhthuan1).text
        print(check_trang_trangcuaban_binhthuan)
        logging.info("Trang - Trang của bạn")
        logging.info("check font-end: Trang của bạn - Bình Thuận")
        logging.info(check_trang_trangcuaban_binhthuan == "Bình Thuận")
        driver.back()
        time.sleep(0.5)
        driver.back()
        time.sleep(0.5)
        #Trường test bản tin
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(1)
        check_trang_trangcuaban_truongtest = driver.find_element(By.XPATH,var.trang_trangcuaban_truongtest1).text
        print(check_trang_trangcuaban_truongtest)
        logging.info("Trang - Trang của bạn")
        logging.info("check font-end: Trang của bạn - Trường test bản tin")
        logging.info(check_trang_trangcuaban_truongtest == "Trường test bản tin")



    def khampha(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(2)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trang_khampha).click()
        #Gợi ý cho bạn
        check_trang_khampha_goiy1 = driver.find_element(By.XPATH,var.check_trang_khampha_goiy1a).text
        print(check_trang_khampha_goiy1)
        logging.info("Trang - Khám phá")
        logging.info("check font-end: Khám phá - Có trang gợi ý hay không")
        logging.info(check_trang_khampha_goiy1 != None)

        khampha_tentranggoiy1 = driver.find_element(By.XPATH,var.khampha_tentranggoiy1aa).text
        print(khampha_tentranggoiy1)
        logging.info("Trang - Khám phá")
        logging.info("check font-end: Hiện tên gợi ý 1 hay không")
        logging.info(khampha_tentranggoiy1)
        logging.info(khampha_tentranggoiy1 != None)

        khampha_linhvuc_tranggoiy1 = driver.find_element(By.XPATH,var.khampha_linhvuc_tranggoiy11).text
        print(khampha_linhvuc_tranggoiy1)
        logging.info("Trang - Khám phá")
        logging.info("check font-end: Hiện lĩnh vực gợi ý 1 hay không")
        logging.info(khampha_linhvuc_tranggoiy1)
        logging.info(khampha_linhvuc_tranggoiy1 != None)

        khampha_luotthich_tranggoiy1 = driver.find_element(By.XPATH,var.khampha_luotthich_tranggoiy11).text
        print(khampha_luotthich_tranggoiy1)
        logging.info("Trang - Khám phá")
        logging.info("check font-end: Hiện lượt thích gợi ý 1 hay không")
        logging.info(khampha_luotthich_tranggoiy1)
        logging.info(khampha_luotthich_tranggoiy1 != None)

        check_trang_khampha_goiy1_trangthai = driver.find_element(By.XPATH,var.check_trang_khampha_goiy1_trangthai1).text
        print(check_trang_khampha_goiy1_trangthai)
        logging.info("Trang - Khám phá")
        logging.info("check font-end: Trạng thái gợi ý 1 - Thích")
        logging.info(check_trang_khampha_goiy1_trangthai == "  Thích")

        #Vào Trang đầu tiên của Gợi ý cho bạn
        driver.find_element(By.XPATH, var.khampha_goiy1).click()
        time.sleep(1)
        check_goiy1_vaotrang_ten = driver.find_element(By.XPATH,var.check_goiy1_vaotrang_ten1).text
        print(check_goiy1_vaotrang_ten)
        logging.info(khampha_tentranggoiy1)
        logging.info(check_goiy1_vaotrang_ten)
        logging.info("Trang - Khám phá")
        logging.info("check font-end: Vào trang gợi ý 1 " + khampha_tentranggoiy1)
        logging.info(check_goiy1_vaotrang_ten == khampha_tentranggoiy1)

        check_goiy1_trangthai = driver.find_element(By.XPATH,var.check_goiy1_trangthai1).text
        print(check_goiy1_trangthai)
        logging.info("Trang - Khám phá")
        logging.info("check font-end: Trạng thái gợi ý 1 khi đã like và vào trang - Thích")
        logging.info(check_goiy1_trangthai == "  Thích")

        # khampha_linhvuc_tranggoiy1_vaotrang = driver.find_element(By.XPATH,var.khampha_linhvuc_tranggoiy1_vaotrang1).text
        # print(khampha_linhvuc_tranggoiy1_vaotrang[2::])
        # logging.info("Trang - Khám phá")
        # logging.info("check font-end: Lĩnh vực gợi ý 1 khi vào trang có giống không?")
        # logging.info(khampha_linhvuc_tranggoiy1_vaotrang[2::])
        # logging.info(khampha_linhvuc_tranggoiy1)
        # logging.info(khampha_linhvuc_tranggoiy1_vaotrang[2::] == khampha_linhvuc_tranggoiy1)

        khampha_luotthich_tranggoiy1_vaotrang = driver.find_element(By.XPATH,var.khampha_luotthich_tranggoiy1_vaotrang1).text
        print(khampha_luotthich_tranggoiy1_vaotrang)
        logging.info("Trang - Khám phá")
        logging.info("check font-end: Lượt thích gợi ý 1 khi vào trang có giống nhau không?")
        logging.info(khampha_luotthich_tranggoiy1_vaotrang)
        logging.info(khampha_luotthich_tranggoiy1)
        logging.info(khampha_luotthich_tranggoiy1_vaotrang[0:10] == khampha_luotthich_tranggoiy1[0:10])
        time.sleep(1)



    def trangdathich(self):
        driver.implicitly_wait(15)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trang_trangdathich).click()
        time.sleep(2)
        #Thích sơm nhất
        trangdathich_tentranggoiy1 = driver.find_element(By.XPATH,var.khampha_tentranggoiy1).text
        print(trangdathich_tentranggoiy1)
        logging.info("Trang - Trang đã thích")
        logging.info("check font-end: Hiện tên trang đã thích 1 hay không")
        logging.info(trangdathich_tentranggoiy1)
        logging.info(trangdathich_tentranggoiy1 != None)

        trangdathich_linhvuc_tranggoiy1 = driver.find_element(By.XPATH,var.trangdathich_linhvuc_tranggoiy1).text
        print(trangdathich_linhvuc_tranggoiy1)
        logging.info("Trang - Trang đã thích")
        logging.info("check font-end: Hiện lĩnh vực trang đã thích 1 hay không")
        logging.info(trangdathich_linhvuc_tranggoiy1)
        logging.info(trangdathich_linhvuc_tranggoiy1 != None)

        driver.find_element(By.XPATH, var.trangdathich_goiy1_icom3cham).click()
        trangdathich_goiy1_trangthai = driver.find_element(By.XPATH,var.trangdathich_goiy1_trangthaibb).text
        print(trangdathich_goiy1_trangthai)
        logging.info("Trang - Trang đã thích")
        logging.info("check font-end: Trạng thái gợi ý 1 - Đã thích")
        logging.info(trangdathich_goiy1_trangthai)
        logging.info(trangdathich_goiy1_trangthai == "Đã thích")
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep).click()
        # Đã thích gần đây trước
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep_dathichgandaytruoc).click()
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep).click()
        try:
            check_sapxep_dathichganday= driver.find_element(By.XPATH,var.check_sapxep_dathichganday).is_displayed()
            logging.info("Trang - Trang đã thích")
            logging.info("check font-end: Đã thích gần đây trước - có icon dấu tích")
            logging.info(check_sapxep_dathichganday)
        except:
            logging.info("Trang - Trang đã thích")
            logging.info("check font-end: Đã thích gần đây trước - có icon dấu tích")
            logging.info("False")
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep).click()

        trangdathich_tentranggoiy2 = driver.find_element(By.XPATH,var.trangdathich_tentranggoiy2).text
        print(trangdathich_tentranggoiy2)
        logging.info("Trang - Trang đã thích")
        logging.info("check font-end: Hiện tên trang sau khi sắp xếp Đã thích gần đây trước không")
        logging.info(trangdathich_tentranggoiy2)
        logging.info(trangdathich_tentranggoiy2 != trangdathich_tentranggoiy1)

        # #Gửi tin nhắn
        driver.find_element(By.XPATH, var.trang_trangdathich_guitinnhan).click()        #ko nhấn được nút
        driver.find_element(By.XPATH, var.trang_trangdathich_guitinnhan_input).send_keys(data['trang']['trangdathich'])
        driver.find_element(By.XPATH, var.trang_trangdathich_guitinnhan_input).send_keys(Keys.ENTER)
        checK_trangdathich_tinnhan = driver.find_element(By.XPATH,var.checK_trangdathich_tinnhan1).text
        print(checK_trangdathich_tinnhan)
        logging.info("Trang - Trang đã thích")
        logging.info("check font-end: Gửi tinn nhắn - Test trang hoy")
        logging.info(checK_trangdathich_tinnhan)
        logging.info(checK_trangdathich_tinnhan =="Test trang hoy")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangdathich_guitinnhan_input).send_keys(data['trang']['trangdathich1'])
        driver.find_element(By.XPATH, var.trang_trangdathich_guitinnhan_input_x).click()
        time.sleep(1)


    def trangdathich_dau3cham(self):
        driver.implicitly_wait(15)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trang_trangdathich).click()
        time.sleep(2)
        # #Theo dõi
        # #Khi chưa click theo dõi
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep).click()
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep_dathichgandaytruoc).click()
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        check_trangthaitrang_benngoai = driver.find_element(By.XPATH, var.check_trangthaitrang_benngoai).text
        driver.find_element(By.XPATH, var.trangdathich_xemtrangdau).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangdathich_icom3cham_bentrong).click()
        check_trangthaitrang_bentrong = driver.find_element(By.XPATH, var.check_trangthaitrang_bentrong).text
        print(check_trangthaitrang_benngoai)
        print(check_trangthaitrang_bentrong)
        logging.info("Trang - Trang đã thích - Theo dõi")
        logging.info("check font-end: Trạng thái bên ngoài trang khi chưa Theo dõi: " + check_trangthaitrang_benngoai)
        logging.info("check font-end: Trạng thái bên trong trang khi chưa theo dõi: " + check_trangthaitrang_bentrong)
        logging.info(check_trangthaitrang_benngoai ==check_trangthaitrang_bentrong)
        #Khi click vào theo dõi
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trang_trangdathich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep).click()
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep_dathichgandaytruoc).click()
        time.sleep(2.5)
        try:
            driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        driver.find_element(By.XPATH, var.trangdathich_icom3cham_theodoi).click()
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        check_trangthaitrang_benngoai = driver.find_element(By.XPATH, var.check_trangthaitrang_benngoai).text
        driver.find_element(By.XPATH, var.trangdathich_xemtrangdau).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangdathich_icom3cham_bentrong).click()
        check_trangthaitrang_bentrong = driver.find_element(By.XPATH, var.check_trangthaitrang_bentrong).text
        # print(check_trangthaitrang_benngoai)
        # print(check_trangthaitrang_bentrong)
        logging.info("Trang - Trang đã thích - Theo dõi")
        logging.info("check font-end: Trạng thái bên ngoài trang khi đã chọn Theo dõi: " + check_trangthaitrang_benngoai)
        logging.info("check font-end: Trạng thái bên trong trang khi đã chọn theo dõi: " + check_trangthaitrang_bentrong)
        logging.info(check_trangthaitrang_benngoai ==check_trangthaitrang_bentrong)

        #Lưu
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trang_trangdathich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep).click()
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep_dathichgandaytruoc).click()
        time.sleep(2.5)
        try:
            driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trang_trangdathich_sapxep).click()
            driver.find_element(By.XPATH, var.trang_trangdathich_sapxep_dathichgandaytruoc).click()
            time.sleep(2.5)
            driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        driver.find_element(By.XPATH, var.luutrang).click()
        driver.find_element(By.XPATH, var.trangdathich_icom3cham_luu_no1).click()
        driver.find_element(By.XPATH, var.xong).click()
        check_message_luutrang = driver.find_element(By.XPATH,var.check_message_luutrang1).text
        print(check_message_luutrang)
        logging.info("Trang - Trang đã thích - Lưu trang")
        logging.info("check font-end: Message - Đã lưu vào no1")
        logging.info(check_message_luutrang == "Đã lưu vào no1")

        #Bỏ lưu
        driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        driver.find_element(By.XPATH, var.boluutrang).click()
        check_message_boluutrang = driver.find_element(By.XPATH,var.check_message_boluutrang1).text
        print(check_message_luutrang)
        logging.info("Trang - Trang đã thích - Lưu trang")
        logging.info("check font-end: Message - Bỏ lưu thành công.")
        logging.info(check_message_boluutrang == "Bỏ lưu thành công.")

        #Chia sẻ
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trang_trangdathich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep).click()
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep_dathichgandaytruoc).click()
        time.sleep(2.5)
        try:
            driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        driver.find_element(By.XPATH, var.trangdathich_icom3cham_chiase).click()
        driver.find_element(By.XPATH, var.trangdathich_icom3cham_chiase_mota).send_keys(data['trang']['chiase_mota'])
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(2)
        checK_trangdathich_chiase_message = driver.find_element(By.XPATH,var.checK_trangdathich_chiase_message1).text
        print(checK_trangdathich_chiase_message)
        logging.info("Trang - Trang đã thích - Chia sẻ")
        logging.info("check font-end: Message - Đăng bài viết thành công.")
        logging.info(checK_trangdathich_chiase_message)
        logging.info(checK_trangdathich_chiase_message =="Đăng bài viết thành công.")
        time.sleep(0.5)
        driver.get("https://cmc-fe.emso.vn/user/truongvck333")
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,1000)", "")

        #Vào trang cá nhân check chia sẻ trang
        checK_chiase_trang_tieude = driver.find_element(By.XPATH,var.checK_chiase_trang_tieude1).text
        print(checK_chiase_trang_tieude)
        logging.info("Trang - Trang đã thích - Chia sẻ")
        logging.info("check font-end: Vào trang cá nhân check Tiêu đề - Ngọc Mai đã chia sẻ trang")
        logging.info(checK_chiase_trang_tieude)
        logging.info(checK_chiase_trang_tieude =="Ngọc Mai\n đã chia sẻ trang")

        checK_chiase_trang_mota = driver.find_element(By.XPATH,var.checK_chiase_trang_mota1).text
        print(checK_chiase_trang_mota)
        logging.info("Trang - Trang đã thích - Chia sẻ")
        logging.info("check font-end: Vào trang cá nhân check Mô tả- " + data['trang']['chiase_mota'])
        logging.info(checK_chiase_trang_mota)
        logging.info(checK_chiase_trang_mota ==data['trang']['chiase_mota'])
        time.sleep(1)

    def loimoi(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trang_trangdathich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep).click()
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep_dathichgandaytruoc).click()
        time.sleep(2.5)
        try:
            driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        driver.find_element(By.XPATH, var.trangdathich_icom3cham_moibanbe).click()
        time.sleep(0.5)
        #Nút chọn người
        driver.find_element(By.XPATH, var.trang_moibanbe_nutchon).click()
        driver.find_element(By.XPATH, var.trang_moibanbe_nutchon_nhom).click()
        check_trang_moibanbe_khongcoban = driver.find_element(By.XPATH,var.check_trang_moibanbe_khongcoban).text
        print(check_trang_moibanbe_khongcoban)
        logging.info("Trang - Trang đã thích - Mời bạn bè")
        logging.info("check font-end: Chọn nhóm không có bạn bè chung")
        logging.info(check_trang_moibanbe_khongcoban)
        logging.info(check_trang_moibanbe_khongcoban == "Không có dữ liệu")

        driver.find_element(By.XPATH, var.trang_moibanbe_nutchon).click()
        driver.find_element(By.XPATH, var.trang_moibanbe_nutchon_namtest).click()
        check_trang_moibanbe_nhomcobanchung = driver.find_element(By.XPATH,var.check_trang_moibanbe_nhomcobanchung).text
        print(check_trang_moibanbe_nhomcobanchung)
        logging.info("Trang - Trang đã thích - Mời bạn bè")
        logging.info("check font-end: Chọn nhóm có bạn bè chung - Ngọc Mai")
        logging.info(check_trang_moibanbe_nhomcobanchung)
        logging.info(check_trang_moibanbe_nhomcobanchung == "Ngọc Mai")

        driver.find_element(By.XPATH, var.trang_moibanbe_nutchon).click()
        driver.find_element(By.XPATH, var.trang_moibanbe_nutchon_tatcabanbe).click()
        check_trang_moibanbe_tatcabanbe = driver.find_element(By.XPATH,var.check_trang_moibanbe_tatcabanbe).text
        print(check_trang_moibanbe_tatcabanbe)
        logging.info("Trang - Trang đã thích - Mời bạn bè")
        logging.info("check font-end: Tất cả bạn bè")
        logging.info(check_trang_moibanbe_tatcabanbe)
        logging.info(check_trang_moibanbe_tatcabanbe == "hue nguyen")

        driver.find_element(By.XPATH, var.huenguyen).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.check_trang_moibanbe_tatcabanbe_guiloimoi).click()
        time.sleep(2)
        check_trang_moibanbe_damoi = driver.find_element(By.XPATH,var.huenguyen).text
        print(check_trang_moibanbe_damoi)
        logging.info("Trang - Trang đã thích - Mời bạn bè")
        logging.info("check font-end: Đã mời - hue nguyen")
        logging.info(check_trang_moibanbe_damoi)
        logging.info(check_trang_moibanbe_damoi == "hue nguyen")
        time.sleep(0.5)

        #Gửi lời mời - đồng ý
        login.login4(self, "nguyenhue608196@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu_iconthongbao).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu_iconthongbao_thongbaodautien).click()
        time.sleep(2)
        check_trang_loimoithichtrang = driver.find_element(By.XPATH,var.check_trang_loimoithichtrang).text
        print(check_trang_loimoithichtrang)
        logging.info("Trang - Trang đã thích - Mời bạn bè - hue nguyen")
        logging.info("check font-end: hue nguyen - xem thống báo - Lời mời thích trang")
        logging.info(check_trang_loimoithichtrang)
        logging.info(check_trang_loimoithichtrang == "Lời mời thích trang")
        time.sleep(0.5)
        driver.find_element(By.XPATH,var.chapnhan).click()
        check_trang_loimoithichtrang_dachapnhan = driver.find_element(By.XPATH,var.check_trang_loimoithichtrang_dachapnhan).text
        print(check_trang_loimoithichtrang_dachapnhan)
        logging.info("Trang - Trang đã thích - Mời bạn bè - hue nguyen")
        logging.info("check font-end: hue nguyen - xem thống báo - Lời mời thích trang - Đã chấp nhận")
        logging.info(check_trang_loimoithichtrang_dachapnhan)
        logging.info(check_trang_loimoithichtrang_dachapnhan == "  Đã chấp nhận")

        driver.find_element(By.XPATH,var.loimoithichtrang_xemtrang).click()
        time.sleep(2)
        driver.find_element(By.XPATH,var.loimoithichtrang_xemtrang_bothich).click()
        time.sleep(3)
        loimoithichtrang_xemtrang_dabothich = driver.find_element(By.XPATH,var.loimoithichtrang_xemtrang_dabothich1).text
        logging.info("Trang - Trang bạn bè mời")
        logging.info("check font-end: đã bỏ thích")
        logging.info(loimoithichtrang_xemtrang_dabothich)
        logging.info(loimoithichtrang_xemtrang_dabothich == "  Thích")
        time.sleep(2)

        #Gửi lời mời -Từ chối
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trang_trangdathich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep).click()
        driver.find_element(By.XPATH, var.trang_trangdathich_sapxep_dathichgandaytruoc).click()
        time.sleep(2.5)
        try:
            driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        except:
            driver.refresh()
            time.sleep(2)
            driver.get("https://cmc-fe.emso.vn/pages/liked")
            driver.find_element(By.XPATH, var.trangdathich_icom3cham).click()
        driver.find_element(By.XPATH, var.trangdathich_icom3cham_moibanbe).click()
        driver.find_element(By.XPATH, var.huenguyen).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.check_trang_moibanbe_tatcabanbe_guiloimoi).click()
        time.sleep(2)
        login.login4(self, "nguyenhue608196@gmail.com", "voncamk22")
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trang_loimoi).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_loimoi_dau3cham).click()
        driver.find_element(By.XPATH, var.tuchoi).click()
        driver.find_element(By.XPATH, var.tuchoi).click()
        time.sleep(2)
        check_trang_loimoithichtrang_datuchoi = driver.find_element(By.XPATH,var.check_trang_loimoithichtrang_datuchoi).text
        print(check_trang_loimoithichtrang_datuchoi)
        logging.info("Trang - Trang đã thích - Mời bạn bè - hue nguyen")
        logging.info("check font-end: hue nguyen - xem thống báo - Lời mời thích trang - Đã từ chối")
        logging.info(check_trang_loimoithichtrang_datuchoi)
        logging.info(check_trang_loimoithichtrang_datuchoi == "  Đã từ chối")
        time.sleep(2)

        driver.find_element(By.XPATH,var.loimoithichtrang_xemtrang).click()
        time.sleep(2)
        loimoithichtrang_xemtrang_dabothich = driver.find_element(By.XPATH,var.loimoithichtrang_xemtrang_dabothich1).text
        logging.info("Trang - Trang bạn bè mời")
        logging.info("check font-end:xem trang - trạng thái đã từ chối")
        logging.info(loimoithichtrang_xemtrang_dabothich)
        logging.info(loimoithichtrang_xemtrang_dabothich == "  Thích")
        time.sleep(2)


    def taotrangmoi(self, tentrang, loaitrang, loaihinhkinhdoanh):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1.5)
        # driver.find_element(By.XPATH, var.trang_taotrangmoi).click()
        button = driver.find_element(By.XPATH, var.trang_taotrangmoi)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Thông tin về trang
        driver.find_element(By.XPATH, var.taotrangmoi_thongtinvetrang).send_keys(tentrang)
        #Loai trang
        driver.find_element(By.XPATH, var.taotrangmoi_loaitrang).click()
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, loaitrang)))
        element.click()
        #Loai hình kinh doanh
        driver.find_element(By.XPATH, var.taotrangmoi_loaihinhkinhdoanh).click()
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, loaihinhkinhdoanh)))
        element.click()
        #Hạng mục
        driver.find_element(By.XPATH, var.taotrangmoi_hangmuc).click()
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.taotrangmoi_hangmuc_giaitri)))
        element.click()

        driver.find_element(By.XPATH, var.taotrangmoi_hangmuc).click()
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.taotrangmoi_hangmuc_thietkethoitrang)))
        element.click()
        #Mô tả
        driver.find_element(By.XPATH, var.taotrangmoi_mota).send_keys(data['trang']['taotrangmoi_mota'])
        #Ảnh đai diện
        button = driver.find_element(By.XPATH, var.taotrangmoi_tailenanhdaidien)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhmes1.exe")
        time.sleep(1)
        #Ảnh bìaa
        button = driver.find_element(By.XPATH, var.taotrangmoi_tailenanhbia)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbiatrang.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.taotrangmoi_taotrang).click()
        time.sleep(7)
        driver.find_element(By.XPATH, var.trang_taotrangmoi_trangchu).click()
        #Check tạo trang
        try:
            taotrangmoi_taotrang_anhdaidien = driver.find_element(By.XPATH,var.taotrangmoi_taotrang_anhdaidien1).is_displayed()
            logging.info("Trang - Tạo trang mới")
            logging.info("check font-end: Loại trang - " + tentrang)
            logging.info("Ảnh đại diện có hiển thị hay không")
            logging.info(taotrangmoi_taotrang_anhdaidien)
        except NoSuchElementException:
            logging.info("Trang - Tạo trang mới")
            logging.info("check font-end: Loại trang - " + tentrang)
            logging.info("Ảnh đại diện có hiển thị hay không")
            logging.info("False")

        # try:
        #     taotrangmoi_taotrang_anhbia = driver.find_element(By.XPATH,var.taotrangmoi_taotrang_anhbia1).is_displayed()
        #     logging.info("Trang - Tạo trang mới")     #chưa viết check ảnh bìa trang
        #     logging.info("check font-end: Loại trang - " + tentrang)
        #     logging.info("Ảnh bìa có hiển thị hay không")
        #     logging.info(taotrangmoi_taotrang_anhbia)
        # except NoSuchElementException:
        #     logging.info("Trang - Tạo trang mới")
        #     logging.info("check font-end: Loại trang - " + tentrang)
        #     logging.info("Ảnh bìa có hiển thị hay không")
        #     logging.info("False")

        driver.implicitly_wait(5)
        try:
            taotrangmoi_taotrang_loaitrang = driver.find_element(By.XPATH,var.taotrangmoi_taotrang_loaitrang).text
            logging.info("Trang - Tạo trang mới")
            logging.info("check font-end: Loại trang - " + tentrang)
            logging.info(taotrangmoi_taotrang_loaitrang)
            logging.info(taotrangmoi_taotrang_loaitrang == tentrang)
            #trang ban hang
            driver.find_element(By.XPATH, var.trang_caidat).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.chung).click()
            time.sleep(1)
            driver.execute_script("window.scrollBy(0,900)", "")
            driver.find_element(By.XPATH, var.molaitrangbanhang).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.xacnhankhoapage).click()
            # driver.find_element(By.XPATH, var.message_xacnhankhoapage).click() #chưa có mesage khóa page
            time.sleep(3)
        except:
            taotrangmoi_taotrang_loaitrang1 = driver.find_element(By.XPATH,var.taotrangmoi_taotrang_loaitrang1).text
            logging.info("Trang - Tạo trang mới")
            logging.info("check font-end: Loại trang - " + tentrang)
            logging.info(taotrangmoi_taotrang_loaitrang1)
            logging.info(taotrangmoi_taotrang_loaitrang1 == tentrang)
            #trang ca nhan, giai tri
            #Xoá trang
            driver.find_element(By.XPATH, var.trang_caidat).click()
            driver.execute_script("window.scrollBy(0,900)", "")
            driver.find_element(By.XPATH, var.trang_caidat_xoatrang).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.trang_caidat_xoatrang_gotrangvinhvien).click()
            time.sleep(3)
        time.sleep(1)
        # if taotrangmoi_taotrang_loaitrang == "Trang bán hàng":
        #     driver.find_element(By.XPATH, var.trang_caidat).click()
        #     time.sleep(1)
        #     driver.find_element(By.XPATH, var.chung).click()
        #     time.sleep(1)
        #     driver.execute_script("window.scrollBy(0,900)", "")
        #     driver.find_element(By.XPATH, var.molaitrangbanhang).click()
        #     time.sleep(1)
        #     driver.find_element(By.XPATH, var.xacnhankhoapage).click()
        #     # driver.find_element(By.XPATH, var.message_xacnhankhoapage).click() #chưa có mesage khóa page
        #     time.sleep(3)
        # else:
        #     #Xoá trang
        #     driver.find_element(By.XPATH, var.trang_caidat).click()
        #     driver.execute_script("window.scrollBy(0,900)", "")
        #     driver.find_element(By.XPATH, var.trang_caidat_xoatrang).click()
        #     time.sleep(1)
        #     driver.find_element(By.XPATH, var.trang_caidat_xoatrang_gotrangvinhvien).click()
        #     time.sleep(3)

    def taotrangmoi_banhang(self):
        trang.taotrangmoi(self, data['trang']['trangbanhang'], var.banhang, var.doanhnghiep)

    def taotrangmoi_bankhoahoc(self):
        trang.taotrangmoi(self, data['trang']['trangbankhoahoc'], var.bankhoahoc, var.canhan)

    def taotrangmoi_trangnoidung(self):
        trang.taotrangmoi(self, data['trang']['trangnoidung'], var.trangnoidung, var.doanhnghiep)


    def trang_gioithieu(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(2)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Trang của bạn
        #Trường test bản tin
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Giới thiệu
        driver.find_element(By.XPATH, var.trang_gioithieu).click()
        #Nhập vị trí
        # driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri).click()
        # time.sleep(1)
        driver.find_element(By.XPATH, var.trang_gioithieu_datlaivitri).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban).send_keys(data['trang']['vitri'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_songtai_hanoi)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        #Mô tả
        driver.find_element(By.XPATH, var.trang_gioithieu_mota).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_mota_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_mota_input).send_keys(data['trang']['gioithieu_mota'])
        driver.find_element(By.XPATH, var.trang_gioithieu_mota_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.icon_x).click()


        #Hạng mục
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(Keys.BACKSPACE)
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(data['trang']['gioithieu_webgiaitri'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.webgiaitri)))
        element.click()
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(Keys.TAB)
        driver.implicitly_wait(5)
        try:
            gioithieu_dautich = driver.find_element(By.XPATH, var.gioithieu_dautich).is_displayed()
            logging.info("Trang - Trang của bạn - Giới thiệu ")
            logging.info("check font-end: Hạng mục - có dấu tích không không")
            logging.info(gioithieu_dautich)
            driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Giới thiệu ")
            logging.info("check font-end: Hạng mục - có dấu tích không không")
            logging.info("False")
        time.sleep(2)
        driver.find_element(By.XPATH, var.icon_x).click()
        driver.implicitly_wait(15)

        #Số điện thoại
        driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input).send_keys(data['trang']['gioithieu_sdt'])
        driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.icon_x).click()

        #Email
        driver.find_element(By.XPATH, var.trang_gioithieu_email).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_email_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_email_input).send_keys(data['trang']['gioithieu_email'])
        driver.find_element(By.XPATH, var.trang_gioithieu_email_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.icon_x).click()

        #Web
        driver.find_element(By.XPATH, var.trang_gioithieu_web).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_web_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_web_input).send_keys(data['trang']['gioithieu_web'])
        driver.find_element(By.XPATH, var.trang_gioithieu_web_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.icon_x).click()

        #Thông tin bổ sung
        driver.find_element(By.XPATH, var.trang_gioithieu_thongtinbosung).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_thongtinbosung_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_thongtinbosung_input).send_keys(data['trang']['gioithieu_thongtinbosung'])
        driver.find_element(By.XPATH, var.trang_gioithieu_thongtinbosung_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)

        #Check giới thiệu
        #Mô tả
        driver.find_element(By.XPATH, var.trang_gioithieu_mota)
        check_trang_gioithieu_mota = driver.find_element(By.XPATH,var.check_trang_gioithieu_mota1).text
        print(check_trang_gioithieu_mota)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Mô tả - "+ data['trang']['gioithieu_mota'])
        logging.info(check_trang_gioithieu_mota == data['trang']['gioithieu_mota'])
        #Hang mục
        check_trang_gioithieu_hangmuc = driver.find_element(By.XPATH,var.check_trang_gioithieu_hangmuc1).text
        print(check_trang_gioithieu_hangmuc)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Hạng mục - "+ data['trang']['gioithieu_webgiaitri'])
        logging.info(check_trang_gioithieu_hangmuc == data['trang']['gioithieu_webgiaitri'])
        #số điện thoại
        check_trang_gioithieu_sodienthoai = driver.find_element(By.XPATH,var.check_trang_gioithieu_sodienthoai1).text
        print(check_trang_gioithieu_sodienthoai)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Số điện thoại - "+ data['trang']['gioithieu_sdt'])
        logging.info(check_trang_gioithieu_sodienthoai == data['trang']['gioithieu_sdt'])
        #email
        check_trang_gioithieu_email = driver.find_element(By.XPATH,var.check_trang_gioithieu_email1).text
        print(check_trang_gioithieu_email)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Email - "+ data['trang']['gioithieu_email'])
        logging.info(check_trang_gioithieu_email == data['trang']['gioithieu_email'])
        #web
        check_trang_gioithieu_web = driver.find_element(By.XPATH,var.check_trang_gioithieu_web1).text
        print(check_trang_gioithieu_web)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Web - "+ data['trang']['gioithieu_web'])
        logging.info(check_trang_gioithieu_web == data['trang']['gioithieu_web'])
        #Thông tin bổ sung
        check_trang_gioithieu_thongtinbosung = driver.find_element(By.XPATH,var.check_trang_gioithieu_thongtinbosung1).text
        print(check_trang_gioithieu_thongtinbosung)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Thông tin bổ sung - "+ data['trang']['gioithieu_thongtinbosung'])
        logging.info(check_trang_gioithieu_thongtinbosung == data['trang']['gioithieu_thongtinbosung'])
        time.sleep(0.5)




    def trang_gioithieu_dulieusai(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Trang của bạn
        #Trường test bản tin
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Giới thiệu
        driver.find_element(By.XPATH, var.trang_gioithieu).click()
        #Nhập vị trí
        driver.find_element(By.XPATH, var.trang_gioithieu_datlaivitri).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban).send_keys(data['trang']['vitri_sai'])
        time.sleep(1.5)
        check_trang_gioithieu_vitrisai = driver.find_element(By.XPATH,var.check_trang_gioithieu_vitrisai1).text
        print(check_trang_gioithieu_vitrisai)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Vị trí sai - "+ check_trang_gioithieu_vitrisai)
        logging.info(check_trang_gioithieu_vitrisai == "Không có dữ liệu...")

        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban).send_keys(data['trang']['vitri1'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_songtai_langson)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        #Mô tả
        driver.find_element(By.XPATH, var.trang_gioithieu_mota).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_mota_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_mota_input).send_keys(data['trang']['gioithieu_mota_dai'])
        driver.find_element(By.XPATH, var.trang_gioithieu_mota_input).send_keys(Keys.TAB)
        time.sleep(1)
        driver.find_element(By.XPATH, var.gioithieu_dautich_sai).click()
        check_trang_gioithieu_motadai = driver.find_element(By.XPATH,var.check_trang_gioithieu_motadai1).text
        print(check_trang_gioithieu_motadai)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Mô tả dài - "+ check_trang_gioithieu_motadai)
        logging.info(check_trang_gioithieu_motadai == "Mô tả không được quá 255 ký tự")

        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_mota_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_mota_input).send_keys(data['trang']['gioithieu_mota1'])
        driver.find_element(By.XPATH, var.trang_gioithieu_mota_input).send_keys(Keys.TAB)
        time.sleep(1)
        driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_gioithieu_mota)

        #Hạng mục
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(Keys.BACKSPACE)
        #hạng mục 1
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(data['trang']['gioithieu_trangphucquanao'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangphucquanao)))
        element.click()
        #hạng mục 2
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(data['trang']['gioithieu_phongcanh'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.phongcanh)))
        element.click()
        #hạng mục 3
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(data['trang']['gioithieu_trangtrinhacua'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangtrinhacua)))
        element.click()
        #hạng mục 4
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys("Cửa hàng phụ tùng ô tô")
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.cuahangphutungoto)))
        element.click()
        check_trang_gioithieu_hangmuc_daiqua3 = driver.find_element(By.XPATH,var.check_trang_gioithieu_hangmuc_daiqua3).text
        print(check_trang_gioithieu_hangmuc_daiqua3)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Hjang mục - "+ check_trang_gioithieu_hangmuc_daiqua3)
        logging.info(check_trang_gioithieu_hangmuc_daiqua3 == "Chỉ được chọn tối đa 3 hạng mục")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(Keys.BACKSPACE)
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(Keys.BACKSPACE)
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(Keys.BACKSPACE)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_gioithieu_webgiaitri_input).send_keys(data['trang']['gioithieu_trangphucquanao'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangphucquanao)))
        element.click()

        driver.implicitly_wait(2)
        try:
            gioithieu_dautich = driver.find_element(By.XPATH, var.gioithieu_dautich).is_displayed()
            logging.info("Trang - Trang của bạn - Giới thiệu ")
            logging.info("check font-end: Hạng mục - có dấu tích không không")
            logging.info(gioithieu_dautich)
            driver.find_element(By.XPATH, var.gioithieu_dautich).click()          #Giới thiệu - hạng mục chưa có dấu tích xanh khi lưu
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Giới thiệu ")
            logging.info("check font-end: Hạng mục - có dấu tích không không")
            logging.info("False")
        driver.implicitly_wait(15)
        driver.find_element(By.XPATH, var.icon_x).click()

        #Số điện thoại
        driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai).click()
        time.sleep(1)
        #Số điện thoại ngắn
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input).send_keys(data['trang']['gioithieu_sdt_ngan'])
        driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich_sai).click()
        check_trang_gioithieu_sodienthoai_ngan = driver.find_element(By.XPATH,var.check_trang_gioithieu_sodienthoai_ngan1).text
        print(check_trang_gioithieu_sodienthoai_ngan)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Số điện thoại ngắn - " + check_trang_gioithieu_sodienthoai_ngan)
        logging.info(check_trang_gioithieu_sodienthoai_ngan == "Số điện thoại không hợp lệ")
        time.sleep(1)
        #Số điện thoại dài
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input).send_keys(data['trang']['gioithieu_sdt_dai'])
        driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich_sai).click()
        check_trang_gioithieu_sodienthoai_dai = driver.find_element(By.XPATH,var.check_trang_gioithieu_sodienthoai_ngan1).text
        print(check_trang_gioithieu_sodienthoai_dai)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Số điện thoại dài - " + check_trang_gioithieu_sodienthoai_dai)
        logging.info(check_trang_gioithieu_sodienthoai_dai == "Số điện thoại không hợp lệ")
        time.sleep(2)
        #Số điện thoại hợp lệ
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input).send_keys(data['trang']['gioithieu_sdt_trung'])
        driver.find_element(By.XPATH, var.trang_gioithieu_sodienthoai_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)
        check_trang_gioithieu_sodienthoai = driver.find_element(By.XPATH,var.check_trang_gioithieu_sodienthoai1).text
        print(check_trang_gioithieu_sodienthoai)

        #Email
        driver.find_element(By.XPATH, var.trang_gioithieu_email).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_email_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_email_input).send_keys(data['trang']['gioithieu_email_dulieusai'])
        driver.find_element(By.XPATH, var.trang_gioithieu_email_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich_sai).click()
        time.sleep(1)
        check_trang_gioithieu_email_dulieusai = driver.find_element(By.XPATH,var.check_trang_gioithieu_email_dulieusai1).text
        print(check_trang_gioithieu_email_dulieusai)
        logging.info("Trang - Trang của bạn - Giới thiệu ")
        logging.info("check font-end: Email dữ liệu sai - "+ check_trang_gioithieu_email_dulieusai)
        logging.info(check_trang_gioithieu_email_dulieusai == "email không hợp lệ")
        time.sleep(1)

        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_email_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_email_input).send_keys(data['trang']['gioithieu_email1'])
        driver.find_element(By.XPATH, var.trang_gioithieu_email_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)

        #Web
        driver.find_element(By.XPATH, var.trang_gioithieu_web).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_web_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_web_input).send_keys(data['trang']['gioithieu_web_dulieusai'])
        driver.find_element(By.XPATH, var.trang_gioithieu_web_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich_sai).click()
        time.sleep(1)
        driver.implicitly_wait(3)
        try:
            check_trang_gioithieu_web_dulieusai = driver.find_element(By.XPATH,var.check_trang_gioithieu_web_dulieusai1).is_displayed()
            logging.info("Trang - Trang của bạn - Giới thiệu ")
            logging.info("check font-end: Web dữ liệu sai - Vui lòng nhập một URL hợp lệ")
            logging.info(check_trang_gioithieu_web_dulieusai)
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Giới thiệu ")
            logging.info("check font-end: Web dữ liệu sai - Vui lòng nhập một URL hợp lệ")
            logging.info("False")
        time.sleep(1)

        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_web_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_web_input).send_keys(data['trang']['gioithieu_web1'])
        driver.find_element(By.XPATH, var.trang_gioithieu_web_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)

        driver.implicitly_wait(15)
        #Thông tin bổ sung
        driver.find_element(By.XPATH, var.trang_gioithieu_thongtinbosung).click()
        time.sleep(1)
        #Dài quá 50k ký tự
        # xoa = driver.find_element(By.XPATH, var.trang_gioithieu_thongtinbosung_input)     #dài quá đơ máy luôn
        # xoa.send_keys(Keys.CONTROL, "a")
        # driver.find_element(By.XPATH, var.trang_gioithieu_thongtinbosung_input).send_keys(data['trang']['gioithieu_thongtinbosung_daiqua50k'])
        # driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        # time.sleep(1)
        # check_trang_gioithieu_thongtinbosung_daiqua50k = driver.find_element(By.XPATH,var.check_trang_gioithieu_thongtinbosung_daiqua50k).text
        # print(check_trang_gioithieu_thongtinbosung_daiqua50k)
        # logging.info("Trang - Trang của bạn - Giới thiệu ")
        # logging.info("check font-end: Thông tin bổ sung dài quá 50k ký tự - "+ check_trang_gioithieu_thongtinbosung_daiqua50k)
        # logging.info(check_trang_gioithieu_thongtinbosung_daiqua50k == "Thông tin bổ sung không được quá 50000 ký tự")
        # time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_thongtinbosung_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_thongtinbosung_input).send_keys(data['trang']['gioithieu_thongtinbosung1'])
        driver.find_element(By.XPATH, var.trang_gioithieu_thongtinbosung_input).send_keys(Keys.TAB)
        driver.find_element(By.XPATH, var.gioithieu_dautich).click()
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(3)


    def anh(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Trang của bạn
        #Trường test bản tin
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Ảnh
        driver.find_element(By.XPATH, var.trang_anh).click()
        driver.execute_script("window.scrollBy(0,400)", "")
        #ảnh
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_anh_xemanhdau).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.thich).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(data['trang']['anh_binhluan'])
        driver.find_element(By.XPATH, var.vietbinhluan).submit()
        time.sleep(2)
        driver.find_element(By.XPATH, var.esc).click()

        #album
        # driver.find_element(By.XPATH, var.trang_anh_album).click()
        button = driver.find_element(By.XPATH, var.trang_anh_album)
        driver.execute_script("arguments[0].click();", button)

        driver.find_element(By.XPATH, var.trang_anh_alum_taomoi).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_ten).send_keys(data['trangcanhan_anh_video']['trangcanhan_anh_anh'])
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_mota).send_keys(data['trangcanhan_anh_video']['trangcanhan_anh_mota'])
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_tailen).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia1.exe")
        time.sleep(2)
        #chọn vị trí
        driver.find_element(By.XPATH, var.trang_alum_chonvitri).click()
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban)
        xoa.send_keys(Keys.CONTROL, "a")

        driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban).send_keys(data['trang']['vitri1'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_songtai_langson)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_anhtailen_chuthich).send_keys(data['trangcanhan_anh_video']['chuthich'])
        # driver.find_element(By.XPATH, var.trangcanhan_anh_album_dang).click()
        time.sleep(3)
        driver.back()
        time.sleep(1)
        # driver.find_element(By.XPATH, var.trangcanhan_anh_album_iconxoa).click()        #Chưa có dấu 3 chấm để xoá và cập nhật album
        # driver.find_element(By.XPATH, var.trangcanhan_anh_album_xoa).click()
        # time.sleep(1.5)
        # driver.find_element(By.XPATH, var.trangcanhan_khoanhkhac_xemvideo_xoa).click()
        # time.sleep(1)


    def music(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Trang của bạn
        #Trường test bản tin
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_music).click()
        driver.execute_script("window.scrollBy(0,400)", "")
        #Phổ biến nhất
        tenbainhac1 = driver.find_element(By.XPATH, var.tenbainhac1).text
        logging.info("Trang - Music - Phổ biến nhất")
        logging.info("check font-end: Tên bài hát phổ biến nhất 1 " + tenbainhac1)
        logging.info(tenbainhac1 == "Forget Me Now - Fishy ft. Trí Dũng「Cukak Remix」/ Audio Lyrics")

        tacgia1 = driver.find_element(By.XPATH, var.tacgia1).text
        logging.info("Trang - Music - Phổ biến nhất")
        logging.info("check font-end: Tác giả bài hát phổ biến nhất 1 " + tacgia1)
        logging.info(tacgia1 == "Trường test bản tin")

        # driver.find_element(By.XPATH, var.trang_music_phobiennhat_nghebai1).click()
        button = driver.find_element(By.XPATH, var.trang_music_phobiennhat_nghebai1)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.implicitly_wait(3)
        #Chọn xem có hiển thị popup nghe không
        try:
            tenbainhac1_dangnghe = driver.find_element(By.XPATH, var.tenbainhac1_dangnghe1).text
            logging.info("Trang - Music - Phổ biến nhất")
            logging.info("check font-end: nghe bài số 1 - Có hiển thị popup nghe không")
            logging.info(tenbainhac1_dangnghe == tenbainhac1)
        except NoSuchElementException:
            logging.info("Trang - Music - Phổ biến nhất")
            logging.info("check font-end: nghe bài số 1 - có hiển thị popup nghe không")
            logging.info("False")

        try:
            tenbainhac1_dangnghe_trangthai = driver.find_element(By.XPATH, var.tenbainhac1_dangnghe_trangthai).is_displayed()
            logging.info("Trang - Music - Phổ biến nhất")
            logging.info("check font-end: nghe bài số 1 - Có nghe được bài số 1 không")
            logging.info(tenbainhac1_dangnghe_trangthai)
        except NoSuchElementException:
            logging.info("Trang - Music - Phổ biến nhất")
            logging.info("check font-end: nghe bài số 1 - Có nghe được bài số 1 không")
            logging.info("False")
        driver.implicitly_wait(15)
        time.sleep(1)

        #Danh sách đĩa nhạc
        trang_music_danhsachdianhac_tenalbum = driver.find_element(By.XPATH,var.trang_music_danhsachdianhac_tenalbum).text

        try:
            check_trang_music_danhsachdianhac = driver.find_element(By.XPATH, var.trang_music_danhsachdianhac_tenalbum).is_displayed()
            logging.info("Trang - Music - Danh sách đĩa nhạc")
            logging.info("check font-end: có hiển thị album hay không")
            logging.info(check_trang_music_danhsachdianhac)
        except NoSuchElementException:
            logging.info("Trang - Music - Danh sách đĩa nhạc")
            logging.info("check font-end: có hiển thị album hay không")
            logging.info("False")

        driver.find_element(By.XPATH, var.trang_music_danhsachdianhac_album1).click()
        time.sleep(2)

        trang_music_danhsachdianhac_tenalbumvaoxem = driver.find_element(By.XPATH, var.trang_music_danhsachdianhac_tenalbumvaoxem).text
        logging.info("Trang - Music - Danh sách đĩa nhạc")
        logging.info("check font-end: Có giống album ở không gian âm nhạc không")
        logging.info(trang_music_danhsachdianhac_tenalbum)
        logging.info(trang_music_danhsachdianhac_tenalbumvaoxem)
        logging.info(trang_music_danhsachdianhac_tenalbum == trang_music_danhsachdianhac_tenalbumvaoxem)
        driver.back()
        time.sleep(1)
        driver.execute_script("window.scrollBy(0,700)", "")

        #Fan cũng thích
        driver.find_element(By.XPATH, var.fancungthich1).click()
        time.sleep(2)
        check_trang_fancungthich = driver.find_element(By.XPATH, var.check_trang_fancungthich1).text
        logging.info("Trang - Music - Fan cũng thích")
        logging.info("check font-end: xem được trang Anh dep không")
        logging.info(check_trang_fancungthich == "Anh dep")
        driver.back()
        time.sleep(3)



    def video(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Trang của bạn
        #Trường test bản tin
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_video).click()
        driver.execute_script("window.scrollBy(0,400)", "")
        #Phổ biến nhất
        try:
            check_trang_video_phobiennhat = driver.find_element(By.XPATH,var.check_trang_video_phobiennhat1).is_displayed()
            print(check_trang_video_phobiennhat)
            logging.info("Trang - Trang của bạn - Video")
            logging.info("check font-end: Phổ biến nhất - Có video hiển thị hay không")
            logging.info(check_trang_video_phobiennhat)
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Video")
            logging.info("check font-end: Phổ biến nhất - Có video hiển thị hay không")
            logging.info("False")

        #Bài đăng
        try:
            check_trang_video_baidang = driver.find_element(By.XPATH,var.check_trang_video_baidang1).is_displayed()
            print(check_trang_video_baidang)
            logging.info("Trang - Trang của bạn - Video")
            logging.info("check font-end: Tất cả video - Bài đăng - Có video hiển thị hay không")
            logging.info(check_trang_video_baidang)
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Video")
            logging.info("check font-end: Tất cả video - Bài đăng - Có video hiển thị hay không")
            logging.info("False")

        #Khoảnh khăc
        driver.find_element(By.XPATH, var.trang_video_khoanhkhac).click()
        time.sleep(2)
        try:
            check_trang_video_khoanhkhac = driver.find_element(By.XPATH,var.check_trang_video_khoanhkhac1).is_displayed()
            print(check_trang_video_khoanhkhac)
            logging.info("Trang - Trang của bạn - Video")
            logging.info("check font-end: Tất cả video - Khoảnh khắc - Có video hiển thị hay không")
            logging.info(check_trang_video_khoanhkhac)
        except  NoSuchElementException:
            logging.info("Trang - Trang của bạn - Video")
            logging.info("check font-end: Tất cả video - Khoảnh khắc - Có video hiển thị hay không")
            logging.info("False")

        #Danh sách phát
        driver.find_element(By.XPATH, var.trang_video_danhsachphat).click()
        time.sleep(2)
        try:
            check_trang_video_danhsachphat = driver.find_element(By.XPATH,var.check_trang_video_danhsachphat1).is_displayed()
            print(check_trang_video_danhsachphat)
            logging.info("Trang - Trang của bạn - Video")
            logging.info("check font-end: Tất cả video - Danh sách phát - Danh sách có hiển thị hay không")
            logging.info(check_trang_video_danhsachphat)
        except  NoSuchElementException:
            logging.info("Trang - Trang của bạn - Video")
            logging.info("check font-end: Tất cả video - Danh sách phát - Danh sách có hiển thị hay không")
            logging.info("False")
        time.sleep(1)

        #Tạo danh sách phát
        driver.find_element(By.XPATH, var.trang_video_taodanhsachphat).click()
        driver.find_element(By.XPATH, var.trang_video_taodanhsachphat_viettieude).send_keys(data['trang']['danhsachphat_tieude'])
        driver.find_element(By.XPATH, var.trang_video_taodanhsachphat_mota).send_keys(data['trang']['danhsachphat_mota'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_video_taodanhsachphat_tailenanhbia).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia2.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(3)

        #Xem danh sách phát vừa tạo
        driver.find_element(By.XPATH, var.danhsachphat_xem).click()         #mất video để thêm vào danh sach
        time.sleep(1)
        #Chọn video 1
        driver.find_element(By.XPATH, var.danhsachphat_xem_themvideo).click()
        driver.find_element(By.XPATH, var.danhsachphat_xem_themvideo_timkiemvideo).send_keys(data['trang']['timkiem_video1'])
        driver.find_element(By.XPATH, var.danhsachphat_xem_themvideo_chonvideo1).click()
        #Chọn video 2
        xoa = driver.find_element(By.XPATH, var.danhsachphat_xem_themvideo_timkiemvideo)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.danhsachphat_xem_themvideo_timkiemvideo).send_keys(data['trang']['timkiem_video2'])
        driver.find_element(By.XPATH, var.danhsachphat_xem_themvideo_chonvideo1).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        check_danhsachphat_themvideo = driver.find_element(By.XPATH, var.check_danhsachphat_themvideo1).text
        logging.info("Trang - Video - Danh sách phát - Thêm video")
        logging.info("check font-end: Message - Thêm video vào danh sách phát thành công")
        logging.info(check_danhsachphat_themvideo == "Thêm video vào danh sách phát thành công")
        #Xoá video
        driver.find_element(By.XPATH, var.danhsachphat_icon3chamxoavideo1).click()
        driver.find_element(By.XPATH, var.danhsachphat_xoavideo1).click()
        driver.find_element(By.XPATH, var.xoa).click()
        check_danhsachphat_xoavideo = driver.find_element(By.XPATH, var.check_danhsachphat_xoavideo1).text
        logging.info("Trang - Video - Danh sách phát - Xoá video")
        logging.info("check font-end: Message - Xoá video thành công")
        logging.info(check_danhsachphat_xoavideo == "Xoá video thành công")

        #Dấu 3 chấm của danh sách phát
        #Chỉnh sửa danh sách phát
        driver.find_element(By.XPATH, var.danhsachphat_icon3cham).click()
        driver.find_element(By.XPATH, var.danhsachphat_icon3cham_chinhsuadanhsachphat).click()
        xoa = driver.find_element(By.XPATH, var.chinhsuadanhsachphat_mota)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuadanhsachphat_mota).send_keys(data['trang']['danhsachphat_mota1'])
        driver.find_element(By.XPATH, var.capnhat).click()
        # driver.find_element(By.XPATH, var.huy).click()
        time.sleep(2)
        #Sao chép ID danh sách phát
        driver.find_element(By.XPATH, var.danhsachphat_icon3cham).click()
        driver.find_element(By.XPATH, var.danhsachphat_icon3cham_saochepid).click()
        check_danhsachphat_saochepid = driver.find_element(By.XPATH, var.check_danhsachphat_saochepid1).text
        logging.info("Trang - Video - Danh sách phát - Sao chép id")
        logging.info("check font-end: Message - Sao chép id thành công")
        logging.info(check_danhsachphat_saochepid == "Sao chép id thành công")
        time.sleep(2)
        #Xoá
        driver.find_element(By.XPATH, var.xoa1).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,800)", "")
        driver.implicitly_wait(3)
        try:
            check_danhsachphat2 = driver.find_element(By.XPATH, var.check_danhsachphat2).is_displayed()
            logging.info("Trang - Video - Danh sách phát - Xem - Xoá")
            logging.info("check font-end: Xoá danh sách phát thành công")
            logging.info("Fasle")
        except NoSuchElementException:
            logging.info("Trang - Video - Danh sách phát - Xem - Xoá")
            logging.info("check font-end: Xoá danh sách phát thành công")
            logging.info("True")
        time.sleep(2)


    def cuahang(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Trang của bạn
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_cuahang).click()
        driver.execute_script("window.scrollBy(0,400)", "")
        time.sleep(4)
        #Mã giảm giá của shop
        try:
            trang_cuahang_magiamgia = driver.find_element(By.XPATH, var.trang_cuahang_magiamgia1).is_displayed()
            logging.info("Trang - Cửa hàng - Mã giảm giá")
            logging.info("check font-end: Shop có mã giảm giá không?")
            logging.info(trang_cuahang_magiamgia)
        except NoSuchElementException:
            logging.info("Trang - Cửa hàng - Mã giảm giá")
            logging.info("check font-end: Shop có mã giảm giá không?")
            logging.info("False")

    def taomoisukien(self, tensukien, quyen, chonquyen):
        #Tên sự kiên
        driver.get("https://cmc-fe.emso.vn/events/create/offline")
        time.sleep(2)
        driver.find_element(By.XPATH, var.taosukien_tensukien).send_keys(tensukien)
        #Ngày bắt đầu
        driver.find_element(By.XPATH, var.taosukien_iconngaybatdau).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukien_ngaybatdau_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukien_ngaybatdau_chonnam_2024).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukien_ngaybatdau_chonngay_20).click()
        #Giờ bắt đầu
        driver.find_element(By.XPATH, var.taosukien_giobatdau_input).send_keys(data['trang']['sukien_giobatdau'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.giobatdau0800)))
        element.click()
        #Ngày kết thúc
        driver.find_element(By.XPATH, var.taosukien_ngayvagioketthuc).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukien_iconngayketthuc).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukien_ngayketthuc_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukien_ngayketthuc_chonnam_2024).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taosukien_ngayketthuc_chonngay_25).click()
        time.sleep(1)
        #Giờ Kết thúc
        driver.find_element(By.XPATH, var.taosukien_gioketthuc_input).send_keys(data['trang']['sukien_gioketthuc'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.gioketthuc1530)))
        element.click()
        #Quyền sự kiện
        driver.find_element(By.XPATH, var.taosukien_quyensukien_input).send_keys(quyen)
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, chonquyen)))
        element.click()
        #Hạng mục
        driver.find_element(By.XPATH, var.taosukien_hangmuc_input).send_keys("Game")
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.hangmuc_chon)))
        element.click()
        #Vị trí
        button = driver.find_element(By.XPATH, var.sukien_iconvitri)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban).send_keys(data['trang']['vitri'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_songtai_hanoi)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        cuon = driver.find_element(By.XPATH, var.anhbia)
        driver.execute_script("arguments[0].scrollIntoView();", cuon)
        #Mô tả
        driver.find_element(By.XPATH, var.sukien_mota).send_keys(data['trang']['sukien_mota'])
        #Ảnh bìa
        button = driver.find_element(By.XPATH, var.sukien_tailenanhbia)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/tft.exe")
        time.sleep(1)
        #Người đồng tổ chức
        driver.find_element(By.XPATH, var.sukien_nguoidongtochuc_input).send_keys(data['trang']['sukien_ngocmai'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.ngocmai)))
        element.click()
        time.sleep(2)
        #Check xem trước sự kiện
        check_xemtruocsukien_thoigian = driver.find_element(By.XPATH,var.check_xemtruocsukien_thoigian1).text
        print(check_xemtruocsukien_thoigian)
        logging.info("Sự kiện - Tạo mới sự kiện")
        logging.info("check font-end: Thời gian - 08:00 Thứ sáu, 20 tháng 12, 2024 - 15:30 Thứ tư, 25 tháng 12, 2024")
        logging.info(check_xemtruocsukien_thoigian == "08:00 Thứ sáu, 20 tháng 12, 2024 - 15:30 Thứ tư, 25 tháng 12, 2024")

        check_xemtruocsukien_tensukien = driver.find_element(By.XPATH,var.check_xemtruocsukien_tensukien1).text
        print(check_xemtruocsukien_tensukien)
        logging.info("Sự kiện - Tạo mới sự kiện")
        logging.info("check font-end: Tên sự kiện - "+ tensukien)
        logging.info(check_xemtruocsukien_tensukien == tensukien)

        check_xemtruocsukien_quyen = driver.find_element(By.XPATH,var.check_xemtruocsukien_quyen1).text
        if quyen == "Công khai":
            print(check_xemtruocsukien_quyen)
            logging.info("Sự kiện - Tạo mới sự kiện")
            logging.info("check font-end: Quyền của sự kiện - "+ quyen)
            logging.info(check_xemtruocsukien_quyen)
            logging.info(check_xemtruocsukien_quyen == "Công khai - Tất cả mọi người")

        if quyen == "Riêng tư":
            print(check_xemtruocsukien_quyen)
            logging.info("Sự kiện - Tạo mới sự kiện")
            logging.info("check font-end: Quyền của sự kiện - "+ quyen)
            logging.info(check_xemtruocsukien_quyen)
            logging.info(check_xemtruocsukien_quyen == "Riêng tư - Chỉ những người được mời")

        if quyen == "Bạn bè":
            print(check_xemtruocsukien_quyen)
            logging.info("Sự kiện - Tạo mới sự kiện")
            logging.info("check font-end: Quyền của sự kiện - "+ quyen)
            logging.info(check_xemtruocsukien_quyen)
            logging.info(check_xemtruocsukien_quyen == "Bạn bè - Bạn bè của bạn")
        #Đăng sự kiện
        driver.find_element(By.XPATH, var.sukien_dangsukien).click()
        time.sleep(3)
        check_taosukienmoi_message_tao = driver.find_element(By.XPATH,var.check_taosukienmoi_message1).text
        print(check_taosukienmoi_message_tao)
        logging.info("Sự kiện - Tạo mới sự kiện")
        logging.info("check font-end: message Tạo - "+ check_taosukienmoi_message_tao)
        logging.info(check_taosukienmoi_message_tao == "Đăng sự kiện thành công")

        #Xoá sự kiện
        time.sleep(1)
        driver.find_element(By.XPATH, var.sukien_dau3cham).click()
        driver.find_element(By.XPATH, var.sukien_dau3cham_huysukien).click()
        driver.find_element(By.XPATH, var.xoa).click()
        check_taosukienmoi_message_xoa = driver.find_element(By.XPATH,var.check_taosukienmoi_message_xoa).text
        print(check_taosukienmoi_message_xoa)
        logging.info("Sự kiện - Giới thiệu")
        logging.info("check font-end: message Xoá sự kiện- "+ check_taosukienmoi_message_xoa)
        logging.info(check_taosukienmoi_message_xoa == "Xóa sự kiện thành công")
        time.sleep(2)


    def xemthem(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Trang của bạn
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_xemthem).click()
        driver.execute_script("window.scrollBy(0,400)", "")

        #Sự kiện
        driver.find_element(By.XPATH, var.trang_xemthem_sukien).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_xemthem_sukien_taosukienmoi).click()
        time.sleep(2)
        trang.taomoisukien(self, data['trang']['tensukien_congkhai'], "Công khai", var.congkhai)
        trang.taomoisukien(self, data['trang']['tensukien_riengtu'], "Riêng tư", var.riengtu)
        trang.taomoisukien(self, data['trang']['tensukien_banbe'], "Bạn bè", var.banbe)
        # time.sleep(2)
        driver.get("https://cmc-fe.emso.vn/page/108277159419223992/event")
        #Sự kiện sắp tới
        check_trang_sukiensaptoi_thoigian = driver.find_element(By.XPATH, var.check_trang_sukiensaptoi_thoigian).text
        logging.info("Trang - Xem thêm - Sự kiện - Sự kiện sắp tới")
        logging.info("check font-end: Thời gian - 08:15, 13 tháng 11, 2024")
        logging.info(check_trang_sukiensaptoi_thoigian)
        logging.info(check_trang_sukiensaptoi_thoigian == "08:15, 13 tháng 11, 2024")

        check_trang_sukiensaptoi_tensukien = driver.find_element(By.XPATH, var.check_trang_sukiensaptoi_tensukien).text
        logging.info("Trang - Xem thêm - Sự kiện - Sự kiện sắp tới")
        logging.info("check font-end: Tên sự kiện - Giao lưu mèo hn")
        logging.info(check_trang_sukiensaptoi_tensukien)
        logging.info(check_trang_sukiensaptoi_tensukien == "Giao lưu mèo hn")

        check_trang_sukiensaptoi_khachmoi = driver.find_element(By.XPATH, var.check_trang_sukiensaptoi_khachmoi).text
        logging.info("Trang - Xem thêm - Sự kiện - Sự kiện sắp tới")
        logging.info("check font-end: Khách mời - "+ check_trang_sukiensaptoi_khachmoi)
        logging.info(check_trang_sukiensaptoi_khachmoi)
        logging.info(check_trang_sukiensaptoi_khachmoi != None)


        #Sự kiện đã qua
        check_trang_sukiendaqua_thoigian = driver.find_element(By.XPATH, var.check_trang_sukiendaqua_thoigian).text
        logging.info("Trang - Xem thêm - Sự kiện - Sự kiện đã qua")
        logging.info("check font-end: Thời gian - 16:00, 13 tháng 11, 2023")
        logging.info(check_trang_sukiendaqua_thoigian)
        logging.info(check_trang_sukiendaqua_thoigian == "16:00, 12 tháng 12, 2023")

        check_trang_sukiendaqua_tensukien = driver.find_element(By.XPATH, var.check_trang_sukiendaqua_tensukien).text
        logging.info("Trang - Xem thêm - Sự kiện - Sự kiện đã qua")
        logging.info("check font-end: Tên sự kiện - test trang")
        logging.info(check_trang_sukiendaqua_tensukien)
        logging.info(check_trang_sukiendaqua_tensukien == "test trang")

        check_trang_sukiendaqua_khachmoi = driver.find_element(By.XPATH, var.check_trang_sukiendaqua_khachmoi).text
        logging.info("Trang - Xem thêm - Sự kiện - Sự kiện đã qua")
        logging.info("check font-end: Khách mời - 0 người quan tâm ·2 người sẽ tham gia")
        logging.info(check_trang_sukiendaqua_khachmoi)
        logging.info(check_trang_sukiendaqua_khachmoi == "0 người quan tâm ·2 người sẽ tham gia")

        time.sleep(2)
        #Nhóm
        driver.execute_script("window.scrollBy(0,400)", "")
        driver.find_element(By.XPATH, var.trang_xemthem).click()
        driver.find_element(By.XPATH, var.trang_xemthem_nhom).click()
        time.sleep(1)

        driver.find_element(By.XPATH, var.trang_xemthem_nhom_huylienket).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_xemthem_nhom_lienketnhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_xemthem_nhom_lienketnhomlienket1).click()
        check_trang_xemthem_nhom_lienket = driver.find_element(By.XPATH,var.check_trang_xemthem_nhom_lienket).text
        time.sleep(1)
        print(check_trang_xemthem_nhom_lienket)
        logging.info("Trang - Xem thêm - Nhóm Liên kết")
        logging.info("check font-end: Message - Bạn có muốn liên kết nam test với Trang của mình là Trường test bản tin không?")
        logging.info(check_trang_xemthem_nhom_lienket == "Bạn có muốn liên kết nam test với Trang của mình là Trường test bản tin không?")
        driver.find_element(By.XPATH, var.lienket).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trang_xemthem_nhom_huylienket).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_xemthem_nhom_lienketnhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_xemthem_nhom_lienketnhomlienket2).click()
        time.sleep(1)

        check_trang_xemthem_nhom_chuaditoi = driver.find_element(By.XPATH, var.check_trang_xemthem_nhom_chuaditoi).text
        check_trang_xemthem_quyennhom_chuaditoi = driver.find_element(By.XPATH, var.check_trang_xemthem_quyennhom_chuaditoi).text
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_xemthem_nhom_ditoi).click()
        check_trang_xemthem_nhom_ditoi = driver.find_element(By.XPATH, var.check_trang_xemthem_nhom_ditoi1).text
        logging.info("Trang - Xem thêm - Nhóm - Đi tới")
        logging.info("check font-end: Đi tới nhóm - " + check_trang_xemthem_nhom_chuaditoi)
        logging.info(check_trang_xemthem_nhom_chuaditoi)
        logging.info(check_trang_xemthem_nhom_ditoi)
        logging.info(check_trang_xemthem_nhom_chuaditoi == check_trang_xemthem_nhom_ditoi)

        check_trang_xemthem_quyennhom_ditoi = driver.find_element(By.XPATH, var.check_trang_xemthem_quyennhom_ditoi).text
        logging.info("Trang - Xem thêm - Nhóm - Đi tới")
        logging.info("check font-end: Quyền của nhóm - " + check_trang_xemthem_quyennhom_ditoi)
        logging.info(check_trang_xemthem_quyennhom_chuaditoi)
        logging.info(check_trang_xemthem_quyennhom_ditoi)
        logging.info(check_trang_xemthem_quyennhom_chuaditoi == check_trang_xemthem_quyennhom_ditoi)
        time.sleep(2)
        driver.back()
        time.sleep(2)

        #Đánh giá
        driver.find_element(By.XPATH, var.trang_xemthem).click()
        driver.find_element(By.XPATH, var.trang_xemthem_danhgia).click()
        time.sleep(1)
        check_trang_xemthem_danhgia_sosao = driver.find_element(By.XPATH, var.check_trang_xemthem_danhgia_sosao1).text
        logging.info("Trang - Xem thêm - Đánh giá")
        logging.info("check font-end: Đề xuất và đánh giá - " + check_trang_xemthem_danhgia_sosao)
        logging.info(check_trang_xemthem_danhgia_sosao)
        logging.info(check_trang_xemthem_danhgia_sosao == "4.7/5")

        check_trang_xemthem_danhgia_songuoidanhgia = driver.find_element(By.XPATH, var.check_trang_xemthem_danhgia_songuoidanhgia).text
        logging.info("Trang - Xem thêm - Đánh giá")
        logging.info(check_trang_xemthem_danhgia_songuoidanhgia)
        logging.info(check_trang_xemthem_danhgia_songuoidanhgia == "Dựa trên đánh giá của 3 người")

        check_trang_xemthem_danhgia1_tennguoidanhgia1 = driver.find_element(By.XPATH, var.check_trang_xemthem_danhgia1_tennguoidanhgia1).text
        logging.info("Trang - Xem thêm - Đánh giá")
        logging.info("check font-end: Tên người đánh giá - " + check_trang_xemthem_danhgia1_tennguoidanhgia1)
        logging.info(check_trang_xemthem_danhgia1_tennguoidanhgia1)
        logging.info(check_trang_xemthem_danhgia1_tennguoidanhgia1 == "Vương Lâm")

        check_trang_xemthem_danhgia1_binhluan1 = driver.find_element(By.XPATH, var.check_trang_xemthem_danhgia1_binhluan1).text
        logging.info("Trang - Xem thêm - Đánh giá")
        logging.info("check font-end: Bình luận người đánh giá 1 - " + check_trang_xemthem_danhgia1_tennguoidanhgia1)
        logging.info(check_trang_xemthem_danhgia1_binhluan1)
        logging.info(check_trang_xemthem_danhgia1_binhluan1 == "4 sao thoy")
        time.sleep(2)

        #Cộng đồng
        driver.find_element(By.XPATH, var.trang_xemthem).click()
        driver.find_element(By.XPATH, var.trang_xemthem_congdong).click()
        time.sleep(1)
        logging.info("Trang - Xem thêm - Cộng đồng")
        logging.info("Chức năng chưa hoạt động")

        #Ưu đãi
        driver.find_element(By.XPATH, var.trang_xemthem).click()
        driver.find_element(By.XPATH, var.trang_xemthem_uudai).click()
        time.sleep(1)
        logging.info("Trang - Xem thêm - Ưu đãi")
        logging.info("Chức năng chưa hoạt động")

    def cuahang(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_binhthuan).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_trangchu1).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_cuahang).click()
        driver.execute_script("window.scrollBy(0,400)", "")

        logging.info("Trang - Trang của bạn - Cửa hàng")
        logging.info("check font-end: Button - Xem thêm")
        logging.info("Chức năng chưa hoạt động")

        #Mã giảm giá
        driver.implicitly_wait(3)
        try:
            check_trang_cuahang_magiamgia = driver.find_element(By.XPATH, var.check_trang_cuahang_magiamgia).is_displayed()
            logging.info("Trang - Trang của bạn - Cửa hàng")
            logging.info("check font-end: Mã giảm giá - Có hiển thị hay không")
            logging.info(check_trang_cuahang_magiamgia)
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Cửa hàng")
            logging.info("check font-end: Mã giảm giá - Có hiển thị hay không")
            logging.info("False")
        driver.implicitly_wait(15)

        #Gợi ý cho bạn
        check_trang_cuahang_goiy_tensp = driver.find_element(By.XPATH, var.check_trang_cuahang_goiy_tensp).text
        logging.info("Trang - Trang của bạn - Cửa hàng")
        logging.info("check font-end: Gợi ý cho bạn - Tên sản phẩm có hiển thị hay không")
        logging.info(check_trang_cuahang_goiy_tensp)
        logging.info(check_trang_cuahang_goiy_tensp != None)

        check_trang_cuahang_goiy_anhsp = driver.find_element(By.XPATH, var.check_trang_cuahang_goiy_anhsp).get_attribute("src")
        logging.info("Trang - Trang của bạn - Cửa hàng")
        logging.info("check font-end: Gợi ý cho bạn - Ảnh sản phẩm có hiển thị hay không")
        logging.info(check_trang_cuahang_goiy_anhsp)
        logging.info(check_trang_cuahang_goiy_anhsp != None)

        #Bán chạy nhất
        driver.execute_script("window.scrollBy(0,1000)", "")
        check_trang_cuahang_banchaynhat_tensp = driver.find_element(By.XPATH, var.check_trang_cuahang_banchaynhat_tensp).text
        logging.info("Trang - Trang của bạn - Cửa hàng")
        logging.info("check font-end: Bán chạy nhất - Tên sản phẩm có hiển thị hya không")
        logging.info(check_trang_cuahang_banchaynhat_tensp)
        logging.info(check_trang_cuahang_banchaynhat_tensp != None)

        check_trang_cuahang_banchaynhat_anhsp = driver.find_element(By.XPATH, var.check_trang_cuahang_banchaynhat_anhsp).get_attribute("src")
        logging.info("Trang - Trang của bạn - Cửa hàng")
        logging.info("check font-end: Bán chạy nhất - Ảnh sản phẩm có hiển thị hay không")
        logging.info(check_trang_cuahang_banchaynhat_anhsp)
        logging.info(check_trang_cuahang_banchaynhat_anhsp != None)

        #Sản phẩm mới
        driver.execute_script("window.scrollBy(0,300)", "")
        check_trang_cuahang_spmoi_tensp = driver.find_element(By.XPATH, var.check_trang_cuahang_spmoi_tensp).text
        logging.info("Trang - Trang của bạn - Cửa hàng")
        logging.info("check font-end: Sản phẩm mới - Tên sản phẩm có hin thị hya không")
        logging.info(check_trang_cuahang_spmoi_tensp)
        logging.info(check_trang_cuahang_spmoi_tensp != None)

        check_trang_cuahang_spmoi_anhsp = driver.find_element(By.XPATH, var.check_trang_cuahang_spmoi_anhsp).get_attribute("src")
        logging.info("Trang - Trang của bạn - Cửa hàng")
        logging.info("check font-end: Sản phẩm mới - Ảnh sản phẩm có hiển thị hay không")
        logging.info(check_trang_cuahang_spmoi_anhsp)
        logging.info(check_trang_cuahang_spmoi_anhsp != None)
        time.sleep(2)

        #Tất cả sản phẩm
        driver.execute_script("window.scrollBy(0,-1500)", "")
        driver.find_element(By.XPATH, var.trang_cuahang_tatcasanpham).click()
        time.sleep(1)
        element1 = driver.find_element(By.XPATH, var.trang_cuahang_tatcasanpham)
        color = element1.value_of_css_property("color")
        logging.info("Trang - Trang của bạn - Cửa hàng")
        logging.info("check font-end: Chuyển màu khi chọn nút - Tất cả sản phẩm")
        logging.info(color)
        logging.info(color == "rgba(243, 132, 52, 1)")

        try:
            check_trang_cuahang_tatcasanpham = driver.find_element(By.XPATH, var.check_trang_cuahang_tatcasanpham).is_displayed()
            logging.info("Trang - Trang của bạn - Cửa hàng")
            logging.info("check font-end: Tất cả sản phẩm - Có hiển thị sản phẩm không")
            logging.info(check_trang_cuahang_tatcasanpham)
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Cửa hàng")
            logging.info("check font-end: Tất cả sản phẩm - Có hiển thị sản phẩm không")
            logging.info("False")

        try:
            check_trang_cuahang_tatcasanpham_anh = driver.find_element(By.XPATH, var.check_trang_cuahang_tatcasanpham_anh).get_attribute("src")
            logging.info("Trang - Trang của bạn - Cửa hàng")
            logging.info("check font-end: Tất cả sản phẩm - Có hiển thị ảnh sản phẩm không")
            logging.info(check_trang_cuahang_tatcasanpham_anh != None)
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Cửa hàng")
            logging.info("check font-end: Tất cả sản phẩm - Có hiển thị ảnh sản phẩm không")
            logging.info("False")


        #Váy
        driver.find_element(By.XPATH, var.trang_cuahang_tatcasanpham_vay).click()
        time.sleep(1.5)
        driver.implicitly_wait(2)
        try:
            check_trang_cuahang_tatcasanpham_vay = driver.find_element(By.XPATH, var.check_trang_cuahang_tatcasanpham).is_displayed()
            logging.info("Trang - Trang của bạn - Cửa hàng")
            logging.info("check font-end: Chọn xem sp Váy - Có hiển thị váy hay không")
            logging.info(check_trang_cuahang_tatcasanpham_vay)
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Cửa hàng")
            logging.info("check font-end: Chọn xem sp Váy - Có hiển thị váy hay không")
            logging.info("False")
        driver.implicitly_wait(15)


        #Đầm
        driver.find_element(By.XPATH, var.trang_cuahang_tatcasanpham_dam).click()
        time.sleep(1.5)
        check_trang_cuahang_tatcasanpham_dam = driver.find_element(By.XPATH, var.check_trang_cuahang_tatcasanpham).text
        logging.info("Trang - Trang của bạn - Cửa hàng")
        logging.info("check font-end: Chọn xem sp Đầm - Có hiển thị Đầm hay không")
        logging.info(check_trang_cuahang_tatcasanpham_dam)
        logging.info(check_trang_cuahang_tatcasanpham_dam == "VÁY ĐẦM BODY ÔM SÁT HÀN QUỐC SANG CHẢNH TÔN DÁNG SIÊU XINH THỜI TRANG 1989")
        time.sleep(2)

        #Giày thể thao
        driver.find_element(By.XPATH, var.trang_cuahang_tatcasanpham_xemthem).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_cuahang_tatcasanpham_giaythethao).click()
        time.sleep(1.5)
        check_trang_cuahang_tatcasanpham_giay = driver.find_element(By.XPATH, var.check_trang_cuahang_tatcasanpham).text
        logging.info("Trang - Trang của bạn - Cửa hàng")
        logging.info("check font-end: Chọn xem sp Giày thể thao - Có hiển thị Giày thẻ thao nữ hay không")
        logging.info(check_trang_cuahang_tatcasanpham_giay)
        logging.info(check_trang_cuahang_tatcasanpham_giay == "Giày thẻ thao nữ")
        time.sleep(2)



    def cacnutkhac(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Thêm mới
        #Tạo khoảnh khắc
        driver.find_element(By.XPATH, var.trang_themmoi).click()
        driver.find_element(By.XPATH, var.trang_themmoi_khoanhkhac).click()
        time.sleep(1)
        try:
            driver.find_element(By.XPATH, var.trang_themmoi_khoanhkhac_nhapnoidung).send_keys(data['trang']['themoi_khoanhkhac'])
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trang_themmoi).click()
            driver.find_element(By.XPATH, var.trang_themmoi_khoanhkhac).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.trang_themmoi_khoanhkhac_nhapnoidung).send_keys(data['trang']['themoi_khoanhkhac'])
        driver.find_element(By.XPATH, var.chontaptin).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/bungnolongdungcam.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.dangbai).click()
        time.sleep(20)
        # driver.find_element(By.XPATH, var.x).click()

        check_trang_themmoi_thoigiankhoanhkhac = driver.find_element(By.XPATH, var.check_trang_themmoi_thoigiankhoanhkhac1).text
        logging.info("Trang - Trang của bạn - Thêm mới - Khoảnh khắc")
        logging.info("check font-end: Thời gian khoảnh khắc - Vài giây trước")
        logging.info(check_trang_themmoi_thoigiankhoanhkhac)
        logging.info(check_trang_themmoi_thoigiankhoanhkhac == "Vài giây trước ")

        check_trang_themmoi_tenkhoanhkhac = driver.find_element(By.XPATH, var.check_trang_themmoi_tenkhoanhkhac).text
        logging.info("Trang - Trang của bạn - Thêm mới - Khoảnh khắc")
        logging.info("check font-end: Tên khoảnh khắc - " + data['trang']['themoi_khoanhkhac'])
        logging.info(check_trang_themmoi_tenkhoanhkhac)
        logging.info(check_trang_themmoi_tenkhoanhkhac == data['trang']['themoi_khoanhkhac'])

        try:
            check_trang_themmoi_videokhoanhkhac = driver.find_element(By.XPATH, var.check_trang_themmoi_videokhoanhkhac).get_attribute("src")
            logging.info("Trang - Trang của bạn - Thêm mới - Khoảnh khắc")
            logging.info("check font-end: Có tạo được video khoảnh khắc không?")
            logging.info(check_trang_themmoi_videokhoanhkhac)
            logging.info(check_trang_themmoi_videokhoanhkhac!= None)
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Thêm mới - Khoảnh khắc")
            logging.info("check font-end: Có tạo được video khoảnh khắc không?")
            logging.info("False")
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,600)", "")

        #Xoá khoảnh khắc trang
        driver.find_element(By.XPATH, var.trang_baivietkhoanhkhac_icon3cham).click()
        driver.find_element(By.XPATH, var.xoavideo).click()
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(1)
        check_message_trang_khoanhkhac_xoa = driver.find_element(By.XPATH, var.check_message_trang_khoanhkhac_xoa1).text
        logging.info("Trang - Trang của bạn - Thêm mới - Khoảnh khắc")
        logging.info("check font-end: Message xoá khoảnh khắc - Bài viết khoảnh khắc của bạn đã bị xóa.")
        logging.info(check_message_trang_khoanhkhac_xoa)
        logging.info(check_message_trang_khoanhkhac_xoa == "Bài viết khoảnh khắc của bạn đã bị xóa.")
        time.sleep(1)

        driver.execute_script("window.scrollBy(0,-600)", "")
        #Nhật ký story
        driver.find_element(By.XPATH, var.trang_themmoi).click()
        driver.find_element(By.XPATH, var.trang_themmoi_nhatkystory).click()
        time.sleep(1)
        logging.info("Trang - Trang của bạn - Thêm mới")
        logging.info("check font-end: Nhật ký (Story)")
        logging.info("Chức năng chưa hoạt động")

        #Tìm kiếm
        driver.find_element(By.XPATH, var.trang_timkiem1).click()
        driver.find_element(By.XPATH, var.trang_timkiem_input).send_keys(data['trang']['timkiem1'])
        driver.find_element(By.XPATH, var.trang_timkiem_input).send_keys(Keys.ENTER)
        time.sleep(2)
        check_trang_timkiemtrongtrang = driver.find_element(By.XPATH, var.check_trang_timkiemtrongtrang).text
        logging.info("Trang - Trang của bạn - Tìm kiếm trang trang")
        logging.info("check font-end: Có tìm kiếm được nutri caffee hay không")
        logging.info(check_trang_timkiemtrongtrang == "nutri caffee")
        time.sleep(1)


        #Hành động khác
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Thích
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham).click()
        time.sleep(1)
        check_trang_trangthai_thich = driver.find_element(By.XPATH, var.check_trang_trangthai_thich1).text
        logging.info("Trang - Trang của bạn - Dấu 3 chấm")
        logging.info("check font-end: Trạng thái thích trang - Đã thích")
        logging.info(check_trang_trangthai_thich == "Đã thích")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham_thichtrang).click()
        time.sleep(5)

        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham).click()
        check_trang_trangthai_thich = driver.find_element(By.XPATH, var.check_trang_trangthai_thich1).text
        logging.info("Trang - Trang của bạn - Dấu 3 chấm")
        logging.info("check font-end: Trạng thái thích trang - Thích")
        logging.info(check_trang_trangthai_thich == "Thích")
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham_thichtrang).click()
        time.sleep(1)

        #theo dõi
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham).click()
        time.sleep(1)
        check_trang_trangthai_datheodoi = driver.find_element(By.XPATH, var.check_trang_trangthai_theodoi1).text
        logging.info("Trang - Trang của bạn - Dấu 3 chấm")
        logging.info("check font-end: Trạng thái Theo dõi trang - Đã theo dõi")
        logging.info(check_trang_trangthai_datheodoi == "Đã theo dõi")
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham_theodoi).click()
        time.sleep(5)

        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham).click()
        time.sleep(1)
        check_trang_trangthai_theodoi = driver.find_element(By.XPATH, var.check_trang_trangthai_theodoi1).text
        logging.info("Trang - Trang của bạn - Dấu 3 chấm")
        logging.info("check font-end: Trạng thái Theo dõi trang - Theo dõi")
        logging.info(check_trang_trangthai_theodoi == "Theo dõi")
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham_theodoi).click()
        time.sleep(1)

        #Tạo trang
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham_taotrang).click()
        time.sleep(1)
        check_trang_truongtest_dau3cham_taotrang = driver.find_element(By.XPATH, var.check_trang_truongtest_dau3cham_taotrang).text
        logging.info("Trang - Trang của bạn - Dấu 3 chấm - Tạo trang")
        logging.info("check font-end: Đang ở popup - "+ check_trang_truongtest_dau3cham_taotrang)
        logging.info(check_trang_truongtest_dau3cham_taotrang == "Tạo Trang")
        time.sleep(1)
        driver.find_element(By.XPATH, var.icon_x1).click()
        time.sleep(2)

        #Mời bạn bè
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham_moibanbe).click()
        time.sleep(1)
        check_trang_truongtest_dau3cham_moibb = driver.find_element(By.XPATH, var.check_trang_truongtest_dau3cham_moibb1).text
        logging.info("Trang - Trang của bạn - Dấu 3 chấm - Mời bạn bè")
        logging.info("check font-end: Popup khi chọn mời bạn bè - "+ check_trang_truongtest_dau3cham_moibb)
        logging.info(check_trang_truongtest_dau3cham_moibb == "Mời bạn thích trang Trường test bản tin")
        driver.find_element(By.XPATH, var.trang_truongtest_dau3cham_moibanbe_x).click()
        time.sleep(2)

        #Chọn cách tương tác
        # driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac).click()
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_tranquangtruong).click()
        driver.implicitly_wait(3)
        time.sleep(1)
        try:
            check_trang_choncactuongtac_user = driver.find_element(By.XPATH,var.check_trang_choncactuongtac_user).is_displayed()
            logging.info("Trang - Trang của bạn - Chọn cách tương tác")
            logging.info("check font-end: Chọn với quyền user")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Chọn cách tương tác")
            logging.info("check font-end: Chọn với quyền user")
            logging.info("True")


        # driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac1).click()
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac1)
        driver.execute_script("arguments[0].click();", button)

        driver.find_element(By.XPATH, var.choncachtuongtac_truongtestbantin).click()
        time.sleep(1)
        try:
            check_trang_choncactuongtac_user = driver.find_element(By.XPATH,var.check_trang_choncactuongtac_user).is_displayed()
            logging.info("Trang - Trang của bạn - Chọn cách tương tác")
            logging.info("check font-end: Chọn với quyền quản trị")
            logging.info(check_trang_choncactuongtac_user)
        except NoSuchElementException:
            logging.info("Trang - Trang của bạn - Chọn cách tương tác")
            logging.info("check font-end: Chọn với quyền quản trị")
            logging.info("False")
        driver.implicitly_wait(15)
        time.sleep(2)


    def themnuthanhdong(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Thêm nút hành động
        #Gửi tin nhắn
        driver.find_element(By.XPATH, var.trang_themnuthanhdong).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut_guitinnhan).click()
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_tranquangtruong).click()
        driver.find_element(By.XPATH, var.trang_guitinnhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_guitinnhan_input).send_keys("alo alo")
        driver.find_element(By.XPATH, var.trang_guitinnhan_input).send_keys(Keys.ENTER)
        time.sleep(2)
        check_trang_themnuthanhdong_guitinnhan = driver.find_element(By.XPATH, var.check_trang_themnuthanhdong_guitinnhan).text
        logging.info("Trang - Thêm nút hành động - Gửi tin nhắn")
        logging.info("check font-end: Gửi đoạn text cho page - "+ check_trang_themnuthanhdong_guitinnhan)
        logging.info(check_trang_themnuthanhdong_guitinnhan == "Trường test bản tin")
        driver.find_element(By.XPATH, var.trang_guitinnhan_input_x).click()
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac1)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_truongtestbantin).click()
        #Theo dõi
        driver.find_element(By.XPATH, var.trang_themnuthanhdong).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut_theodoi).click()
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_tranquangtruong).click()
        check_trang_themnuthanhdong_theodoi = driver.find_element(By.XPATH, var.check_trang_themnuthanhdong_theodoi).text
        logging.info("Trang - Thêm nút hành động - Theo dõi")
        logging.info("check font-end: Trạng thái theo dõi hiện tại - "+ check_trang_themnuthanhdong_theodoi)
        logging.info(check_trang_themnuthanhdong_theodoi == "Đang theo dõi")
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac1)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_truongtestbantin).click()
        #Đăng ký
        driver.find_element(By.XPATH, var.trang_themnuthanhdong).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut_dangky).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut_dangky_input).send_keys(data['trang']['dangky_sai'])
        check_trang_themnuthanhdong_dangky_sai = driver.find_element(By.XPATH, var.check_trang_themnuthanhdong_dangky_sai1).text
        logging.info("Trang - Thêm nút hành động - Đăng ký")
        logging.info("check font-end: Nhập dữ liệu sai - "+ check_trang_themnuthanhdong_dangky_sai)
        logging.info(check_trang_themnuthanhdong_dangky_sai == "Vui lòng nhập đúng dạng http:// hoặc https://")
        xoa = driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut_dangky_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut_dangky_input).send_keys(data['trang']['dangky_dung'])
        driver.find_element(By.XPATH, var.capnhat).click()      #Nút chưa chuyển màu (Trang - thêm hnh động - đăng ký/tìm hiểu thêm)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_tranquangtruong).click()
        check_trang_themnuthanhdong_theodoi = driver.find_element(By.XPATH, var.check_trang_themnuthanhdong_theodoi).text
        logging.info("Trang - Thêm nút hành động - Đăng ký")
        logging.info("check font-end: Trạng thái nút thêm hành động hiẹn tại - "+ check_trang_themnuthanhdong_theodoi)
        logging.info(check_trang_themnuthanhdong_theodoi == "Đăng ký")
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac1)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_truongtestbantin).click()
        #Tìm hiểu thêm
        driver.find_element(By.XPATH, var.trang_themnuthanhdong).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut_timhieuthem).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut_timhieuthem_input).send_keys(data['trang']['dangky_sai'])
        check_trang_themnuthanhdong_timhieuthem_sai = driver.find_element(By.XPATH, var.check_trang_themnuthanhdong_dangky_sai1).text
        logging.info("Trang - Thêm nút hành động - Tìm hiểu thêm")
        logging.info("check font-end: Nhập dữ liệu sai - "+ check_trang_themnuthanhdong_timhieuthem_sai)
        logging.info(check_trang_themnuthanhdong_timhieuthem_sai == "Vui lòng nhập đúng dạng http:// hoặc https://")
        xoa = driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut_timhieuthem_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_chinhsuanut_timhieuthem_input).send_keys(data['trang']['dangky_dung'])
        driver.find_element(By.XPATH, var.capnhat).click()      #Nút chưa chuyển màu
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_tranquangtruong).click()
        check_trang_themnuthanhdong_theodoi = driver.find_element(By.XPATH, var.check_trang_themnuthanhdong_theodoi).text
        logging.info("Trang - Thêm nút hành động - Tìm hiểu thêm")
        logging.info("check font-end: Trạng thái nút thêm hành động Tìm hiểu thêm - "+ check_trang_themnuthanhdong_theodoi)
        logging.info(check_trang_themnuthanhdong_theodoi == "Tìm hiểu thêm")
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac1)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_truongtestbantin).click()
        #Xoá nút
        driver.find_element(By.XPATH, var.trang_themnuthanhdong).click()
        driver.find_element(By.XPATH, var.trang_themnuthanhdong_xoanut).click()
        time.sleep(1.5)
        check_trang_themnuthanhdong_xoanut = driver.find_element(By.XPATH, var.check_trang_themnuthanhdong_xoanut).text
        logging.info("Trang - Thêm nút hành động - Xoá nút")
        logging.info("check font-end: Xoá nút thêm hành động")
        logging.info(check_trang_themnuthanhdong_xoanut == "Thêm nút hành động")
        time.sleep(2)

    def anhdaidien_anhbia(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Ảnh đại diện
        driver.find_element(By.XPATH, var.trang_anhdaidien_icontailen).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_taianhlen).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_icon_taianhlen).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhmes1.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_mota).send_keys(data['trang']['anhdaidien_mota'])
        driver.find_element(By.XPATH, var.capnhatanhdaidien_luu).click()
        time.sleep(7)
        driver.refresh()
        time.sleep(2)
        check_trang_thayanhdaidien = driver.find_element(By.XPATH, var.check_trang_thayanhdaidien1).text
        logging.info("Trang - Thay ảnh đại diện")
        logging.info("check font-end: Có hiện bài viết trên bảng tin không")
        logging.info(check_trang_thayanhdaidien)
        logging.info(check_trang_thayanhdaidien == "Trường test bản tin\n đã cập nhật ảnh đại diện")

        check_trang_thayanhdaidien_anh = driver.find_element(By.XPATH, var.check_trang_thayanhdaidien_anh).get_attribute("src")
        logging.info("Trang - Thay ảnh đại diện")
        logging.info("check font-end: Có hiện ảnh đại diện trên bảng tin không")
        logging.info(check_trang_thayanhdaidien_anh)
        logging.info(check_trang_thayanhdaidien_anh != None)
        #Ảnh bìa
        driver.execute_script("window.scrollBy(0,-100)", "")
        driver.find_element(By.XPATH, var.trang_anhbia_icontailen).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.tailentuthietbi).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbiatrang.exe")
        time.sleep(1)
        cap_nhat = driver.find_element(By.XPATH, var.capnhat)
        cap_nhat.click()
        time.sleep(3.5)
        driver.refresh()
        time.sleep(2)
        check_trang_thayanhbia = driver.find_element(By.XPATH, var.check_trang_thayanhbia).text
        logging.info("Trang - Thay ảnh bìa")
        logging.info("check font-end: Có hiện bài viết trên bảng tin không")
        logging.info(check_trang_thayanhbia)
        logging.info(check_trang_thayanhbia == "Trường test bản tin\n đã cập nhật ảnh bìa của họ")

        check_trang_thayanhbia_anh = driver.find_element(By.XPATH, var.check_trang_thayanhbia_anh).get_attribute("src")
        logging.info("Trang - Thay ảnh đại diện")
        logging.info("check font-end: Có hiện ảnh bìa trên bảng tin không")
        logging.info(check_trang_thayanhbia_anh)
        logging.info(check_trang_thayanhbia_anh != None)
        # chọn khung yeu thuong vn
        driver.find_element(By.XPATH, var.trang_anhdaidien_icontailen).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_themkhung).click()
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_yeuthuongvn).click()
        del driver.requests
        time.sleep(1)
        driver.find_element(By.XPATH, var.capnhatanhdaidien_chonkhung_luu2).click()
        time.sleep(3.5)

    def taobaiviet(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Ảnh video
        driver.find_element(By.XPATH, var.trang_taobaiviet_anhvideo).click()
        check_trang_taobaiviet_anhvideo = driver.find_element(By.XPATH, var.check_trang_taobaiviet_anhvideo).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Ảnh/video")
        logging.info("check font-end: Popup - Tạo bài viết")
        logging.info(check_trang_taobaiviet_anhvideo)
        logging.info(check_trang_taobaiviet_anhvideo == "Tạo bài viết")
        driver.find_element(By.XPATH, var.x).click()

        #Khoảnh khắc
        driver.find_element(By.XPATH, var.trang_taobaiviet_khoanhkhac).click()
        time.sleep(1)
        check_trang_taobaiviet_anhvideo = driver.find_element(By.XPATH, var.check_trang_taobaiviet_anhvideo).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Khoảnh khắc")
        logging.info("check font-end: Popup - Tạo bài khoảnh khắc")
        logging.info(check_trang_taobaiviet_anhvideo)
        logging.info(check_trang_taobaiviet_anhvideo == "Tạo bài khoảnh khắc")
        driver.find_element(By.XPATH, var.icon_x).click()

        #Cảm xúc/hoạt động
        driver.find_element(By.XPATH, var.trang_taobaiviet_camxuchoatdong).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_camxuc).click()
        check_trang_taobaiviet_anhvideo = driver.find_element(By.XPATH, var.check_trang_taobaiviet_anhvideo).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Cảm xúc/Hoạt động")
        logging.info("check font-end: Popup - Bạn đang cảm thấy thế nào?")
        logging.info(check_trang_taobaiviet_anhvideo)
        logging.info(check_trang_taobaiviet_anhvideo == "Bạn đang cảm thấy thế nào?")
        time.sleep(0.5)

        driver.find_element(By.XPATH, var.trang_taobaiviet_hoatdong).click()
        check_trang_taobaiviet_anhvideo = driver.find_element(By.XPATH, var.check_trang_taobaiviet_anhvideo).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Cảm xúc/Hoạt động")
        logging.info("check font-end: Popup - Bạn đang làm gì?")
        logging.info(check_trang_taobaiviet_anhvideo)
        logging.info(check_trang_taobaiviet_anhvideo == "Bạn đang làm gì?")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)

        # Tạo bài viết - ảnh video
        driver.find_element(By.XPATH, var.trang_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_anhvideo).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_mota).send_keys(data['trang']['taobaiviet_anh_mota'])
        driver.find_element(By.XPATH, var.taobaiviet_tailenanhvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/tothichcau.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(1)
        check_trang_taobaiviet_video = driver.find_element(By.XPATH, var.check_trang_taobaiviet_video).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Ảnh/video")
        logging.info("check font-end: Message - Video của bạn đang trong quá trình xử lý, chúng tôi sẽ thông báo cho bạn khi video đã sẵn sàng.")
        logging.info(check_trang_taobaiviet_video)
        logging.info(check_trang_taobaiviet_video == "Video của bạn đang trong quá trình xử lý, chúng tôi sẽ thông báo cho bạn khi video đã sẵn sàng.")
        time.sleep(2)
        # check_trang_baiviet_anhvideo_mota = driver.find_element(By.XPATH, var.check_trang_baiviet_anhvideo_mota1).text      #đăng video lên trang nhưng  ko hiển thị ở dòng tời gian
        # logging.info("Trang - Tạo bài viết - Ảnh/Video - Dòng thời gian")
        # logging.info("check font-end: Mô tả - " + check_trang_baiviet_anhvideo_mota)
        # logging.info(check_trang_baiviet_anhvideo_mota)
        # logging.info(check_trang_baiviet_anhvideo_mota == "")
        #
        # check_trang_baiviet_anhvideo_video = driver.find_element(By.XPATH, var.check_trang_baiviet_anhvideo_video).get_attribute("src")
        # logging.info("Trang - Tạo bài viết - Ảnh/Video - Dòng thời gian")
        # logging.info("check font-end: Có hiển thị video lên dòng thời gian trang không")
        # logging.info(check_trang_baiviet_anhvideo_video)
        # logging.info(check_trang_baiviet_anhvideo_video != None)


        # Tạo bài viết - Video trực tiếp
        driver.find_element(By.XPATH, var.trang_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_videotructiep).click()
        check_trang_taobaiviet_videotructiep = driver.find_element(By.XPATH, var.check_trang_taobaiviet_videotructiep).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Video trực tiếp")
        logging.info("check font-end: Chuyển tới màn hình - Tạo video trực tiếp")
        logging.info(check_trang_taobaiviet_videotructiep)
        logging.info(check_trang_taobaiviet_videotructiep == "Tạo video trực tiếp")
        driver.back()
        time.sleep(2)

        #Tạo bài viết - Cảm xúc/Hoạt động
        driver.find_element(By.XPATH, var.trang_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_camxuchoatdong).click()
        driver.find_element(By.XPATH, var.camxuchoatdong_xinhxan).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_mota).send_keys(data['trang']['taobaiviet_camxuchoatdong_mota'])
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(2)
        check_trang_taobaiviet_hoatdongcamxuc = driver.find_element(By.XPATH, var.check_trang_taobaiviet_hoatdongcamxuc).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Cảm xúc/Hoạt động")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_trang_taobaiviet_hoatdongcamxuc)
        logging.info(check_trang_taobaiviet_hoatdongcamxuc == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)
        check_trang_baiviet_camxuchoatdong_tieude = driver.find_element(By.XPATH, var.check_trang_baiviet_camxuchoatdong_tieude).text
        logging.info("Trang - Tạo bài viết - Cảm xúc/Hoạt động - Dòng thời gian")
        logging.info("check font-end: Tiêu đề - "+  check_trang_baiviet_camxuchoatdong_tieude)
        logging.info(check_trang_baiviet_camxuchoatdong_tieude)
        logging.info(check_trang_baiviet_camxuchoatdong_tieude == "Trường test bản tin\n đang  cảm thấy buồn bã")

        check_trang_baiviet_camxuchoatdong_mota = driver.find_element(By.XPATH, var.check_trang_baiviet_camxuchoatdong_mota).text
        logging.info("Trang - Tạo bài viết - Cảm xúc/Hoạt động - Dòng thời gian")
        logging.info("check font-end: Mô tả - " + check_trang_baiviet_camxuchoatdong_mota)
        logging.info(check_trang_baiviet_camxuchoatdong_mota)
        logging.info(check_trang_baiviet_camxuchoatdong_mota == data['trang']['taobaiviet_camxuchoatdong_mota'])

        #Tạo bài viết - Gắn thẻ người khác
        driver.find_element(By.XPATH, var.trang_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_ganthenguoikhac).click()
        driver.find_element(By.XPATH, var.ganthenguoikhac_input).send_keys("Mạnh Cường")
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.taobaiviet_manhcuong)))
        element.click()
        driver.find_element(By.XPATH, var.xong).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_mota).send_keys(data['trang']['taobaiviet_ganthebaiviet_mota'])
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(1)
        check_trang_taobaiviet_ganthenguoikhac = driver.find_element(By.XPATH, var.check_trang_taobaiviet_ganthenguoikhac).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Gắn thẻ người khác")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_trang_taobaiviet_ganthenguoikhac)
        logging.info(check_trang_taobaiviet_ganthenguoikhac == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)
        check_trang_baiviet_ganthenguoikhac_tieude = driver.find_element(By.XPATH, var.check_trang_baiviet_camxuchoatdong_tieude).text
        logging.info("Trang - Tạo bài viết - Gắn thẻ người khác - Dòng thời gian")
        logging.info("check font-end: Tiêu đề - Trường test bản tin cùng với Mạnh Cường")
        logging.info(check_trang_baiviet_ganthenguoikhac_tieude)
        logging.info(check_trang_baiviet_ganthenguoikhac_tieude == "Trường test bản tin\n cùng với \nMạnh Cường")

        check_trang_baiviet_ganthenguoikhac_mota = driver.find_element(By.XPATH, var.check_trang_baiviet_camxuchoatdong_mota).text
        logging.info("Trang - Tạo bài viết - Gắn thẻ người khác - Dòng thời gian")
        logging.info("check font-end: Mô tả - " + check_trang_baiviet_ganthenguoikhac_mota)
        logging.info(check_trang_baiviet_ganthenguoikhac_mota)
        logging.info(check_trang_baiviet_ganthenguoikhac_mota == data['trang']['taobaiviet_ganthebaiviet_mota'])

        #Tạo bài viết - Check in
        driver.find_element(By.XPATH, var.trang_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_checkin).click()
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban).send_keys("Cần Thơ")
        time.sleep(1)
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trang_vitri_cantho)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_taobaiviet_mota).send_keys(data['trang']['taobaiviet_checkin_mota'])
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(2)
        check_trang_taobaiviet_checkin = driver.find_element(By.XPATH, var.check_trang_taobaiviet_checkin1).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Check in")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_trang_taobaiviet_checkin)
        logging.info(check_trang_taobaiviet_checkin == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)
        check_trang_baiviet_checkin_tieude = driver.find_element(By.XPATH, var.check_trang_baiviet_camxuchoatdong_tieude).text
        logging.info("Trang - Tạo bài viết - Check in - Dòng thời gian")
        logging.info("check font-end: Tiêu đề - Trường test bản tin đang ở Cần Thơ")
        logging.info(check_trang_baiviet_checkin_tieude)
        logging.info(check_trang_baiviet_checkin_tieude == "Trường test bản tin\n đang ở Cần Thơ")

        check_trang_baiviet_ganthenguoikhac_mota = driver.find_element(By.XPATH, var.check_trang_baiviet_camxuchoatdong_mota).text
        logging.info("Trang - Tạo bài viết - Check in - Dòng thời gian")
        logging.info("check font-end: Mô tả - " + check_trang_baiviet_ganthenguoikhac_mota)
        logging.info(check_trang_baiviet_ganthenguoikhac_mota)
        logging.info(check_trang_baiviet_ganthenguoikhac_mota == data['trang']['taobaiviet_checkin_mota'])

        check_trang_baiviet_checkin_thanhpho = driver.find_element(By.XPATH, var.check_trang_baiviet_checkin_thanhpho).text
        logging.info("Trang - Tạo bài viết - Check in - Dòng thời gian")
        logging.info("check font-end: Thành phố - " + check_trang_baiviet_checkin_thanhpho)
        logging.info(check_trang_baiviet_checkin_thanhpho)
        logging.info(check_trang_baiviet_checkin_thanhpho == "Cần Thơ")

        #Xoá bài viết
        driver.execute_script("window.scrollBy(0,400)", "")
        driver.find_element(By.XPATH, var.trang_baiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_baiviet_dau3cham_xoabaiviet).click()
        driver.find_element(By.XPATH, var.xoa).click()
        check_trang_baiviet_xoa = driver.find_element(By.XPATH, var.check_trang_baiviet_xoa).text
        logging.info("Trang - Trang của bạn - Chỉnh sửa bài viết")
        logging.info("check font-end: Xoá - Message - Bài viết của bạn đã bị xóa.")
        logging.info(check_trang_baiviet_xoa)
        logging.info(check_trang_baiviet_xoa == "Bài viết của bạn đã bị xóa.")
        time.sleep(1)


        #Tạo bài viết - Gif
        driver.find_element(By.XPATH, var.trang_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_filegif).click()
        driver.find_element(By.XPATH, var.trang_gif_input).send_keys("meo")
        time.sleep(5)

        driver.implicitly_wait(0.1)
        count = 0
        n = 0
        while (count < 25):
            n = int(n)
            n = n + 1
            count = count + 1
            n = str(n)
            trang_taobaiviet_gif_anh = "/html/body/div[" + n + "]/div[3]/div/div[2]/div/div/img[1]"
            print(trang_taobaiviet_gif_anh)
            try:
                button = driver.find_element(By.XPATH, trang_taobaiviet_gif_anh)
                driver.execute_script("arguments[0].click();", button)
                if driver.find_element(By.XPATH, var.check_taobaiviet_gif).is_displayed():
                    break
            except:
                pass
        driver.implicitly_wait(15)

        # try:
        #     wait = WebDriverWait(driver, 10)
        #     element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trang_taobaiviet_gif_anh1)))
        #     element.click()
        # except:
        #     wait = WebDriverWait(driver, 10)
        #     element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trang_taobaiviet_gif_anh2)))
        #     element.click()

        driver.find_element(By.XPATH, var.trang_taobaiviet_mota).send_keys(data['trang']['taobaiviet_gif_mota'])
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(1)
        check_trang_taobaiviet_gif = driver.find_element(By.XPATH, var.check_trang_taobaiviet_gif).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Gif")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_trang_taobaiviet_gif)
        logging.info(check_trang_taobaiviet_gif == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)

        check_trang_baiviet_gif_anh = driver.find_element(By.XPATH, var.check_trang_baiviet_anhvideo_video).get_attribute("src")
        logging.info("Trang - Tạo bài viết - Gif - Dòng thời gian")
        logging.info("check font-end: Có hiển thị Gif lên dòng thời gian trang không")
        logging.info(check_trang_baiviet_gif_anh)
        logging.info(check_trang_baiviet_gif_anh != None)

        check_trang_baiviet_gif_mota = driver.find_element(By.XPATH, var.check_trang_baiviet_camxuchoatdong_mota).text
        logging.info("Trang - Tạo bài viết - Gif - Dòng thời gian")
        logging.info("check font-end: Mô tả - " + check_trang_baiviet_gif_mota)
        logging.info(check_trang_baiviet_gif_mota)
        logging.info(check_trang_baiviet_gif_mota == data['trang']['taobaiviet_gif_mota'])

        # Tạo bài viết - Tổ chức buổi H&Đ
        driver.find_element(By.XPATH, var.trang_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_tochucbuoihd).click()
        driver.find_element(By.XPATH, var.datcauhoi_input).send_keys(data['trang']['taobaiviet_datcauhoi_input'])
        driver.find_element(By.XPATH, var.datcauhoi).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trang_taobaiviet_mota).send_keys(data['trang']['taobaiviet_tochucbuoihd_mota'])
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.dang).click()
        try:
            check_trang_taobaiviet_tochucbuoihd = driver.find_element(By.XPATH, var.check_trang_taobaiviet_tochucbuoihd).text
            logging.info("Trang - Trang của bạn - Tạo bài viết - Tổ chức buổi H&Đ")
            logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
            logging.info(check_trang_taobaiviet_tochucbuoihd)
            logging.info(check_trang_taobaiviet_tochucbuoihd == "Đăng bài viết thành công.")
            driver.find_element(By.XPATH, var.icon_x).click()
        except:
            logging.info("Trang - Trang của bạn - Tạo bài viết - Tổ chức buổi H&Đ")
            logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
            logging.info("False")
        time.sleep(2)
        check_trang_baiviet_tochucbuoihoatdong_mota = driver.find_element(By.XPATH, var.check_trang_baiviet_camxuchoatdong_mota).text
        logging.info("Trang - Tạo bài viết - Tổ chức buổi H&Đ - Dòng thời gian")
        logging.info("check font-end: Mô tả - " + check_trang_baiviet_tochucbuoihoatdong_mota)
        logging.info(check_trang_baiviet_tochucbuoihoatdong_mota)
        logging.info(check_trang_baiviet_tochucbuoihoatdong_mota == data['trang']['taobaiviet_tochucbuoihd_mota'])

        check_trang_baiviet_tochucbuoihoatdong_cauhoi = driver.find_element(By.XPATH, var.check_trang_baiviet_tochucbuoihoatdong_cauhoi).text
        logging.info("Trang - Tạo bài viết - Tổ chức buổi H&Đ - Dòng thời gian")
        logging.info("check font-end: Câu hỏi - " + check_trang_baiviet_tochucbuoihoatdong_cauhoi)
        logging.info(check_trang_baiviet_tochucbuoihoatdong_cauhoi)
        logging.info(check_trang_baiviet_tochucbuoihoatdong_cauhoi == data['trang']['taobaiviet_datcauhoi_input'])


        # Tạo bài viết - Công bố mục tiêu
        driver.find_element(By.XPATH, var.trang_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_congbomuctieu).click()
        driver.find_element(By.XPATH, var.congbomuctieu_input).send_keys(data['trang']['taobaiviet_muctieu_input'])
        driver.find_element(By.XPATH, var.datmuctieu).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.trang_taobaiviet_mota).send_keys(data['trang']['taobaiviet_congbomuctieu_mota'])
        driver.find_element(By.XPATH, var.dang).click()
        check_trang_taobaiviet_congbomuctieu = driver.find_element(By.XPATH, var.check_trang_taobaiviet_congbomuctieu).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Tổ chức buổi H&Đ")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_trang_taobaiviet_congbomuctieu)
        logging.info(check_trang_taobaiviet_congbomuctieu == "Đăng bài viết thành công.")
        # driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)

        check_trang_baiviet_congbomuctieu_mota = driver.find_element(By.XPATH, var.check_trang_baiviet_camxuchoatdong_mota).text
        logging.info("Trang - Tạo bài viết - Công bố mục tiêu - Dòng thời gian")
        logging.info("check font-end: Mô tả - " + check_trang_baiviet_congbomuctieu_mota)
        logging.info(check_trang_baiviet_congbomuctieu_mota)
        logging.info(check_trang_baiviet_congbomuctieu_mota == data['trang']['taobaiviet_congbomuctieu_mota'])

        check_trang_baiviet_congbomuctieu_cauhoi = driver.find_element(By.XPATH, var.check_trang_baiviet_tochucbuoihoatdong_cauhoi).text
        logging.info("Trang - Tạo bài viết - Công bố mục tiêu - Dòng thời gian")
        logging.info("check font-end: Câu hỏi - " + check_trang_baiviet_congbomuctieu_cauhoi)
        logging.info(check_trang_baiviet_congbomuctieu_cauhoi)
        logging.info(check_trang_baiviet_congbomuctieu_cauhoi == data['trang']['taobaiviet_muctieu_input'])
        time.sleep(2)

        #Xem chi tiết bài viết
        driver.execute_script("window.scrollBy(0,200)", "")
        driver.find_element(By.XPATH, var.trang_baiviet_xemchitiet).click()
        check_trang_baiviet_xemchitiet = driver.find_element(By.XPATH, var.check_trang_baiviet_xemchitiet).text
        logging.info("Trang - Bài viết - Xem chi tiết")
        logging.info("check font-end: Popup - " + check_trang_baiviet_xemchitiet)
        logging.info(check_trang_baiviet_xemchitiet)
        logging.info(check_trang_baiviet_xemchitiet == "Thông tin chi tiết về bài viết")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)


    def hopthu(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Gửi tin nhắn cho trang
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_tranquangtruong).click()
        driver.find_element(By.XPATH, var.trang_guitinnhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_guitinnhan_input).send_keys("alo alo")
        driver.find_element(By.XPATH, var.trang_guitinnhan_input).send_keys(Keys.ENTER)
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_guitinnhan_input_x).click()
        button = driver.find_element(By.XPATH, var.trang_truongtest_choncachtuongtac1)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.choncachtuongtac_truongtestbantin).click()
        #Hộp thử
        driver.find_element(By.XPATH, var.trang_hopthu).click()
        # driver.find_element(By.XPATH, var.hover_hopthu1).click()
        button = driver.find_element(By.XPATH, var.hover_hopthu1)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)

        #Check danh bạ
        check_trang_hopthu_danhba_sdt = driver.find_element(By.XPATH, var.check_trang_hopthu_danhba_sdt1).text
        logging.info("Trang - Hộp thư - Danh bạ")
        logging.info("check font-end: Số điện thoại - 0359667777")
        logging.info(check_trang_hopthu_danhba_sdt)
        logging.info(check_trang_hopthu_danhba_sdt == "0359667777")

        check_trang_hopthu_danhba_email = driver.find_element(By.XPATH, var.check_trang_hopthu_danhba_email1).text
        logging.info("Trang - Hộp thư - Danh bạ")
        logging.info("check font-end: Email - truongvck33@gmail.com")
        logging.info(check_trang_hopthu_danhba_email)
        logging.info(check_trang_hopthu_danhba_email == "truongvck33@gmail.com")

        check_trang_hopthu_danhba_ngaysinh = driver.find_element(By.XPATH, var.check_trang_hopthu_danhba_ngaysinh1).text
        logging.info("Trang - Hộp thư - Danh bạ")
        logging.info("check font-end: Ngày sinh - 3/10/2000")
        logging.info(check_trang_hopthu_danhba_ngaysinh)
        logging.info(check_trang_hopthu_danhba_ngaysinh == "3/10/2000")

        check_trang_hopthu_danhba_songtai = driver.find_element(By.XPATH, var.check_trang_hopthu_danhba_songtai1).text
        logging.info("Trang - Hộp thư - Danh bạ")
        logging.info("check font-end: Sống tại - Sống tại Hưng Yên")
        logging.info(check_trang_hopthu_danhba_songtai)
        logging.info(check_trang_hopthu_danhba_songtai == "Sống tại: Hưng Yên")

        check_trang_hopthu_danhba_dentu = driver.find_element(By.XPATH, var.check_trang_hopthu_danhba_dentu1).text
        logging.info("Trang - Hộp thư - Danh bạ")
        logging.info("check font-end: Đến từ - Quê quán : Thái Nguyên")
        logging.info(check_trang_hopthu_danhba_dentu)
        logging.info(check_trang_hopthu_danhba_dentu == "Quê quán : Thái Nguyên")

        check_trang_hopthu_danhba_hocvan = driver.find_element(By.XPATH, var.check_trang_hopthu_danhba_hocvan1).text
        logging.info("Trang - Hộp thư - Danh bạ")
        logging.info("check font-end: Học vấn - Đã tốt nghiệp : Đại học bách khoa hà nội")
        logging.info(check_trang_hopthu_danhba_hocvan)
        logging.info(check_trang_hopthu_danhba_hocvan == "Đã tốt nghiệp : Đại học bách khoa hà nội")

        check_trang_hopthu_danhba_congviec = driver.find_element(By.XPATH, var.check_trang_hopthu_danhba_congviec1).text
        logging.info("Trang - Hộp thư - Danh bạ")
        logging.info("check font-end: Công việc - Làm việc tại Quán trà đá")
        logging.info(check_trang_hopthu_danhba_congviec)
        logging.info(check_trang_hopthu_danhba_congviec == "Làm việc tại Quán trà đá")

        #Thêm ghi chú
        driver.find_element(By.XPATH, var.trang_hopthu_themghichu).click()
        driver.find_element(By.XPATH, var.trang_hopthu_themghichu_input).send_keys(data['trang']['ghichu1'])
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        #Chỉnh sửa ghi chú
        driver.find_element(By.XPATH, var.trang_hopthu_suaghichu).click()
        xoa = driver.find_element(By.XPATH, var.trang_hopthu_themghichu_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_hopthu_themghichu_input).send_keys(data['trang']['ghichu2'])
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        check_trang_hopthu_danhba_suaghichu = driver.find_element(By.XPATH, var.check_trang_hopthu_danhba_suaghichu1).text
        logging.info("Trang - Hộp thư - Danh bạ")
        logging.info("check font-end: Thêm mới ghi chú - sửa ghi chú - Test sửa ghi chú")
        logging.info(check_trang_hopthu_danhba_suaghichu)
        logging.info(check_trang_hopthu_danhba_suaghichu == "Test sửa ghi chú")

        #Xoá ghi chú
        driver.implicitly_wait(2)
        driver.find_element(By.XPATH, var.trang_hopthu_xoaghichu).click()
        try:
            check_trang_hopthu_danhba_suaghichu = driver.find_element(By.XPATH,var.check_trang_hopthu_danhba_suaghichu1).is_displayed()
            logging.info("Trang - Hộp thư - Danh bạ")
            logging.info("check font-end: Thêm mới ghi chú - sửa ghi chú - Test sửa ghi chú")
            logging.info("Fasle")
        except NoSuchElementException:
            logging.info("Trang - Hộp thư - Danh bạ")
            logging.info("check font-end: Thêm mới ghi chú - sửa ghi chú - Test sửa ghi chú")
            logging.info("True")
        driver.implicitly_wait(15)

        # Đính kèm file
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icontailenfile).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhmes1.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_input).send_keys(Keys.ENTER)
        # Gif
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icongif).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icongif_input).send_keys(data['trangchu_tinnhan']['tinnhanmoi_meo'])
        time.sleep(5)
        try:
            driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icongif_anh1a).click()
        except:
            driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icongif_anh1).click()
            # driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_icongif_anh4).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_input).send_keys(Keys.ENTER)
        time.sleep(1)

        # Nhãn dán
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan_2).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan_3).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan_4).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan_1).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_nhandan_1_chonnhandan).click()
        # input
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_input).send_keys(data['trangchu_tinnhan']['tinnhanmoi_input'])
        # Biểu tượng cảm xúc
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page2).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page3).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page4).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page5).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page6).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page7).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page8).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_page9).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_lichsu).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc_chon).click()
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_input).send_keys(Keys.ENTER)
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_bieutuongcamxuc).click()
        # Like
        driver.find_element(By.XPATH, var.trangchu_tinnhanmoi_like).click()
        time.sleep(3)


        # Tùy chọn nhóm
        ## Đánh dáu là chưa đọc      #Trang -  hộp thư - Chức năng chưa hoạt động
        actions = ActionChains(driver)
        # hover_hopthu = driver.find_element(By.XPATH, var.hover_hopthu1)
        # actions.move_to_element(hover_hopthu).perform()
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.trang_tuychonnhom).click()
        # driver.find_element(By.XPATH, var.tuychonnhom_danhdaulachuadoc).click()
        # time.sleep(1)
        # element1 = driver.find_element(By.XPATH, var.check_color_mautennhom)
        # color = element1.value_of_css_property("color")
        # logging.info("Trang - Hộp thư - Cài đặt cuộc hội thoại")
        # logging.info("check font-end: Đánh dấu là chưa đọc")
        # logging.info(color)
        # logging.info(color == "rgba(53, 120, 229, 1)")
        # print(color)
        # # Đánh dáu là đã đọc
        # hover_hopthu = driver.find_element(By.XPATH, var.hover_hopthu1)
        # actions.move_to_element(hover_hopthu).perform()
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.trang_tuychonnhom).click()
        # driver.find_element(By.XPATH, var.tuychonnhom_danhdauladadoc).click()
        # time.sleep(1)
        # element1 = driver.find_element(By.XPATH, var.check_color_mautennhom)
        # color = element1.value_of_css_property("color")
        # logging.info("Trang - Hộp thư - Cài đặt cuộc hội thoại")
        # logging.info("check font-end: Đánh dấu là đã đọc")
        # logging.info(color)
        # logging.info(color == "rgba(101, 103, 107, 1)")
        # print(color)

        # Tắt thông báo     #Trang -  hộp thư - Chức năng chưa hoạt động
        # hover_hopthu = driver.find_element(By.XPATH, var.hover_hopthu1)
        # actions.move_to_element(hover_hopthu).perform()
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.trang_tuychonnhom).click()
        # driver.find_element(By.XPATH, var.tuychonnhom_tatthongbao).click()
        # time.sleep(1)
        # try:
        #     check_nhomcaidat_tatthongbao = driver.find_element(By.XPATH,var.check_nhomcaidat_tatthongbao1).is_displayed()
        #     logging.info("Trang - Hộp thư - Cài đặt cuộc hội thoại")
        #     logging.info("check font-end: Tắt thông báo - có icon tắt thông báo")
        #     logging.info(check_nhomcaidat_tatthongbao)
        # except NoSuchElementException:
        #     logging.info("Trang - Hộp thư - Cài đặt cuộc hội thoại")
        #     logging.info("check font-end: Tắt thông báo - có icon tắt thông báo")
        #     logging.info("False")
        # # Bật thông báo
        # hover_hopthu = driver.find_element(By.XPATH, var.hover_hopthu1)
        # actions.move_to_element(hover_hopthu).perform()
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.trang_tuychonnhom).click()
        # check_nhomcaidat_batthongbao = driver.find_element(By.XPATH, var.check_nhomcaidat_batthongbao1).text
        # logging.info("Trang - Hộp thư - Cài đặt cuộc hội thoại")
        # logging.info("check font-end: Bật thông báo")
        # logging.info(check_nhomcaidat_batthongbao == "Bật thông báo")
        # driver.find_element(By.XPATH, var.tuychonnhom_batthongbao).click()
        # time.sleep(2)

        # #Thêm Nhãn dán
        # hover_hopthu = driver.find_element(By.XPATH, var.hover_hopthu1)
        # actions.move_to_element(hover_hopthu).perform()
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.trang_tuychonnhom).click()
        # driver.find_element(By.XPATH, var.tuychonnhom_iconthemnhan).click()
        # driver.find_element(By.XPATH, var.hopthu_themnhanmoi).send_keys(data['trang']['nhandanmoi'])
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.luu).click()      #Mất nút lưu khi thêm nhãn dán trang 6591
        # time.sleep(1)
        # try:
        #     driver.find_element(By.XPATH, var.hopthu_xoanhan1).click()
        #     driver.find_element(By.XPATH, var.trora).click()
        # except:
        #     driver.find_element(By.XPATH, var.hopthu_xoanhan1).click()
        #     driver.find_element(By.XPATH, var.trora).click()
        # time.sleep(1)

        # Xoá đoạn chat
        hover_hopthu = driver.find_element(By.XPATH, var.hover_hopthu1)
        actions.move_to_element(hover_hopthu).perform()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_tuychonnhom).click()
        button = driver.find_element(By.XPATH, var.tuychonnhom_xoadoanchat)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        check_trang_hopthu_xoadoanchat = driver.find_element(By.XPATH, var.check_trang_hopthu_xoadoanchat).text
        logging.info("Trang - Hộp thư - Cài đặt cuộc hội thoại")
        logging.info("check font-end: Message - Xoá đoạn chat thành công.")
        logging.info(check_trang_hopthu_xoadoanchat == "Xoá đoạn chat thành công.")
        time.sleep(2)

        #Bình luận trên Emso
        driver.find_element(By.XPATH, var.trang_hopthu_binhluantrenemso).click()
        check_trang_hopthu_binhluantrenemso1 = driver.find_element(By.XPATH, var.check_trang_hopthu_binhluantrenemso1).text
        logging.info("Trang - Hộp thư - Bình luận trên Emso")
        logging.info("check font-end: Đánh giá và bình luận về bài viết của bạn - Không có bình luận")
        logging.info(check_trang_hopthu_binhluantrenemso1 != "Không có bình luận")

        check_trang_hopthu_binhluantrenemso2 = driver.find_element(By.XPATH, var.check_trang_hopthu_binhluantrenemso2).text
        logging.info("Trang - Hộp thư - Bình luận trên Emso")
        logging.info("check font-end: Bên phải - Chưa có bình luận nào trên trang của bạn")
        logging.info(check_trang_hopthu_binhluantrenemso2 != "Chưa có bình luận nào trên trang của bạn")
        time.sleep(2)



    def thongbao(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_thongbao).click()
        time.sleep(1)
        #Nhật ký kiểm duyệt
        driver.find_element(By.XPATH, var.trang_thongbao_nhatkykiemduyet).click()
        time.sleep(1)
        check_trang_thongbao = driver.find_element(By.XPATH, var.check_trang_thongbao).text
        logging.info("Trang - Thông báo")
        logging.info("check font-end: Nhật ký kiểm duyệt - Có thông báo hay không?")    #6595 chức năng ko haoht động
        logging.info(check_trang_thongbao != "Không có thông báo kiểm duyệt nào")

    def chatluongtrang(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Chất lượng trang
        driver.find_element(By.XPATH, var.trang_thongbao).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_chatluongtrang).click()
        check_trang_chatluongtrang = driver.find_element(By.XPATH, var.check_trang_chatluongtrang).text
        logging.info("Trang - Thông báo")
        logging.info("check font-end: Chất lượng trang - Trang không có vấn đề")
        logging.info(check_trang_chatluongtrang == "Trang không có vấn đề")
        #Trang này
        driver.find_element(By.XPATH, var.trang_chatluongtrang_quanlyquantrivien).click()
        time.sleep(1)
        check_trang_chatluongtrang_vaitrotrentrang = driver.find_element(By.XPATH, var.check_trang_chatluongtrang_vaitrotrentrang).text
        logging.info("Trang - Thông báo")
        logging.info("check font-end: Chất lượng trang - Trang này - Vai trò trên Trang")
        logging.info(check_trang_chatluongtrang_vaitrotrentrang == "Vai trò trên Trang")
        time.sleep(2)

    def baivietdalenlich(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Lên lịch bài viết
        driver.find_element(By.XPATH, var.trang_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_mota).send_keys(data['trang']['lenlich'])
        driver.find_element(By.XPATH, var.trang_taobaiviet_iconlenlich).click()

        xoa = driver.find_element(By.XPATH, var.trang_taobaiviet_lenlich_ngay)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_taobaiviet_lenlich_ngay).send_keys("08-12-2024")
        driver.find_element(By.XPATH, var.trang_taobaiviet_lenlich).click()

        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(3)
        check_trang_taobaiviet_lenlich = driver.find_element(By.XPATH,var.check_trang_taobaiviet_lenlich).text
        logging.info("Trang - Trang của bạn - Tạo bài viết - Lên lịch")
        logging.info("check font-end: Message tạo bài viết - Đã lên lịch cho bài viết của bạn.")
        logging.info(check_trang_taobaiviet_lenlich)
        logging.info(check_trang_taobaiviet_lenlich == "Đã lên lịch cho bài viết của bạn.")


        #Bài viết đã lên lịch
        driver.find_element(By.XPATH, var.check_trang_baivietdalenlich).click()
        time.sleep(2)
        check_trang_baivietdalenlich = driver.find_element(By.XPATH, var.check_trang_baivietdalenlich1).text
        logging.info("Trang - Bài viết đã lên lịch")
        logging.info("check font-end: Tiêu đề bài viết đợi đăng - "+ data['trang']['lenlich'])
        logging.info(check_trang_baivietdalenlich == data['trang']['lenlich'])

        driver.find_element(By.XPATH, var.trang_baivietdalenlich_doilich).click()
        xoa = driver.find_element(By.XPATH, var.trang_baivietdalenlich_doilichinput)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_baivietdalenlich_doilichinput).send_keys("20-11-2024")
        driver.find_element(By.XPATH, var.trang_baivietdalenlich_doilich_capnhat).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.dangngay).click()
        check_trang_baivietdalenlich_message = driver.find_element(By.XPATH, var.check_trang_baivietdalenlich_message).text
        logging.info("Trang - Bài viết đã lên lịch")
        logging.info("check font-end: Message đăng ngay - Đăng bài viết thành công.")
        logging.info(check_trang_baivietdalenlich_message == "Đăng bài viết thành công.")
        time.sleep(5)

        driver.find_element(By.XPATH, var.trang_trangchu).click()
        driver.refresh()
        time.sleep(2)
        check_trang_baiviet_lenlich = driver.find_element(By.XPATH, var.check_trang_baiviet_camxuchoatdong_mota).text
        logging.info("Trang - Tạo bài viết - Lên lịch - Dòng thời gian")
        logging.info("check font-end: Mô tả - " + check_trang_baiviet_lenlich)
        logging.info(check_trang_baiviet_lenlich)
        logging.info(check_trang_baiviet_lenlich == data['trang']['lenlich'])
        time.sleep(2)

    def chinhsuathongtintrang(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        #Chỉnh sửa thông tin trang
        driver.find_element(By.XPATH, var.trang_chinhsuathongtintrang).click()
        #Tên trang
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_tentrang)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_tentrang).send_keys(data['trang']['chinhsua_tentrang1'])

        #Tên người dùng
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_tennguoidung)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_tennguoidung).send_keys(data['trang']['chinhsua_tennguoidung1'])

        #Mô tả
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_mota)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_mota).send_keys(data['trang']['chinhsua_mota1'])

        #Hạng mục
        # driver.find_element(By.XPATH, var.chinhsuathongtintrang_hangmuc).send_keys()    #chưa viét code chỉnh sửa trang

        #Số điện thoại
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_sodienthoai)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_sodienthoai).send_keys(data['trang']['chinhsua_sdt_sai'])
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_email).click()

        check_trang_chinhsuathongtin_sodienthoai_sai = driver.find_element(By.XPATH,var.check_trang_chinhsuathongtin_sodienthoai_sai).text
        logging.info(check_trang_chinhsuathongtin_sodienthoai_sai)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Validate Số điện thoại - Số điện thoại không hợp lệ")
        logging.info(check_trang_chinhsuathongtin_sodienthoai_sai == "Số điện thoại không hợp lệ")

        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_sodienthoai)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_sodienthoai).send_keys(data['trang']['chinhsua_sdt_dung'])

        #Email
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_email)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_email).send_keys(data['trang']['chinhsua_email_sai'])
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_sodienthoai).click()

        check_trang_chinhsuathongtin_email_sai = driver.find_element(By.XPATH,var.check_trang_chinhsuathongtin_email_sai).text
        logging.info(check_trang_chinhsuathongtin_email_sai)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Validate Email - email không hợp lệ")
        logging.info(check_trang_chinhsuathongtin_email_sai == "email không hợp lệ")

        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_email)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_email).send_keys(data['trang']['chinhsua_email_dung'])

        #Web
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_web)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_web).send_keys(data['trang']['chinhsua_web_sai'])
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_sodienthoai).click()


        check_trang_chinhsuathongtin_web_sai = driver.find_element(By.XPATH,var.check_trang_chinhsuathongtin_web_sai).text
        logging.info(check_trang_chinhsuathongtin_web_sai)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Validate Web - Vui lòng nhập một URL hợp lệ")
        logging.info(check_trang_chinhsuathongtin_web_sai == "Vui lòng nhập một URL hợp lệ")

        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_web)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_web).send_keys(data['trang']['chinhsua_web_dung'])

        #Địa chỉ
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_diachi)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_diachi).send_keys(data['trang']['chinhsua_diachi1'])

        #Tỉnh/Thành phố
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_tinhtp)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_tinhtp).send_keys("Hà Nội")
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.hanoi)))
        element.click()

        #Quận huyện
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_quanhuyen)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_quanhuyen).send_keys("Quận Hoàng Mai")
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.quanhoangmai)))
        element.click()

        #Phường xã
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_phuongxa)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_phuongxa).send_keys("Phường Hoàng Văn Thụ")
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.phuonghoangvanthu)))
        element.click()

        #Mã zip
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_mazip)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_mazip).send_keys(data['trang']['chinhsua_mazip1'])

        #Giờ mở cửa - Không có giờ làm việc
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_giomocua1).click()

        #Thay đổi tạm thời về dịch vụ - tạm đóng cửa
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_thaydoitamthoidv1).click()

        #Chính sách quyền riêng tư
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_chinhsachquuyenriengtu)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_chinhsachquuyenriengtu).send_keys(data['trang']['chinhsua_chinhsachquuyenriengtu_sai'])
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_thaydoitamthoidv1).click()


        check_trang_chinhsuathongtin_chinhsachquuyenriengtu_sai = driver.find_element(By.XPATH,var.check_trang_chinhsuathongtin_chinhsachquuyenriengtu_sai).text
        logging.info(check_trang_chinhsuathongtin_chinhsachquuyenriengtu_sai)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Vlidate Chính sách quyền riêng tư - Liên kết không hợp lệ")
        logging.info(check_trang_chinhsuathongtin_chinhsachquuyenriengtu_sai == "Liên kết không hợp lệ")

        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_chinhsachquuyenriengtu)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_chinhsachquuyenriengtu).send_keys(data['trang']['chinhsua_chinhsachquuyenriengtu_dung'])

        #Tuyên bố quyền sở hữu và quyền tác giả
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_quyensohuu)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_quyensohuu).send_keys(data['trang']['chinhsua_quyensohuu1'])

        #San phẩm
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_sanpham)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_sanpham).send_keys(data['trang']['chinhsua_sanpham1'])

        #Thông tin bổ sung
        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_thongtinbosung)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_thongtinbosung).send_keys(data['trang']['chinhsua_thongtinbosung1'])
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_sanpham).click()
        time.sleep(2.5)

        #Check dữ liệu
        driver.refresh()
        time.sleep(2)
        # Tên trang
        check_trang_chinhsuathongtin_tentrang = driver.find_element(By.XPATH,var.chinhsuathongtintrang_tentrang).text
        logging.info(check_trang_chinhsuathongtin_tentrang)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Tên Trang - " + data['trang']['chinhsua_tentrang1'])
        logging.info(check_trang_chinhsuathongtin_tentrang == data['trang']['chinhsua_tentrang1'])

        xoa = driver.find_element(By.XPATH, var.chinhsuathongtintrang_tentrang)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_tentrang).send_keys(data['trang']['chinhsua_tentrang'])
        driver.find_element(By.XPATH, var.chinhsuathongtintrang_tennguoidung).click()

        # Tên người dùng
        check_trang_chinhsuathongtin_tennguoidung = driver.find_element(By.XPATH,var.chinhsuathongtintrang_tennguoidung).text
        logging.info(check_trang_chinhsuathongtin_tennguoidung)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Tên Nguời dùng - " + data['trang']['chinhsua_tennguoidung1'])
        logging.info(check_trang_chinhsuathongtin_tennguoidung == data['trang']['chinhsua_tennguoidung1'])

        # Mô tả
        check_trang_chinhsuathongtin_mota = driver.find_element(By.XPATH,var.chinhsuathongtintrang_mota).text
        logging.info(check_trang_chinhsuathongtin_mota)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Mô tả - " + data['trang']['chinhsua_mota1'])
        logging.info(check_trang_chinhsuathongtin_mota == data['trang']['chinhsua_mota1'])

        # Hạng mục
        # driver.find_element(By.XPATH, var.chinhsuathongtintrang_hangmuc).send_keys()

        # Số điện thoại
        check_trang_chinhsuathongtin_sodienthoai = driver.find_element(By.XPATH,var.chinhsuathongtintrang_sodienthoai).text
        logging.info(check_trang_chinhsuathongtin_sodienthoai)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Mô tả - " + data['trang']['chinhsua_sdt_dung'])
        logging.info(check_trang_chinhsuathongtin_sodienthoai == data['trang']['chinhsua_sdt_dung'])

        # Email
        check_trang_chinhsuathongtin_email = driver.find_element(By.XPATH,var.chinhsuathongtintrang_email).text
        logging.info(check_trang_chinhsuathongtin_email)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Mô tả - " + data['trang']['chinhsua_email_dung'])
        logging.info(check_trang_chinhsuathongtin_email == data['trang']['chinhsua_email_dung'])

        # Web
        check_trang_chinhsuathongtin_web = driver.find_element(By.XPATH,var.chinhsuathongtintrang_web).text
        logging.info(check_trang_chinhsuathongtin_web)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Mô tả - " + data['trang']['chinhsua_web_dung'])
        logging.info(check_trang_chinhsuathongtin_web == data['trang']['chinhsua_web_dung'])


        # Địa chỉ
        check_trang_chinhsuathongtin_diachi = driver.find_element(By.XPATH,var.chinhsuathongtintrang_diachi).text
        logging.info(check_trang_chinhsuathongtin_diachi)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Mô tả - " + data['trang']['chinhsua_diachi1'])
        logging.info(check_trang_chinhsuathongtin_diachi == data['trang']['chinhsua_diachi1'])

        # Tỉnh/Thành phố
        check_trang_chinhsuathongtin_tinhtp = driver.find_element(By.XPATH,var.chinhsuathongtintrang_tinhtp).get_attribute("value")
        logging.info(check_trang_chinhsuathongtin_tinhtp)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Tỉnh thành phố - Hà Nội")
        logging.info(check_trang_chinhsuathongtin_tinhtp)
        logging.info(check_trang_chinhsuathongtin_tinhtp == "Hà Nội")

        # Quận huyện
        check_trang_chinhsuathongtin_quanhuyen = driver.find_element(By.XPATH,var.chinhsuathongtintrang_quanhuyen).get_attribute("value")
        logging.info(check_trang_chinhsuathongtin_quanhuyen)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Quận huyện - Quận Hoàng Mai")
        logging.info(check_trang_chinhsuathongtin_quanhuyen)
        logging.info(check_trang_chinhsuathongtin_quanhuyen == "Quận Hoàng Mai")

        # Phường xã
        check_trang_chinhsuathongtin_phuongxa = driver.find_element(By.XPATH,var.chinhsuathongtintrang_phuongxa).get_attribute("value")
        logging.info(check_trang_chinhsuathongtin_phuongxa)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Phường xã - Phường Hoàng Văn Thụ")
        logging.info(check_trang_chinhsuathongtin_phuongxa)
        logging.info(check_trang_chinhsuathongtin_phuongxa == "Phường Hoàng Văn Thụ")

        # Mã zip
        check_trang_chinhsuathongtin_mazip = driver.find_element(By.XPATH,var.chinhsuathongtintrang_mazip).text
        logging.info(check_trang_chinhsuathongtin_mazip)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Mã zip - " + data['trang']['chinhsua_mazip1'])
        logging.info(check_trang_chinhsuathongtin_mazip == data['trang']['chinhsua_mazip1'])

        # Giờ mở cửa - Không có giờ làm việc
        check_trang_chinhsuathongtin_giomocua = driver.find_element(By.XPATH, var.check_trang_chinhsuathongtin_giomocua).is_selected()
        check_trang_chinhsuathongtin_giomocua2 =driver.find_element(By.XPATH, var.check_trang_chinhsuathongtin_giomocua2).is_selected()
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("Check font-end: Giờ mở cửa - Không có giờ làm việc, có được chọn không?")
        logging.info(check_trang_chinhsuathongtin_giomocua)
        logging.info(check_trang_chinhsuathongtin_giomocua2)

        # Thay đổi tạm thời về dịch vụ - tạm đóng cửa
        check_trang_chinhsuathongtin_thaydoi = driver.find_element(By.XPATH, var.check_trang_chinhsuathongtin_thaydoitamthoidv1).is_selected()
        check_trang_chinhsuathongtin_thaydoi2 = driver.find_element(By.XPATH, var.check_trang_chinhsuathongtin_thaydoitamthoidv2).is_selected()
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("Check font-end: Thay đổi tạm thời về dịch vụ - tạm đóng cửa, có được chọn không?")
        logging.info(check_trang_chinhsuathongtin_thaydoi)
        # logging.info(check_trang_chinhsuathongtin_thaydoi2)

        # Chính sách quyền riêng tư
        check_trang_chinhsuathongtin_chinhsachquuyenriengtu = driver.find_element(By.XPATH,var.chinhsuathongtintrang_chinhsachquuyenriengtu).text
        logging.info(check_trang_chinhsuathongtin_chinhsachquuyenriengtu)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Chính sách quyền riêng tư - " + data['trang']['chinhsua_chinhsachquuyenriengtu_dung'])
        logging.info(check_trang_chinhsuathongtin_chinhsachquuyenriengtu == data['trang']['chinhsua_chinhsachquuyenriengtu_dung'])

        # Tuyên bố quyền sở hữu và quyền tác giả
        check_trang_chinhsuathongtin_quyensohuu = driver.find_element(By.XPATH,var.chinhsuathongtintrang_quyensohuu).text
        logging.info(check_trang_chinhsuathongtin_quyensohuu)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Tuyên bố quyền sở hữu và quyền tác giả - " + data['trang']['chinhsua_quyensohuu1'])
        logging.info(check_trang_chinhsuathongtin_quyensohuu == data['trang']['chinhsua_quyensohuu1'])

        # San phẩm
        check_trang_chinhsuathongtin_sanpham = driver.find_element(By.XPATH,var.chinhsuathongtintrang_sanpham).text
        logging.info(check_trang_chinhsuathongtin_sanpham)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Sản phẩm - " + data['trang']['chinhsua_sanpham1'])
        logging.info(check_trang_chinhsuathongtin_sanpham == data['trang']['chinhsua_sanpham1'])

        # Thông tin bổ sung
        check_trang_chinhsuathongtin_thongtinbosung = driver.find_element(By.XPATH,var.chinhsuathongtintrang_thongtinbosung).text
        logging.info(check_trang_chinhsuathongtin_thongtinbosung)
        logging.info("Trang - Chỉnh sửa thông tin trang")
        logging.info("check font-end: Thông tin bổ sung - " + data['trang']['chinhsua_thongtinbosung1'])
        logging.info(check_trang_chinhsuathongtin_thongtinbosung == data['trang']['chinhsua_thongtinbosung1'])
        time.sleep(2)

    def caidat_chung(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        # Cài đặt
        driver.find_element(By.XPATH, var.trang_caidat).click()
        # Cài đặt - Chung
        check_caidatchung_gioihannguoidungdangbai = driver.find_element(By.XPATH, var.caidatchung_gioihannguoidungdangbai).text
        logging.info("Trang - Cài đặt - Chung")
        logging.info("check font-end: Giới hạn người dùng đăng bài - Không cho phép đăng bài")
        logging.info(check_caidatchung_gioihannguoidungdangbai)
        logging.info(check_caidatchung_gioihannguoidungdangbai == "Không cho phép đăng bài")

        #Giới hạn tuổi
        xoa = driver.find_element(By.XPATH, var.caidatchung_gioihandotuoi)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.caidatchung_gioihandotuoi).send_keys(data['trang']['caidatchung_gioihantuoi'])

        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.gioihantuoi18)))
        element.click()

        #Kiểm duyệt nội dung
        driver.implicitly_wait(2)
        driver.find_element(By.XPATH, var.xoatatca).click()
        try:
            check_caidatchung_kiemduyetnoidung1 = driver.find_element(By.XPATH,var.check_caidatchung_kiemduyetnoidung1).is_displayed()
            logging.info("Trang - Cài đặt - Chung")
            logging.info("check font-end: Xoá tất cả kiểm duyệt nội dung 1")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Trang - Cài đặt - Chung")
            logging.info("check font-end: Xoá tất cả kiểm duyệt nội dung 1")
            logging.info("True")

        driver.find_element(By.XPATH, var.caidatchung_kiemduyetnoidung).send_keys(data['trang']['caidatchung_kiemduyetnoidung2'])
        driver.find_element(By.XPATH, var.them).click()
        check_caidatchung_kiemduyetnoidung1 = driver.find_element(By.XPATH, var.check_caidatchung_kiemduyetnoidung1).text
        logging.info("Trang - Cài đặt - Chung")
        logging.info("check font-end: Thêm kiểm duyệt nội dung 1 - Đằng Hoá Nguyên")
        logging.info(check_caidatchung_kiemduyetnoidung1)
        logging.info(check_caidatchung_kiemduyetnoidung1 == "Đằng Hoá Nguyên")

        #Xoá kiểm duyệt nội dung 1
        driver.find_element(By.XPATH, var.caidatchung_kiemduyetnoidung_xoa1).click()
        try:
            check_caidatchung_kiemduyetnoidung1 = driver.find_element(By.XPATH,var.check_caidatchung_kiemduyetnoidung1).is_displayed()
            logging.info("Trang - Cài đặt - Chung")
            logging.info("check font-end: Xoá kiểm duyệt nội dung 1")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Trang - Cài đặt - Chung")
            logging.info("check font-end: Xoá kiểm duyệt nội dung 1")
            logging.info("True")

        driver.implicitly_wait(15)
        driver.find_element(By.XPATH, var.caidatchung_kiemduyetnoidung).send_keys(data['trang']['caidatchung_kiemduyetnoidung1'])
        driver.find_element(By.XPATH, var.them).click()
        driver.find_element(By.XPATH, var.luuthaydoi).click()
        time.sleep(3)

        #Gỡ trang
        driver.find_element(By.XPATH, var.caidatchung_gotrang).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.huy).click()
        time.sleep(2)


    def caidat_thongbao(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        # Trang - Cài đặt - Thông báo
        driver.find_element(By.XPATH, var.caidat_thongbao).click()

        # Cài đặt thông báo toàn cầu
        time.sleep(1)
        driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbaotoancau).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbaotoancau).click()
        check_caidat_thongbao_chophepthongbaotoancau = driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbaotoancau).is_enabled()
        logging.info("Trang - Cài đặt - Thông báo")
        logging.info("Check font-end: Cho phép thông báo toàn cầu - có bật không?")
        logging.info(check_caidat_thongbao_chophepthongbaotoancau)

        #Tương tác với trang
        #Cho phép thông báo
        time.sleep(1)
        driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbao).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbao).click()
        check_caidat_thongbao_chophepthongbao = driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbao).is_enabled()
        logging.info("Trang - Cài đặt - Thông báo")
        logging.info("Check font-end: Cho phép thông báo - có bật không?")
        logging.info(check_caidat_thongbao_chophepthongbao)

        #Cho phép thông báo qua email
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbaoquaemail).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbaoquaemail).click()
        time.sleep(0.5)
        check_caidat_thongbao_chophepthongbaoquamail = driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbaoquaemail).is_enabled()
        logging.info("Trang - Cài đặt - Thông báo")
        logging.info("Check font-end: Cho phép thông báo qua email - có bật không?")
        logging.info(check_caidat_thongbao_chophepthongbaoquamail)

        #Cho phép thông báo qua sms
        driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbaoquasms).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbaoquasms).click()
        time.sleep(0.5)
        check_caidat_thongbao_chophepthongbaoquasms = driver.find_element(By.XPATH, var.caidat_thongbao_chophepthongbaoquasms).is_enabled()
        logging.info("Trang - Cài đặt - Thông báo")
        logging.info("Check font-end: Cho phép thông báo qua sms - có bật không?")
        logging.info(check_caidat_thongbao_chophepthongbaoquasms)

        #Thông báo khi người check in page
        driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhinguoicheckinpage).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhinguoicheckinpage).click()
        time.sleep(0.5)
        check_caidat_thongbao_thongbaokhiconguoicheckinpage = driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhinguoicheckinpage).is_enabled()
        logging.info("Trang - Cài đặt - Thông báo")
        logging.info("Check font-end: Thông báo khi người check in page - có bật không?")
        logging.info(check_caidat_thongbao_thongbaokhiconguoicheckinpage)

        #Thông báo khi được nhắc tới
        driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhiduocnhactoi).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhiduocnhactoi).click()
        time.sleep(0.5)
        check_caidat_thongbao_thongbaokhiduocnhactoi = driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhiduocnhactoi).is_enabled()
        logging.info("Trang - Cài đặt - Thông báo")
        logging.info("Check font-end: Thông báo khi được nhắc tới - có bật không?")
        logging.info(check_caidat_thongbao_thongbaokhiduocnhactoi)

        #Thông báo khi có review mới
        driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhicoreviewmoi).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhicoreviewmoi).click()
        time.sleep(0.5)
        check_caidat_thongbao_thongbaokhicoreviewmoi = driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhicoreviewmoi).is_enabled()
        logging.info("Trang - Cài đặt - Thông báo")
        logging.info("Check font-end: Thông báo khi có review mới - có bật không?")
        logging.info(check_caidat_thongbao_thongbaokhicoreviewmoi)

        #Thông báo khi có comment
        driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhicocomment).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhicocomment).click()
        time.sleep(0.5)
        check_caidat_thongbao_thongbaokhicocomment = driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhicocomment).is_enabled()
        logging.info("Trang - Cài đặt - Thông báo")
        logging.info("Check font-end: Thông báo khi có comment - có bật không?")
        logging.info(check_caidat_thongbao_thongbaokhicocomment)

        #Thông báo khi like mới
        driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhicolikemoi).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhicolikemoi).click()
        time.sleep(0.5)
        check_caidat_thongbao_thongbaokhicolikemoi = driver.find_element(By.XPATH, var.caidat_thongbao_thongbaokhicolikemoi).is_enabled()
        logging.info("Trang - Cài đặt - Thông báo")
        logging.info("Check font-end: Thông báo khi like mới - có bật không?")
        logging.info(check_caidat_thongbao_thongbaokhicolikemoi)
        time.sleep(2)


    def caidat_vaitrotrentrang(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        # Trang - Cài đặt - Thông báo
        driver.find_element(By.XPATH, var.caidat_vaitrotrentrang).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.caidat_vaitrotrentrang_timbanbeinput).send_keys("hue nguyen")
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.huenguyen)))
        element.click()

        driver.find_element(By.XPATH, var.caidat_vaitrotrentrang_chonquyen).click()
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.nguoikiemduyet)))
        element.click()
        driver.find_element(By.XPATH, var.them).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(2)

        #hue nguyen chấp nhận lời mời làm qtv
        login.login4(self, "nguyenhue608196@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangchu_iconthongbao).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu_iconthongbao_thongbaodautien).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.chapnhan).click()
        time.sleep(3.5)
        check_trang_loimoilamqtv = driver.find_element(By.XPATH, var.check_trang_loimoilamqtv).text
        logging.info(check_trang_loimoilamqtv)
        logging.info("Trang - Cài đặt - Vai trò trên trang - Mời bạn bè - hue nguyen")
        logging.info("check font-end: hue nguyen - xem thống báo - Trần Quang Trường đã mời bạn làm Người kiểm duyệt một trang")
        logging.info(check_trang_loimoilamqtv == "  Đã chấp nhận")
        time.sleep(1)

        #Huy thêm qtv
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(2)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangcuaban).click()
        driver.find_element(By.XPATH, var.trang_trangcuaban_truongtest).click()
        time.sleep(2)
        # Cài đặt
        driver.find_element(By.XPATH, var.trang_caidat).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.caidat_vaitrotrentrang).click()
        time.sleep(1)
        #Chỉnh sửa vai trò
        ten_qtv = driver.find_element(By.XPATH, var.ten_qtv).text
        if ten_qtv == "hue nguyen":
            #Đổi quyền thành quản trị viên
            driver.find_element(By.XPATH, var.caidat_vaitrotrentrang_chinhsuavaitro).click()
            driver.find_element(By.XPATH, var.caidat_vaitrotrentrang_chinhsuavaitro_chonquyen).click()
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var.quantrivien1)))
            element.click()
            driver.find_element(By.XPATH, var.luu).click()
            # driver.find_element(By.XPATH, var.huy).click()
            time.sleep(2)
            #Đổi quyền thành người kiểm duyệt
            driver.find_element(By.XPATH, var.caidat_vaitrotrentrang_chinhsuavaitro).click()
            driver.find_element(By.XPATH, var.caidat_vaitrotrentrang_chinhsuavaitro_chonquyen).click()
            element = wait.until(EC.element_to_be_clickable((By.XPATH, var.nguoikiemduyet1)))
            element.click()
            driver.find_element(By.XPATH, var.luu).click()
            time.sleep(2)

            # Gỡ
            driver.find_element(By.XPATH, var.caidat_vaitrotrentrang_go).click()
            driver.find_element(By.XPATH, var.go).click()
            time.sleep(3)
        else:
            logging.info("Trang - Cài đặt - Vai trò trên trang - Mời bạn bè - hue nguyen")
            logging.info("check font-end: hue nguyen chấp chận - Có hiển thị hue nguyen làm người kiểm duyệt không")
            logging.info("False")

        # Huỷ lời mời
        driver.find_element(By.XPATH, var.caidat_vaitrotrentrang_timbanbeinput).send_keys("hue nguyen")
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.huenguyen)))
        element.click()

        driver.find_element(By.XPATH, var.caidat_vaitrotrentrang_chonquyen).click()
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.nguoikiemduyet)))
        element.click()
        driver.find_element(By.XPATH, var.them).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.huyloimoi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(3)


    def caidat_chutaikhoantrang(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        # Trang - Cài đặt - Chủ tai khoản trang
        driver.find_element(By.XPATH, var.caidat_chutaikhoantrang).click()
        time.sleep(1)
        # #Tài khoản chưa xác minh danh tính        #Chưa add thêm người chưa xác minh danh tính
        # driver.find_element(By.XPATH, var.chutaikhoantrang_capnhatchutaikhoan).click()
        # wait = WebDriverWait(driver, 10)
        # element = wait.until(EC.element_to_be_clickable((By.XPATH, var.chonchutaikhoan_ngocmai)))
        # element.click()
        # driver.find_element(By.XPATH, var.capnhat).click()
        #
        # check_trang_chutaikhoantrang_sai = driver.find_element(By.XPATH, var.check_trang_chutaikhoantrang_sai).text
        # logging.info(check_trang_chutaikhoantrang_sai)
        # logging.info("Trang - Cài đặt - Chủ tài khoản trang")
        # logging.info("check font-end: Tài khoản chưa xác minh - Tài khoản chưa xác minh danh tính")
        # logging.info(check_trang_chutaikhoantrang_sai == "Tài khoản chưa xác minh danh tính")
        # time.sleep(1)

        #Tài khoản đã xác minh danh tính
        driver.find_element(By.XPATH, var.chutaikhoantrang_capnhatchutaikhoan).click()
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.chonchutaikhoan_tranquangtruong)))
        element.click()
        driver.find_element(By.XPATH, var.capnhat).click()

        check_trang_chutaikhoantrang_dung = driver.find_element(By.XPATH, var.check_trang_chutaikhoantrang_dung).text
        logging.info(check_trang_chutaikhoantrang_dung)
        logging.info("Trang - Cài đặt - Chủ tài khoản trang")
        logging.info("check font-end: Tài khoản đã xác minh - Cập nhật chủ tài khoản thành công")
        logging.info(check_trang_chutaikhoantrang_dung == "Cập nhật chủ tài khoản thành công")
        time.sleep(1)

        check_trang_chutaikhoantrang = driver.find_element(By.XPATH, var.check_trang_chutaikhoantrang).text
        logging.info(check_trang_chutaikhoantrang)
        logging.info("Trang - Cài đặt - Chủ tài khoản trang")
        logging.info("check font-end: Chủ tài khoản của Trang - Trần Quang Trường")
        logging.info(check_trang_chutaikhoantrang == "Trần Quang Trường")
        time.sleep(2)


    def caidat_nguoivatrangkhac(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        # Trang - Cài đặt - Người và trang khác
        driver.find_element(By.XPATH, var.caidat_nguoivatrangkhac).click()
        #Những người thích trang này
        # driver.find_element(By.XPATH, var.nguoivatrangkhac_chonloai).click()
        # driver.find_element(By.XPATH, var.nguoivatrangkhac_chonloai_nguoithichtrang).click()
        time.sleep(1)
        check_trang_nguoivatrangkhac_nguoithichtrang = driver.find_element(By.XPATH, var.check_trang_nguoivatrangkhac_nguoithichtrang).text
        logging.info(check_trang_nguoivatrangkhac_nguoithichtrang)
        logging.info("Trang - Cài đặt - Người và trang khác")
        logging.info("check font-end: Những người thích Trang này - Có hiển thị tên người thích trang không?")
        logging.info(check_trang_nguoivatrangkhac_nguoithichtrang != None)
        time.sleep(1)

        # Thêm người vào danh sách cấm trên trang
        driver.find_element(By.XPATH, var.nguoivatrangkhac_timkiem).send_keys("Vương Lâm")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nguoivatrangkhac_chonnguoi1).click()
        driver.find_element(By.XPATH, var.nguoivatrangkhac_iconcaidat).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.camtrentrang).click()

        check_trang_nguoivatrangkhac_camtrentrang = driver.find_element(By.XPATH,var.check_trang_nguoivatrangkhac_camtrentrang).text
        logging.info(check_trang_nguoivatrangkhac_camtrentrang)
        logging.info("Trang - Cài đặt - Người và trang khác")
        logging.info("check font-end: Cấm trên trang Message - Cấm thành công.")
        logging.info(check_trang_nguoivatrangkhac_camtrentrang == "Cấm thành công.")

        time.sleep(1)
        driver.refresh()
        time.sleep(2)

        # Những người theo dõi Trang này
        driver.find_element(By.XPATH, var.nguoivatrangkhac_chonloai).click()
        driver.find_element(By.XPATH, var.nguoivatrangkhac_chonloai_nguoitheodoitrang).click()
        time.sleep(3)
        check_trang_nguoivatrangkhac_nguoitheodoitrang = driver.find_element(By.XPATH,var.check_trang_nguoivatrangkhac_nguoithichtrang).text
        logging.info(check_trang_nguoivatrangkhac_nguoitheodoitrang)
        logging.info("Trang - Cài đặt - Người và trang khác")
        logging.info("check font-end: Những người theo dõi Trang này - Có hiển thị tên người theo dõi trang không?")
        logging.info(check_trang_nguoivatrangkhac_nguoithichtrang != None)

        # Những người bị cấm trên Trang này
        driver.find_element(By.XPATH, var.nguoivatrangkhac_chonloai).click()
        driver.find_element(By.XPATH, var.nguoivatrangkhac_chonloai_nguoibicamtrentrang).click()
        time.sleep(3)
        check_trang_nguoivatrangkhac_nguoibicamtrentrang = driver.find_element(By.XPATH,var.check_trang_nguoivatrangkhac_nguoithichtrang).text
        logging.info(check_trang_nguoivatrangkhac_nguoibicamtrentrang)
        logging.info("Trang - Cài đặt - Người và trang khác")
        logging.info("check font-end: Những người bị cấm trên Trang này - Có hiển thị tên người bị cấm trên trang không?")
        logging.info(check_trang_nguoivatrangkhac_nguoibicamtrentrang == "Vương Lâm")
        time.sleep(1)

        # Bỏ cấm khỏi trang
        driver.find_element(By.XPATH, var.nguoivatrangkhac_chonnguoi1).click()
        driver.find_element(By.XPATH, var.nguoivatrangkhac_iconcaidat).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.bocamtrentrang).click()
        time.sleep(1)

        check_trang_nguoivatrangkhac_bocamtrentrang = driver.find_element(By.XPATH,var.check_trang_nguoivatrangkhac_bocamtrentrang).text
        logging.info(check_trang_nguoivatrangkhac_bocamtrentrang)
        logging.info("Trang - Cài đặt - Người và trang khác")
        logging.info("check font-end: Bỏ cấm trên trang Message - Bỏ cấm thành công")
        logging.info(check_trang_nguoivatrangkhac_bocamtrentrang == "Bỏ cấm thành công")
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            driver.find_element(By.XPATH, var.vuonglam).is_displayed()
            logging.info("Trang - Cài đặt - Người và trang khác")
            logging.info("Bỏ cấm khỏi trang - Vương Lâm")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Trang - Cài đặt - Người và trang khác")
            logging.info("Bỏ cấm khỏi trang - Vương Lâm")
            logging.info("True")
        time.sleep(2)


    def caidat_nhatkyhoatdong(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        # Trang - Cài đặt - Nhật ký hoạt động
        driver.find_element(By.XPATH, var.caidat_nhatkyhoatdong).click()
        #Tất cả hành động - Mọi người - Mới nhất
        driver.find_element(By.XPATH, var.nhatkyhoatdong_loai).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_loai_tatcahanhdong).click()

        driver.find_element(By.XPATH, var.nhatkyhoatdong_nguoi).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_nguoi_moinguoi).click()

        driver.find_element(By.XPATH, var.nhatkyhoatdong_thoigian).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_thoigian_moinhat).click()
        time.sleep(2)

        check_trang_nhatkyhoatdong_thongbao1 = driver.find_element(By.XPATH,var.check_trang_nhatkyhoatdong_thongbao1).text
        logging.info("Trang - Cài đặt - Nhật ký hoạt động")
        logging.info("check font-end: Tất cả hành động - Mọi người - Mới nhất , có thông báo hay không")
        logging.info(check_trang_nhatkyhoatdong_thongbao1)
        logging.info(check_trang_nhatkyhoatdong_thongbao1 != None)

        #Thông tin trên trang - Ngọc Mai - Mới nhất
        driver.find_element(By.XPATH, var.nhatkyhoatdong_loai).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_loai_thongtintrentrang).click()

        driver.find_element(By.XPATH, var.nhatkyhoatdong_nguoi).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_nguoi_ngocmai).click()

        driver.find_element(By.XPATH, var.nhatkyhoatdong_thoigian).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_thoigian_moinhat).click()
        time.sleep(2)

        check_trang_nhatkyhoatdong_thongbao1 = driver.find_element(By.XPATH,var.check_trang_nhatkyhoatdong_thongbao1).text
        logging.info("Trang - Cài đặt - Nhật ký hoạt động")
        logging.info("check font-end: Thông tin trên trang - Ngọc Mai - Mới nhất , có thông báo hay không")
        logging.info(check_trang_nhatkyhoatdong_thongbao1)
        logging.info(check_trang_nhatkyhoatdong_thongbao1 != None)

        # Vai trò trên Trang - Trần Quang Trường - Cũ nhất
        driver.find_element(By.XPATH, var.nhatkyhoatdong_loai).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_loai_vaitrotrentrang).click()

        driver.find_element(By.XPATH, var.nhatkyhoatdong_nguoi).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_nguoi_tranquangtruong).click()

        driver.find_element(By.XPATH, var.nhatkyhoatdong_thoigian).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_thoigian_cunhat).click()
        time.sleep(2)

        check_trang_nhatkyhoatdong_thongbao1 = driver.find_element(By.XPATH,var.check_trang_nhatkyhoatdong_thongbao1).text
        logging.info("Trang - Cài đặt - Nhật ký hoạt động")
        logging.info("check font-end: Vai trò trên Trang - Trần Quang Trường - Cũ nhất , có thông báo hay không")
        logging.info(check_trang_nhatkyhoatdong_thongbao1)
        logging.info(check_trang_nhatkyhoatdong_thongbao1 != None)

        # Nhóm - Mọi người - Mới nhất
        driver.find_element(By.XPATH, var.nhatkyhoatdong_loai).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_loai_nhom).click()

        driver.find_element(By.XPATH, var.nhatkyhoatdong_nguoi).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_nguoi_moinguoi).click()

        driver.find_element(By.XPATH, var.nhatkyhoatdong_thoigian).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_thoigian_moinhat).click()
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            check_trang_nhatkyhoatdong_thongbao1 = driver.find_element(By.XPATH,var.check_trang_nhatkyhoatdong_thongbao1).is_displayed()
            logging.info("Trang - Cài đặt - Nhật ký hoạt động")
            logging.info("check font-end: Nhóm - Mọi người - Mới nhất , có thông báo hay không")
            logging.info(check_trang_nhatkyhoatdong_thongbao1)
        except NoSuchElementException:
            logging.info("Trang - Cài đặt - Nhật ký hoạt động")
            logging.info("check font-end: Nhóm - Mọi người - Mới nhất , có thông báo hay không")
            logging.info("False")


    def caidat_caidathopthu(self):
        driver.implicitly_wait(15)
        time.sleep(1)
        # Trang - Cài đặt - Cài đặt hộp thư
        driver.find_element(By.XPATH, var.caidat_caidathopthu).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.caidat_caidathopthu_datlai).click()
        time.sleep(1.5)

        driver.implicitly_wait(2)
        try:
            check_trang_caidathopthu_datlai = driver.find_element(By.XPATH,var.check_trang_caidathopthu_datlai).is_displayed()
            logging.info("Trang - Cài đặt - Cài đặt hộp thư")
            logging.info("check font-end: Đặt lại - Có xoá được các câu hỏi trước đó hay không")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Trang - Cài đặt - Cài đặt hộp thư")
            logging.info("check font-end: Đặt lại - Có xoá được các câu hỏi trước đó hay không")
            logging.info("True")
        driver.implicitly_wait(15)

        #Thêm câu hỏi 1 để sửa xoá
        driver.find_element(By.XPATH, var.caicathopthu_themluachon).click()
        driver.find_element(By.XPATH, var.caicathopthu_cauhoi1input).send_keys(data['trang']['caidathopthu_cauhoi1'])
        driver.find_element(By.XPATH, var.caicathopthu_cautraloi1input).send_keys(data['trang']['caidathopthu_cautraloi1'])
        driver.find_element(By.XPATH, var.caicathopthu_luuthaydoi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.caicathopthu_iconthugon).click()

        check_trang_caidathopthu_luuthaydoi = driver.find_element(By.XPATH,var.check_trang_caidathopthu_luuthaydoi).text
        logging.info("Trang - Cài đặt - Cài đặt hộp thư")
        logging.info("check font-end: Message Lưu thay đổi - Lưu thành công.")
        logging.info(check_trang_caidathopthu_luuthaydoi)
        logging.info(check_trang_caidathopthu_luuthaydoi == "Lưu thành công.")

        #Chỉnh sửa câu hỏi
        driver.find_element(By.XPATH, var.caicathopthu_iconchinhsua).click()
        xoa = driver.find_element(By.XPATH, var.caicathopthu_cauhoi1input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.caicathopthu_cauhoi1input).send_keys(data['trang']['caidathopthu_cauhoi1a'])
        driver.find_element(By.XPATH, var.caicathopthu_luuthaydoi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.caicathopthu_iconthugon).click()
        check_trang_caidathopthu_chinhsuacauhoi = driver.find_element(By.XPATH,var.check_trang_caidathopthu_chinhsuacauhoi).text
        logging.info("Trang - Cài đặt - Cài đặt hộp thư")
        logging.info("check font-end: Chỉnh sửa câu hỏi - Câu hỏi 1:Tôi muốn mua hàng.")
        logging.info(check_trang_caidathopthu_chinhsuacauhoi)
        logging.info(check_trang_caidathopthu_chinhsuacauhoi == "Câu hỏi 1:\nTôi muốn mua hàng.")

        #Xoa câu hỏi
        driver.find_element(By.XPATH, var.caicathopthu_iconchinhsua).click()
        driver.find_element(By.XPATH, var.caicathopthu_icon_x).click()

        driver.implicitly_wait(2)
        try:
            check_trang_caidathopthu_xoacauhoi = driver.find_element(By.XPATH,var.check_trang_caidathopthu_datlai).is_displayed()
            logging.info("Trang - Cài đặt - Cài đặt hộp thư")
            logging.info("check font-end: Xoá câu hỏi - Có xoá được câu hỏi vừa chọn không")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Trang - Cài đặt - Cài đặt hộp thư")
            logging.info("check font-end: Xoá câu hỏi - Có xoá được câu hỏi vừa chọn không")
            logging.info("True")
        driver.implicitly_wait(15)

        # Thêm câu hỏi 1
        driver.find_element(By.XPATH, var.caicathopthu_themluachon).click()
        driver.find_element(By.XPATH, var.caicathopthu_cauhoi1input).send_keys(data['trang']['caidathopthu_cauhoi1a'])
        driver.find_element(By.XPATH, var.caicathopthu_cautraloi1input).send_keys(data['trang']['caidathopthu_cautraloi1'])
        driver.find_element(By.XPATH, var.caicathopthu_luuthaydoi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.caicathopthu_iconthugon).click()

        #Thêm câu hỏi 2
        driver.find_element(By.XPATH, var.caicathopthu_themluachon).click()
        driver.find_element(By.XPATH, var.caicathopthu_cauhoi2input).send_keys(data['trang']['caidathopthu_cauhoi2'])
        driver.find_element(By.XPATH, var.caicathopthu_cautraloi2input).send_keys(data['trang']['caidathopthu_cautraloi2'])
        driver.find_element(By.XPATH, var.caicathopthu_luuthaydoi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.caicathopthu_iconthugon).click()
        time.sleep(2)

        #Check câu hỏi tự động với user
        login.login4(self, "nguyenhue608196@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_trang)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.trang_trangdathich).click()
        time.sleep(2)

        # Check câu hỏi tự động bên ngoài trang
        driver.find_element(By.XPATH, var.trang_trangdathich_guitinnhan).click()
        driver.implicitly_wait(2)
        time.sleep(2)
        try:
            check_trang_trangdathich_cauhoitudong = driver.find_element(By.XPATH,var.check_trang_trangdathich_cauhoitudong).is_displayed()
            logging.info("Trang - Cài đặt - Cài đặt hộp thư - Thêm câu hỏi tự động")
            logging.info("Login vào user khác - Trang -Trang đã thích - Gửi tin nhắn - Có câu hỏi tự động không")
            logging.info(check_trang_trangdathich_cauhoitudong)
        except NoSuchElementException:
            logging.info("Trang - Cài đặt - Cài đặt hộp thư - Thêm câu hỏi tự động")
            logging.info("Login vào user khác - Trang -Trang đã thích - Gửi tin nhắn - Có câu hỏi tự động không")
            logging.info("False")
        driver.implicitly_wait(15)

        #Check câu hỏi tự động bên trong trang
        driver.find_element(By.XPATH, var.truongtetbantin).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.guitinnhan).click()
        time.sleep(2)
        check_trang_trangdathich_cauhoitudong = driver.find_element(By.XPATH,var.check_trang_trangdathich_cauhoitudong).text
        logging.info("Trang - Cài đặt - Cài đặt hộp thư - Thêm câu hỏi tự động")
        logging.info("Login vào user khác - Trang - Trường test bn tin - Có câu hỏi tự động không")
        logging.info(check_trang_trangdathich_cauhoitudong)
        logging.info(check_trang_trangdathich_cauhoitudong == data['trang']['caidathopthu_cauhoi1a'])
        time.sleep(1)

        #Check câu trả lời của của câu hỏi tự động
        driver.find_element(By.XPATH, var.trang_hopthoai_clicktoimuonmuahang).click()
        time.sleep(1)
        check_trang_trangdathich_cautraloitudong = driver.find_element(By.XPATH,var.check_trang_trangdathich_cautraloitudong).text
        logging.info("Trang - Cài đặt - Cài đặt hộp thư - Thêm câu hỏi tự động")
        logging.info("Login vào user khác - Trang - Trường test bn tin - Có câu trả lời tự động không")
        logging.info(check_trang_trangdathich_cautraloitudong)
        logging.info(check_trang_trangdathich_cautraloitudong == data['trang']['caidathopthu_cautraloi1'])
        time.sleep(1)

        #Xoá hộp thư
        driver.find_element(By.XPATH, var.trang_hopthoai_truongtestbantin).click()
        driver.find_element(By.XPATH, var.trang_hopthoai_truongtestbantin_xoacuoctrochuyen).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(2)



class nhom():
    def timkiem(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Tìm kiếm nhóm - tất cả
        driver.find_element(By.XPATH, var.nhom_timkiem).send_keys(data['nhom']['timkiem'])
        driver.find_element(By.XPATH, var.nhom_timkiem).send_keys(Keys.ENTER)
        check_nhom_timkiem = driver.find_element(By.XPATH,var.check_nhom_timkiem).text
        print(check_nhom_timkiem)
        logging.info("Nhóm - Tìm kiếm")
        logging.info("check font-end: Tìm kiếm - "+ data['nhom']['timkiem'])
        logging.info(check_nhom_timkiem == data['nhom']['timkiem'])
        time.sleep(1)

        #Bộ lọc - nhóm
        driver.find_element(By.XPATH, var.nhom_timkiem_iconlocnhom).click()
        time.sleep(1)
        #Quyền công khai - bật
        driver.find_element(By.XPATH, var.iconlocnhom_nhomcongkhaibat).click()
        time.sleep(2)
        r = 0
        chedonhom = driver.find_elements(By.XPATH, var.nhom_timkiem_chedonhom)
        for chedo in chedonhom:
            r = int(r)
            r += 1
            r = str(r)
            tenchedo = chedo.text
            logging.info("Nhóm - Tìm kiếm")
            logging.info("check font-end: Quyền công khai - Nhóm số " + r)
            logging.info(tenchedo =="Nhóm Công khai")
        driver.find_element(By.XPATH, var.iconlocnhom_nhomcongkhaibat).click()
        time.sleep(2)

        # Nhóm của tôi - bật
        driver.find_element(By.XPATH, var.iconlocnhom_nhomcuatoibat).click()
        check_nhom_timkiem_nhomcuatoi = driver.find_element(By.XPATH,var.check_nhom_timkiem_nhomcuatoi).text
        print(check_nhom_timkiem_nhomcuatoi)
        logging.info("Nhóm - Tìm kiếm")
        logging.info("check font-end: Nhóm của tôi - nam test")
        logging.info(check_nhom_timkiem_nhomcuatoi == data['nhom']['timkiem'])
        driver.find_element(By.XPATH, var.iconlocnhom_nhomcuatoibat).click()
        time.sleep(2)

        #Bộ lọc - bài viết nhóm
        driver.find_element(By.XPATH, var.nhom_timkiem_iconlocbaivietnhom).click()
        time.sleep(2)
        check_nhom_timkiem_baivietnhom = driver.find_element(By.XPATH, var.check_nhom_timkiem_baivietnhom).text
        logging.info(check_nhom_timkiem_baivietnhom)
        logging.info("Nhóm - Tìm kiếm")
        logging.info("check font-end: Bài viết nhóm, Có hiển thị bài viết không")
        logging.info(check_nhom_timkiem_baivietnhom != None)

        #Bài viết gần đây - bật
        driver.find_element(By.XPATH, var.iconlocnhom_baivietgandaybat).click()
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            check_nhom_timkiem_baivietganday = driver.find_element(By.XPATH,var.check_nhom_timkiem_baivietnhom).is_displayed()
            logging.info(check_nhom_timkiem_baivietganday)
            logging.info("Nhóm - Tìm kiếm")
            logging.info("check font-end: Bài viết gần đây, có hiển thị hay không")
            logging.info(check_nhom_timkiem_baivietganday)
        except NoSuchElementException:
            logging.info("Nhóm - Tìm kiếm")
            logging.info("check font-end: Bài viết gần đây, có hiển thị hay không")
            logging.info("False")
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconlocnhom_baivietgandaybat).click()

        #Bài viết nhóm - chọn ngày
        driver.find_element(By.XPATH, var.iconlocnhom_iconngaydang).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.iconlocnhom_chonngay_2023).click()
        time.sleep(2)

        try:
            check_nhom_timkiem_ngaydang = driver.find_element(By.XPATH,var.check_nhom_timkiem_baivietnhom).text     #Không hiện thị bài viết 6525
            logging.info(check_nhom_timkiem_ngaydang)
            logging.info("Nhóm - Tìm kiếm")
            logging.info("check font-end: Ngày đăng 2023, tiêu đề bài viết số 1 - test")
            logging.info(check_nhom_timkiem_ngaydang == "test")
        except NoSuchElementException:
            logging.info("Nhóm - Tìm kiếm")
            logging.info("check font-end: Ngày đăng 2023, tiêu đề bài viết số 1 - test")
            logging.info("False")
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconlocnhom_chonngay_datlai).click()


    def bangtincuaban(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Nhóm - Bảng tin của bạn
        driver.find_element(By.XPATH, var.nhom_bangtincuaban).click()
        #Bài viết của nhóm
        driver.implicitly_wait(2)
        try:
            check_nhom_bangtin_baivietcuanhom = driver.find_element(By.XPATH,var.check_nhom_bangtin_baivietcuanhom1).is_displayed()
            logging.info(check_nhom_bangtin_baivietcuanhom)
            logging.info("Nhóm - Bảng tin của bạn")
            logging.info("check font-end: Có hiển thị bài viết nhóm không?")
            logging.info(check_nhom_bangtin_baivietcuanhom)
        except NoSuchElementException:
            logging.info("Nhóm - Bảng tin của bạn")
            logging.info("check font-end: Có hiển thị bài viết nhóm không?")
            logging.info("False")

        # Bài viết của nhóm - bật thông báo
        driver.find_element(By.XPATH, var.baivietcuanhom_dau3cham).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baivietcuanhom_dau3cham_batthongbao).click()
        check_nhom_baivietcuanhom_batthongbao = driver.find_element(By.XPATH, var.check_nhom_baivietcuanhom_batthongbao).text
        logging.info(check_nhom_baivietcuanhom_batthongbao)
        logging.info("Nhóm - Bảng tin của bạn - Bài viết của nhóm")
        logging.info("check font-end: Message bật thông báo - Đã bật thông báo")
        logging.info(check_nhom_baivietcuanhom_batthongbao == "Đã bật thông báo")
        driver.find_element(By.XPATH, var.icon_x).click()

        # Bài viết của nhóm - tắt thông báo
        driver.find_element(By.XPATH, var.baivietcuanhom_dau3cham).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baivietcuanhom_dau3cham_tatthongbao).click()
        check_nhom_baivietcuanhom_tatthongbao = driver.find_element(By.XPATH, var.check_nhom_baivietcuanhom_tatthongbao).text
        logging.info(check_nhom_baivietcuanhom_batthongbao)
        logging.info("Nhóm - Bảng tin của bạn - Bài viết của nhóm")
        logging.info("check font-end: Message tắt thông báo - Đã tắt thông báo")
        logging.info(check_nhom_baivietcuanhom_tatthongbao == "Đã tắt thông báo")
        driver.find_element(By.XPATH, var.icon_x).click()

        #Cộng đồng của bạn bè
        try:
            check_nhom_bangtin_congdongcuabanbe_tennhom1 = driver.find_element(By.XPATH,var.check_nhom_bangtin_congdongcuabanbe_tennhom1).is_displayed()
            logging.info(check_nhom_bangtin_congdongcuabanbe_tennhom1)
            logging.info("Nhóm - Bảng tin của bạn")
            logging.info("check font-end: Cộng đồng của bạn bè - Có hiển thị nhóm không?")
            logging.info(check_nhom_bangtin_congdongcuabanbe_tennhom1)
        except NoSuchElementException:
            logging.info("Nhóm - Bảng tin của bạn")
            logging.info("check font-end: Cộng đồng của bạn bè - Có hiển thị nhóm không?")
            logging.info("False")
        congdongcuabanbe_tennhom1 = driver.find_element(By.XPATH,var.check_nhom_bangtin_congdongcuabanbe_tennhom1).text

        check_nhom_congdongcuabanbe_thamgia = driver.find_element(By.XPATH, var.check_nhom_congdongcuabanbe_thamgia).text
        logging.info(check_nhom_congdongcuabanbe_thamgia)
        logging.info("Nhóm - Bảng tin của bạn")
        logging.info("check font-end: Cộng đồng của bạn bè - Tham gia nhóm")
        logging.info(check_nhom_congdongcuabanbe_thamgia == "Tham gia nhóm")

        #Cộng đồng của bạn bè - Xem thêm
        driver.find_element(By.XPATH, var.xemthem).click()
        time.sleep(1)
        congdongcuabanbe_xemthem_tennhom2 = driver.find_element(By.XPATH, var.congdongcuabanbe_xemthem_tennhom2).text
        logging.info(congdongcuabanbe_xemthem_tennhom2)
        logging.info("Nhóm - Bảng tin của bạn - Cộng đồng của bạn bè")
        logging.info("check font-end: Xem thêm - Tên nhóm 2 có hiển thị hay không?")
        logging.info("Tên nhóm 1: " + congdongcuabanbe_tennhom1)
        logging.info("Tên nhóm 2: " + congdongcuabanbe_xemthem_tennhom2)
        logging.info(congdongcuabanbe_xemthem_tennhom2 != congdongcuabanbe_tennhom1)

        #Cộng đồng của bạn bè - Gợi ý
        congdongcuabanbe_xemthem_goiy = driver.find_element(By.XPATH, var.congdongcuabanbe_xemthem_goiy).text
        logging.info(congdongcuabanbe_xemthem_goiy)
        logging.info("Nhóm - Bảng tin của bạn - Cộng đồng của bạn bè")
        logging.info("check font-end: Xem thêm - Gợi ý - Có gợi ý nhóm hay không?")
        logging.info(congdongcuabanbe_xemthem_goiy != "Không có dữ liệu để hiển thị")
        time.sleep(1)
        driver.back()

        #Vào trang Cộng đồng của bạn bè để check
        time.sleep(1)
        driver.find_element(By.XPATH, var.check_nhom_bangtin_congdongcuabanbe_tennhom1).click()
        time.sleep(2)
        congdongcuabanbe_vaotrang_tennhom1 =driver.find_element(By.XPATH, var.congdongcuabanbe_vaotrang_tennhom1).text
        logging.info("Nhóm - Bảng tin của bạn")
        logging.info("check font-end: Cộng đồng của bạn bè - tên bên ngoài và sau khi click vào trang")
        logging.info("Tên nhóm bên ngoài: " + congdongcuabanbe_tennhom1)
        logging.info("Tên nhóm bên trong: " + congdongcuabanbe_vaotrang_tennhom1)
        logging.info(congdongcuabanbe_vaotrang_tennhom1 == congdongcuabanbe_tennhom1)
        time.sleep(2)

    def khampha(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Nhóm của bạn bè
        driver.find_element(By.XPATH, var.nhom_khampha).click()
        time.sleep(2)
        nhom_khampha_tennhom1 = driver.find_element(By.XPATH, var.nhom_khampha_tennhom1).text
        nhom_khampha_tennhom2 = driver.find_element(By.XPATH, var.nhom_khampha_tennhom2).text
        logging.info("Nhóm - Khám phá - Nhóm của bạn bè")
        logging.info("check font-end: Xem thêm - Danh sách các nhóm có hiển thị hay không?")
        logging.info("Tên nhóm 1: " + nhom_khampha_tennhom1)
        logging.info("Tên nhóm 2: " + nhom_khampha_tennhom2)
        logging.info(nhom_khampha_tennhom1 != nhom_khampha_tennhom2)

        #Gợi ý khác
        check_khampha_goiykhac = driver.find_element(By.XPATH, var.congdongcuabanbe_xemthem_goiy).text
        logging.info(check_khampha_goiykhac)
        logging.info("Nhóm - Bảng tin của bạn - Cộng đồng của bạn bè")
        logging.info("check font-end: Xem thêm - Gợi ý - Có gợi ý nhóm hay không?")
        logging.info(check_khampha_goiykhac != "Không có dữ liệu để hiển thị")
        time.sleep(1)

    def moithamgianhom(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Mời tham gia nhóm
        driver.find_element(By.XPATH, var.nhom_moithamgianhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_namtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_moi).click()
        time.sleep(1)

        driver.find_element(By.XPATH, var.manhcuong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.check_trang_moibanbe_tatcabanbe_guiloimoi).click()
        time.sleep(2)
        check_nhom_moibanbe_damoi1 = driver.find_element(By.XPATH,var.check_nhom_moibanbe_damoi1).text
        check_nhom_moibanbe_damoi = driver.find_element(By.XPATH,var.manhcuong).text
        print(check_nhom_moibanbe_damoi)
        logging.info("Nhóm - Nhóm của bạn - Mời bạn bè")
        logging.info("check font-end: Đã mời - Mạnh Cường")
        logging.info(check_nhom_moibanbe_damoi)
        logging.info(check_nhom_moibanbe_damoi1)
        logging.info(check_nhom_moibanbe_damoi == "Mạnh Cường")
        time.sleep(0.5)

        # Gửi lời mời tham gia nhóm - xóa
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.nhom_loimoi_xoa)
        driver.execute_script("arguments[0].click();", button)
        # driver.find_element(By.XPATH, var.nhom_loimoi_xoa).click()
        time.sleep(3)

        # Gửi lời mời tham gia nhóm - Chấp nhận
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Mời tham gia nhóm
        driver.find_element(By.XPATH, var.nhom_namtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_moi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.manhcuong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.check_trang_moibanbe_tatcabanbe_guiloimoi).click()
        time.sleep(2)
        check_nhom_moibanbe_damoi1 = driver.find_element(By.XPATH,var.check_nhom_moibanbe_damoi1).text
        time.sleep(1)

        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_iconthongbao).click()
        time.sleep(1)
        check_nhom_loimoithamgianhom = driver.find_element(By.XPATH,var.trangchu_iconthongbao_thongbaodautien1).text
        logging.info("Nhóm - Mời tham gia nhóm - Thông báo")
        logging.info("check font-end: hue nguyen - xem thống báo - Lời mời thích trang")
        logging.info(check_nhom_loimoithamgianhom)
        logging.info(check_nhom_loimoithamgianhom == "Ngọc Mai đã mời bạn tham gia nhóm nam test")
        driver.find_element(By.XPATH, var.trangchu_iconthongbao_thongbaodautien).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_loimoi_thamgianhom).click()
        time.sleep(1.5)

        # check_nhom_message_thamgianhom = driver.find_element(By.XPATH, var.check_nhom_message_thamgianhom).text     #Chưa có message tham gia nhóm
        # logging.info("Nhóm - Mời tham gia nhóm - Tham gia")
        # logging.info("check font-end: Message tham gia nhóm - ")
        # logging.info(check_nhom_message_thamgianhom)
        # logging.info(check_nhom_message_thamgianhom == "")

        check_nhom_loimoi_chapnhan = driver.find_element(By.XPATH, var.check_nhom_loimoi_chapnhan1).text
        logging.info("Nhóm - Mời tham gia nhóm")
        logging.info("check font-end: user 2 xem thông báo chấp nhận - Đã tham gia")
        logging.info(check_nhom_loimoi_chapnhan)
        logging.info(check_nhom_loimoi_chapnhan == "Đã tham gia")
        time.sleep(0.5)

        #Lời mời làm quản trị viên nhóm
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Lời mời làm quản trị viên nhóm - nhóm - thành viên
        driver.find_element(By.XPATH, var.nhom_namtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_thanhvien).click()
        time.sleep(1)
        #Nhóm -Tìm kiếm - Thanh viên - Mạnh cường
        driver.find_element(By.XPATH, var.nhom_thanhvien_timkiem).send_keys(data['nhom']['timkiem_thanhvien'])
        time.sleep(1)
        driver.execute_script("window.scrollBy(0,300)", "")
        driver.find_element(By.XPATH, var.nhom_thanhvien_timkiem_dau3cham).click()
        driver.find_element(By.XPATH, var.nhom_thanhvien_themlamnguoikiemduyet).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(2)

        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(2)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Mời tham gia nhóm
        driver.find_element(By.XPATH, var.nhom_moithamgianhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chapnhan).click()
        driver.implicitly_wait(3)
        driver.find_element(By.XPATH, var.nhom_moithamgianhom_loimoi1).click()
        time.sleep(2)
        try:
            check_nhom_loimoilamqtv_quanlynhom = driver.find_element(By.XPATH, var.check_nhom_loimoilamqtv_quanlynhom).is_displayed()
            logging.info("Nhóm - Mời làm người kiểm duyệt nhóm")
            logging.info("check font-end: User 2 đăng nhập - Chấp nhận - Có module Quản lý nhóm")
            logging.info(check_nhom_loimoilamqtv_quanlynhom)
        except NoSuchElementException:
            logging.info("Nhóm - Mời làm người kiểm duyệt nhóm")
            logging.info("check font-end: User 2 đăng nhập - Chấp nhận - Có module Quản lý nhóm")
            logging.info("False")

        #Rời khỏi nhóm
        driver.implicitly_wait(15)
        driver.find_element(By.XPATH, var.check_nhom_loimoi_chapnhan1).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.roikhoinhom).click()
        time.sleep(1)

        # check_nhom_message_roikhoinhom = driver.find_element(By.XPATH, var.check_nhom_message_roikhoinhom).text     #Chưa có message rời khỏi nhóm
        # logging.info("Nhóm - Mời tham gia nhóm - Đã tham gia - Rời khỏi nhóm")
        # logging.info("check font-end: Message rời khỏi nhóm - ")
        # logging.info(check_nhom_message_roikhoinhom)
        # logging.info(check_nhom_message_roikhoinhom == "")

        check_popup_roikhoinhom = driver.find_element(By.XPATH, var.check_popup_roikhoinhom).text
        logging.info("Nhóm - Rời khỏi nhóm")
        logging.info("check font-end: popup Rời khỏi nhóm Có hiện thị không?")
        logging.info(check_popup_roikhoinhom)
        logging.info(check_popup_roikhoinhom == "Rời khỏi nhóm ?")
        driver.find_element(By.XPATH, var.roikhoinhom).click()
        time.sleep(1)
        try:
            check_roikhoinhom_daroikhoinhom = driver.find_element(By.XPATH, var.roikhoinhom_daroikhoinhom).is_displayed()
            logging.info("Nhóm - Rời khỏi nhóm")
            logging.info("check font-end: Đã rời nhóm thành công")
            logging.info(check_roikhoinhom_daroikhoinhom)
        except NoSuchElementException:
            logging.info("Nhóm - Rời khỏi nhóm")
            logging.info("check font-end: Đã rời nhóm thành công")
            logging.info("False")
        time.sleep(1)



    def taonhommoi(self, tennhom, mota, chedo):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Tạo nhóm mới
        driver.find_element(By.XPATH, var.nhom_taonhommoi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_taonhommoi_tennhom).send_keys(tennhom)
        driver.find_element(By.XPATH, var.nhom_taonhommoi_mota).send_keys(mota)
        driver.find_element(By.XPATH, var.nhom_taonhommoi_danhmuc).send_keys(data['nhom']['taonhom_danhmucnhom'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.mucdichxahoi)))
        element.click()
        driver.find_element(By.XPATH, var.nhom_taonhommoi_tag).send_keys(data['nhom']['taonhom_tag'])
        driver.find_element(By.XPATH, var.nhom_taonhommoi_chedocaidatnhom).click()
        element = wait.until(EC.element_to_be_clickable((By.XPATH, chedo)))
        element.click()

        cuon = driver.find_element(By.XPATH, var.nhom_taonhommoi_tailenanhbia)
        driver.execute_script("arguments[0].scrollIntoView();", cuon)
        button = driver.find_element(By.XPATH, var.nhom_taonhommoi_tailenanhbia)
        driver.execute_script("arguments[0].click();", button)
        # driver.find_element(By.XPATH, var.nhom_taonhommoi_tailenanhbia).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/tft.exe")
        time.sleep(6)
        driver.find_element(By.XPATH, var.nhom_taonhommoi_taonhom).click()
        time.sleep(3)

        # check_nhom_taonhommoi_message = driver.find_element(By.XPATH, var.check_nhom_taonhommoi_message).text
        # logging.info("Nhóm - Tạo nhóm mới")
        # logging.info("check font-end: Chế độ cài đặt nhóm: "+ tennhom)      #Tạo nhóm không có message
        # logging.info("check font-end: Message -  "+ check_nhom_taonhommoi_message)
        # logging.info(check_nhom_taonhommoi_message == "")

        check_nhom_taonhommoi_tennhom = driver.find_element(By.XPATH, var.check_nhom_taonhommoi_tennhom).text
        logging.info("Nhóm - Tạo nhóm mới")
        logging.info("check font-end: Tên nhóm -  "+ tennhom)
        logging.info(check_nhom_taonhommoi_tennhom == tennhom)
        time.sleep(1)

        driver.find_element(By.XPATH, var.check_nhom_loimoi_chapnhan1).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.roikhoinhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xoanhom).click()
        time.sleep(2)

    def nhombanquanly(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Nhóm bạn quản lý
        check_nhom_nhombanquanly_namtest = driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).text
        logging.info("Nhóm - Nhóm bạn quản lý")
        logging.info("check font-end: Có hiển thị tên nhóm quản lý hay không")
        logging.info("check font-end: Tên nhóm - "+ check_nhom_nhombanquanly_namtest)
        logging.info(check_nhom_nhombanquanly_namtest == "nam test")

        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        check_nhom_nhombanquanly_vaotrang_namtest = driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_vaotrang_namtest).text
        logging.info("Nhóm - Nhóm bạn quản lý")
        logging.info("check font-end: Check tên khi click vào nhóm bạn quản lý")
        logging.info("check font-end: Tên nhóm - "+ check_nhom_nhombanquanly_vaotrang_namtest)
        logging.info(check_nhom_nhombanquanly_vaotrang_namtest == "nam test")
        time.sleep(2)


    def nhombanthamgia(self):
        driver.implicitly_wait(3)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Nhóm bạn tham gia
        try:
            check_nhom_nhombanthamgia = driver.find_element(By.XPATH, var.check_nhom_nhombanthamgia).is_displayed()
            logging.info("Nhóm - Nhóm bạn tham gia")
            logging.info("check font-end: Có hiển thị tên nhóm bạn tham gia không")
            logging.info(check_nhom_nhombanthamgia)
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm bạn tham gia")
            logging.info("check font-end: Có hiển thị tên nhóm bạn tham gia không")
            logging.info("False")

        driver.find_element(By.XPATH, var.check_nhom_nhombanthamgia).click()
        time.sleep(2)
        check_nhom_nhombanthamgia_vaotrang1 = driver.find_element(By.XPATH, var.check_nhom_nhombanthamgia_vaotrang1).text
        logging.info("Nhóm - Nhóm bạn tham gia")
        logging.info("check font-end: Check tên khi click vào nhóm bạn tham gia")
        logging.info("check font-end: Tên nhóm - "+ check_nhom_nhombanthamgia_vaotrang1)
        logging.info(check_nhom_nhombanthamgia_vaotrang1 != None)
        time.sleep(2)


    def thaoluan(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Nhóm bạn quản lý
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        #Thảo luận - Tạo bài viết
        #Nhóm -  Ảnh video
        driver.find_element(By.XPATH, var.trang_taobaiviet_anhvideo).click()
        check_nhom_taobaiviet_anhvideo = driver.find_element(By.XPATH, var.check_trang_taobaiviet_anhvideo).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Ảnh/video")
        logging.info("check font-end: Popup - Tạo bài viết")
        logging.info(check_nhom_taobaiviet_anhvideo)
        logging.info(check_nhom_taobaiviet_anhvideo == "Tạo bài viết")
        driver.find_element(By.XPATH, var.taobaiviet_x1).click()

        #Nhóm -  Khoảnh khắc
        driver.find_element(By.XPATH, var.taobaiviet_khoanhkhac).click()
        time.sleep(1)
        check_nhom_taobaiviet_khoanhkhac = driver.find_element(By.XPATH, var.check_trang_taobaiviet_anhvideo).text
        logging.info("Nhóm - Nhsom của bạn - Tạo Khoảnh khắc")
        logging.info("check font-end: Popup - Tạo bài khoảnh khắc")
        logging.info(check_nhom_taobaiviet_khoanhkhac)
        logging.info(check_nhom_taobaiviet_khoanhkhac == "Tạo bài khoảnh khắc")
        # driver.find_element(By.XPATH, var.icon_x).click()

        try:
            driver.find_element(By.XPATH, var.trang_themmoi_khoanhkhac_nhapnoidung).send_keys(data['nhom']['themoi_khoanhkhac'])
        except:
            driver.refresh()
            time.sleep(2)
            driver.find_element(By.XPATH, var.trang_themmoi).click()
            driver.find_element(By.XPATH, var.trang_themmoi_khoanhkhac).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.trang_themmoi_khoanhkhac_nhapnoidung).send_keys(data['nhom']['themoi_khoanhkhac'])
        driver.find_element(By.XPATH, var.chontaptin).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/bnldc.exe")
        time.sleep(3)
        driver.find_element(By.XPATH, var.dangbai).click()
        # driver.find_element(By.XPATH, var.taokhoanhkhac_x).click()      #đăng bài lâu quá 1p

        wait = WebDriverWait(driver, 60)
        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH, var.check_nhom_taobaiviet_message))).text
        logging.info("Nhóm - Nhóm của bạn - Tạo khoảnh khắc")
        logging.info("check font-end: Message Tạo khoảnh khắc - Đăng khoảnh khắc thành công")
        logging.info(check_nhom_taobaiviet_messsage)
        logging.info(check_nhom_taobaiviet_messsage == "Đăng khoảnh khắc thành công")
        time.sleep(2)
        check_nhom_themmoi_tenkhoanhkhac = driver.find_element(By.XPATH, var.check_nhom_themmoi_tenkhoanhkhac).text
        logging.info("Nhóm - Nhóm của bạn - Thêm mới - Khoảnh khắc")
        logging.info("check font-end: Tên khoảnh khắc - " + data['nhom']['themoi_khoanhkhac'])
        logging.info(check_nhom_themmoi_tenkhoanhkhac)
        logging.info(check_nhom_themmoi_tenkhoanhkhac == data['nhom']['themoi_khoanhkhac'])
        time.sleep(2)


        # #Nhóm -  Cảm xúc/hoạt động
        driver.find_element(By.XPATH, var.nhom_taobaiviet_camxuchoatdong).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_camxuc).click()
        check_nhom_taobaiviet_anhvideo = driver.find_element(By.XPATH, var.check_trang_taobaiviet_anhvideo).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Cảm xúc/Hoạt động")
        logging.info("check font-end: Popup - Bạn đang cảm thấy thế nào?")
        logging.info(check_nhom_taobaiviet_anhvideo)
        logging.info(check_nhom_taobaiviet_anhvideo == "Bạn đang cảm thấy thế nào?")
        time.sleep(0.5)

        driver.find_element(By.XPATH, var.trang_taobaiviet_hoatdong).click()
        check_nhom_taobaiviet_anhvideo = driver.find_element(By.XPATH, var.check_trang_taobaiviet_anhvideo).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Cảm xúc/Hoạt động")
        logging.info("check font-end: Popup - Bạn đang làm gì?")
        logging.info(check_nhom_taobaiviet_anhvideo)
        logging.info(check_nhom_taobaiviet_anhvideo == "Bạn đang làm gì?")
        driver.find_element(By.XPATH, var.camxuchoatdong_x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)

        # #Nhóm -  Tạo bài viết - ảnh video
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_anhvideo).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['taobaiviet_anh_mota'])
        driver.find_element(By.XPATH, var.taobaiviet_tailenanhvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/tothichcau.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.dang).click()
        wait = WebDriverWait(driver, 20)
        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH, var.check_nhom_taobaiviet_video))).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Ảnh/video")
        logging.info("check font-end: Message - Video của bạn đang trong quá trình xử lý, chúng tôi sẽ thông báo cho bạn khi video đã sẵn sàng.")
        logging.info(check_nhom_taobaiviet_messsage)
        logging.info(check_nhom_taobaiviet_messsage == "Video của bạn đang trong quá trình xử lý, chúng tôi sẽ thông báo cho bạn khi video đã sẵn sàng.")
        if check_nhom_taobaiviet_messsage != "Video của bạn đang trong quá trình xử lý, chúng tôi sẽ thông báo cho bạn khi video đã sẵn sàng.":
            driver.refresh()
            time.sleep(2)

        #Nhóm -  Tạo bài viết - Video trực tiếp
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_dau3cham_videotructiep).click()
        check_nhom_taobaiviet_videotructiep = driver.find_element(By.XPATH,var.check_trang_taobaiviet_videotructiep).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Video trực tiếp")
        logging.info("check font-end: Chuyển tới màn hình - Tạo video trực tiếp")
        logging.info(check_nhom_taobaiviet_videotructiep)
        logging.info(check_nhom_taobaiviet_videotructiep == "Tạo video trực tiếp")
        driver.back()
        time.sleep(2)

        # Nhóm - Tạo bài viết - Cảm xúc/Hoạt động
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_camxuchoatdong).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.camxuc).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.camxuchoatdong_xinhxan).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['taobaiviet_camxuchoatdong_mota'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.dang).click()

        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH,var.check_nhom_taobaiviet_message))).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Cảm xúc/Hoạt động")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_nhom_taobaiviet_messsage)
        logging.info(check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)

        if check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.":
            check_nhom_baiviet_camxuchoatdong_tieude = driver.find_element(By.XPATH,var.check_nhom_baiviet_camxuchoatdong_tieude).text
            logging.info("Nhóm - Tạo bài viết - Cảm xúc/Hoạt động - Dòng thời gian")
            logging.info("check font-end: Tiêu đề - Ngọc Mai đang  cảm thấy buồn bã")   #bài viết ko có cảm xúc đã chọn 6603
            logging.info(check_nhom_baiviet_camxuchoatdong_tieude)
            logging.info(check_nhom_baiviet_camxuchoatdong_tieude == "Ngọc Mai\n đang  cảm thấy buồn bã")

            check_nhom_baiviet_camxuchoatdong_mota = driver.find_element(By.XPATH,var.check_nhom_baiviet_camxuchoatdong_mota).text
            logging.info("Nhóm - Tạo bài viết - Cảm xúc/Hoạt động - Dòng thời gian")
            logging.info("check font-end: Mô tả - " + data['nhom']['taobaiviet_camxuchoatdong_mota'])
            logging.info(check_nhom_baiviet_camxuchoatdong_mota)
            logging.info(check_nhom_baiviet_camxuchoatdong_mota == data['nhom']['taobaiviet_camxuchoatdong_mota'])
        else:
            driver.refresh()
            time.sleep(2)


        # Nhóm - Tạo bài viết - Gắn thẻ người khác
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_ganthenguoikhac).click()
        driver.find_element(By.XPATH, var.ganthenguoikhac_input).send_keys("Mạnh Cường")
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.taobaiviet_manhcuong)))
        element.click()
        driver.find_element(By.XPATH, var.xong).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['taobaiviet_ganthebaiviet_mota'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.dang).click()

        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH,var.check_nhom_taobaiviet_message))).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Gắn thẻ người khác")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_nhom_taobaiviet_messsage)
        logging.info(check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)
        if check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.":
            check_nhom_baiviet_ganthenguoikhac_tieude = driver.find_element(By.XPATH,var.check_nhom_baiviet_camxuchoatdong_tieude).text
            logging.info("Nhóm - Tạo bài viết - Gắn thẻ người khác - Dòng thời gian")
            logging.info("check font-end: Tiêu đề - Ngọc Mai cùng với Mạnh Cường")
            logging.info(check_nhom_baiviet_ganthenguoikhac_tieude)
            logging.info(check_nhom_baiviet_ganthenguoikhac_tieude == "Ngọc Mai\n cùng với \nMạnh Cường")

            check_nhom_baiviet_ganthenguoikhac_mota = driver.find_element(By.XPATH,var.check_nhom_baiviet_camxuchoatdong_mota).text
            logging.info("Nhóm - Tạo bài viết - Gắn thẻ người khác - Dòng thời gian")
            logging.info("check font-end: Mô tả - " + check_nhom_baiviet_ganthenguoikhac_mota)
            logging.info(check_nhom_baiviet_ganthenguoikhac_mota)
            logging.info(check_nhom_baiviet_ganthenguoikhac_mota == data['nhom']['taobaiviet_ganthebaiviet_mota'])
        else:
            driver.refresh()
            time.sleep(2)

        #Nhóm -  Tạo bài viết - Check in
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_checkin).click()
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban).send_keys("Cần Thơ")
        time.sleep(1)
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trang_vitri_cantho)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['taobaiviet_checkin_mota'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.dang).click()

        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH, var.check_nhom_taobaiviet_message))).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Check in")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_nhom_taobaiviet_messsage)
        logging.info(check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)

        if check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.":
            check_nhom_baiviet_checkin_tieude = driver.find_element(By.XPATH,var.check_nhom_baiviet_checkin_tieude).text
            logging.info("Nhóm - Tạo bài viết - Check in - Dòng thời gian")
            logging.info("check font-end: Tiêu đề - Ngọc Mai đang ở Cần Thơ")
            logging.info(check_nhom_baiviet_checkin_tieude)
            logging.info(check_nhom_baiviet_checkin_tieude == "Ngọc Mai\n đang ở Cần Thơ")

            check_nhom_baiviet_checkin_mota = driver.find_element(By.XPATH,var.check_nhom_baiviet_checkin_mota).text
            logging.info("Nhóm - Tạo bài viết - Check in - Dòng thời gian")
            logging.info("check font-end: Mô tả - " + data['nhom']['taobaiviet_checkin_mota'])
            logging.info(check_nhom_baiviet_checkin_mota)
            logging.info(check_nhom_baiviet_checkin_mota == data['nhom']['taobaiviet_checkin_mota'])

            check_nhom_baiviet_checkin_thanhpho = driver.find_element(By.XPATH,var.check_nhom_baiviet_checkin_thanhpho).text
            logging.info("Nhóm - Tạo bài viết - Check in - Dòng thời gian")
            logging.info("check font-end: Thành phố - Cần Thơ")
            logging.info(check_nhom_baiviet_checkin_thanhpho)
            logging.info(check_nhom_baiviet_checkin_thanhpho == "Cần Thơ")

            # Xoá bài viết
            driver.execute_script("window.scrollBy(0,400)", "")
            driver.find_element(By.XPATH, var.nhom_baiviet_dau3cham).click()
            driver.find_element(By.XPATH, var.trang_baiviet_dau3cham_xoabaiviet).click()
            driver.find_element(By.XPATH, var.xoa).click()
            time.sleep(2)
            check_nhom_baiviet_xoa = driver.find_element(By.XPATH, var.check_nhom_taobaiviet_message).text
            logging.info("Nhóm - Nhóm của bạn - Chỉnh sửa bài viết")
            logging.info("check font-end: Xoá - Message - Bài viết của bạn đã bị xóa.")
            logging.info(check_nhom_baiviet_xoa)
            logging.info(check_nhom_baiviet_xoa == "Bài viết của bạn đã bị xóa.")
            time.sleep(1)
        else:
            driver.refresh()
            time.sleep(2)

        #Nhóm - Tạo bài viết - Sự kiện trong đời
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_dau3cham_sukientrongdoi).click()

        driver.find_element(By.XPATH, var.sothichvahoatdong).click()
        driver.find_element(By.XPATH, var.nhacuavadoisong_tailenfile).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien1.exe")
        time.sleep(1)
        sukientrongdoi_clear1 = driver.find_element(By.XPATH, var.sothichvahoatdong_tieude_input)
        sukientrongdoi_clear1.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.sothichvahoatdong_tieude_input).send_keys(data['nhom']['hoatdongvasothich_tieude'])
        driver.find_element(By.XPATH, var.sothichvahoatdong_thoigian).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.sothichvahoatdong_thoigian_nam).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.sothichvahoatdong_thoigian_nam_2023).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.sothichvahoatdong_thoigian_thang).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.sothichvahoatdong_thoigian_thang_9).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.sothichvahoatdong_thoigian_ngay).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.sothichvahoatdong_thoigian_ngay_11).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['hoatdongvasothich_mota'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.dang).click()

        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH, var.check_nhom_taobaiviet_message))).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Sự kiện trong đời")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_nhom_taobaiviet_messsage)
        logging.info(check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)

        if check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.":
            check_nhom_baiviet_sukientrongdoi_anh = driver.find_element(By.XPATH,var.check_nhom_baiviet_sukientrongdoi_anh).get_attribute("src")
            logging.info("Nhóm - Tạo bài viết - Sự kiện trong đời - Dòng thời gian")
            logging.info("check font-end: Có hiển thị ảnh Sự kiện trong đời lên dòng thời gian nhóm không")
            logging.info(check_nhom_baiviet_sukientrongdoi_anh)
            logging.info(check_nhom_baiviet_sukientrongdoi_anh != None)

            check_nhom_baiviet_sukientrongdoi_mota = driver.find_element(By.XPATH, var.check_nhom_baiviet_sukientrongdoi_mota).text
            logging.info("Nhóm - Tạo bài viết - Sự kiện trong đời - Dòng thời gian")
            logging.info("check font-end: Mô tả - " + data['nhom']['hoatdongvasothich_mota'])
            logging.info(check_nhom_baiviet_sukientrongdoi_mota)
            logging.info(check_nhom_baiviet_sukientrongdoi_mota == data['nhom']['hoatdongvasothich_mota'])
        else:
            driver.refresh()
            time.sleep(2)

        #Nhóm - Tạo bài viết - Gif
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_dau3cham_filegif).click()
        driver.find_element(By.XPATH, var.trang_gif_input).send_keys("meo")
        time.sleep(5)
        driver.implicitly_wait(0.1)
        count = 0
        n = 0
        while (count < 25):
            n = int(n)
            n = n + 1
            count = count + 1
            n = str(n)
            nhom_taobaiviet_gif_anh = "/html/body/div[" + n + "]/div[3]/div/div[2]/div/div/img[1]"
            print(nhom_taobaiviet_gif_anh)
            try:
                button = driver.find_element(By.XPATH, nhom_taobaiviet_gif_anh)
                driver.execute_script("arguments[0].click();", button)
                if driver.find_element(By.XPATH, var.check_taobaiviet_gif).is_displayed():
                    break
            except:
                pass
        driver.implicitly_wait(15)
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['taobaiviet_gif_mota'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.dang).click()

        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH, var.check_nhom_taobaiviet_message))).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Gif")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_nhom_taobaiviet_messsage)
        logging.info(check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)

        if check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.":
            check_nhom_baiviet_gif_anh = driver.find_element(By.XPATH,var.check_nhom_baiviet_gif_anh).get_attribute("src")
            logging.info("Nhóm - Tạo bài viết - Gif - Dòng thời gian")
            logging.info("check font-end: Có hiển thị Gif lên dòng thời gian nhóm không")
            logging.info(check_nhom_baiviet_gif_anh)
            logging.info(check_nhom_baiviet_gif_anh != None)

            check_nhom_baiviet_gif_mota = driver.find_element(By.XPATH, var.check_nhom_baiviet_gif_mota).text
            logging.info("Nhóm - Tạo bài viết - Gif - Dòng thời gian")
            logging.info("check font-end: Mô tả - "+ data['nhom']['taobaiviet_gif_mota'])
            logging.info(check_nhom_baiviet_gif_mota)
            logging.info(check_nhom_baiviet_gif_mota == data['nhom']['taobaiviet_gif_mota'])
        else:
            driver.refresh()
            time.sleep(2)

        #Nhóm - Tạo bài viết - Tổ chức buổi H&Đ
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_dau3cham_tochucbuoihd).click()
        driver.find_element(By.XPATH, var.datcauhoi_input).send_keys(data['nhom']['taobaiviet_datcauhoi_input'])
        driver.find_element(By.XPATH, var.datcauhoi).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['taobaiviet_tochucbuoihd_mota'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.dang).click()
        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH, var.check_nhom_taobaiviet_message))).text

        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Tổ chức buổi H&Đ")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_nhom_taobaiviet_messsage)
        logging.info(check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(2)

        if check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.":
            check_nhom_baiviet_tochucbuoihoatdong_mota = driver.find_element(By.XPATH,var.check_nhom_baiviet_tochucbuoihoatdong_mota).text
            logging.info("Nhóm - Tạo bài viết - Tổ chức buổi H&Đ - Dòng thời gian")
            logging.info("check font-end: Mô tả - " + data['nhom']['taobaiviet_tochucbuoihd_mota'])
            logging.info(check_nhom_baiviet_tochucbuoihoatdong_mota)
            logging.info(check_nhom_baiviet_tochucbuoihoatdong_mota == data['nhom']['taobaiviet_tochucbuoihd_mota'])

            check_nhom_baiviet_tochucbuoihoatdong_cauhoi = driver.find_element(By.XPATH,var.check_nhom_baiviet_tochucbuoihoatdong_cauhoi).text
            logging.info("Nhóm - Tạo bài viết - Tổ chức buổi H&Đ - Dòng thời gian")
            logging.info("check font-end: Câu hỏi - " + data['nhom']['taobaiviet_datcauhoi_input'])
            logging.info(check_nhom_baiviet_tochucbuoihoatdong_cauhoi)
            logging.info(check_nhom_baiviet_tochucbuoihoatdong_cauhoi == data['nhom']['taobaiviet_datcauhoi_input'])
        else:
            driver.refresh()
            time.sleep(2)

        # Nhóm - Tạo bài viết - Thăm dò ý kiến
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_dau3cham_thamdoykien).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['thamdoykien_mota'])
        driver.find_element(By.XPATH, var.cuocthamdoykien_luachon1).send_keys(data['nhom']['cuocthamdoykien_luachon1'])
        driver.find_element(By.XPATH, var.cuocthamdoykien_luachon2).send_keys(data['nhom']['cuocthamdoykien_luachon2'])
        driver.find_element(By.XPATH, var.themluachon).click()
        driver.find_element(By.XPATH, var.cuocthamdoykien_luachon3).send_keys(data['nhom']['cuocthamdoykien_luachon3'])
        driver.find_element(By.XPATH, var.themluachon).click()
        driver.find_element(By.XPATH, var.cuocthamdoykien_luachon4).send_keys(data['nhom']['cuocthamdoykien_luachon4'])
        driver.find_element(By.XPATH, var.cuocthamdoykien_luachon4_x).click()
        driver.find_element(By.XPATH, var.cuocthamdoykien_chonthoigian).click()
        driver.find_element(By.XPATH, var.cuocthamdoykien_chonthoigian_30p).click()
        driver.find_element(By.XPATH, var.cuocthamdoykien_chonnhieu).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.dang).click()
        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH, var.check_nhom_taobaiviet_message))).text

        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Thăm dò ý kiến")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_nhom_taobaiviet_messsage)
        logging.info(check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.")
        driver.find_element(By.XPATH, var.icon_x).click()
        if check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.":
            check_nhom_baiviet_thamdoykien_mota = driver.find_element(By.XPATH,var.check_nhom_baiviet_thamdoykien_mota).text
            logging.info("Nhóm - Tạo bài viết - Thăm dò ý kiến - Dòng thời gian")
            logging.info("check font-end: Mô tả - " + data['nhom']['thamdoykien_mota'])
            logging.info(check_nhom_baiviet_thamdoykien_mota)
            logging.info(check_nhom_baiviet_thamdoykien_mota == data['nhom']['thamdoykien_mota'])

            check_nhom_baiviet_thamdoykien_luachon1 = driver.find_element(By.XPATH,var.check_nhom_baiviet_thamdoykien_luachon1).text
            logging.info("Nhóm - Tạo bài viết - Thăm dò ý kiến - Dòng thời gian")
            logging.info("check font-end: Lựa chọn 1 - " + data['nhom']['cuocthamdoykien_luachon1'])
            logging.info(check_nhom_baiviet_thamdoykien_luachon1)
            logging.info(check_nhom_baiviet_thamdoykien_luachon1 == data['nhom']['cuocthamdoykien_luachon1'])

            driver.find_element(By.XPATH, var.cuocthamdoykien_luachon1_chon).click()
            time.sleep(2)
            check_nhom_baiviet_thamdoykien_luachon1_dachon = driver.find_element(By.XPATH,var.check_nhom_baiviet_thamdoykien_luachon1_dachon).text
            logging.info("Nhóm - Tạo bài viết - Thăm dò ý kiến - Dòng thời gian")
            logging.info("check font-end: Lựa chọn 1 click chọn  - 1 lượt bình chọn")
            logging.info(check_nhom_baiviet_thamdoykien_luachon1_dachon)
            logging.info(check_nhom_baiviet_thamdoykien_luachon1_dachon == "1 lượt bình chọn")
            time.sleep(1)
        else:
            driver.refresh()
            time.sleep(2)

        #Nhóm -  Tạo bài viết - Công bố mục tiêu
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_dau3cham_congbomuctieu).click()
        driver.find_element(By.XPATH, var.congbomuctieu_input).send_keys(data['nhom']['taobaiviet_muctieu_input'])
        driver.find_element(By.XPATH, var.datmuctieu).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['taobaiviet_congbomuctieu_mota'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.dang).click()
        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH, var.check_nhom_taobaiviet_message))).text

        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Công bố mục tiêu")
        logging.info("check font-end: Message tạo bài viết - Đăng bài viết thành công.")
        logging.info(check_nhom_taobaiviet_messsage)
        logging.info(check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.")
        if check_nhom_taobaiviet_messsage == "Đăng bài viết thành công.":
            check_nhom_baiviet_congbomuctieu_mota = driver.find_element(By.XPATH,var.check_nhom_baiviet_congbomuctieu_mota).text
            logging.info("Nhóm - Tạo bài viết - Công bố mục tiêu - Dòng thời gian")
            logging.info("check font-end: Mô tả - " + data['nhom']['taobaiviet_congbomuctieu_mota'])
            logging.info(check_nhom_baiviet_congbomuctieu_mota)
            logging.info(check_nhom_baiviet_congbomuctieu_mota == data['nhom']['taobaiviet_congbomuctieu_mota'])

            check_nhom_baiviet_congbomuctieu_cauhoi = driver.find_element(By.XPATH,var.check_nhom_baiviet_congbomuctieu_cauhoi).text
            logging.info("Nhóm - Tạo bài viết - Công bố mục tiêu - Dòng thời gian")
            logging.info("check font-end: Câu hỏi - " + data['nhom']['taobaiviet_muctieu_input'])
            logging.info(check_nhom_baiviet_congbomuctieu_cauhoi)
            logging.info(check_nhom_baiviet_congbomuctieu_cauhoi == data['nhom']['taobaiviet_muctieu_input'])
            time.sleep(2)
        else:
            driver.refresh()
            time.sleep(2)

        try:
            check_nhom_thaoluan_filemoichiase1 = driver.find_element(By.XPATH, var.check_nhom_thaoluan_filemoichiase11).is_displayed()
            logging.info("Nhóm - Thảo luận - File phương tiện mới chia sẻ")
            logging.info("check font-end: Có hiển thị File phương tiện mới chia sẻ không")
            logging.info(check_nhom_thaoluan_filemoichiase1)
        except NoSuchElementException:
            logging.info("Nhóm - Thảo luận - File phương tiện mới chia sẻ")
            logging.info("check font-end: Có hiển thị File phương tiện mới chia sẻ không")
            logging.info("False")
        # driver.find_element(By.XPATH, var.nhom_thaoluan_filphuongtienmoichiase).click()
        button = driver.find_element(By.XPATH, var.nhom_thaoluan_filphuongtienmoichiase)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)

        check_nhom_thaoluan_filemoichiase1_xemtatca = driver.find_element(By.XPATH,var.check_nhom_thaoluan_filemoichiase1_xemtatca).text
        logging.info("Nhóm - Thảo luận - File phương tiện mới chia sẻ - Xem tất cả")
        logging.info("check font-end: Chuyển tới trang - File phương tiện")
        logging.info(check_nhom_thaoluan_filemoichiase1_xemtatca)
        logging.info(check_nhom_thaoluan_filemoichiase1_xemtatca == "File phương tiện")
        time.sleep(2)

    def dau3chambaiviet(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        #Tạo post 1
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['taobaiviet_gobaiviet1'])
        driver.find_element(By.XPATH, var.dang).click()
        driver.find_element(By.XPATH, var.check_nhom_taobaiviet_message)
        time.sleep(1)
        #Tạo post 2
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['taobaiviet_gobaiviet2'])
        driver.find_element(By.XPATH, var.dang).click()
        driver.find_element(By.XPATH, var.check_nhom_taobaiviet_message)
        time.sleep(1)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        #Dấu 3 chám bài viết nhóm
        #Ghim bài viết
        driver.find_element(By.XPATH, var.nhom_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.ghimbaiviet).click()
        check_nhom_message_ghimbaiviet = driver.find_element(By.XPATH,var.check_nhom_message_ghimbaiviet).text
        logging.info("Nhóm - Thảo luận - Bài viết 1")
        logging.info("check font-end: Message Ghim bài viết - Đã ghim bài viết của bạn")
        logging.info(check_nhom_message_ghimbaiviet)
        logging.info(check_nhom_message_ghimbaiviet == "Đã ghim bài viết của bạn")
        #Bỏ ghim bài viết
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_taobaiviet_dau3cham1).click()
        driver.find_element(By.XPATH, var.boghimbaiviet).click()
        check_nhom_message_boghimbaiviet = driver.find_element(By.XPATH,var.check_nhom_message_boghimbaiviet).text
        logging.info("Nhóm - Thảo luận - Bài viết 1")
        logging.info("check font-end: Message Bo ghim bài viết - Đã bỏ ghim bài viết của bạn")
        logging.info(check_nhom_message_boghimbaiviet)
        logging.info(check_nhom_message_boghimbaiviet == "Đã bỏ ghim bài viết của bạn")
        # time.sleep(1)   nhóm_dau 3 cham mowsi viest xong ghim




    def thanhvien(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        #Thành viên
        driver.find_element(By.XPATH, var.trang_trangcuaban_thanhvien).click()
        #Check dữ liệu thành viên
        try:
            check_nhom_tv_tenbanthan = driver.find_element(By.XPATH, var.check_nhom_tv_tenbanthan)
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có hiển thị tên của bản thân không?")
            logging.info(check_nhom_tv_tenbanthan.is_displayed())
            logging.info(check_nhom_tv_tenbanthan.text)
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có hiển thị tên của bản thân không?")
            logging.info("False")

        try:
            check_nhom_qtv_nguoikiemduyet = driver.find_element(By.XPATH, var.check_nhom_qtv_nguoikiemduyet)
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có hiển thị tên Quản trị viên và người kiểm duyệt không?")
            logging.info(check_nhom_qtv_nguoikiemduyet.is_displayed())
            logging.info(check_nhom_qtv_nguoikiemduyet.text)
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có hiển thị tên Quản trị viên và người kiểm duyệt không?")
            logging.info("False")

        # Nhóm - Thanh viên - Mời tham gia nhóm -  Mạnh cường
        driver.find_element(By.XPATH, var.nhom_moi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.manhcuong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.check_trang_moibanbe_tatcabanbe_guiloimoi).click()
        time.sleep(2)
        check_nhom_moibanbe_damoi1 = driver.find_element(By.XPATH, var.check_nhom_moibanbe_damoi1).text
        time.sleep(1)
        #Mạnh Cường đồng ý tham gia nhóm
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(2)
        driver.find_element(By.XPATH, var.trangchu_iconthongbao).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangchu_iconthongbao_thongbaodautien).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_loimoi_thamgianhom).click()
        time.sleep(1.5)
        check_nhom_loimoi_chapnhan = driver.find_element(By.XPATH, var.check_nhom_loimoi_chapnhan1).text
        time.sleep(0.5)

        #Mời Mạnh Cường làm Quản trị viên
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_trangcuaban_thanhvien).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.nhom_thanhvien_timkiem).send_keys(data['nhom']['timkiem_thanhvien'])
        time.sleep(1)
        driver.execute_script("window.scrollBy(0,300)", "")
        driver.find_element(By.XPATH, var.nhom_thanhvien_timkiem_dau3cham).click()
        driver.find_element(By.XPATH, var.nhom_thanhvien_themlamquantrivien).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
        try:
            check_nhom_qtv_nguoikiemduyetdamoi = driver.find_element(By.XPATH,var.check_nhom_qtv_nguoikiemduyetdamoi)
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có hiển thị tên Quản trị viên và người kiểm duyệt đã mời không?")
            logging.info(check_nhom_qtv_nguoikiemduyetdamoi.is_displayed())
            logging.info(check_nhom_qtv_nguoikiemduyetdamoi.text)
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có hiển thị tên Quản trị viên và người kiểm duyệt đã mời không?")
            logging.info("False")

        #Hủy lời mời làm qtv
        driver.find_element(By.XPATH, var.nhom_thanhvien_qtv_dau3cham).click()
        driver.find_element(By.XPATH, var.huyloimoilamqtv).click()
        driver.find_element(By.XPATH, var.xacnhan).click()
        try:
            check_nhom_thanhvien_huyloimoilamqtv = driver.find_element(By.XPATH,var.nhom_thanhvien_qtv_dau3cham).is_displayed()
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Hủy lời làm làm Quản trị viên ")
            logging.info(check_nhom_thanhvien_huyloimoilamqtv)
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Hủy lời làm làm Quản trị viên ")
            logging.info("False")

        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_thanhvien_timkiem).send_keys(data['nhom']['timkiem_thanhvien'])
        time.sleep(1)
        driver.execute_script("window.scrollBy(0,300)", "")
        driver.find_element(By.XPATH, var.nhom_thanhvien_timkiem_dau3cham).click()
        driver.find_element(By.XPATH, var.nhom_thanhvien_themlamquantrivien).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(2)


        # Mạnh Cường đồng ý làm qtv
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(2)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_moithamgianhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chapnhan).click()
        driver.implicitly_wait(3)
        driver.find_element(By.XPATH, var.nhom_moithamgianhom_loimoi1).click()
        time.sleep(2)


        #Gỡ vai trò qtv của Mạnh Cường
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.trang_trangcuaban_thanhvien).click()
        try:
            check_nhom_thanhvien_banbe = driver.find_element(By.XPATH, var.check_nhom_thanhvien_banbe)
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có hiển thị tên bạn bè không?")
            logging.info(check_nhom_thanhvien_banbe.is_displayed())
            logging.info(check_nhom_thanhvien_banbe.text)
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có hiển thị tên bạn bè không?")
            logging.info("False")

        try:
            check_nhom_thanhvien_thanhvien = driver.find_element(By.XPATH, var.check_nhom_thanhvien_thanhvien)
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có hiển thị tên thành viên không?")
            logging.info(check_nhom_thanhvien_thanhvien.is_displayed())
            logging.info(check_nhom_thanhvien_thanhvien.text)
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có hiển thị tên thành viên không?")
            logging.info("False")


        driver.execute_script("window.scrollBy(0,600)", "")
        driver.find_element(By.XPATH, var.nhom_thanhvien_banbe_dau3cham).click()
        driver.find_element(By.XPATH, var.nhom_thanhvien_qtv_dau3cham_goqtv).click()
        check_nhom_thanhvien_goqtv = driver.find_element(By.XPATH,var.check_nhom_thanhvien_goqtv).text
        logging.info("Nhóm - Nhóm của bạn - Thành viên")
        logging.info("check font-end: Message gỡ qtv - Bạn sắp gỡ vai trò quản trị viên nhóm của Mạnh Cường trong nam test.")
        logging.info(check_nhom_thanhvien_goqtv)
        logging.info(check_nhom_thanhvien_goqtv == "Bạn sắp gỡ vai trò quản trị viên nhóm của Mạnh Cường trong nam test.")

        driver.find_element(By.XPATH, var.xacnhan).click()
        driver.implicitly_wait(2)
        time.sleep(2)
        try:
            check_nhom_thanhvien_qtv_daxoamanhcuong = driver.find_element(By.XPATH, var.check_nhom_thanhvien_qtv_daxoamanhcuong).is_displayed()
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có gỡ quyền Quản trị viên của Mạnh Cường được không?")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có gỡ quyền Quản trị viên của Mạnh Cường được không?")
            logging.info("True")
        driver.implicitly_wait(15)

        #Xóa Mạnh Cường khỏi nhóm
        driver.find_element(By.XPATH, var.nhom_thanhvien_banbe_dau3cham).click()
        driver.find_element(By.XPATH, var.xoakhoinhom).click()
        time.sleep(1)
        check_nhom_thanhvien_banbe_xoakhoinhom = driver.find_element(By.XPATH,var.check_nhom_thanhvien_banbe_xoakhoinhom).text
        logging.info("Nhóm - Nhóm của bạn - Thành viên")
        logging.info("check font-end: Message xóa khỏi nhóm - Bạn chắc chắn muốn xóa Mạnh Cường khỏi nhóm chứ ?")
        logging.info(check_nhom_thanhvien_banbe_xoakhoinhom)
        logging.info(check_nhom_thanhvien_banbe_xoakhoinhom == "Bạn chắc chắn muốn xóa Mạnh Cường khỏi nhóm chứ ?")
        driver.find_element(By.XPATH, var.xacnhan).click()
        driver.implicitly_wait(3)
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
        try:
            check_nhom_thanhvien_banbe_xoamanhcuongkhoinhom = driver.find_element(By.XPATH, var.check_nhom_thanhvien_banbe_xoamanhcuongkhoinhom).is_displayed()
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có xóa Mạnh Cường khỏi nhóm được không?")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Có xóa Mạnh Cường khỏi nhóm được không?")
            logging.info("True")

        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(2)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879/member")
        time.sleep(2)
        check_nhom_chanthanhvien = driver.find_element(By.XPATH,var.check_nhom_chanthanhvien).text
        logging.info("Nhóm - Nhóm của bạn - Thành viên")
        logging.info("check font-end: Chặn Mạnh Cường khỏi nhóm nam test - Trang này không hiển thị")
        logging.info(check_nhom_chanthanhvien)
        logging.info(check_nhom_chanthanhvien == "Trang này không hiển thị")
        time.sleep(1)

        #Danh sách thành viên bị chặn
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(2)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879/member")
        time.sleep(2)
        try:
            check_nhom_thanhvien_thanhvienbichan = driver.find_element(By.XPATH, var.check_nhom_thanhvien_thanhvienbichan)
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Danh sách thành viên bị chặn Có hiển thị không")
            logging.info(check_nhom_thanhvien_thanhvienbichan.is_displayed())
            logging.info(check_nhom_thanhvien_thanhvienbichan.text)
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Danh sách thành viên bị chặn Có hiển thị không")
            logging.info("False")
        # driver.find_element(By.XPATH, var.nhom_thanhvien_dsbichan_dau3cham).click()
        button = driver.find_element(By.XPATH, var.nhom_thanhvien_dsbichan_dau3cham)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.bochan).click()
        time.sleep(2)

        try:
            check_nhom_thanhvien_thanhvienbichan = driver.find_element(By.XPATH, var.check_nhom_thanhvien_thanhvienbichan).is_displayed()
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Bỏ chặn Mạnh Cường khỏi Danh sách thành viên bị chặn")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thành viên")
            logging.info("check font-end: Bỏ chặn Mạnh Cường khỏi Danh sách thành viên bị chặn")
            logging.info("True")
        time.sleep(2)



    def filephuongtien(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        #File phương tiện
        driver.find_element(By.XPATH, var.trang_trangcuaban_filephuongtien).click()

    def anh(self):
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_filephuongtien_anh_xemanhdau).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.thich).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(data['nhom']['anh_binhluan'])
        driver.find_element(By.XPATH, var.vietbinhluan).submit()
        time.sleep(3)
        driver.implicitly_wait(0.1)
        count = 0
        n = 0
        while (count < 15):
            n = int(n)
            n = n + 1
            count = count + 1
            n = str(n)
            check_filephuongtien_anh_binhluan1 = "//*[@class='scrollbar-container ps']/div[" + n + "]/ul/li[1]/div[1]/div[1]/div[2]/div/span/div/div/p"
            print(check_filephuongtien_anh_binhluan1)
            try:
                check_filephuongtien_anh_binhluan = driver.find_element(By.XPATH,var.check_filephuongtien_anh_binhluan1).text
                logging.info("Nhóm - Nhóm của bạn - File phương tiện")
                logging.info("check font-end: Ảnh - Xem ảnh đầu - Bình luận - " + data['nhom']['anh_binhluan'])
                logging.info(check_filephuongtien_anh_binhluan == data['nhom']['anh_binhluan'])
                break
            except:
                pass
        driver.implicitly_wait(15)
        driver.find_element(By.XPATH, var.esc).click()
        time.sleep(1)
        #Xóa ảnh
        nhom_filephuongtien_linkanh1 = driver.find_element(By.XPATH, var.nhom_filephuongtien_anh_xemanhdau).get_attribute("style")
        driver.find_element(By.XPATH, var.nhom_iconxoaanh).click()
        driver.find_element(By.XPATH, var.xoafilephuongtien).click()
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(2)
        nhom_filephuongtien_linkanh2 = driver.find_element(By.XPATH, var.nhom_filephuongtien_anh_xemanhdau).get_attribute("style")
        logging.info("Nhóm - Nhóm của bạn - File phương tiện")
        logging.info("check font-end: Ảnh - Xóa ảnh")
        logging.info(nhom_filephuongtien_linkanh1)
        logging.info(nhom_filephuongtien_linkanh2)
        logging.info(nhom_filephuongtien_linkanh1 != nhom_filephuongtien_linkanh2)

    def video(self):
        driver.find_element(By.XPATH, var.nhom_filephuongtien_video).click()
        driver.find_element(By.XPATH, var.nhom_filephuongtien_anh_xemvideodau).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.thich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(data['nhom']['video_binhluan'])
        driver.find_element(By.XPATH, var.vietbinhluan).submit()
        time.sleep(3)
        driver.implicitly_wait(0.1)
        count = 0
        n = 0
        while (count < 15):
            n = int(n)
            n = n + 1
            count = count + 1
            n = str(n)
            check_filephuongtien_video_binhluan = "//*[@class='scrollbar-container ps']/div[" + n + "]/ul/li[1]/div[1]/div[1]/div[2]/div/span/div/div/p"
            print(check_filephuongtien_video_binhluan)
            try:
                check_filephuongtien_video_binhluan = driver.find_element(By.XPATH,check_filephuongtien_video_binhluan).text
                logging.info("Nhóm - Nhóm của bạn - File phương tiện")
                logging.info("check font-end: Ảnh - Xem video đầu - Bình luận - " + data['nhom']['video_binhluan'])
                logging.info(check_filephuongtien_video_binhluan == data['nhom']['video_binhluan'])
                break
            except:
                pass
        driver.implicitly_wait(15)
        driver.find_element(By.XPATH, var.xemvideo_x).click()
        #Xóa video
        nhom_filephuongtien_linkvideo1 = driver.find_element(By.XPATH, var.nhom_filephuongtien_anh_xemvideodau).get_attribute("src")
        driver.find_element(By.XPATH, var.nhom_iconxoavideo).click()
        driver.find_element(By.XPATH, var.xoafilephuongtien).click()
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(3)
        nhom_filephuongtien_linkvideo2 = driver.find_element(By.XPATH, var.nhom_filephuongtien_anh_xemvideodau).get_attribute("src")
        logging.info("Nhóm - Nhóm của bạn - File phương tiện")
        logging.info("check font-end: Video - Xóa video")
        logging.info(nhom_filephuongtien_linkvideo1)
        logging.info(nhom_filephuongtien_linkvideo2)
        logging.info(nhom_filephuongtien_linkvideo1 != nhom_filephuongtien_linkvideo2)

    def khoanhkhac(self):
        driver.find_element(By.XPATH, var.nhom_filephuongtien_khoanhkhac).click()
        driver.find_element(By.XPATH, var.nhom_filephuongtien_khoanhkhac_xemkhoanhkhacdau).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.thich).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(data['nhom']['khoanhkhac_binhluan'])
        driver.find_element(By.XPATH, var.vietbinhluan).submit()
        time.sleep(3)
        driver.implicitly_wait(0.1)
        count = 0
        n = 0
        while (count < 25):
            n = int(n)
            n = n + 1
            count = count + 1
            n = str(n)
            check_nhom_khoanhkhac_binhluan = "//*[@class='scrollbar-container ps']/div[" + n + "]/ul/li[1]/div[1]/div[1]/div[2]/div/span/div/div/p"
            print(check_nhom_khoanhkhac_binhluan)
            try:
                check_filephuongtien_khoanhkhac_binhluan1 = driver.find_element(By.XPATH, check_nhom_khoanhkhac_binhluan).text
                logging.info("Nhóm - Nhóm của bạn - File phương tiện")
                logging.info("check font-end: Khoảnh khắc - Xem khoảnh khắc đầu - Bình luận - " + data['nhom']['khoanhkhac_binhluan'])
                logging.info(check_filephuongtien_khoanhkhac_binhluan1 == data['nhom']['khoanhkhac_binhluan'])
                break
            except:
                pass
        driver.implicitly_wait(15)
        driver.find_element(By.XPATH, var.xemvideo_x).click()
        #Xóa Khoảnh khắc
        nhom_filephuongtien_linkkhoanhkhac1 = driver.find_element(By.XPATH, var.nhom_filephuongtien_khoanhkhac_xemkhoanhkhacdau).get_attribute("src")
        driver.find_element(By.XPATH, var.nhom_iconxoakhoanhkhac).click()
        driver.find_element(By.XPATH, var.xoafilephuongtien).click()
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
        nhom_filephuongtien_linkkhoanhkhac2 = driver.find_element(By.XPATH, var.nhom_filephuongtien_khoanhkhac_xemkhoanhkhacdau).get_attribute("src")
        logging.info("Nhóm - Nhóm của bạn - File phương tiện")
        logging.info("check font-end: Khoảnh khắc - Xóa khoảnh khắc")
        logging.info(nhom_filephuongtien_linkkhoanhkhac1)
        logging.info(nhom_filephuongtien_linkkhoanhkhac2)
        logging.info(nhom_filephuongtien_linkkhoanhkhac1 != nhom_filephuongtien_linkkhoanhkhac2)
        time.sleep(1)

    def album(self):
        button = driver.find_element(By.XPATH, var.nhom_filephuongtien_album)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.trang_anh_alum_taomoi).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_ten).send_keys(data['nhom']['album_ten'])
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_mota).send_keys(data['nhom']['album_mota'])
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_tailen).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbiatrang1.exe")
        time.sleep(2)
        #chọn vị trí
        driver.find_element(By.XPATH, var.trang_alum_chonvitri).click()
        xoa = driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_gioithieu_nhapvitri_chonvitricuaban).send_keys(data['trang']['vitri1'])
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.trangcanhan_gioithieu_songtai_langson)))
        element.click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_anhtailen_chuthich).send_keys(data['nhom']['album_chuthich'])
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_dang).click()
        time.sleep(3)
        check_filephuongtien_album_taomoi = driver.find_element(By.XPATH,var.check_filephuongtien_album_taomoi).text
        logging.info("Nhóm - Nhóm của bạn - File phương tiện")
        logging.info("check font-end: Album - Tạo album mới - " + data['nhom']['album_ten'])
        logging.info(check_filephuongtien_album_taomoi)
        logging.info(check_filephuongtien_album_taomoi ==  data['nhom']['album_ten'])
        # #dấu 3 chấm album
        #Chỉnh sửa album
        driver.find_element(By.XPATH, var.nhom_filephuongtien_album_icondau3cham).click()
        driver.find_element(By.XPATH, var.capnhatalbum).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.trangcanhan_anh_album_ten).send_keys(data['nhom']['album_ten1'])
        driver.find_element(By.XPATH, var.luuthaydoi).click()
        time.sleep(2)
        #Xóa album
        nhom_filephuongtien_linkalbum1 = driver.find_element(By.XPATH, var.nhom_filephuongtien_linkalbum1).get_attribute("style")
        driver.find_element(By.XPATH, var.nhom_filephuongtien_album_icondau3cham).click()
        driver.find_element(By.XPATH, var.xoaalbum).click()
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(2)
        # #Xóa album
        nhom_filephuongtien_linkalbum2 = driver.find_element(By.XPATH, var.nhom_filephuongtien_linkalbum1).get_attribute("style")
        logging.info("Nhóm - Nhóm của bạn - File phương tiện")
        logging.info("check font-end: Album - xóa album")
        logging.info(nhom_filephuongtien_linkalbum1)
        logging.info(nhom_filephuongtien_linkalbum2)
        logging.info(nhom_filephuongtien_linkalbum1 != nhom_filephuongtien_linkalbum2)
        time.sleep(2)



    def yeucaulamthanhvien(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_exinanh).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_exinanh_yeucaulamthanhvien).click()
        time.sleep(1)
        #Tìm kiếm
        driver.find_element(By.XPATH, var.nhom_yeucaulamthanhvien_timkiem).send_keys("Trần Quang Trường")   #Không tìm kiếm được
        driver.find_element(By.XPATH, var.nhom_yeucaulamthanhvien_timkiem).send_keys(Keys.TAB)
        time.sleep(2)
        check_yeucaulamthanhvien_timkiem = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_timkiem).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Tìm kiếm - Trần Quang Trường")
        logging.info(check_yeucaulamthanhvien_timkiem)
        logging.info(check_yeucaulamthanhvien_timkiem == "Trần Quang Trường")
        # driver.find_element(By.XPATH, var.nhom_yeucaulamthanhvien_timkiem).clear()
        # xoa = driver.find_element(By.XPATH, var.nhom_yeucaulamthanhvien_timkiem)
        # xoa.send_keys(Keys.CONTROL, "a")
        driver.refresh()
        time.sleep(2)

        #Ngày tham gia
        #Ngày tham gia - Chưa đến 3 tháng trước
        driver.find_element(By.XPATH, var.ngaythamgia).click()
        driver.find_element(By.XPATH, var.chuaden3thangtruoc).click()
        time.sleep(2)
        check_yeucaulamthanhvien_chuaden3thangtruoc = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Ngày tham gia - Chưa đến 3 tháng trước - Có dữ liệu không?")
        logging.info(check_yeucaulamthanhvien_chuaden3thangtruoc)
        logging.info(check_yeucaulamthanhvien_chuaden3thangtruoc != None)
        driver.find_element(By.XPATH, var.chuaden3thangtruoc).click()

        #Ngày tham gia - Chưa đến 6 tháng trước
        driver.find_element(By.XPATH, var.chuaden6thangtruoc).click()
        time.sleep(2)
        check_yeucaulamthanhvien_chuaden6thangtruoc = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Ngày tham gia - Chưa đến 6 tháng trước - Có dữ liệu không?")
        logging.info(check_yeucaulamthanhvien_chuaden6thangtruoc)
        logging.info(check_yeucaulamthanhvien_chuaden6thangtruoc != None)
        driver.find_element(By.XPATH, var.chuaden6thangtruoc).click()

        #Ngày tham gia - Hơn 1 năm trước
        driver.find_element(By.XPATH, var.hon1namtruoc).click()
        time.sleep(2)
        check_yeucaulamthanhvien_hon1namtruoc = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieutrong).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Ngày tham gia - Hơn 1 năm trước - Không có thành viên đang chờ nào")
        logging.info(check_yeucaulamthanhvien_hon1namtruoc)
        logging.info(check_yeucaulamthanhvien_hon1namtruoc == "Không có thành viên đang chờ nào")
        driver.find_element(By.XPATH, var.hon1namtruoc).click()

        #Ngày tham gia - Hơn 2 năm trước
        driver.find_element(By.XPATH, var.hon2namtruoc).click()
        time.sleep(2)
        check_yeucaulamthanhvien_hon2namtruoc = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieutrong).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Ngày tham gia - Hơn 2 năm trước - Không có thành viên đang chờ nào")
        logging.info(check_yeucaulamthanhvien_hon2namtruoc)
        logging.info(check_yeucaulamthanhvien_hon2namtruoc == "Không có thành viên đang chờ nào")
        driver.find_element(By.XPATH, var.xoaboloc).click()
        time.sleep(2)

        #Thời gian yêu cầu
        #Thời gian yêu cầu- Dưới 1 tuần
        driver.find_element(By.XPATH, var.thoigianyeucau).click()
        driver.find_element(By.XPATH, var.duoi1tuan).click()
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            check_yeucaulamthanhvien_duoi1tuan = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
            logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
            logging.info("check font-end: Thời gian yêu cầu - Dưới 1 tuần")
            logging.info("Flase")
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
            logging.info("check font-end: Thời gian yêu cầu - Dưới 1 tuần")
            logging.info("True")

        driver.implicitly_wait(15)
        driver.find_element(By.XPATH, var.duoi1tuan).click()

        #Thời gian yêu cầu - Dưới 3 tháng
        driver.find_element(By.XPATH, var.duoi3thang).click()
        time.sleep(2)
        check_yeucaulamthanhvien_duoi3thang = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Thời gian yêu cầu - Dưới 3 tháng - Có dữ liệu không?")
        logging.info(check_yeucaulamthanhvien_duoi3thang)
        logging.info(check_yeucaulamthanhvien_duoi3thang != None)
        driver.find_element(By.XPATH, var.duoi3thang).click()

        #Thời gian yêu cầu - Hơn 3 tháng trước
        driver.find_element(By.XPATH, var.hon3thangtruoc).click()
        time.sleep(2)
        check_yeucaulamthanhvien_hon3thangtruoc = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieutrong).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Thời gian yêu cầu - Hơn 3 tháng trước - Không có thành viên đang chờ nào")
        logging.info(check_yeucaulamthanhvien_hon3thangtruoc)
        logging.info(check_yeucaulamthanhvien_hon3thangtruoc == "Không có thành viên đang chờ nào")
        driver.find_element(By.XPATH, var.hon3thangtruoc).click()

        #Thời gian yêu cầu - Hơn 6 tháng trước
        driver.find_element(By.XPATH, var.hon6thangtruoc).click()
        time.sleep(2)
        check_yeucaulamthanhvien_hon6thangtruoc = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieutrong).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Thời gian yêu cầu - Hơn 6 tháng trước - Không có thành viên đang chờ nào")
        logging.info(check_yeucaulamthanhvien_hon6thangtruoc)
        logging.info(check_yeucaulamthanhvien_hon6thangtruoc == "Không có thành viên đang chờ nào")
        driver.find_element(By.XPATH, var.hon6thangtruoc).click()

        #Thời gian yêu cầu - Hơn 1 năm trước
        driver.find_element(By.XPATH, var.hon1namtruoc).click()
        time.sleep(2)
        check_yeucaulamthanhvien_hon1namtruoc = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieutrong).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Thời gian yêu cầu - Hơn 1 năm trước - Không có thành viên đang chờ nào")
        logging.info(check_yeucaulamthanhvien_hon1namtruoc)
        logging.info(check_yeucaulamthanhvien_hon1namtruoc == "Không có thành viên đang chờ nào")
        driver.find_element(By.XPATH, var.xoaboloc).click()
        time.sleep(2)

        #Giới tính
        #Giới tính - Nữ
        driver.find_element(By.XPATH, var.gioitinh).click()
        driver.find_element(By.XPATH, var.nu).click()
        time.sleep(2)
        check_yeucaulamthanhvien_nu = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Giới tính - Nữ - Có dữ liệu không")
        logging.info(check_yeucaulamthanhvien_nu)
        logging.info(check_yeucaulamthanhvien_nu != None)
        logging.info(check_yeucaulamthanhvien_nu == "Ngọc Mai")
        driver.find_element(By.XPATH, var.nu).click()

        #Giới tính - Nam
        # driver.find_element(By.XPATH, var.gioitinh).click()
        driver.find_element(By.XPATH, var.nam).click()
        time.sleep(2)
        check_yeucaulamthanhvien_nam = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Giới tính - Nam - Có dữ liệu không")
        logging.info(check_yeucaulamthanhvien_nam)
        logging.info(check_yeucaulamthanhvien_nam != None)

        #Giới tính - Khác
        driver.find_element(By.XPATH, var.nam).click()
        driver.find_element(By.XPATH, var.khac).click()
        time.sleep(2)
        check_yeucaulamthanhvien_nam = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Giới tính - Khác - Có dữ liệu không")
        logging.info(check_yeucaulamthanhvien_nam)
        logging.info(check_yeucaulamthanhvien_nam != None)
        logging.info(check_yeucaulamthanhvien_nam == "hue nguyen")
        driver.find_element(By.XPATH, var.xoaboloc).click()
        time.sleep(2)

        #Ảnh đại diện
        #Ảnh đại diện - Có ảnh đại diện
        driver.find_element(By.XPATH, var.anhdaidien).click()
        driver.find_element(By.XPATH, var.coanhdaidien).click()
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            check_yeucaulamthanhvien_coanhdaidien = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_anhdaidien).get_attribute("src")
            logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
            logging.info("check font-end: Ảnh đại diện - Có ảnh đại diện")
            logging.info(check_yeucaulamthanhvien_coanhdaidien != "https://cmc-sn.emso.vn/avatars/original/missing.png")
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
            logging.info("check font-end: Ảnh đại diện - Có ảnh đại diện")
            logging.info("False")

        #Ảnh đại diện - Không có ảnh đại diện
        driver.find_element(By.XPATH, var.coanhdaidien).click()
        driver.find_element(By.XPATH, var.khongcoanhdaidien).click()
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            check_yeucaulamthanhvien_coanhdaidien = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_anhdaidien).get_attribute("src")
            logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
            logging.info("check font-end: Ảnh đại diện - Không có ảnh đại diện")
            logging.info(check_yeucaulamthanhvien_coanhdaidien == "https://cmc-sn.emso.vn/avatars/original/missing.png")
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
            logging.info("check font-end: Ảnh đại diện - Không có ảnh đại diện")
            logging.info("False")
        time.sleep(2)


        #Yêu cầu làm thành viên - Phê duyệt
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(2)
        driver.get("https://cmc-fe.emso.vn/group/111561259300587836")
        driver.find_element(By.XPATH, var.thamgianhom).click()
        try:
            check_yeucaulamthanhvien_thamgianhom = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_thamgianhom).text
            logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
            logging.info("check font-end: Message tham gia nhóm - Yêu cầu của bạn đã được gửi thành công")
            logging.info(check_yeucaulamthanhvien_thamgianhom)
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
            logging.info("check font-end: Message tham gia nhóm - Yêu cầu của bạn đã được gửi thành công")
            logging.info("False")
        time.sleep(1)

        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(2)
        driver.get("https://cmc-fe.emso.vn/group/111561259300587836/request_member")
        time.sleep(2)
        check_yeucaulamthanhvien_guiyeucau = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
        logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
        logging.info("check font-end: Mạnh Cường yêu cầu tham gia nhóm E xin anh")
        logging.info(check_yeucaulamthanhvien_guiyeucau)
        logging.info(check_yeucaulamthanhvien_guiyeucau == "Mạnh Cường")
        if check_yeucaulamthanhvien_guiyeucau == "Mạnh Cường":
            driver.find_element(By.XPATH, var.yeucauthamgianhom_manhcuong_pheduyet).click()
            time.sleep(3)
            check_yeucaulamthanhvien_guiyeucau = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
            if check_yeucaulamthanhvien_guiyeucau != "Mạnh Cường":
                logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
                logging.info("check font-end: Mạnh Cường - Phê duyệt")
                logging.info("True")
            else:
                logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
                logging.info("check font-end: Mạnh Cường - Phê duyệt")
                logging.info("False")
        else:
            logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
            logging.info("check font-end: Mạnh Cường - Gửi yêu cầu tham gia nhóm có hiển thị trong danh sách yêu cầu không?")
            logging.info("False")
        time.sleep(2)
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(2)
        driver.get("https://cmc-fe.emso.vn/group/111561259300587836")
        time.sleep(2)
        driver.find_element(By.XPATH, var.dathamgia).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.roikhoinhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.roikhoinhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.thamgianhom).click()
        time.sleep(2)
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(2)
        driver.get("https://cmc-fe.emso.vn/group/111561259300587836/request_member")
        time.sleep(2)
        check_yeucaulamthanhvien_guiyeucau = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
        if check_yeucaulamthanhvien_guiyeucau == "Mạnh Cường":
            driver.find_element(By.XPATH, var.yeucauthamgianhom_manhcuong_tuchoi).click()
            time.sleep(3)
            check_yeucaulamthanhvien_guiyeucau = driver.find_element(By.XPATH,var.check_yeucaulamthanhvien_dulieu1).text
            if check_yeucaulamthanhvien_guiyeucau != "Mạnh Cường":
                logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
                logging.info("check font-end: Mạnh Cường - Từ chối")
                logging.info("True")
            else:
                logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
                logging.info("check font-end: Mạnh Cường - Từ chối")
                logging.info("False")
        else:
            logging.info("Nhóm - Nhóm của bạn - Yêu cầu làm thành viên")
            logging.info("check font-end: Mạnh Cường - Gửi yêu cầu tham gia nhóm có hiển thị trong danh sách yêu cầu không?")
            logging.info("False")
        time.sleep(2)


    def baivietdalenlich(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        #Nhóm - Lên lịch bài viết
        driver.find_element(By.XPATH, var.nhom_taobaiviet).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['lenlich'])
        driver.find_element(By.XPATH, var.trang_taobaiviet_iconlenlich).click()

        xoa = driver.find_element(By.XPATH, var.trang_taobaiviet_lenlich_ngay)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_taobaiviet_lenlich_ngay).send_keys("08-12-2024")
        driver.find_element(By.XPATH, var.trang_taobaiviet_lenlich).click()
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(1)
        check_nhom_taobaiviet_lenlich = driver.find_element(By.XPATH,var.check_trang_taobaiviet_lenlich).text
        logging.info("Nhóm - Nhóm của bạn - Tạo bài viết - Lên lịch")
        logging.info("check font-end: Message tạo bài viết - Đã lên lịch cho bài viết của bạn.")
        logging.info(check_nhom_taobaiviet_lenlich)
        logging.info(check_nhom_taobaiviet_lenlich == "Đã lên lịch cho bài viết của bạn.")
        time.sleep(2)

        #Bài viết đã lên lịch
        driver.refresh()
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            check_baivietdalenlich_thongbao = driver.find_element(By.XPATH,var.check_baivietdalenlich_thongbao)
            logging.info("Nhóm - Nhóm của bạn - Bài viết đã lên lịch")
            logging.info("check font-end: Có thông báo ở module hay không")
            logging.info(check_baivietdalenlich_thongbao.is_displayed())
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Bài viết đã lên lịch")
            logging.info("check font-end: Có thông báo ở module hay không")
            logging.info("False")
        driver.find_element(By.XPATH, var.nhom_baivietdalenlich).click()
        check_nhom_baivietdalenlich = driver.find_element(By.XPATH, var.check_nhom_baivietdalenlich).text
        logging.info("Nhóm - Bài viết đã lên lịch")
        logging.info("check font-end: Tiêu đề bài viết đợi đăng - "+ data['nhom']['lenlich'])
        logging.info(check_nhom_baivietdalenlich == data['nhom']['lenlich'])

        driver.find_element(By.XPATH, var.trang_baivietdalenlich_doilich).click()
        xoa = driver.find_element(By.XPATH, var.trang_baivietdalenlich_doilichinput)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trang_baivietdalenlich_doilichinput).send_keys("20-11-2024")
        driver.find_element(By.XPATH, var.trang_baivietdalenlich_doilich_capnhat).click()
        time.sleep(1)
        driver.refresh()
        time.sleep(2)
        driver.find_element(By.XPATH, var.dangngay).click()     #Không đăng ngay được
        try:
            check_trang_baivietdalenlich_message = driver.find_element(By.XPATH, var.check_trang_baivietdalenlich_message).text
            logging.info("Nhóm - Bài viết đã lên lịch")
            logging.info("check font-end: Message - Đăng bài viết thành công.")
            logging.info(check_trang_baivietdalenlich_message == "Đăng bài viết thành công.")
        except NoSuchElementException:
            logging.info("Nhóm - Bài viết đã lên lịch")
            logging.info("check font-end: Message - Đăng bài viết thành công.")
            logging.info("False")
        time.sleep(5)
        driver.implicitly_wait(15)
        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        check_nhom_baiviet_lenlich = driver.find_element(By.XPATH, var.check_nhom_baiviet_lenlich).text
        logging.info("Nhóm - Tạo bài viết - Lên lịch - Dòng thời gian")
        logging.info("check font-end: Mô tả - " + data['nhom']['lenlich'])
        logging.info(check_nhom_baiviet_lenlich)
        logging.info(check_nhom_baiviet_lenlich == data['nhom']['lenlich'])
        time.sleep(2)




    def nhatkyhoatdong(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhatkyhoatdong).click()
        #Quản trị viên và người kiểm duyệt
        driver.find_element(By.XPATH, var.quantrivienvanguoikiemduyet).click()
        driver.find_element(By.XPATH, var.tranquangtruong).click()
        driver.implicitly_wait(2)
        time.sleep(2)
        try:
            check_nhatkyhoatdong_qtvvanguoikiemduyet_tranquangtruong = driver.find_element(By.XPATH,var.check_nhatkyhoatdong_dulieu1)
            logging.info("Nhóm - Nhóm của bạn - Nhật ký hoạt động")
            logging.info("check font-end: Quản trị viên và người kiểm duyệt - Trần Quang Trường")
            logging.info(check_nhatkyhoatdong_qtvvanguoikiemduyet_tranquangtruong.text)
            logging.info(check_nhatkyhoatdong_qtvvanguoikiemduyet_tranquangtruong.is_displayed())
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Nhật ký hoạt động")
            logging.info("check font-end: Quản trị viên và người kiểm duyệt - Trần Quang Trường")
            logging.info("False")
        driver.find_element(By.XPATH, var.xoaboloc).click()
        time.sleep(1)

        #Thành viên
        driver.find_element(By.XPATH, var.thanhvien).click()
        driver.find_element(By.XPATH, var.nhatkyhoatdong_thanhvien_timkiem).send_keys("Vương Lâm")
        time.sleep(3)
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.nhatkyhoatdong_thanhvien_timkiem_vuonglam)))
        element.click()
        time.sleep(2)
        try:
            check_nhatkyhoatdong_thanhvien_vuonglam = driver.find_element(By.XPATH,var.check_nhatkyhoatdong_dulieu1)
            logging.info("Nhóm - Nhóm của bạn - Nhật ký hoạt động")
            logging.info("check font-end: Thành viên - Vương Lâm")
            logging.info(check_nhatkyhoatdong_thanhvien_vuonglam.text)
            logging.info(check_nhatkyhoatdong_thanhvien_vuonglam.is_displayed())
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Nhật ký hoạt động")
            logging.info("check font-end: Thành viên - Vương Lâm")
            logging.info("False")
        driver.find_element(By.XPATH, var.xoaboloc).click()
        time.sleep(1)

        # Khác - Kiểm duyệt nội dung
        driver.find_element(By.XPATH, var.khac).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.kieuhoatdong).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.kiemduyetnoidung).click()
        time.sleep(2)
        try:
            check_nhatkyhoatdong_khac_kiemduyetnoidung = driver.find_element(By.XPATH,var.check_nhatkyhoatdong_dulieu1)
            logging.info("Nhóm - Nhóm của bạn - Nhật ký hoạt động")
            logging.info("check font-end: Khác - Kiểm duyệt nội dung")
            logging.info(check_nhatkyhoatdong_khac_kiemduyetnoidung.text)
            logging.info(check_nhatkyhoatdong_khac_kiemduyetnoidung.is_displayed())
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Nhật ký hoạt động")
            logging.info("check font-end: Khác - Kiểm duyệt nội dung")
            logging.info("False")


        # Khác - Thành viên
        driver.find_element(By.XPATH, var.kieuhoatdong).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.kiemduyetnoidung_thanhvien).click()
        time.sleep(2)
        try:
            check_nhatkyhoatdong_khac_thanhvien = driver.find_element(By.XPATH,var.check_nhatkyhoatdong_dulieu1)
            logging.info("Nhóm - Nhóm của bạn - Nhật ký hoạt động")
            logging.info("check font-end: Khác - Thành viên")
            logging.info(check_nhatkyhoatdong_khac_thanhvien.text)
            logging.info(check_nhatkyhoatdong_khac_thanhvien.is_displayed())
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Nhật ký hoạt động")
            logging.info("check font-end: Khác - thành viên")
            logging.info("False")


        # Khác - Cài đặt nhóm
        driver.find_element(By.XPATH, var.kieuhoatdong).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.kiemduyetnoidung_caidat).click()
        time.sleep(2)
        try:
            check_nhatkyhoatdong_khac_caidat = driver.find_element(By.XPATH,var.check_nhatkyhoatdong_dulieu1)
            logging.info("Nhóm - Nhóm của bạn - Nhật ký hoạt động")
            logging.info("check font-end: Khác - Cài đặt")
            logging.info(check_nhatkyhoatdong_khac_caidat.text)
            logging.info(check_nhatkyhoatdong_khac_caidat.is_displayed())
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Nhật ký hoạt động")
            logging.info("check font-end: Khác - Cài đặt")
            logging.info("False")
        time.sleep(2)



    def quytacnhom(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.quytacnhom).click()
        #Tạo quy tắc nhóm
        driver.find_element(By.XPATH, var.batdau).click()
        driver.find_element(By.XPATH, var.haytutelichsu).click()
        time.sleep(1)
        quytacnhom_haytutevalichsu_tieude = driver.find_element(By.XPATH, var.quytacnhom_haytutevalichsu_tieude).get_attribute("value")
        logging.info("Nhóm - Quy tắc nhóm")
        logging.info("check font-end: Tiêu đề Hãy tử tế và lịch sự - Gợi ý - Hãy tử tế và lịch sự")
        logging.info(quytacnhom_haytutevalichsu_tieude)
        logging.info(quytacnhom_haytutevalichsu_tieude == "Hãy tử tế và lịch sự")

        quytacnhom_haytutevalichsu_mota = driver.find_element(By.XPATH, var.quytacnhom_haytutevalichsu_mota).text
        logging.info("Nhóm - Quy tắc nhóm")
        logging.info("check font-end: Mô tả Hãy tử tế và lịch sự - Gợi ý -Tất cả chúng ta cùng có mặt ở đây để tạo nên một môi trường thân thiện. Hãy tôn trọng tất cả mọi người. Tranh luận lành mạnh là điều hết sức tự nhiên nhưng cũng cần tử tế.")
        logging.info(quytacnhom_haytutevalichsu_mota)
        logging.info(quytacnhom_haytutevalichsu_mota == "Tất cả chúng ta cùng có mặt ở đây để tạo nên một môi trường thân thiện. Hãy tôn trọng tất cả mọi người. Tranh luận lành mạnh là điều hết sức tự nhiên nhưng cũng cần tử tế.")

        driver.find_element(By.XPATH, var.tao).click()
        time.sleep(1)
        check_taoquytacnhom_haytutelichsu_tieude = driver.find_element(By.XPATH, var.check_taoquytacnhom_haytutelichsu_tieude).text
        logging.info("Nhóm - Quy tắc nhóm")
        logging.info("check font-end: Tạo quy tắc nhóm - " + quytacnhom_haytutevalichsu_tieude)
        logging.info(check_taoquytacnhom_haytutelichsu_tieude)
        logging.info(check_taoquytacnhom_haytutelichsu_tieude == quytacnhom_haytutevalichsu_tieude)

        check_taoquytacnhom_haytutelichsu_mota = driver.find_element(By.XPATH, var.check_taoquytacnhom_haytutelichsu_mota).text
        logging.info("Nhóm - Quy tắc nhóm")
        logging.info("check font-end: Tạo quy tắc nhóm - " + quytacnhom_haytutevalichsu_mota)
        logging.info(check_taoquytacnhom_haytutelichsu_mota)
        logging.info(check_taoquytacnhom_haytutelichsu_mota == quytacnhom_haytutevalichsu_mota)
        
        time.sleep(1)
        driver.find_element(By.XPATH, var.taoquytac).click()
        driver.find_element(By.XPATH, var.quytacnhom_haytutevalichsu_tieude).send_keys(data['nhom']['quytacnhom_tieude'])
        driver.find_element(By.XPATH, var.quytacnhom_haytutevalichsu_mota).send_keys(data['nhom']['quytacnhom_mota'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.tao).click()
        time.sleep(1)
        
        #Check quy tắc nhóm - Trang chủ - Giới thiệu
        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        driver.find_element(By.XPATH, var.gioithieu).click()
        driver.execute_script("window.scrollBy(0,700)", "")
        check_nhom_gioithieu_quytacnhom_tieude = driver.find_element(By.XPATH,var.check_nhom_gioithieu_quytacnhom_tieude).text
        logging.info("Nhóm - Quy tắc nhóm")
        logging.info("check font-end: Trang chủ - Giới thiệu - Quy tắc nhóm - Tiêu đề - " + check_taoquytacnhom_haytutelichsu_tieude)
        logging.info(check_nhom_gioithieu_quytacnhom_tieude[1::])
        logging.info(check_nhom_gioithieu_quytacnhom_tieude[1::] == check_taoquytacnhom_haytutelichsu_tieude)

        check_nhom_gioithieu_quytacnhom_mota = driver.find_element(By.XPATH,var.check_nhom_gioithieu_quytacnhom_mota).text
        logging.info("Nhóm - Quy tắc nhóm")
        logging.info("check font-end: Trang chủ - Giới thiệu - Quy tắc nhóm - Mô tả - " + check_taoquytacnhom_haytutelichsu_mota)
        logging.info(check_nhom_gioithieu_quytacnhom_mota)
        logging.info(check_nhom_gioithieu_quytacnhom_mota == check_taoquytacnhom_haytutelichsu_mota)
        time.sleep(2)

        #Xóa quy tắc nhóm
        driver.find_element(By.XPATH, var.quytacnhom).click()
        driver.find_element(By.XPATH, var.quytacnhom_quytac1_dau3cham).click()
        driver.find_element(By.XPATH, var.xoaquytac).click()
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(2)
        check_taoquytacnhom_haytutelichsu_xoa = driver.find_element(By.XPATH, var.check_taoquytacnhom_haytutelichsu_tieude).text
        logging.info("Nhóm - Quy tắc nhóm")
        logging.info("check font-end: Xóa quy tắc - Hãy tử tế và lịch sự")
        logging.info(check_taoquytacnhom_haytutelichsu_xoa != "Hãy tử tế và lịch sự")

        #Chỉnh sửa quy tắc
        driver.find_element(By.XPATH, var.quytacnhom_quytac1_dau3cham).click()
        driver.find_element(By.XPATH, var.chinhsuaquytac).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.quytacnhom_chinhsuaquytac_tieude)
        xoa.send_keys(Keys.CONTROL, "a")

        driver.find_element(By.XPATH, var.quytacnhom_chinhsuaquytac_tieude).send_keys(data['nhom']['quytacnhom_tieudesua'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(2)
        check_taoquytacnhom_haytutelichsu_tieudesua = driver.find_element(By.XPATH, var.check_taoquytacnhom_haytutelichsu_tieude).text
        logging.info("Nhóm - Quy tắc nhóm")
        logging.info("check font-end: Chỉnh sửa quy tắc nhóm - " + data['nhom']['quytacnhom_tieudesua'])
        logging.info(check_taoquytacnhom_haytutelichsu_tieude)
        logging.info(check_taoquytacnhom_haytutelichsu_tieudesua)
        logging.info(check_taoquytacnhom_haytutelichsu_tieudesua == data['nhom']['quytacnhom_tieudesua'])

        driver.find_element(By.XPATH, var.quytacnhom_quytac1_dau3cham).click()
        driver.find_element(By.XPATH, var.xoaquytac).click()
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(2)


    def noidungthanhvienbaocao(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        #Tạo post 1
        driver.find_element(By.XPATH, var.nhom_taobaiviet1).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys("Post này để test báo cáo quản trị viên - Giữ lại")
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(2.5)

        #Nội dung thành viên báo cáo - Giữ lại
        #Tạo post 2
        driver.find_element(By.XPATH, var.nhom_taobaiviet1).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham).click()
        driver.find_element(By.XPATH, var.trang_taobaiviet_dau3cham_anhvideo).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['taobaiviet_anh_mota'])
        driver.find_element(By.XPATH, var.taobaiviet_tailenanhvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien3.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.dang).click()
        wait = WebDriverWait(driver, 10)
        check_nhom_taobaiviet_messsage = wait.until(EC.element_to_be_clickable((By.XPATH, var.check_nhom_taobaiviet_message))).text
        driver.execute_script("window.scrollBy(0,700)", "")
        driver.find_element(By.XPATH, var.nhom_baiviet_binhluan).click()
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(data['nhom']['noidungbaocao'])
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(Keys.ENTER)
        time.sleep(2)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_baiviet_dau3cham).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocaovoiquantrivien).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.noidungkhonglienquan).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.xong).click()
        time.sleep(1)
        driver.refresh()
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            check_noidungthanhvienbaocao_thongbao = driver.find_element(By.XPATH,var.check_noidungthanhvienbaocao_thongbao)
            logging.info("Nhóm - Nhóm của bạn - Nội dung thành viên báo cáo")
            logging.info("check font-end: Có thông báo ở module hay không")
            logging.info(check_noidungthanhvienbaocao_thongbao.is_displayed())
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Nội dung thành viên báo cáo")
            logging.info("check font-end: Có thông báo ở module hay không")
            logging.info("False")
        driver.find_element(By.XPATH, var.noidungthanhvienbaocao).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.noidungthanhvienbaocao_guilai1).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        time.sleep(2)
        check_nhom_noidungthanhvienbaocao_giulai = driver.find_element(By.XPATH, var.check_nhom_baiviet_camxuchoatdong_mota).text
        logging.info("Nhóm - Nhóm của bạn - Nội dung thành viên báo cáo")
        logging.info("check font-end: Giữ lại - Có giữ lại bài viết vừa báo cáo không")
        logging.info(check_nhom_noidungthanhvienbaocao_giulai)
        logging.info(check_nhom_noidungthanhvienbaocao_giulai == data['nhom']['taobaiviet_anh_mota'])

        #Nội dung thành viên báo cáo - Gỡ bài
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        #Tạo post 1
        driver.find_element(By.XPATH, var.nhom_taobaiviet1).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['noidungbaocao_gobai'])
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(2.5)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_baiviet_dau3cham).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocaovoiquantrivien).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.noidungkhonglienquan).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.xong).click()
        time.sleep(1)
        driver.refresh()
        time.sleep(2)
        driver.find_element(By.XPATH, var.noidungthanhvienbaocao).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.noidungthanhvienbaocao_gobai1).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        time.sleep(2)
        check_nhom_noidungthanhvienbaocao_gobai = driver.find_element(By.XPATH, var.check_nhom_baiviet_camxuchoatdong_mota).text
        if check_nhom_noidungthanhvienbaocao_gobai == data['nhom']['noidungbaocao_gobai']:
            logging.info("Nhóm - Nhóm của bạn - Nội dung thành viên báo cáo")
            logging.info("check font-end: Gỡ bài - Có gỡ bài viết vừa báo cáo không")
            logging.info("False")
        else:
            logging.info("Nhóm - Nhóm của bạn - Nội dung thành viên báo cáo")
            logging.info("check font-end: Gỡ bài - Có gỡ bài viết vừa báo cáo không")
            logging.info("True")


        #Nội dung thành viên báo cáo - Dấu 3 chấm
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        #Tạo post 1
        driver.find_element(By.XPATH, var.nhom_taobaiviet1).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['noidungbaocao_dau3cham'])
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(2.5)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_baiviet_dau3cham).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.baocaovoiquantrivien).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.noidungkhonglienquan).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.xong).click()
        time.sleep(1)
        driver.refresh()
        time.sleep(2)
        driver.find_element(By.XPATH, var.noidungthanhvienbaocao).click()
        time.sleep(1)

        #Nội dung thành viên báo cáo - Dấu 3 chấm - Sao chép liên kết để chia sẻ với quản trị viên
        driver.find_element(By.XPATH, var.noidungthanhvienbaocao_dau3cham).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.noidungthanhvienbaocao_dau3cham_saocheplienket).click()
        time.sleep(1)

        #Nội dung thành viên báo cáo - Dấu 3 chấm - Gỡ bài viết và chặn tác giả
        driver.find_element(By.XPATH, var.noidungthanhvienbaocao_dau3cham).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.noidungthanhvienbaocao_dau3cham_gobaivietvachantacgia).click()
        time.sleep(3)
        driver.refresh()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        check_nhom_noidungthanhvienbaocao_gobaivachantacgia = driver.find_element(By.XPATH, var.check_nhom_baiviet_camxuchoatdong_mota).text
        if check_nhom_noidungthanhvienbaocao_gobaivachantacgia == data['nhom']['noidungbaocao_dau3cham']:
            logging.info("Nhóm - Nhóm của bạn - Nội dung thành viên báo cáo")
            logging.info("check font-end: Gỡ bài viết và chặn tác giả - Có gỡ bài viết vừa báo cáo không")
            logging.info("False")
        else:
            logging.info("Nhóm - Nhóm của bạn - Nội dung thành viên báo cáo")
            logging.info("check font-end: Gỡ bài viết và chặn tác giả - Có gỡ bài viết vừa báo cáo không")
            logging.info("True")

        driver.find_element(By.XPATH, var.nhom_thanhvien).click()
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            check_nhom_thanhvien_thanhvienbichan = driver.find_element(By.XPATH, var.check_nhom_thanhvien_thanhvienbichan)
            logging.info("Nhóm - Nhóm của bạn - Nội dung thành viên báo cáo")
            logging.info("check font-end: Gỡ bài viết và chặn tác giả - Danh sách thành viên bị chặn Có hiển thị không")
            logging.info(check_nhom_thanhvien_thanhvienbichan.is_displayed())
            logging.info(check_nhom_thanhvien_thanhvienbichan.text)
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Nội dung thành viên báo cáo")
            logging.info("check font-end: Gỡ bài viết và chặn tác giả - Danh sách thành viên bị chặn Có hiển thị không")
            logging.info("False")
        button = driver.find_element(By.XPATH, var.nhom_thanhvien_dsbichan_dau3cham)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.bochan).click()
        time.sleep(2)


    def thongbaokiemduyet(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_thongbaokiemduyet).click()
        driver.find_element(By.XPATH, var.chinhsuathongbao).click()
        time.sleep(1)
        #Xóa từ khóa
        driver.find_element(By.XPATH, var.nhom_thongbaokiemduyet_xoatukhoa1).click()
        time.sleep(1)
        driver.implicitly_wait(2)
        try:
            check_nhom_thongbaokiemduyet_xoatukhoa1 = driver.find_element(By.XPATH, var.nhom_thongbaokiemduyet_xoatukhoa1).is_displayed()
            logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
            logging.info("check font-end: Xóa từ khóa 1")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
            logging.info("check font-end: Xóa từ khóa 1")
            logging.info("True")

        #Thêm từ khóa
        driver.find_element(By.XPATH, var.nhom_thongbaokiemduyet_themtukhoainput).send_keys(data['nhom']['thongbaokiemduyet'])
        driver.find_element(By.XPATH, var.them).click()
        time.sleep(1)
        try:
            check_nhom_thongbaokiemduyet_themtukhoa1 = driver.find_element(By.XPATH, var.check_nhom_thongbaokiemduyet_themtukhoa1).text
            logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
            logging.info("check font-end: Thêm từ khóa 1 - " + data['nhom']['thongbaokiemduyet'])
            logging.info(check_nhom_thongbaokiemduyet_themtukhoa1)
            logging.info(data['nhom']['thongbaokiemduyet'])
            logging.info(check_nhom_thongbaokiemduyet_themtukhoa1 == data['nhom']['thongbaokiemduyet'])
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
            logging.info("check font-end: Thêm từ khóa 1 - " + data['nhom']['thongbaokiemduyet'])
            logging.info("False")
        time.sleep(1)

        driver.implicitly_wait(15)
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        #Tạo post 1
        driver.find_element(By.XPATH, var.nhom_taobaiviet1).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys("Post này để test thông báo kiểm duyệt - Giữ lại comment")
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_baiviet_binhluan).click()
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(data['nhom']['thongbaokiemduyet'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(Keys.ENTER)
        time.sleep(2)

        check_nhom_baiviet_binhluan_thongbaokiemduyet = driver.find_element(By.XPATH,var.nhom_baiviet_binhluan_thongbaokiemduyet).text
        logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
        logging.info("check font-end: Bình luận vào bài viết  - " + data['nhom']['thongbaokiemduyet'])
        logging.info(check_nhom_baiviet_binhluan_thongbaokiemduyet)
        logging.info(check_nhom_baiviet_binhluan_thongbaokiemduyet == data['nhom']['thongbaokiemduyet'])

        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879/report_censorship")
        time.sleep(2)
        driver.implicitly_wait(3)
        try:
            check_thongbaokiemduyet_thongbao = driver.find_element(By.XPATH,var.check_noidungthanhvienbaocao_thongbao)
            logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
            logging.info("check font-end: Có thông báo ở module hay không")
            logging.info(check_thongbaokiemduyet_thongbao.is_displayed())
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
            logging.info("check font-end: Có thông báo ở module hay không")
            logging.info("False")

        check_thongbaokiemduyet_tukhoabaocao = driver.find_element(By.XPATH,var.check_thongbaokiemduyet_tukhoabaocao).text
        logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
        logging.info("check font-end: Từ khóa được báo cáo  - " + data['nhom']['thongbaokiemduyet'])
        logging.info(check_thongbaokiemduyet_tukhoabaocao)
        logging.info(check_thongbaokiemduyet_tukhoabaocao == data['nhom']['thongbaokiemduyet'])

        #Thông báo kiểm duyệt - Giữ lại
        driver.find_element(By.XPATH, var.giulai).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_baiviet_binhluan).click()
        try:
            check_thongbaokiemduyet_giulai = driver.find_element(By.XPATH,var.nhom_baiviet_binhluan_thongbaokiemduyet).text
            logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
            logging.info("check font-end: Giữ lại - " + data['nhom']['thongbaokiemduyet'])
            logging.info(check_thongbaokiemduyet_giulai)
            logging.info(check_thongbaokiemduyet_giulai == data['nhom']['thongbaokiemduyet'])
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
            logging.info("check font-end: Giữ lại - " + data['nhom']['thongbaokiemduyet'])
            logging.info("False")

        #Thông báo kiểm duyệt - Gỡ bình luận
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        #Tạo post 1
        driver.find_element(By.XPATH, var.nhom_taobaiviet1).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys("Post này để test thông báo kiểm duyệt - Gỡ commnet")
        driver.find_element(By.XPATH, var.dang).click()
        time.sleep(2.5)
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_baiviet_binhluan).click()
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(data['nhom']['thongbaokiemduyet'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(Keys.ENTER)
        time.sleep(2)
        check_nhom_baiviet_binhluan_thongbaokiemduyet = driver.find_element(By.XPATH,var.nhom_baiviet_binhluan_thongbaokiemduyet)

        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879/report_censorship")
        time.sleep(2)
        driver.find_element(By.XPATH, var.gobinhluan).click()
        time.sleep(3)
        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        time.sleep(2)
        check_nhom_baiviet_mota = driver.find_element(By.XPATH, var.check_nhom_baiviet_mota).text
        driver.implicitly_wait(3)
        if check_nhom_baiviet_mota == "Post này để test thông báo kiểm duyệt - Gỡ commnet":
            driver.find_element(By.XPATH, var.nhom_baiviet_binhluan).click()
            try:
                check_thongbaokiemduyet_gobinhluan = driver.find_element(By.XPATH,var.nhom_baiviet_binhluan_thongbaokiemduyet).is_displayed()
                logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
                logging.info("check font-end: Gỡ bình luận - " + data['nhom']['thongbaokiemduyet'])
                logging.info("False")
            except NoSuchElementException:
                logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
                logging.info("check font-end: Gỡ bình luận - " + data['nhom']['thongbaokiemduyet'])
                logging.info("True")
        else:
            logging.info("Nhóm - Nhóm của bạn - Thông báo kiểm duyệt")
            logging.info("check font-end: Gỡ bình luận - Có còn bài viết không?")
            logging.info("False")
        time.sleep(2)

    def cauhoichonthanhvien(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_cauhoichonthanhvien).click()

        #Thêm câu hỏi 1 - Ô để đánh dấu
        driver.find_element(By.XPATH, var.themcauhoi).click()
        driver.find_element(By.XPATH, var.taomoicauhoi_input).send_keys(data['nhom']['odedanhdau'])
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon).click()
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon1).send_keys(data['nhom']['odedanhdau_luachon1'])
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon).click()
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon2).send_keys(data['nhom']['odedanhdau_luachon2'])
        driver.find_element(By.XPATH, var.tao).click()
        time.sleep(1)
        
        check_cauhoichonthanhvien_tencauhoi1 = driver.find_element(By.XPATH,var.check_cauhoichonthanhvien_tencauhoi1).text
        logging.info("Nhóm - Nhóm của bạn - Câu hỏi chọn thành viên")
        logging.info("check font-end: Câu hỏi 1 - " + data['nhom']['odedanhdau'])
        logging.info(check_cauhoichonthanhvien_tencauhoi1)
        logging.info(check_cauhoichonthanhvien_tencauhoi1 == data['nhom']['odedanhdau'])

        check_cauhoichonthanhvien_cautraloicauhoi1 = driver.find_element(By.XPATH,var.check_cauhoichonthanhvien_cautraloicauhoi1).text
        logging.info("Nhóm - Nhóm của bạn - Câu hỏi chọn thành viên")
        logging.info("check font-end: Câu trả lời của câu hỏi 1 - " + data['nhom']['odedanhdau_luachon1'])
        logging.info(check_cauhoichonthanhvien_cautraloicauhoi1)
        logging.info(check_cauhoichonthanhvien_cautraloicauhoi1 == data['nhom']['odedanhdau_luachon1'])

        #Chỉnh sửa câu hỏi
        driver.find_element(By.XPATH, var.chinhsua).click()
        xoa = driver.find_element(By.XPATH, var.taomoicauhoi_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.taomoicauhoi_input).send_keys(data['nhom']['odedanhdau_sua'])
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon).click()
        xoa = driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon1)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon1).send_keys(data['nhom']['odedanhdau_luachon1_sua'])
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon).click()
        xoa = driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon2)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon2).send_keys(data['nhom']['odedanhdau_luachon2_sua'])
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon).click()
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon3).send_keys(data['nhom']['odedanhdau_luachon3_sua'])
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon3_x).click()
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon3_x).click()
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon3_x).click()
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)

        #Thêm câu hỏi 2 - Trắc nghiệm
        driver.find_element(By.XPATH, var.taocauhoi).click()
        driver.find_element(By.XPATH, var.odedanhdau).click()
        driver.find_element(By.XPATH, var.tracnghiem).click()
        driver.find_element(By.XPATH, var.taomoicauhoi_input).send_keys(data['nhom']['trachnghiem'])
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon).click()
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon1).send_keys(data['nhom']['trachnghiem_luachon1'])
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon).click()
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_themluachon_luachon2).send_keys(data['nhom']['trachnghiem_luachon2'])
        driver.find_element(By.XPATH, var.tao).click()

        #Thêm câu hỏi 3 - Trắc nghiệm
        driver.find_element(By.XPATH, var.taocauhoi).click()
        driver.find_element(By.XPATH, var.odedanhdau).click()
        driver.find_element(By.XPATH, var.cauhoitudo).click()
        driver.find_element(By.XPATH, var.taomoicauhoi_input).send_keys(data['nhom']['cauhoitudo'])
        driver.find_element(By.XPATH, var.tao).click()
        time.sleep(1)

        #Login check câu hỏi
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        driver.find_element(By.XPATH, var.thamgianhom).click()
        time.sleep(2)
        check_thamgianhom_cauhoichonthanhvien_tencauhoi1 = driver.find_element(By.XPATH,var.check_thamgianhom_cauhoichonthanhvien_tencauhoi1).text
        check_thamgianhom_cauhoichonthanhvien_tencauhoi2 = driver.find_element(By.XPATH,var.check_thamgianhom_cauhoichonthanhvien_tencauhoi2).text
        check_thamgianhom_cauhoichonthanhvien_tencauhoi3 = driver.find_element(By.XPATH,var.check_thamgianhom_cauhoichonthanhvien_tencauhoi3).text

        logging.info("Nhóm - Nhóm của bạn - Câu hỏi chọn thành viên")
        logging.info("check font-end:  Tên câu hỏi 1 - " + data['nhom']['odedanhdau_sua'])
        logging.info(check_thamgianhom_cauhoichonthanhvien_tencauhoi1[11::]== data['nhom']['odedanhdau_sua'])

        logging.info("check font-end:  Tên câu hỏi 2 - " + data['nhom']['trachnghiem'])
        logging.info(check_thamgianhom_cauhoichonthanhvien_tencauhoi2[11::]== data['nhom']['trachnghiem'])

        logging.info("check font-end:  Tên câu hỏi 3 - " + data['nhom']['cauhoitudo'])
        logging.info(check_thamgianhom_cauhoichonthanhvien_tencauhoi3[11::]== data['nhom']['cauhoitudo'])

        driver.find_element(By.XPATH, var.traloicauhoi_cautraloi1).click()
        driver.find_element(By.XPATH, var.traloicauhoi_cautraloi2).click()
        driver.find_element(By.XPATH, var.traloicauhoi_cautraloi3).send_keys(data['nhom']['traloicauhoi_cautraloi3'])
        driver.find_element(By.XPATH, var.gui).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.dathamgia).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.roikhoinhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.roikhoinhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.thamgianhom)
        time.sleep(1)

        #Xóa câu hỏi
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879/member_question")
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_cauhoi1_xoa).click()
        driver.find_element(By.XPATH, var.xoacauhoi_xoa).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_cauhoi1_xoa).click()
        driver.find_element(By.XPATH, var.xoacauhoi_xoa).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.cauhoichonthanhvien_cauhoi1_xoa).click()
        driver.find_element(By.XPATH, var.xoacauhoi_xoa).click()
        time.sleep(1)
        driver.implicitly_wait(2)
        try:
            check_cauhoichonthanhvien_cauhoi1_xoa = driver.find_element(By.XPATH,var.cauhoichonthanhvien_cauhoi1_xoa).is_displayed()
            logging.info("Nhóm - Nhóm của bạn - Câu hỏi chọn thành viên")
            logging.info("check font-end: Có xóa được câu hỏi chọn thành viên không?")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Câu hỏi chọn thành viên")
            logging.info("check font-end: Có xóa được câu hỏi chọn thành viên không?")
            logging.info("True")
        time.sleep(2)


    def caidatnhom(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.check_nhom_nhombanquanly_namtest).click()
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_caidatnhom).click()

    def thietlapnhom(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        #Tên và mô tả
        driver.find_element(By.XPATH, var.nhom_caidatnhom_tenvamota).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.tenvamota_tennhom)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.tenvamota_tennhom).send_keys(data['nhom']['tenvamota_tennhom'])

        xoa = driver.find_element(By.XPATH, var.tenvamota_nhomnaynoive)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.tenvamota_nhomnaynoive).send_keys(data['nhom']['tenvamota_nhomnaynoive'])
        driver.find_element(By.XPATH, var.luu).click()

        check_caidatnhom_tenvamota_message = driver.find_element(By.XPATH,var.check_caidatnhom_tenvamota_message).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message khi lưu Tên và mô tả - Cập nhật dữ liệu thành công")
        logging.info(check_caidatnhom_tenvamota_message == "Cập nhật dữ liệu thành công")
        time.sleep(1)

        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        driver.find_element(By.XPATH, var.gioithieu).click()
        driver.execute_script("window.scrollBy(0,700)", "")
        check_nhom_gioithieu_motanhom = driver.find_element(By.XPATH,var.check_nhom_gioithieu_motanhom).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end: Trang chủ - Giới thiệu - Mô tả nhóm - " + data['nhom']['tenvamota_nhomnaynoive'])
        logging.info(check_nhom_gioithieu_motanhom)
        logging.info(check_nhom_gioithieu_motanhom == data['nhom']['tenvamota_nhomnaynoive'])

        check_nhom_tennhom = driver.find_element(By.XPATH,var.check_nhom_tennhom).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end: Đổi tên nhóm - " + data['nhom']['tenvamota_tennhom'])
        logging.info(check_nhom_tennhom)
        logging.info(check_nhom_tennhom == data['nhom']['tenvamota_tennhom'])
        time.sleep(1)

        #Sửa lại tên nhóm
        driver.find_element(By.XPATH, var.nhom_caidatnhom).click()
        driver.find_element(By.XPATH, var.nhom_caidatnhom_tenvamota).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.tenvamota_tennhom)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.tenvamota_tennhom).send_keys(data['nhom']['tenvamota_tennhom1'])

        xoa = driver.find_element(By.XPATH, var.tenvamota_nhomnaynoive)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.tenvamota_nhomnaynoive).send_keys(data['nhom']['tenvamota_nhomnaynoive1'])
        driver.find_element(By.XPATH, var.luu).click()
        check_caidatnhom_tenvamota_message = driver.find_element(By.XPATH,var.check_caidatnhom_tenvamota_message).text
        time.sleep(1)

        #Danh mục và tag
        driver.find_element(By.XPATH, var.nhom_caidatnhom_danhmucvatag).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.nhom_caidatnhom_danhmuc)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.nhom_caidatnhom_danhmuc).send_keys("Âm nhạc")
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.amnhac)))
        element.click()

        xoa = driver.find_element(By.XPATH, var.nhom_caidatnhom_tag)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.nhom_caidatnhom_tag).send_keys(data['nhom']['danhmucvatag_tag'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        check_caidatnhom_danhmucvatag_message = driver.find_element(By.XPATH,var.check_caidatnhom_tenvamota_message).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message khi lưu Danh mục và tag - Cập nhật dữ liệu thành công")
        logging.info(check_caidatnhom_danhmucvatag_message == "Cập nhật dữ liệu thành công")
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_danhmucvatag).click()
        time.sleep(1)

        check_caidatnhom_danhmucvatag_danhmuc = driver.find_element(By.XPATH,var.check_caidatnhom_danhmucvatag_danhmuc).get_attribute("value")
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end: Danh mục và tag - Danh mục - Âm nhạc")
        logging.info(check_caidatnhom_danhmucvatag_danhmuc)
        logging.info(check_caidatnhom_danhmucvatag_danhmuc == "Âm nhạc")

        check_caidatnhom_danhmucvatag_tag = driver.find_element(By.XPATH,var.check_caidatnhom_danhmucvatag_tag).get_attribute("value")
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end: Danh mục và tag - Tag - " + data['nhom']['danhmucvatag_tag'])
        logging.info(check_caidatnhom_danhmucvatag_tag)
        logging.info(check_caidatnhom_danhmucvatag_tag == data['nhom']['danhmucvatag_tag'])
        time.sleep(1)

        #Sửa lại Danh mục và tag
        xoa = driver.find_element(By.XPATH, var.nhom_caidatnhom_danhmuc)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.nhom_caidatnhom_danhmuc).send_keys("Làm đẹp")
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, var.lamdep)))
        element.click()

        xoa = driver.find_element(By.XPATH, var.nhom_caidatnhom_tag)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.nhom_caidatnhom_tag).send_keys(data['nhom']['danhmucvatag_tag_sua'])
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        check_caidatnhom_danhmucvatag_message = driver.find_element(By.XPATH,var.check_caidatnhom_tenvamota_message).text
        time.sleep(1)

        #Quyền riêng tư - Công khai
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        check_nhomcongkhai_quyenriengtu = driver.find_element(By.XPATH,var.check_nhom_quyenriengtu).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm - Nhóm công khai")
        logging.info("check font-end: Quyền riêng tư nam test - Nhóm công khai")
        logging.info(check_nhomcongkhai_quyenriengtu)
        logging.info(check_nhomcongkhai_quyenriengtu == "Nhóm công khai")

        #Tạo post khi chưa tham gia nhóm
        driver.find_element(By.XPATH, var.nhom_taobaiviet2).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys("Post này để test thành viên chưa tham gia nhom nhưng tạo bài viết")
        driver.find_element(By.XPATH, var.dang).click()
        check_nhomcongkhai_dangbailoi = driver.find_element(By.XPATH,var.check_nhomcongkhai_dangbailoi).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm - Nhóm công khai")
        logging.info("check font-end: Message Đăng bài khi chưa tham gia nhóm - Vui lòng tham gia nhóm để có thể đăng nội dung.")
        logging.info(check_nhomcongkhai_dangbailoi == "Vui lòng tham gia nhóm để có thể đăng nội dung.")
        time.sleep(1)
        driver.find_element(By.XPATH, var.icon_x).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.icon_x).click()

        #Bình luận khi chưa tham gia nhóm
        driver.find_element(By.XPATH, var.nhom_baiviet_binhluan).click()
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys("chưa tham gia nhóm nhưng vẫn bình luận được")
        driver.find_element(By.XPATH, var.vietbinhluan).send_keys(Keys.ENTER)
        time.sleep(2)

        #Quyền riêng tư - Riêng tư - Ẩn nhóm
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111561259300587836/setting_group")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_quyenriengtu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_quyenriengtu_chon).click()
        driver.find_element(By.XPATH, var.nhom_caidatnhom_quyenriengtu_chon_annhom).click()
        driver.find_element(By.XPATH, var.luu).click()
        check_caidatnhom_quyenriengtu_message = driver.find_element(By.XPATH,var.check_caidatnhom_tenvamota_message).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message khi lưu quyền riêng tư ẩn nhóm - Cập nhật dữ liệu thành công")
        logging.info(check_caidatnhom_quyenriengtu_message == "Cập nhật dữ liệu thành công")
        time.sleep(1)

        #Login để tìm kiếm nhóm riêng tư ẩn
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Tìm kiếm
        driver.find_element(By.XPATH, var.nhom_timkiem).send_keys(data['nhom']['nhomriengtu_timkiem'])
        driver.find_element(By.XPATH, var.nhom_timkiem).send_keys(Keys.ENTER)
        check_nhom_riengtu_an_timkiem = driver.find_element(By.XPATH,var.check_nhom_timkiem).text
        print(check_nhom_riengtu_an_timkiem)
        logging.info("Nhóm - Tìm kiếm- Nhóm riêng tư")
        logging.info("check font-end: Tìm kiếm nhóm ẩn - "+ data['nhom']['nhomriengtu_timkiem'])
        logging.info(check_nhom_riengtu_an_timkiem != data['nhom']['nhomriengtu_timkiem'])
        time.sleep(1)
        #Chuyển  tới trang riêng tư- ẩn
        driver.get("https://cmc-fe.emso.vn/group/111561259300587836")
        time.sleep(2)
        check_nhom_riengtu_an_linktoitrang = driver.find_element(By.XPATH,var.nhomnaykhonghienthi).text
        print(check_nhom_riengtu_an_linktoitrang)
        logging.info("Nhóm - Tìm kiếm- Nhóm riêng tư")
        logging.info("check font-end: Link tới nhóm riêng tư ẩn - Nhóm này không hiển thị")
        logging.info(check_nhom_riengtu_an_linktoitrang == "Nhóm này không hiển thị")
        time.sleep(1)

        # Quyền riêng tư - Riêng tư - Ẩn nhóm
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111561259300587836/setting_group")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_quyenriengtu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_quyenriengtu_chon).click()
        driver.find_element(By.XPATH, var.nhom_caidatnhom_quyenriengtu_chon_hienthi).click()
        driver.find_element(By.XPATH, var.luu).click()
        check_caidatnhom_quyenriengtu_message = driver.find_element(By.XPATH,var.check_caidatnhom_tenvamota_message).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message khi lưu quyền riêng tư hiển thị - Cập nhật dữ liệu thành công")
        logging.info(check_caidatnhom_quyenriengtu_message == "Cập nhật dữ liệu thành công")
        time.sleep(1)
        #Login để tìm kiếm nhóm riêng tư hiển thị
        login.login4(self, "truongvck22@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.icon_nhom)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Tìm kiếm
        driver.find_element(By.XPATH, var.nhom_timkiem).send_keys(data['nhom']['nhomriengtu_timkiem'])
        driver.find_element(By.XPATH, var.nhom_timkiem).send_keys(Keys.ENTER)
        check_nhom_riengtu_an_timkiem = driver.find_element(By.XPATH,var.check_nhom_timkiem).text
        print(check_nhom_riengtu_an_timkiem)
        logging.info("Nhóm - Tìm kiếm- Nhóm riêng tư")
        logging.info("check font-end: Tìm kiếm nhóm hiển thị - "+ data['nhom']['nhomriengtu_timkiem'])
        logging.info(check_nhom_riengtu_an_timkiem == data['nhom']['nhomriengtu_timkiem'])
        time.sleep(1)
        #Chuyển  tới trang riêng tư- hiển thị
        driver.find_element(By.XPATH, var.check_nhom_timkiem).click()
        time.sleep(2)
        check_nhom_riengtu_hienthi_linktoitrang = driver.find_element(By.XPATH,var.check_nhom_riengtu_hienthi_linktoitrang).text
        print(check_nhom_riengtu_hienthi_linktoitrang)
        logging.info("Nhóm - Tìm kiếm- Nhóm riêng tư")
        logging.info("check font-end: Chọn nhóm riêng tư hiển thị - Đây là nhóm riêng tư")
        logging.info(check_nhom_riengtu_hienthi_linktoitrang == "Đây là nhóm riêng tư")
        time.sleep(1)


    def tuychinhnhom(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879/setting_group")
        time.sleep(2)
        #Tùy chỉnh nhóm
        driver.find_element(By.XPATH, var.nhom_caidatnhom_diachiweb).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_diachiweb_input).send_keys("namtest")
        driver.find_element(By.XPATH, var.luu).click()
        check_caidatnhom_diachiweb_message = driver.find_element(By.XPATH,var.check_caidatnhom_tenvamota_message).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message khi lưu địa chỉ web - Cập nhật dữ liệu thành công")
        logging.info(check_caidatnhom_diachiweb_message == "Cập nhật dữ liệu thành công")
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_diachiweb).click()
        time.sleep(1)

        check_nhom_caidat_diachiweb = driver.find_element(By.XPATH,var.check_nhom_caidat_diachiweb).text
        print(check_nhom_caidat_diachiweb)
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end: Địa chỉ web sau khi lưu - https://cmc-fe.emso.vn/group/namtest")
        logging.info(check_nhom_caidat_diachiweb)
        logging.info(check_nhom_caidat_diachiweb == "https://cmc-fe.emso.vn/group/namtest")

        #check link sau khi thay đổi địa chỉ web
        driver.get("https://cmc-fe.emso.vn/group/namtest")
        time.sleep(2)
        driver.implicitly_wait(5)
        try:
            check_nhom_tennhom = driver.find_element(By.XPATH,var.check_nhom_tennhom).text
            print(check_nhom_tennhom)
            logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
            logging.info("check font-end: Chuyển tới Địa chỉ web sau khi lưu - Tên nhóm - nam test")
            logging.info(check_nhom_tennhom)
            logging.info(check_nhom_tennhom == "nam test")
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
            logging.info("check font-end: Chuyển tới Địa chỉ web sau khi lưu - Tên nhóm - nam test")
            logging.info("False")
        driver.implicitly_wait(15)

        driver.get("https://cmc-fe.emso.vn/group/namtest/setting_group")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_diachiweb).click()
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.nhom_caidatnhom_diachiweb_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.nhom_caidatnhom_diachiweb_input).send_keys("111504666936394879")
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(2)


    def quanlynoidungthaoluan(self):
        driver.implicitly_wait(15)
        #Quản lý nội dung thảo luận
        driver.find_element(By.XPATH, var.nhom_caidatnhom_quanlynoidungthaoluan).click()
        time.sleep(1)
        #Ai có thể đăng - Chỉ có quản trị viên
        driver.find_element(By.XPATH, var.nhom_caidatnhom_aicothedang).click()
        driver.find_element(By.XPATH, var.nhom_caidatnhom_aicothedang_chicoquantrivien).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        check_caidatnhom_aicothedang_message = driver.find_element(By.XPATH,var.check_caidatnhom_tenvamota_message).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message khi lưu Ai có thể đăng - Chỉ có quản trị viên - Cập nhật dữ liệu thành công")
        logging.info(check_caidatnhom_aicothedang_message == "Cập nhật dữ liệu thành công")
        time.sleep(1)

        #Check Ai có thể đăng - Chỉ có quản trị viên
        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        time.sleep(1)
        check_aicothedang_chicoqtv_qtv = driver.find_element(By.XPATH,var.check_aicothedang_chicoqtv_qtv).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end: Ai có thể đăng - Chỉ có quản trị viên - QTV check tạo bài viết")
        logging.info(check_aicothedang_chicoqtv_qtv == "Mai ơi, bạn đang nghĩ gì thế?")

        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        driver.implicitly_wait(3)
        try:
            check_aicothedang_chicoqtv_qtv = driver.find_element(By.XPATH, var.check_aicothedang_chicoqtv_thanhvien).is_displayed()
            logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
            logging.info("check font-end: Ai có thể đăng - Chỉ có quản trị viên - Thành viên nhóm check tạo bài viết")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
            logging.info("check font-end: Ai có thể đăng - Chỉ có quản trị viên - Thành viên nhóm check tạo bài viết")
            logging.info("True")


        #Ai có thể đăng - Bất cứ ai trong nhóm - Phê duyệt mọi bài viết của thành viên - Bài viết của thành viên phải được quản trị viên hoặc người kiểm duyệt phê duyệt
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879/setting_group")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_quanlynoidungthaoluan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_aicothedang).click()
        driver.find_element(By.XPATH, var.nhom_caidatnhom_aicothedang_batcuaitrongnhom).click()
        time.sleep(1)
        # Ai có thể đăng - Bất cứ ai trong nhóm - Phê duyệt mọi bài viết của thành viên
        driver.find_element(By.XPATH, var.nhom_caidatnhom_pheduyetmoibaivietcuathanhvien).click()
        driver.find_element(By.XPATH, var.baivietphaiduocpheduyet).click()
        driver.find_element(By.XPATH, var.luu).click()
        check_caidatnhom_pheduyetmoibaiviet_message = driver.find_element(By.XPATH,var.check_caidatnhom_tenvamota_message).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message khi lưu Phê duyệt mọi bài viết của thành viên - Bài viết của thành viên phải được quản trị viên hoặc người kiểm duyệt phê duyệt - Cập nhật dữ liệu thành công")
        logging.info(check_caidatnhom_pheduyetmoibaiviet_message == "Cập nhật dữ liệu thành công")
        time.sleep(1)

        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_taobaiviet1).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['baivietdangcho'])
        driver.find_element(By.XPATH, var.dang).click()
        check_nhom_baivietdangcho_message = driver.find_element(By.XPATH,var.check_nhom_baivietdangcho_message).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message Bài viết đang chờ - Cảm ơn bạn đã đăng bài! Hệ thống đã gửi bài viết cho quản trị viên nhóm phê duyệt.")
        logging.info(check_nhom_baivietdangcho_message == "Cảm ơn bạn đã đăng bài! Hệ thống đã gửi bài viết cho quản trị viên nhóm phê duyệt.")
        time.sleep(1)

        #Bài viết đang chờ - phê duyệt
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879/waiting_post")
        time.sleep(2)
        try:
            check_baivietdangcho_thongbao = driver.find_element(By.XPATH,var.check_baivietdangcho_thongbao)
            logging.info("Nhóm - Nhóm của bạn - Bài viết đang chờ")
            logging.info("check font-end: Có thông báo ở module hay không")
            logging.info(check_baivietdangcho_thongbao.is_displayed())
        except NoSuchElementException:
            logging.info("Nhóm - Nhóm của bạn - Bài viết đang chờ")
            logging.info("check font-end: Có thông báo ở module hay không")
            logging.info("False")
        driver.find_element(By.XPATH, var.pheduyet).click()
        check_nhom_baivietdangcho_pheduyet_message = driver.find_element(By.XPATH,var.check_nhom_baivietdangcho_pheduyet_message).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message Bài viết đang chờ - Phê duyệt - Đã phê duyệt bài viết thành công")
        logging.info(check_nhom_baivietdangcho_pheduyet_message == "Đã phê duyệt bài viết thành công")
        time.sleep(1)

        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        time.sleep(1)
        check_nhom_baivietdangcho_pheduyet_baiviet = driver.find_element(By.XPATH,var.nhom_baiviet1_mota).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message Bài viết đang chờ - Phê duyệt")
        logging.info("check font-end:  check trên dòng thời gian nhóm - "+ data['nhom']['baivietdangcho'])
        logging.info(check_nhom_baivietdangcho_pheduyet_baiviet == data['nhom']['baivietdangcho'])
        time.sleep(1)

        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_taobaiviet1).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['baivietdangcho'])
        driver.find_element(By.XPATH, var.dang).click()

        #Bài viết đang chờ - từ chối
        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_taobaiviet1).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['baivietdangcho_tuchoi'])
        driver.find_element(By.XPATH, var.dang).click()
        check_nhom_baivietdangcho_message = driver.find_element(By.XPATH,var.check_nhom_baivietdangcho_message).text
        time.sleep(1)

        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879/waiting_post")
        time.sleep(2)
        driver.find_element(By.XPATH, var.tuchoi).click()   #Phải từ chối 2l mới đc
        check_nhom_baivietdangcho_tuchoi_message = driver.find_element(By.XPATH,var.check_nhom_baivietdangcho_tuchoi_message).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message Bài viết đang chờ - Từ chối - Đã phê duyệt bài viết thành công")
        logging.info(check_nhom_baivietdangcho_tuchoi_message == "Đã từ chối bài viết thành công")
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_trangchu).click()
        time.sleep(1)
        check_nhom_baivietdangcho_tuchoi_baiviet = driver.find_element(By.XPATH,var.nhom_baiviet1_mota).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Message Bài viết đang chờ - Phê duyệt")
        logging.info("check font-end:  check trên dòng thời gian nhóm - "+ data['nhom']['baivietdangcho_tuchoi'])
        logging.info(check_nhom_baivietdangcho_tuchoi_baiviet != data['nhom']['baivietdangcho_tuchoi'])
        time.sleep(1)

        #Ai có thể đăng - Bất cứ ai trong nhóm - Thành viên có thể trực tiếp đăng bài lên nhóm
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879/setting_group")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_quanlynoidungthaoluan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_aicothedang).click()
        driver.find_element(By.XPATH, var.nhom_caidatnhom_aicothedang_batcuaitrongnhom).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.nhom_caidatnhom_pheduyetmoibaivietcuathanhvien).click()
        driver.find_element(By.XPATH, var.thanhviencothedangbaitructieplennhom).click()
        driver.find_element(By.XPATH, var.luu).click()
        check_caidatnhom_aicothedang_message = driver.find_element(By.XPATH,var.check_caidatnhom_tenvamota_message).text
        time.sleep(1)

        login.login4(self, "truongvck222@gmail.com", "voncamk22")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/group/111504666936394879")
        time.sleep(2)
        driver.find_element(By.XPATH, var.nhom_taobaiviet1).click()
        driver.find_element(By.XPATH, var.nhom_taobaiviet_mota).send_keys(data['nhom']['thanhviencothetructiepdangbai'])
        driver.find_element(By.XPATH, var.dang).click()
        driver.find_element(By.XPATH, var.dangbaivietthanhcong)
        time.sleep(2.5)

        check_nhom_thanhviencothedangtructieplen_baiviet = driver.find_element(By.XPATH,var.nhom_baiviet1_mota).text
        logging.info("Nhóm - Nhóm của bạn - Cài đặt nhóm")
        logging.info("check font-end:  Ai có thể đăng - Thành viên có thể trực tiếp đăng bài lên nhóm")
        logging.info("check font-end:  check trên dòng thời gian nhóm - "+ data['nhom']['thanhviencothetructiepdangbai'])
        logging.info(check_nhom_thanhviencothedangtructieplen_baiviet == data['nhom']['thanhviencothetructiepdangbai'])
        time.sleep(1)










class quatrinhmuahang():
    def timkiemsanpham(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.timkiemsanpham).send_keys(data['market']['timkiemtrangchu'])
        driver.find_element(By.XPATH, var.timkiemsanpham).send_keys(Keys.ENTER)
        time.sleep(1)

        # check_market_timkiem_shop_anhdaidien = driver.find_element(By.XPATH,var.check_market_timkiem_shop_anhdaidien).get_attribute("src")
        # logging.info("Market - Tìm kiếm")
        # logging.info("check font-end:  Shop - Ảnh đại diện - Có hiển thị ảnh đại diện không?")
        # logging.info(check_market_timkiem_shop_anhdaidien)
        # logging.info(check_market_timkiem_shop_anhdaidien != "/static/media/page.8308143d.svg")
        #
        # check_market_timkiem_shop_sao = driver.find_element(By.XPATH,var.check_market_timkiem_shop_sao).text
        # logging.info("Market - Tìm kiếm")
        # logging.info("check font-end:  Shop - Số sao - Có hiển thị số sao không?")
        # logging.info(check_market_timkiem_shop_sao)
        # logging.info(check_market_timkiem_shop_sao != "0.00")
        #
        # check_market_timkiem_shop_follow = driver.find_element(By.XPATH,var.check_market_timkiem_shop_follow).text
        # logging.info("Market - Tìm kiếm")
        # logging.info("check font-end:  Shop - Số follow - Có hiển thị số follow không?")
        # logging.info(check_market_timkiem_shop_follow)
        # logging.info(check_market_timkiem_shop_follow != "0 Followers")
        #
        # check_market_timkiem_shop_tenshopmarket = driver.find_element(By.XPATH,var.check_market_timkiem_shop_tenshopmarket).text
        # #Vào trang check tên tranng
        # driver.find_element(By.XPATH, var.check_market_timkiem_shop_tenshopmarket).click()
        # time.sleep(1)
        # check_market_timkiem_shop_tenshoptrang = driver.find_element(By.XPATH,var.check_market_timkiem_shop_tenshoptrang).text
        # logging.info("Market - Tìm kiếm")
        # logging.info("check font-end:  Shop - Tên Shop - Tên shop khi tìm kiếm và khi click vào có giống nhau không")
        # logging.info(check_market_timkiem_shop_tenshopmarket)
        # logging.info(check_market_timkiem_shop_tenshoptrang)
        # logging.info(check_market_timkiem_shop_tenshopmarket == check_market_timkiem_shop_tenshoptrang)
        # driver.back()
        # time.sleep(1)
        driver.find_element(By.XPATH, var.timkiemsanpham_chonxem).click()
        time.sleep(1)
        #Mua ngay
        driver.find_element(By.XPATH, var.muangay).click()
        check_giohang_tenspmuangay = driver.find_element(By.XPATH, var.check_giohang_tensanpham1).text
        if check_giohang_tenspmuangay == data['market']['timkiemtrangchu']:
            driver.find_element(By.XPATH, var.giohang_xoasp1).click()
            time.sleep(2)
            check_giohang_xoasanpham = driver.find_element(By.XPATH, var.check_giohang_tensanpham1).text
            if check_giohang_xoasanpham == data['market']['timkiemtrangchu']:
                logging.info("Người mua - Giỏ hàng - giỏ hàng của bạn")
                logging.info("check font-end: Xóa sản phẩm 1 - Có xóa được không")
                logging.info("False")
            else:
                logging.info("Người mua - Giỏ hàng - giỏ hàng của bạn")
                logging.info("check font-end: Xóa sản phẩm 1 - Có xóa được không")
                logging.info("True")
        else:
            logging.info("Người mua - Tìm kiếm sản phẩm - Mua ngay")
            logging.info("check font-end: Có chuyển tới Giỏ hàng và hiển thị sản phẩm không")
            logging.info("False")
        driver.back()

        #Thêm vào giỏ hàng
        driver.find_element(By.XPATH, var.themvaogiohang).click()
        message_themvaogiohang = driver.find_element(By.XPATH, var.message_themvaogiohang)
        #Giỏ hàng
        driver.find_element(By.XPATH, var.icongiohang).click()
        time.sleep(1)
        check_giohang_tenspthemvaogiohang = driver.find_element(By.XPATH, var.check_giohang_tensanpham1).text
        if check_giohang_tenspthemvaogiohang == data['market']['timkiemtrangchu']:
            driver.find_element(By.XPATH, var.giohangcuaban_chonsp1).click()
            driver.find_element(By.XPATH, var.xacnhandonhang).click()
            time.sleep(1)
        else:
            logging.info("Người mua - Tìm kiếm sản phẩm - Thêm vào giỏ hàng")
            logging.info("check font-end: Có thêm sản phẩm đã chọn vào gio hàng không")
            logging.info("False")
        #Thanh toán
        driver.find_element(By.XPATH, var.capnhatdiachi).click()
        driver.find_element(By.XPATH, var.luu).click()
        driver.execute_script("window.scrollBy(0,500)", "")
        driver.find_element(By.XPATH, var.thanhtoan_loinhan).send_keys(data['market']['loinhan'])
        driver.find_element(By.XPATH, var.dathang).click()
        time.sleep(1)
        #Đơn hàng của tôi
        button = driver.find_element(By.XPATH, var.donhangcuatoi_chothanhtoan)
        driver.execute_script("arguments[0].click();", button)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_chothanhtoan_tensp1 = driver.find_element(By.XPATH,var.check_donhangcuatoi_chothanhtoan_sp1).text
            logging.info("Người mua - Đơn hàng của tôi - Đặt hàng - Chờ thanh toán")
            logging.info("check font-end: Tên sản phẩm 1 - "+ data['market']['timkiemtrangchu'])
            logging.info(check_donhangcuatoi_chothanhtoan_tensp1)
            logging.info(check_donhangcuatoi_chothanhtoan_tensp1 == data['market']['timkiemtrangchu'])
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi - Đặt hàng - Chờ thanh toán")
            logging.info("check font-end: Tên sản phẩm 1 - "+ data['market']['timkiemtrangchu'])
            logging.info("False")

        try:
            check_donhangcuatoi_chothanhtoan_trangthaisp1 = driver.find_element(By.XPATH, var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Chờ thanh toán - Trạng thái đơn hàng - Chờ thanh toán")
            logging.info(check_donhangcuatoi_chothanhtoan_trangthaisp1 == "Chờ thanh toán")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Chờ thanh toán - Trạng thái đơn hàng - Chờ thanh toán")
            logging.info("False")

    def xacnhandon(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang).click()
        check_quanlydonhang_tatca_sp1 = driver.find_element(By.XPATH,var.check_quanlydonhang_sp1).text
        logging.info("Người bán - Quản lý đơn hàng - Tất cả")
        logging.info("check font-end: Tên sản phẩm 1 - "+ data['market']['timkiemtrangchu'])
        logging.info(check_quanlydonhang_tatca_sp1)
        logging.info(check_quanlydonhang_tatca_sp1 == data['market']['timkiemtrangchu'])

        #Chờ xác nhận
        driver.find_element(By.XPATH, var.quanlydonhang_choxacnhan).click()
        time.sleep(1.5)
        check_quanlydonhang_choxacnhan_sp1 = driver.find_element(By.XPATH,var.check_quanlydonhang_sp1).text
        if check_quanlydonhang_choxacnhan_sp1 == data['market']['timkiemtrangchu']:
            driver.find_element(By.XPATH, var.quanlydonhang_choxacnhan_xacnhandon1).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.xacnhan).click()
            driver.implicitly_wait(2)
            try:
                check_message_xacnhandon = driver.find_element(By.XPATH,var.check_message_quanlydonhang)
                logging.info("Người bán - Quản lý đơn hàng")
                logging.info("check font-end: Có Message Xác nhận đơn Không?")
                logging.info(check_message_xacnhandon.is_displayed())
                logging.info(check_message_xacnhandon.text)
            except NoSuchElementException:
                logging.info("Người bán - Quản lý đơn hàng")
                logging.info("check font-end: Có Message Xác nhận đơn Không?")
                logging.info("False")
        else:
            logging.info("Người bán - Quản lý đơn hàng")
            logging.info("check font-end: Có đơn hàng mới từ người mua chuyển qua không")
            logging.info("False")

        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.donhangcuatoi_chothanhtoan)
        driver.execute_script("arguments[0].click();", button)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_chothanhtoan_trangthai = driver.find_element(By.XPATH, var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Chờ thanh toán - Trạng thái đơn hàng - Đã xác nhận")
            logging.info(check_donhangcuatoi_chothanhtoan_trangthai == "Đã xác nhận")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Chờ thanh toán - Trạng thái đơn hàng - Đã xác nhận")
            logging.info("False")
        time.sleep(1)

    def chuanbihang(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang).click()
        #Chờ lấy hàng
        driver.find_element(By.XPATH, var.quanlydonhang_cholayhang).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang_daxacnhan).click()
        time.sleep(1)
        check_quanlydonhang_cholayhang_sp1 = driver.find_element(By.XPATH,var.check_quanlydonhang_daxacnhan_sp1).text
        logging.info("Người bán - Quản lý đơn hàng - Chờ lấy hàng")
        logging.info("check font-end: Đã xác nhận - Tên sản phẩm 1 - "+ data['market']['timkiemtrangchu'])
        logging.info(check_quanlydonhang_cholayhang_sp1)
        logging.info(check_quanlydonhang_cholayhang_sp1 == data['market']['timkiemtrangchu'])

        driver.find_element(By.XPATH, var.quanlydonhang_cholayhang_chuanbihang).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        driver.implicitly_wait(2)
        try:
            check_message_chuanbihang = driver.find_element(By.XPATH, var.check_message_quanlydonhang)
            logging.info("Người bán - Quản lý đơn hàng")
            logging.info("check font-end: Có Message Chuẩn bị hàng Không?")
            logging.info(check_message_chuanbihang.is_displayed())
            logging.info(check_message_chuanbihang.text)
        except NoSuchElementException:
            logging.info("Người bán - Quản lý đơn hàng")
            logging.info("check font-end: Có Message Chuẩn bị hàng Không?")
            logging.info("False")
        driver.implicitly_wait(15)

        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.donhangcuatoi_vanchuyen)
        driver.execute_script("arguments[0].click();", button)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_vanchuyen_trangthai = driver.find_element(By.XPATH, var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Vận chuyển - Trạng thái đơn hàng - Chờ lấy hàng")
            logging.info(check_donhangcuatoi_vanchuyen_trangthai == "Chờ lấy hàng")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Vận chuyển - Trạng thái đơn hàng - Chờ lấy hàng")
            logging.info("False")
        time.sleep(1)

    def donvivanchuyendanglayhang(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang).click()
        #Chờ Chờ ĐVVC đến lấy hàng
        driver.find_element(By.XPATH, var.quanlydonhang_cholayhang).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang_chodvvcdenlayhang).click()
        time.sleep(1)
        check_quanlydonhang_cholayhang_sp1 = driver.find_element(By.XPATH,var.check_quanlydonhang_daxacnhan_sp1).text
        logging.info("Người bán - Quản lý đơn hàng - Chờ lấy hàng")
        logging.info("check font-end: Chờ ĐVVC đến lấy hàng - Tên sản phẩm 1 - "+ data['market']['timkiemtrangchu'])
        logging.info(check_quanlydonhang_cholayhang_sp1)
        logging.info(check_quanlydonhang_cholayhang_sp1 == data['market']['timkiemtrangchu'])

        driver.find_element(By.XPATH, var.quanlydonhang_cholayhang_donvivanchuyendanglayhang).click()
        driver.implicitly_wait(2)
        try:
            check_message_chuanbihang = driver.find_element(By.XPATH, var.check_message_quanlydonhang)
            logging.info("Người bán - Quản lý đơn hàng")
            logging.info("check font-end: Có Message Đơn vị vận chuyển đang lấy hàng Không?")
            logging.info(check_message_chuanbihang.is_displayed())
            logging.info(check_message_chuanbihang.text)
        except NoSuchElementException:
            logging.info("Người bán - Quản lý đơn hàng")
            logging.info("check font-end: Có Message Đơn vị vận chuyển đang lấy hàng Không?")
            logging.info("False")
        driver.implicitly_wait(15)
        time.sleep(1)

        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.donhangcuatoi_vanchuyen)
        driver.execute_script("arguments[0].click();", button)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_vanchuyen_trangthai = driver.find_element(By.XPATH, var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Vận chuyển - Trạng thái đơn hàng - Đã lấy hàng")
            logging.info(check_donhangcuatoi_vanchuyen_trangthai == "Đã lấy hàng")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Vận chuyển - Trạng thái đơn hàng - Đã lấy hàng")
            logging.info("False")
        time.sleep(1)

    def donvivanchuyendalayhang(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang).click()
        #ĐVVC đã lấy hàng
        driver.find_element(By.XPATH, var.quanlydonhang_cholayhang).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang_dvvcdalayhang).click()
        time.sleep(1)
        check_quanlydonhang_cholayhang_sp1 = driver.find_element(By.XPATH,var.check_quanlydonhang_daxacnhan_sp1).text
        logging.info("Người bán - Quản lý đơn hàng - Chờ lấy hàng")
        logging.info("check font-end: ĐVVC đã lấy hàng - Tên sản phẩm 1 - "+ data['market']['timkiemtrangchu'])
        logging.info(check_quanlydonhang_cholayhang_sp1)
        logging.info(check_quanlydonhang_cholayhang_sp1 == data['market']['timkiemtrangchu'])

        driver.find_element(By.XPATH, var.quanlydonhang_cholayhang_donvivanchuyendalayhang).click()
        driver.implicitly_wait(2)
        try:
            check_message_chuanbihang = driver.find_element(By.XPATH, var.check_message_quanlydonhang)
            logging.info("Người bán - Quản lý đơn hàng")
            logging.info("check font-end: Có Message Đơn vị vận chuyển đã lấy hàng Không?")
            logging.info(check_message_chuanbihang.is_displayed())
            logging.info(check_message_chuanbihang.text)
        except NoSuchElementException:
            logging.info("Người bán - Quản lý đơn hàng")
            logging.info("check font-end: Có Message Đơn vị vận chuyển đã lấy hàng Không?")
            logging.info("False")
        driver.implicitly_wait(15)
        time.sleep(1)

        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.donhangcuatoi_danggiao)
        driver.execute_script("arguments[0].click();", button)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_danggiao_trangthai = driver.find_element(By.XPATH,var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Đang giao - Trạng thái đơn hàng - Đang giao")
            logging.info(check_donhangcuatoi_danggiao_trangthai == "Đang giao")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Đang giao - Trạng thái đơn hàng - Đang giao")
            logging.info("False")
        time.sleep(1)

    def danggiaohang(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang).click()
        #Đang giao
        driver.find_element(By.XPATH, var.quanlydonhang_danggiao).click()
        time.sleep(1)
        check_quanlydonhang_dangiao_sp1 = driver.find_element(By.XPATH,var.check_quanlydonhang_sp1).text
        logging.info("Người bán - Quản lý đơn hàng - Đang giao")
        logging.info("check font-end: Đang giao hàng - Tên sản phẩm 1 - "+ data['market']['timkiemtrangchu'])
        logging.info(check_quanlydonhang_dangiao_sp1)
        logging.info(check_quanlydonhang_dangiao_sp1 == data['market']['timkiemtrangchu'])

        driver.find_element(By.XPATH, var.quanlydonhang_dangiao_dangiaohang).click()
        driver.implicitly_wait(2)
        try:
            check_message_dangiaohang = driver.find_element(By.XPATH, var.check_message_quanlydonhang)
            logging.info("Người bán - Quản lý đơn hàng")
            logging.info("check font-end: Có Message Đơn vị vận chuyển đã lấy hàng Không?")
            logging.info(check_message_dangiaohang.is_displayed())
            logging.info(check_message_dangiaohang.text)
        except NoSuchElementException:
            logging.info("Người bán - Quản lý đơn hàng")
            logging.info("check font-end: Có Message Đơn vị vận chuyển đã lấy hàng Không?")
            logging.info("False")
        driver.implicitly_wait(15)
        time.sleep(1)


        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.donhangcuatoi_danggiao)
        driver.execute_script("arguments[0].click();", button)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_danggiao_trangthai = driver.find_element(By.XPATH,var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Đang giao - Trạng thái đơn hàng - Đã giao")
            logging.info(check_donhangcuatoi_danggiao_trangthai == "Đã giao")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Đang giao - Trạng thái đơn hàng - Đã giao")
            logging.info("False")
        time.sleep(1)

    def danhanhang(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        #Đang giao
        button = driver.find_element(By.XPATH, var.donhangcuatoi_danggiao)      #ko hiện thị sp ở đơn hàng của tôi- đang giao
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.donhangcuatoi_danggiao_danhanhang).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        #Đánh giá sản phẩm
        driver.find_element(By.XPATH, var.donhangcuatoi_hoanthanh_danhgia).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.danhgiasanpham_chon4sao)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.danhgiasanpham_input).send_keys(data['market']['danhgiasanpham'])
        driver.find_element(By.XPATH, var.danhgiasanpham_themanhvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhdaidien2.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.gui).click()
        driver.implicitly_wait(2)
        try:
            check_message_danhgiasanpham = driver.find_element(By.XPATH, var.check_message_danhgiasanpham)
            logging.info("Người Mua - Đơn hàng của tôi")
            logging.info("check font-end: Message đánh giá sản phẩm - Đánh giá thành công!")
            logging.info(check_message_danhgiasanpham.is_displayed())
            logging.info(check_message_danhgiasanpham.text)
        except NoSuchElementException:
            logging.info("Người Mua - Đơn hàng của tôi")
            logging.info("check font-end: Message đánh giá sản phẩm - Đánh giá thành công!")
            logging.info("False")
        driver.implicitly_wait(15)

        #Hoàn thành
        button = driver.find_element(By.XPATH, var.donhangcuatoi_hoanthanh)
        driver.execute_script("arguments[0].click();", button)
        #Mua lại
        driver.find_element(By.XPATH, var.hoanthanh_mualai).click()
        time.sleep(1)
        check_donhangcuatoi_hoanthanh_mualai = driver.find_element(By.XPATH,var.check_donhangcuatoi_hoanthanh_mualai).text
        logging.info("Người Mua - Đơn hàng của tôi - Hoàn thành")
        logging.info("check font-end: Mua lại - Chuyển tới trang Giỏ hàng của bạn")
        logging.info(check_donhangcuatoi_hoanthanh_mualai)
        logging.info(check_donhangcuatoi_hoanthanh_mualai == "Giỏ hàng của bạn")
        driver.back()
        #Liên hệ người bán
        driver.find_element(By.XPATH, var.donhangcuatoi_lienhenguoiban).click()
        time.sleep(1)
        try:
            check_donhangcuatoi_hoanthanh_lienhenguoiban = driver.find_element(By.XPATH, var.check_donhangcuatoi_hoanthanh_lienhenguoiban).text
            logging.info("Người Mua - Đơn hàng của tôi - Hoàn thành")
            logging.info("check font-end: Liên hệ người bán - Có hiện hộp chat chat với người bán không?")
            logging.info(check_donhangcuatoi_hoanthanh_lienhenguoiban)
            logging.info(check_donhangcuatoi_hoanthanh_lienhenguoiban == "Bình Thuận")
            driver.find_element(By.XPATH, var.hoanthanh_lienhenguoiban_x).click()

        except NoSuchElementException:
            logging.info("Người Mua - Đơn hàng của tôi - Hoàn thành")
            logging.info("check font-end: Liên hệ người bán - Có hiện hộp chat chat với người bán không?")
            logging.info("False")
        driver.implicitly_wait(15)

        #Xem đánh giá
        button = driver.find_element(By.XPATH, var.hoanthanh_xemdanhgia)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        check_donhangcuatoi_hoanthanh_danhgiasanpham = driver.find_element(By.XPATH,var.check_donhangcuatoi_hoanthanh_danhgiasanpham)
        logging.info("Người Mua - Đơn hàng của tôi - Hoàn thành")
        logging.info("check font-end: Đánh giá sản phẩm - Có hiển thị các đánh giá không")
        logging.info(check_donhangcuatoi_hoanthanh_danhgiasanpham)
        logging.info(check_donhangcuatoi_hoanthanh_danhgiasanpham != None)
        driver.find_element(By.XPATH, var.x).click()
        time.sleep(1)

    def chothanhtoan_dahuy(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        #Đang giao
        button = driver.find_element(By.XPATH, var.donhangcuatoi_chothanhtoan)      #ko hiện thị sp ở đơn hàng của tôi- đang giao
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)

        #Xem chi tiết
        driver.find_element(By.XPATH, var.donhangcuatoi_chothanhtoan_xemchitiet).click()
        time.sleep(2)
        try:
            check_donhangcuatoi_chothanhtoan_xemchitiet = driver.find_element(By.XPATH,var.check_donhangcuatoi_chothanhtoan_xemchitiet)
            logging.info("Người Mua - Đơn hàng của tôi - Chờ thanh toán")
            logging.info("check font-end: Xem chi tiết - Có chuyển tới trang xem chi tiết sản phẩm không")
            logging.info(check_donhangcuatoi_chothanhtoan_xemchitiet.text)
            logging.info(check_donhangcuatoi_chothanhtoan_xemchitiet.is_displayed())
        except NoSuchElementException:
            logging.info("Người Mua - Đơn hàng của tôi - Chờ thanh toán")
            logging.info("check font-end: Xem chi tiết - Có chuyển tới trang xem chi tiết sản phẩm không")
            logging.info("False")
        time.sleep(1)
        driver.back()
        time.sleep(1)

        #Liên hệ người bán
        # driver.find_element(By.XPATH, var.donhangcuatoi_lienhenguoiban).click()     #Tự bật popup liên hệ người bán
        # time.sleep(1)
        # try:
        #     check_donhangcuatoi_chothanhtoan_lienhenguoiban = driver.find_element(By.XPATH, var.check_donhangcuatoi_hoanthanh_lienhenguoiban).text
        #     logging.info("Người Mua - Đơn hàng của tôi - Chờ thanh toán")
        #     logging.info("check font-end: Liên hệ người bán - Có hiện hộp chat chat với người bán không?")
        #     logging.info(check_donhangcuatoi_chothanhtoan_lienhenguoiban)
        #     logging.info(check_donhangcuatoi_chothanhtoan_lienhenguoiban == "Bình Thuận")
        #     driver.find_element(By.XPATH, var.hoanthanh_lienhenguoiban_x).click()
        #
        # except NoSuchElementException:
        #     logging.info("Người Mua - Đơn hàng của tôi - Chờ thanh toán")
        #     logging.info("check font-end: Liên hệ người bán - Có hiện hộp chat chat với người bán không?")
        #     logging.info("False")
        # driver.implicitly_wait(15)

        #Hủy đơn hàng
        driver.find_element(By.XPATH, var.donhangcuatoi_huydonhang).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.khongphaibaygio).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.donhangcuatoi_huydonhang).click()
        time.sleep(1)
        driver.implicitly_wait(2)
        try:
            driver.find_element(By.XPATH, var.lydohuy_toimuonthaydoima).click()
        except NoSuchElementException:
            logging.info("Người Mua - Đơn hàng của tôi- Đã hủy")
            logging.info("check font-end: Có hiển thị lý do hủy đơn không?")
            logging.info("False")
        driver.find_element(By.XPATH, var.huydonhang).click()
        # driver.implicitly_wait(2)
        # try:
        #     check_message_donhangcuatoi = driver.find_element(By.XPATH, var.check_message_donhangcuatoi)    #ko hủy được đơn hàng
        #     logging.info("Người Mua - Đơn hàng của tôi")
        #     logging.info("check font-end: Message Hủy đơn hàng")
        #     logging.info(check_message_donhangcuatoi.is_displayed())
        #     logging.info(check_message_donhangcuatoi.text)
        # except NoSuchElementException:
        #     logging.info("Người Mua - Đơn hàng của tôi")
        #     logging.info("check font-end: Message Hủy đơn hàng")
        #     logging.info("False")
        # driver.implicitly_wait(15)
        time.sleep(2)
        driver.refresh()
        time.sleep(2)
        #Đã hủy
        button = driver.find_element(By.XPATH, var.donhangcuatoi_dahuy)
        driver.execute_script("arguments[0].click();", button)
        try:
            check_donhangcuatoi_dahuydon_trangthai = driver.find_element(By.XPATH,var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Đang giao - Trạng thái đơn hàng - Đã hủy")
            logging.info(check_donhangcuatoi_dahuydon_trangthai == "Đã hủy")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Đang giao - Trạng thái đơn hàng - Đã hủy")
            logging.info("False")
        time.sleep(1)
        #Mua lại
        driver.find_element(By.XPATH, var.donhangcuatoi_mualai).click()
        time.sleep(1)
        try:
            check_donhangcuatoi_dahuy_mualai = driver.find_element(By.XPATH,var.check_donhangcuatoi_dahuy_mualai).text
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Đã hủy - Mua lại có hiển thị sản phẩm "+ data['market']['timkiemtrangchu']+ "Không?")
            logging.info(check_donhangcuatoi_dahuy_mualai)
            logging.info(check_donhangcuatoi_dahuy_mualai == data['market']['timkiemtrangchu'])
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Đã hủy - Mua lại có hiển thị sản phẩm "+ data['market']['timkiemtrangchu']+ "Không?")
            logging.info("False")
        driver.back()
        time.sleep(1)

        #Chi tiết đơn hủy
        driver.find_element(By.XPATH, var.donhangcuatoi_chitietdonhuy).click()
        time.sleep(1)
        try:
            check_chitietdonhuy_lydo = driver.find_element(By.XPATH,var.check_chitietdonhuy_lydo).text
            logging.info("Người mua - Đơn hàng của tôi - Đã hủy")
            logging.info("check font-end: Lý do hủy đơn: Tôi muốn thêm/thay đổi Mã giảm giá")
            logging.info(check_chitietdonhuy_lydo)
            logging.info(check_chitietdonhuy_lydo == "Lý do hủy đơn: Tôi muốn thêm/thay đổi Mã giảm giá")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi - Đã hủy")
            logging.info("check font-end: Lý do hủy đơn: Tôi muốn thêm/thay đổi Mã giảm giá")
            logging.info("False")
        driver.find_element(By.XPATH, var.quaylaimuasam).click()
        time.sleep(2)

    def shop_het_hang(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.timkiemsanpham).send_keys(data['market']['sanpham_hethang'])
        driver.find_element(By.XPATH, var.timkiemsanpham).send_keys(Keys.ENTER)
        time.sleep(2)
        driver.find_element(By.XPATH, var.timkiemsanpham_chonxem1).click()
        time.sleep(1)
        driver.implicitly_wait(2)
        # Mua ngay
        driver.find_element(By.XPATH, var.sanpham_muangay).click()
        try:
            time.sleep(1)
            check_giohang_tenspmuangay = driver.find_element(By.XPATH, var.check_giohang_tensanpham1).text
            logging.info("Người mua - Tìm kiếm sản phẩm - Mua ngay")
            logging.info("check font-end: Sản phẩm đã hết hàng - đã disable nút mua ngay chưa")
            logging.info("False")
            driver.back()
        except NoSuchElementException:
            logging.info("Người mua - Tìm kiếm sản phẩm - Mua ngay")
            logging.info("check font-end: Sản phẩm đã hết hàng - đã disable nút mua ngay chưa")
            logging.info("True")

        # Thêm vào giỏ hàng
        driver.find_element(By.XPATH, var.themvaogiohang).click()
        try:
            message_themvaogiohang = driver.find_element(By.XPATH, var.message_themvaogiohang)
            logging.info("Người mua - Tìm kiếm sản phẩm - Thêm vào giỏ hàng")
            logging.info("check font-end: Sản phẩm đã hết hàng - đã disable nút thêm vào giỏ hàng chưa chưa")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Người mua - Tìm kiếm sản phẩm - Thêm vào giỏ hàng")
            logging.info("check font-end: Sản phẩm đã hết hàng - đã disable nút thêm vào giỏ hàng chưa chưa")
            logging.info("True")

        check_sanpham_hethang = driver.find_element(By.XPATH,var.check_sanpham_hethang).text
        logging.info("Người mua - Tìm kiếm sản phẩm - Xem sản phẩm")
        logging.info("check font-end: Sản phẩm đã hết hàng - nhãn dán - Hết hàng")
        logging.info(check_sanpham_hethang == "Hết Hàng")
        time.sleep(1)

    def trahanghoantien_dongy(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        #Đang giao
        button = driver.find_element(By.XPATH, var.donhangcuatoi_danggiao)      #ko hiện thị sp ở đơn hàng của tôi- đang giao
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Yêu cầu trả hàng/hoàn tiền
        driver.find_element(By.XPATH, var.quanlydonhang_danggiao_trahanghoantien).click()
        #Lý do
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_hangcovande).click()
        driver.find_element(By.XPATH, var.tieptuc).click()
        time.sleep(1)
        #Chọn sản phẩm
        check_trahanghoantien_chonsanpham = driver.find_element(By.XPATH,var.check_trahanghoantien_chonsanpham).text
        logging.info("Người mua - Quản lý đơn hàng - Đang giao - Trả hàng hoàn tiền")
        logging.info("check font-end: Chọn sản phẩm - Tên sản phẩm 1 - "+ data['market']['timkiemtrangchu'])
        logging.info(check_trahanghoantien_chonsanpham)
        logging.info(check_trahanghoantien_chonsanpham == data['market']['timkiemtrangchu'])
        driver.find_element(By.XPATH, var.trahanghoantien_chonsanpham1).click()
        driver.find_element(By.XPATH, var.tieptuc).click()
        time.sleep(1)
        #Yêu cầu
        driver.find_element(By.XPATH, var.trahanghoantien_lydo).click()
        driver.find_element(By.XPATH, var.trahanghoantien_lydo_hethansudung).click()

        xoa = driver.find_element(By.XPATH, var.trahanghoantien_hoanlai_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.trahanghoantien_hoanlai_input).send_keys(data['market']['trahanghoantien_hoanlai'])
        driver.find_element(By.XPATH, var.trahanghoantien_mota_input).send_keys(data['market']['trahanghoantien_mota'])
        driver.find_element(By.XPATH, var.trahanghoantien_themhinhanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/anhbia1.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_themvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/vay1.exe")
        time.sleep(2)
        driver.find_element(By.XPATH, var.tieptuc).click()
        time.sleep(1)
        #Hình thức vận chuyển
        #Lấy hàng tại địa chỉ yêu cầu
        driver.find_element(By.XPATH, var.gui).click()      #Chết luồng, ko gửi đươc yêu cầu hoàn tiền 6771
        time.sleep(3)
        #Trả hàng hoàn tiền
        button = driver.find_element(By.XPATH, var.donhangcuatoi_trahanghoantien)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_trahanghoantien_trangthai = driver.find_element(By.XPATH, var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Trả hàng/Hoàn tiền - Trạng thái đơn hàng - Shop đang xử lý")
            logging.info(check_donhangcuatoi_trahanghoantien_trangthai == "Shop đang xử lý")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end: Trả hàng/Hoàn tiền - Trạng thái đơn hàng - Shop đang xử lý")
            logging.info("False")
        time.sleep(1)
        driver.refresh()
        time.sleep(2)
        #Quản lý đơn
        driver.find_element(By.XPATH, var.quanlydonhang_trahanghoantien_quanlydon).click()
        time.sleep(1)
        driver.execute_script("window.scrollBy(0,1500)", "")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_vietthaoluan).send_keys(data['market']['trahanghoantien_vietthaoluan1'])
        driver.find_element(By.XPATH, var.trahanghoantien_iconanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/iconbuon.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_vietthaoluan).send_keys(Keys.ENTER)   #ko bình luận được
        time.sleep(3)
        driver.implicitly_wait(2)
        try:
            check_trahanghoantien_quanlydon_thaoluan1 = driver.find_element(By.XPATH,var.check_trahanghoantien_xemchitiet_thaoluan1).text
            logging.info("Người Mua - Trả hàng/Hoàn tiền - Quản lý đơn")
            logging.info("check font-end: Thảo luận - Bình luận 1 - "+ data['market']['trahanghoantien_vietthaoluan1'])
            logging.info(check_trahanghoantien_quanlydon_thaoluan1 == data['market']['trahanghoantien_vietthaoluan1'])
        except NoSuchElementException:
            logging.info("Người Mua - Trả hàng/Hoàn tiền - Quản lý đơn")
            logging.info("check font-end: Thảo luận - Bình luận 1 - "+ data['market']['trahanghoantien_vietthaoluan1'])
            logging.info("False")

        try:
            check_trahanghoantien_xemchitiet_anh1 = driver.find_element(By.XPATH,var.check_trahanghoantien_xemchitiet_anh1)
            logging.info("Người Mua - Trả hàng/Hoàn tiền - Quản lý đơn")
            logging.info("check font-end: Thảo luận - Có hiển thị ảnh không")
            logging.info(check_trahanghoantien_xemchitiet_anh1)
        except NoSuchElementException:
            logging.info("Người Mua - Trả hàng/Hoàn tiền - Quản lý đơn")
            logging.info("check font-end: Thảo luận - Có hiển thị ảnh không")
            logging.info("False")
        time.sleep(1)
        driver.implicitly_wait(15)


        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang).click()
        # Trả hàng/hoàn tiền
        driver.find_element(By.XPATH, var.quanlydonhang_trahanghoantien).click()
        time.sleep(1)
        check_quanlydonhang_trahanghoantien_sp1 = driver.find_element(By.XPATH,var.check_quanlydonhang_daxacnhan_sp1).text
        logging.info("Người bán - Quản lý đơn hàng - Trả hàng/hoàn tiền")
        logging.info("check font-end: Tên sản phẩm 1 - " + data['market']['timkiemtrangchu'])
        logging.info(check_quanlydonhang_trahanghoantien_sp1)
        logging.info(check_quanlydonhang_trahanghoantien_sp1 == data['market']['timkiemtrangchu'])

        driver.find_element(By.XPATH, var.quanlydonhang_trahanghoantien_xemchitiet).click()
        time.sleep(2)
        handles = driver.window_handles
        for handle in handles:
            driver.switch_to.window(handle)
        time.sleep(1)
        driver.execute_script("window.scrollBy(0,1500)", "")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_vietthaoluan).send_keys(data['market']['trahanghoantien_vietthaoluan_dongy'])
        driver.find_element(By.XPATH, var.trahanghoantien_iconanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/vay2.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_vietthaoluan).send_keys(Keys.ENTER)   #Ko bình luận được
        time.sleep(3)
        driver.implicitly_wait(2)
        try:
            check_trahanghoantien_xemchitiet_thaoluan1 = driver.find_element(By.XPATH,var.check_trahanghoantien_xemchitiet_thaoluan1).text
            logging.info("Người Bán - Trả hàng/Hoàn tiền - Xem chi tiết")
            logging.info("check font-end: Thảo luận - Bình luận 1 - "+ data['market']['trahanghoantien_vietthaoluan_dongy'])
            logging.info(check_trahanghoantien_xemchitiet_thaoluan1 == data['market']['trahanghoantien_vietthaoluan_dongy'])
        except NoSuchElementException:
            logging.info("Người Bán - Trả hàng/Hoàn tiền - Xem chi tiết")
            logging.info("check font-end: Thảo luận - Bình luận 1 - "+ data['market']['trahanghoantien_vietthaoluan_dongy'])
            logging.info("False")
        try:
            check_trahanghoantien_xemchitiet_anh1 = driver.find_element(By.XPATH,var.check_trahanghoantien_xemchitiet_anh1)
            logging.info("Người Bán - Trả hàng/Hoàn tiền - Xem chi tiết")
            logging.info("check font-end: Thảo luận - Có hiển thị ảnh không")
            logging.info(check_trahanghoantien_xemchitiet_anh1)
        except NoSuchElementException:
            logging.info("Người Bán - Trả hàng/Hoàn tiền - Xem chi tiết")
            logging.info("check font-end: Thảo luận - Có hiển thị ảnh không")
            logging.info("False")

        driver.execute_script("window.scrollBy(0,-800)", "")
        button = driver.find_element(By.XPATH, var.dongyhoanhangvahoantien)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.chapnhan).click()
        time.sleep(1)
        check_popup_trahanghoantien_dongy = driver.find_element(By.XPATH,var.check_popup_trahanghoantien_dongy).text
        logging.info("Người bán - Quản lý đơn hàng - Trả hàng/hoàn tiền")
        logging.info("check font-end: Popup đồng ý - Gửi khiếu nại thành công")
        logging.info(check_popup_trahanghoantien_dongy)
        logging.info(check_popup_trahanghoantien_dongy == "Gửi khiếu nại thành công")
        time.sleep(1)
        driver.find_element(By.XPATH, var.dong).click()
        time.sleep(1)

        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        # Trả hàng/Hoàn tiền
        button = driver.find_element(By.XPATH,var.donhangcuatoi_trahanghoantien)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_trahanghoantien_trangthai = driver.find_element(By.XPATH, var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi - Trả hàng/Hoàn tiền")
            logging.info("check font-end: Shop đồng ý hoàn tiền - Trạng thái đơn hàng - Shop đồng ý")
            logging.info(check_donhangcuatoi_trahanghoantien_trangthai == "Shop đồng ý")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi - Trả hàng/Hoàn tiền")
            logging.info("check font-end: Shop đồng ý hoàn tiền - Trạng thái đơn hàng - Shop đồng ý")
            logging.info("False")
        time.sleep(1)

    def trahanghoantien_khongdongy(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        #Đang giao
        button = driver.find_element(By.XPATH, var.donhangcuatoi_danggiao)      #
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Yêu cầu trả hàng/hoàn tiền
        driver.find_element(By.XPATH, var.quanlydonhang_danggiao_trahanghoantien).click()
        # Lý do
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_toichuanhanduochang).click()
        driver.find_element(By.XPATH, var.tieptuc).click()
        time.sleep(1)
        # Chọn sản phẩm
        check_trahanghoantien_chonsanpham = driver.find_element(By.XPATH, var.check_trahanghoantien_chonsanpham).text
        logging.info("Người mua - Quản lý đơn hàng - Đang giao - Trả hàng hoàn tiền2")
        logging.info("check font-end: Chọn sản phẩm - Tên sản phẩm 1 - " + data['market']['timkiemtrangchu'])
        logging.info(check_trahanghoantien_chonsanpham)
        logging.info(check_trahanghoantien_chonsanpham == data['market']['timkiemtrangchu'])
        # driver.find_element(By.XPATH, var.trahanghoantien_chontatcasanpham).click()     6758#trả hàng hoàn tiền ko chọn được icon tất cả
        driver.find_element(By.XPATH, var.trahanghoantien_chonsanpham1).click()
        driver.find_element(By.XPATH, var.tieptuc).click()
        time.sleep(1)
        # Yêu cầu
        driver.find_element(By.XPATH, var.trahanghoantien_lydo).click()
        driver.find_element(By.XPATH, var.trahanghoantien_lydo_thieuhang).click()
        driver.find_element(By.XPATH, var.trahanghoantien_mota_input).send_keys(data['market']['trahanghoantien_mota1'])
        driver.find_element(By.XPATH, var.trahanghoantien_themhinhanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/tft.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_themvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/ao2.exe")
        time.sleep(2)
        driver.find_element(By.XPATH, var.tieptuc).click()
        time.sleep(1)
        # Hình thức vận chuyển
        # Gửi hàng tại bưu cục
        driver.find_element(By.XPATH, var.guihangtaibuucuc).click()
        driver.find_element(By.XPATH, var.gui).click()
        time.sleep(2)
        # Trả hàng hoàn tiền
        button = driver.find_element(By.XPATH, var.donhangcuatoi_trahanghoantien)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_trahanghoantien_trangthai = driver.find_element(By.XPATH,var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi2")
            logging.info("check font-end: Trả hàng/Hoàn tiền - Trạng thái đơn hàng - Shop đang xử lý")
            logging.info(check_donhangcuatoi_trahanghoantien_trangthai == "Shop đang xử lý")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi2")
            logging.info("check font-end: Trả hàng/Hoàn tiền - Trạng thái đơn hàng - Shop đang xử lý")
            logging.info("False")
        time.sleep(1)
        driver.refresh()
        time.sleep(2)
        # Quản lý đơn
        driver.find_element(By.XPATH, var.quanlydonhang_trahanghoantien_quanlydon).click()
        time.sleep(1)
        driver.execute_script("window.scrollBy(0,1500)", "")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_vietthaoluan).send_keys(data['market']['trahanghoantien_vietthaoluan2'])
        driver.find_element(By.XPATH, var.trahanghoantien_iconanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/iconbuon.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_vietthaoluan).send_keys(Keys.ENTER)  # ko bình luận được
        time.sleep(3)
        driver.implicitly_wait(2)
        try:
            check_trahanghoantien_quanlydon_thaoluan1 = driver.find_element(By.XPATH,var.check_trahanghoantien_xemchitiet_thaoluan1).text
            logging.info("Người Mua - Trả hàng/Hoàn tiền - Quản lý đơn2")
            logging.info("check font-end: Thảo luận - Bình luận 1 - "+ data['market']['trahanghoantien_vietthaoluan2'])
            logging.info(check_trahanghoantien_quanlydon_thaoluan1 == data['market']['trahanghoantien_vietthaoluan2'])
        except NoSuchElementException:
            logging.info("Người Mua - Trả hàng/Hoàn tiền - Quản lý đơn2")
            logging.info("check font-end: Thảo luận - Bình luận 1 - "+ data['market']['trahanghoantien_vietthaoluan2'])
            logging.info("False")

        try:
            check_trahanghoantien_xemchitiet_anh1 = driver.find_element(By.XPATH,var.check_trahanghoantien_xemchitiet_anh1)
            logging.info("Người Mua - Trả hàng/Hoàn tiền - Quản lý đơn2")
            logging.info("check font-end: Thảo luận - Có hiển thị ảnh không")
            logging.info(check_trahanghoantien_xemchitiet_anh1)
        except NoSuchElementException:
            logging.info("Người Mua - Trả hàng/Hoàn tiền - Quản lý đơn2")
            logging.info("check font-end: Thảo luận - Có hiển thị ảnh không")
            logging.info("False")


        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang).click()
        # Trả hàng/hoàn tiền
        driver.find_element(By.XPATH, var.quanlydonhang_trahanghoantien).click()
        time.sleep(1)
        check_quanlydonhang_trahanghoantien_sp1 = driver.find_element(By.XPATH,var.check_quanlydonhang_daxacnhan_sp1).text
        logging.info("Người bán - Quản lý đơn hàng - Trả hàng/hoàn tiền2")
        logging.info("check font-end: Tên sản phẩm 1 - " + data['market']['timkiemtrangchu'])
        logging.info(check_quanlydonhang_trahanghoantien_sp1)
        logging.info(check_quanlydonhang_trahanghoantien_sp1 == data['market']['timkiemtrangchu'])

        driver.find_element(By.XPATH, var.quanlydonhang_trahanghoantien_xemchitiet).click()
        time.sleep(2)
        handles = driver.window_handles
        for handle in handles:
            driver.switch_to.window(handle)
        time.sleep(1)
        driver.execute_script("window.scrollBy(0,1500)", "")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_vietthaoluan).send_keys(data['market']['trahanghoantien_vietthaoluan_khongdongy'])
        driver.find_element(By.XPATH, var.trahanghoantien_iconanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/vay2.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_vietthaoluan).send_keys(Keys.ENTER)  # Ko bình luận được
        time.sleep(3)
        driver.implicitly_wait(2)
        try:
            check_trahanghoantien_xemchitiet_thaoluan1 = driver.find_element(By.XPATH,var.check_trahanghoantien_xemchitiet_thaoluan1).text
            logging.info("Người Bán - Trả hàng/Hoàn tiền - Xem chi tiết2")
            logging.info("check font-end: Thảo luận - Bình luận 1 - "+ data['market']['trahanghoantien_vietthaoluan_khongdongy'])
            logging.info(check_trahanghoantien_xemchitiet_thaoluan1 == data['market']['trahanghoantien_vietthaoluan_khongdongy'])
        except NoSuchElementException:
            logging.info("Người Bán - Trả hàng/Hoàn tiền - Xem chi tiết2")
            logging.info("check font-end: Thảo luận - Bình luận 1 - "+ data['market']['trahanghoantien_vietthaoluan_khongdongy'])
            logging.info("False")

        try:
            check_trahanghoantien_xemchitiet_anh1 = driver.find_element(By.XPATH,var.check_trahanghoantien_xemchitiet_anh1)
            logging.info("Người Bán - Trả hàng/Hoàn tiền - Xem chi tiết2")
            logging.info("check font-end: Thảo luận - Có hiển thị ảnh không")
            logging.info(check_trahanghoantien_xemchitiet_anh1)
        except NoSuchElementException:
            logging.info("Người Bán - Trả hàng/Hoàn tiền - Xem chi tiết2")
            logging.info("check font-end: Thảo luận - Có hiển thị ảnh không")
            logging.info("False")
        driver.implicitly_wait(15)
        driver.execute_script("window.scrollBy(0,-800)", "")
        button = driver.find_element(By.XPATH, var.khongdongyhoanhangvahoantien)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(0.5)

        #Popup khiếu nại đơn hàng
        driver.find_element(By.XPATH, var.khieunaidonhang_mota).send_keys(data['market']['khieunaidonhang_mota'])
        driver.find_element(By.XPATH, var.khieunaidonhang_themhinhanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/donghang.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.khieunaidonhang_themvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/vay1.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.gui).click()      #6765 Không gửi được khiếu nại

        check_message_trahanghoantien_khongdongy = driver.find_element(By.XPATH, var.check_message_trahanghoantien_khongdongy).text
        logging.info("Người bán - Quản lý đơn hàng - Trả hàng/hoàn tiền2")
        logging.info("check font-end: Mesage không đồng ý - Gửi khiếu nại thành công")
        logging.info(check_message_trahanghoantien_khongdongy)
        logging.info(check_message_trahanghoantien_khongdongy == "Gửi khiếu nại thành công")
        time.sleep(1)


        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        # Trả hàng/Hoàn tiền
        button = driver.find_element(By.XPATH, var.donhangcuatoi_trahanghoantien)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_trahanghoantien_trangthai = driver.find_element(By.XPATH,var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi - Trả hàng/Hoàn tiền2")
            logging.info("check font-end: Shop đồng ý hoàn tiền - Trạng thái đơn hàng - Shop Không đồng ý")
            logging.info(check_donhangcuatoi_trahanghoantien_trangthai == "Shop không đồng ý")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi - Trả hàng/Hoàn tiền2")
            logging.info("check font-end: Shop đồng ý hoàn tiền - Trạng thái đơn hàng - Shop không đồng ý")
            logging.info("False")
        time.sleep(1)

    def shophuydon(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlydonhang).click()
        #Chờ xác nhận
        driver.find_element(By.XPATH, var.quanlydonhang_choxacnhan).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.quanlydonhang_choxacnhan_xemmadonhang).click()
        time.sleep(1)
        handles = driver.window_handles
        for handle in handles:
            driver.switch_to.window(handle)
        time.sleep(1)
        #Ghi chú
        driver.execute_script("window.scrollBy(0,1500)", "")
        button = driver.find_element(By.XPATH, var.them)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.xemchitietdonhang_ghichu_input).send_keys(data['market']['xemchitietdonhang_ghichu'])
        driver.find_element(By.XPATH, var.them).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.capnhat).click()
        xoa = driver.find_element(By.XPATH, var.xemchitietdonhang_ghichu_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.xemchitietdonhang_ghichu_input).send_keys(data['market']['xemchitietdonhang_ghichu1'])
        driver.find_element(By.XPATH, var.capnhat).click()
        time.sleep(1)
        check_xemchitietdonhang_ghichu1 = driver.find_element(By.XPATH, var.check_xemchitietdonhang_ghichu1).text
        logging.info("Người Bán - Quản lý đơn hàng - Chờ xác nhận - Mã đơn hàng")
        logging.info("check font-end: Ghi chú 1 - " + data['market']['xemchitietdonhang_ghichu1'])
        logging.info(check_xemchitietdonhang_ghichu1)
        logging.info(check_xemchitietdonhang_ghichu1 == data['market']['xemchitietdonhang_ghichu1'])

        #Hủy đơn hàng
        driver.execute_script("window.scrollBy(0,-1500)", "")
        button = driver.find_element(By.XPATH, var.huydonhang1)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.huydonhang_chonlydo).click()
        driver.find_element(By.XPATH, var.hethang).click()
        driver.find_element(By.XPATH, var.huydonhangnay).click()
        check_message_shophuydonhang = driver.find_element(By.XPATH, var.check_message_shophuydonhang).text
        logging.info("Người Bán - Quản lý đơn hàng - Chờ xác nhận - Mã đơn hàng")
        logging.info("check font-end: Message Hủy đơn hàng - Huỷ đơn hàng thành công")
        logging.info(check_message_shophuydonhang == "Huỷ đơn hàng thành công")
        time.sleep(1)

        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        #Đã Hủy
        button = driver.find_element(By.XPATH, var.donhangcuatoi_dahuy)
        driver.execute_script("arguments[0].click();", button)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_dahuy_trangthai = driver.find_element(By.XPATH,var.check_donhangcuatoi_trangthai1).text
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end:Đã hủy - Trạng thái đơn hàng khi shop hủy đơn - Đã hủy")
            logging.info(check_donhangcuatoi_dahuy_trangthai == "Đã hủy")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi")
            logging.info("check font-end:Đã hủy - Trạng thái đơn hàng khi shop hủy đơn - Đã hủy")
            logging.info("False")
        time.sleep(1)




    def trahanghoantien_nguoimuahuy(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck333@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.icontaikhoan).click()
        driver.find_element(By.XPATH, var.donhangcuatoi).click()
        time.sleep(1)
        #Đang giao
        button = driver.find_element(By.XPATH, var.donhangcuatoi_danggiao)      #
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        #Yêu cầu trả hàng/hoàn tiền
        driver.find_element(By.XPATH, var.quanlydonhang_danggiao_trahanghoantien).click()
        # Lý do
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_toichuanhanduochang).click()
        driver.find_element(By.XPATH, var.tieptuc).click()
        time.sleep(1)
        # Chọn sản phẩm
        check_trahanghoantien_chonsanpham = driver.find_element(By.XPATH, var.check_trahanghoantien_chonsanpham).text
        logging.info("Người mua - Quản lý đơn hàng - Đang giao - Trả hàng hoàn tiền3")
        logging.info("check font-end: Chọn sản phẩm - Tên sản phẩm 1 - " + data['market']['timkiemtrangchu'])
        logging.info(check_trahanghoantien_chonsanpham)
        logging.info(check_trahanghoantien_chonsanpham == data['market']['timkiemtrangchu'])
        driver.find_element(By.XPATH, var.trahanghoantien_chonsanpham1).click()
        driver.find_element(By.XPATH, var.tieptuc).click()
        time.sleep(1)
        # Yêu cầu
        driver.find_element(By.XPATH, var.trahanghoantien_lydo).click()
        driver.find_element(By.XPATH, var.trahanghoantien_lydo_thieuhang).click()
        driver.find_element(By.XPATH, var.trahanghoantien_mota_input).send_keys(data['market']['trahanghoantien_mota1'])
        driver.find_element(By.XPATH, var.trahanghoantien_themhinhanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/tft.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.trahanghoantien_themvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/ao2.exe")
        time.sleep(2)
        driver.find_element(By.XPATH, var.tieptuc).click()
        time.sleep(1)
        # Hình thức vận chuyển
        # Gửi hàng tại bưu cục
        driver.find_element(By.XPATH, var.guihangtaibuucuc).click()
        driver.find_element(By.XPATH, var.gui).click()
        time.sleep(2)
        # Trả hàng hoàn tiền
        button = driver.find_element(By.XPATH, var.donhangcuatoi_trahanghoantien)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.quanlydonhang_trahanghoantien_huyyeucau).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(2)
        driver.implicitly_wait(2)
        try:
            check_donhangcuatoi_trahanghoantien_trangthai = driver.find_element(By.XPATH,var.check_donhangcuatoi_trangthai1).text   #7209, f5 load lại mới đổi trạng thái
            logging.info("Người mua - Đơn hàng của tôi3")
            logging.info("check font-end: Trả hàng/Hoàn tiền - Trạng thái đơn hàng - Người mua huỷ hoàn hàng")
            logging.info(check_donhangcuatoi_trahanghoantien_trangthai == "Người mua huỷ hoàn hàng")
        except NoSuchElementException:
            logging.info("Người mua - Đơn hàng của tôi3")
            logging.info("check font-end: Trả hàng/Hoàn tiền - Trạng thái đơn hàng - Người mua huỷ hoàn hàng")
            logging.info("False")
        time.sleep(1)






class quanlysanpham():
    def themsanphammoi(self, guipheduyet_luunhap):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlysanpham).click()
        driver.find_element(By.XPATH, var.themsanphammoi).click()
        time.sleep(1)
        #Thông tin cơ bản
        driver.find_element(By.XPATH, var.themsanpham_tensanpham).send_keys(data['quanlysanpham']['themspmoi_tensanpham'])
        time.sleep(2)
        driver.find_element(By.XPATH, var.themsanpham_nghanhhang).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.thoitrangnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.hoodie_aoni).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.hoodie).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.themsanpham_themhinhanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/hoodie1.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.themsanpham_themvideo).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/hoodie5.exe")
        time.sleep(1)
        driver.find_element(By.XPATH, var.themsanpham_mota).send_keys(data['quanlysanpham']['themspmoi_tensanpham'])
        #Thông tin chi tiết
        driver.execute_script("window.scrollBy(0,400)", "")
        button = driver.find_element(By.XPATH, var.themsanpham_xemthem)
        driver.execute_script("arguments[0].click();", button)
        # driver.find_element(By.XPATH, var.themsanpham_xemthem).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.themsanpham_thuonghieu).click()
        driver.find_element(By.XPATH, var.no_brand).click()

        button = driver.find_element(By.XPATH, var.themsanpham_xuatxu)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.hanquoc).click()

        button = driver.find_element(By.XPATH, var.themsanpham_tall_fit)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.co).click()

        button = driver.find_element(By.XPATH, var.themsanpham_ratlon)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.co).click()

        button = driver.find_element(By.XPATH, var.themsanpham_phongcach)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.thethao).click()
        time.sleep(1)
        # button = driver.find_element(By.XPATH, var.themsanpham_phongcach)
        # driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.duongpho).click()

        button = driver.find_element(By.XPATH, var.themsanpham_mua)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.muadong).click()

        button = driver.find_element(By.XPATH, var.themsanpham_mau)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.hoatiet).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.hoa).click()

        button = driver.find_element(By.XPATH, var.themsanpham_chieudaitayao)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.daitay).click()

        button = driver.find_element(By.XPATH, var.themsanpham_chatlieu)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.ni).click()
        driver.find_element(By.XPATH, var.chinos).click()

        button = driver.find_element(By.XPATH, var.themsanpham_tentochuc)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.dangcapnhat).click()

        button = driver.find_element(By.XPATH, var.themsanpham_diachitochuc)
        driver.execute_script("arguments[0].click();", button)
        driver.find_element(By.XPATH, var.dangcapnhat).click()

        #Thông tin bán hàng
        driver.find_element(By.XPATH, var.themsanpham_themnhomphanloai).click()
        #Phân loại hàng
        #Nhóm phân loại
        driver.find_element(By.XPATH, var.themsanpham_nhomphanloai1).send_keys(data['quanlysanpham']['themspmoi_nhomphanloai1'])
        driver.find_element(By.XPATH, var.phanloaihanga).send_keys(data['quanlysanpham']['themspmoi_phanloaihang1a'])
        driver.find_element(By.XPATH, var.phanloaihangb).send_keys(data['quanlysanpham']['themspmoi_phanloaihang1b'])
        driver.find_element(By.XPATH, var.phanloaihangc).send_keys(data['quanlysanpham']['themspmoi_phanloaihang1C'])
        time.sleep(1)

        #Nhóm phân loại 2
        driver.find_element(By.XPATH, var.themsanpham_themnhomphanloai2).click()
        driver.find_element(By.XPATH, var.themsanpham_nhomphanloai2).send_keys(data['quanlysanpham']['themspmoi_nhomphanloai2'])
        driver.find_element(By.XPATH, var.phanloaihang2a).send_keys(data['quanlysanpham']['themspmoi_phanloaihang2a'])
        driver.find_element(By.XPATH, var.phanloaihang2b).send_keys(data['quanlysanpham']['themspmoi_phanloaihang2b'])
        driver.find_element(By.XPATH, var.phanloaihang2c).send_keys(data['quanlysanpham']['themspmoi_phanloaihang2c'])

        #Danh sách phân loại
        driver.find_element(By.XPATH, var.themsanpham_gia).send_keys(data['quanlysanpham']['themspmoi_gia'])
        driver.find_element(By.XPATH, var.themsanpham_tonkho).send_keys(data['quanlysanpham']['themspmoi_tonkho'])
        driver.find_element(By.XPATH, var.themsanpham_sku).send_keys(data['quanlysanpham']['themspmoi_sku'])
        driver.find_element(By.XPATH, var.apdungchotatca).click()
        time.sleep(5)
        #Xám
        driver.find_element(By.XPATH, var.dsphanloai_tailenanh_xam).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/hoodie1.exe")
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.dsphanloai_xam_gia_l)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.dsphanloai_xam_gia_l).send_keys(data['quanlysanpham']['dsphanloai_gia_l'])
        xoa = driver.find_element(By.XPATH, var.dsphanloai_xam_gia_xl)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.dsphanloai_xam_gia_xl).send_keys(data['quanlysanpham']['dsphanloai_gia_xl'])
        #Lục
        driver.find_element(By.XPATH, var.dsphanloai_tailenanh_luc).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/hoodie3.exe")
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.dsphanloai_luc_gia_l)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.dsphanloai_luc_gia_l).send_keys(data['quanlysanpham']['dsphanloai_gia_l'])
        xoa = driver.find_element(By.XPATH, var.dsphanloai_luc_gia_xl)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.dsphanloai_luc_gia_xl).send_keys(data['quanlysanpham']['dsphanloai_gia_xl'])
        #Xanh
        driver.find_element(By.XPATH, var.dsphanloai_tailenanh_xanh).click()
        time.sleep(1)
        subprocess.Popen("C:/Users/Admin/PycharmProjects/pythonProject/import/hoodie4.exe")
        time.sleep(1)
        xoa = driver.find_element(By.XPATH, var.dsphanloai_xanh_gia_l)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.dsphanloai_xanh_gia_l).send_keys(data['quanlysanpham']['dsphanloai_gia_l'])
        xoa = driver.find_element(By.XPATH, var.dsphanloai_xanh_gia_xl)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.dsphanloai_xanh_gia_xl).send_keys(data['quanlysanpham']['dsphanloai_gia_xl'])

        #Vận chuyển
        driver.find_element(By.XPATH, var.themsanpham_cannang).send_keys(data['quanlysanpham']['themspmoi_cannang'])
        driver.find_element(By.XPATH, var.themsanpham_kichthuocdonggoi_r).send_keys(data['quanlysanpham']['themspmoi_kichthuoc_r'])
        driver.find_element(By.XPATH, var.themsanpham_kichthuocdonggoi_d).send_keys(data['quanlysanpham']['themspmoi_kichthuoc_d'])
        driver.find_element(By.XPATH, var.themsanpham_kichthuocdonggoi_c).send_keys(data['quanlysanpham']['themspmoi_kichthuoc_c'])

        #Thông tin khác
        driver.find_element(By.XPATH, var.themsanpham_hangdattruoc_dongy).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.themsanpham_hangdattruoc_khongdongy).click()

        driver.find_element(By.XPATH, var.themsanpham_tinhtrang_daquasudung).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.themsanpham_tinhtrang_moi).click()
        time.sleep(1)
        #Gửi phê duyệt - luu nhap
        driver.find_element(By.XPATH, guipheduyet_luunhap).click()

    def themsanphammoi_guipheduyet(self):
        quanlysanpham.themsanphammoi(self, var.guipheduyet)
        driver.implicitly_wait(7)
        try:
            check_message_themmoisanpham = driver.find_element(By.XPATH, var.check_message_themmoisanpham).text
            logging.info("Người bán - Quản lý sản phẩm - Thêm sản phẩm mới")
            logging.info("check font-end: Message Gửi phê duyệt - Tạo sản phẩm thành công!")
            logging.info(check_message_themmoisanpham == "Tạo sản phẩm thành công!")
        except NoSuchElementException:
            check_message_themmoisanpham = driver.find_element(By.XPATH, var.check_message_themmoisanpham).text
            logging.info("Người bán - Quản lý sản phẩm - Thêm sản phẩm mới")
            logging.info("check font-end: Message Gửi phê duyệt - Tạo sản phẩm thành công!")
            logging.info("False")
        time.sleep(2)

    def themsanphammoi_luunhap(self):
        quanlysanpham.themsanphammoi(self, var.luunhap)
        driver.implicitly_wait(7)
        try:
            check_message_themmoisanpham_luunhap = driver.find_element(By.XPATH,
                                                                       var.check_message_themmoisanpham_luunhap)
            logging.info("Người bán - Quản lý sản phẩm - Thêm sản phẩm mới")
            logging.info("check font-end: Message Lưu nháp - Lưu sản phẩm nháp thành công!")
            logging.info(check_message_themmoisanpham_luunhap.is_displayed())
            logging.info(check_message_themmoisanpham_luunhap.text == "Lưu sản phẩm nháp thành công!")
        except NoSuchElementException:
            logging.info("Người bán - Quản lý sản phẩm - Thêm sản phẩm mới")
            logging.info("check font-end: Message Lưu nháp - Lưu sản phẩm nháp thành công!")
            logging.info("False")
        time.sleep(2)
        driver.find_element(By.XPATH, var.danhsachsanpham).click()
        driver.find_element(By.XPATH, var.roikhoi).click()
        time.sleep(2)
        check_danhsachsanpham_tensp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_tensp1).text
        check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_trangthaisp1).text
        if check_danhsachsanpham_tensp1 == data['quanlysanpham'][
            'themspmoi_tensanpham'] and check_danhsachsanpham_trangthaisp1 == "Nháp":
            driver.find_element(By.XPATH, var.danhsachsanpham_sanpham1_dau3cham).click()
            driver.find_element(By.XPATH, var.danhsachsanpham_sanpham1_dau3cham_xoa).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.danhsachsanpham_sanpham1_dau3cham_xoa_xoa).click()
            check_message_danhsachsanpham_xoasp = driver.find_element(By.XPATH,
                                                                      var.check_message_danhsachsanpham_xoasp).text
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Message Xóa Nháp - Xóa sản phẩm thành công")
            logging.info(check_message_danhsachsanpham_xoasp == "Xóa sản phẩm thành công")
        else:
            logging.info("Người bán - Quản lý sản phẩm - Danh sách sản phẩm")
            logging.info("check font-end: Có hiển thị bản nháp không")
            logging.info("False")
        time.sleep(1)

    def danhsachsanpham(self):
        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlysanpham).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham).click()
        time.sleep(1)
        # Check các trạng thái
        # Tất cả
        check_danhsachsanpham_tensp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_tensp1)
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - có hiển thị sản phẩm không?")
        logging.info(check_danhsachsanpham_tensp1.text)
        logging.info(check_danhsachsanpham_tensp1.is_displayed())

        check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_trangthaisp1)
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - có hiển thị trạng thái sản phẩm không")
        logging.info(check_danhsachsanpham_trangthaisp1.text)
        logging.info(check_danhsachsanpham_trangthaisp1.is_displayed())
        time.sleep(1)

        # Đang hoạt động
        driver.find_element(By.XPATH, var.danhsachsanpham_danghoatdong).click()
        check_danhsachsanpham_tensp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_tensp1)
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Đang hoạt động - có hiển thị sản phẩm không?")
        logging.info(check_danhsachsanpham_tensp1.text)
        logging.info(check_danhsachsanpham_tensp1.is_displayed())

        check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH,
                                                                 var.check_danhsachsanpham_trangthaisp1_danghoatdong).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Đang hoạt động - Trạng thái - Đang hoạt động")
        logging.info(check_danhsachsanpham_trangthaisp1)
        logging.info(check_danhsachsanpham_trangthaisp1 == "Đang hoạt động")
        time.sleep(1)

        # Chờ duyệt
        driver.find_element(By.XPATH, var.danhsachsanpham_choduyet).click()
        check_danhsachsanpham_tensp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_tensp1)
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Chờ duyệt - có hiển thị sản phẩm không?")
        logging.info(check_danhsachsanpham_tensp1.text)
        logging.info(check_danhsachsanpham_tensp1.is_displayed())

        check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH,var.check_danhsachsanpham_trangthaisp1_choduyet).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Chờ duyệt - Trạng thái - Chờ duyệt")
        logging.info(check_danhsachsanpham_trangthaisp1)
        logging.info(check_danhsachsanpham_trangthaisp1 == "Chờ duyệt")
        time.sleep(1)

        # Trả lại
        driver.find_element(By.XPATH, var.danhsachsanpham_tralai).click()
        check_danhsachsanpham_tensp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_tensp1)
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Trả lại - có hiển thị sản phẩm không?")
        logging.info(check_danhsachsanpham_tensp1.text)
        logging.info(check_danhsachsanpham_tensp1.is_displayed())

        check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH,var.check_danhsachsanpham_trangthaisp1_tralai).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Trả lại - Trạng thái - Trả lại")
        logging.info(check_danhsachsanpham_trangthaisp1)
        logging.info(check_danhsachsanpham_trangthaisp1 == "Trả lại")
        time.sleep(1)

        # Nháp
        driver.find_element(By.XPATH, var.danhsachsanpham_nhap).click()
        check_danhsachsanpham_tensp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_tensp1)
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Nháp - có hiển thị sản phẩm không?")
        logging.info(check_danhsachsanpham_tensp1.text)
        logging.info(check_danhsachsanpham_tensp1.is_displayed())

        check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH,var.check_danhsachsanpham_trangthaisp1_nhap).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Nháp - Trạng thái - Nháp")
        logging.info(check_danhsachsanpham_trangthaisp1)
        logging.info(check_danhsachsanpham_trangthaisp1 == "Nháp")
        time.sleep(1)

        # Đã ẩn
        driver.find_element(By.XPATH, var.danhsachsanpham_daan).click()
        check_danhsachsanpham_tensp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_tensp1)
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Đã ẩn - có hiển thị sản phẩm không?")
        logging.info(check_danhsachsanpham_tensp1.text)
        logging.info(check_danhsachsanpham_tensp1.is_displayed())

        check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH,var.check_danhsachsanpham_trangthaisp1_daan).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Đã ẩn - Trạng thái - Đã ẩn")
        logging.info(check_danhsachsanpham_trangthaisp1)
        logging.info(check_danhsachsanpham_trangthaisp1 == "Đã ẩn")
        time.sleep(1)

        # Vi phạm
        driver.find_element(By.XPATH, var.danhsachsanpham_vipham).click()
        time.sleep(1.5)
        # Hạn chế hiển thị
        driver.find_element(By.XPATH, var.danhsachsanpham_vipham_hanchehienthi).click()
        time.sleep(1.5)
        check_danhsachsanpham_vipham_tensp1 = driver.find_element(By.XPATH,var.check_danhsachsanpham_vipham_tensp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm - Vi phạm")
        logging.info(
            "check font-end: Hạn chế hiển thị - tên sp - 🌸Kẹp Tóc Hình Trái Tim Màu Hồng Ngọt Ngào Dễ Thương Cho Y2K 2023 YINSAF🌸")
        logging.info(check_danhsachsanpham_vipham_tensp1)
        logging.info(
            check_danhsachsanpham_vipham_tensp1 == "🌸Kẹp Tóc Hình Trái Tim Màu Hồng Ngọt Ngào Dễ Thương Cho Y2K 2023 YINSAF🌸")

        # Đã tạm khóa
        driver.find_element(By.XPATH, var.danhsachsanpham_vipham_datamkhoa).click()
        time.sleep(2.5)
        check_danhsachsanpham_vipham_tensp1 = driver.find_element(By.XPATH,var.check_danhsachsanpham_vipham_tensp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm - Vi phạm")
        logging.info("check font-end: Đã tạm khóa - tên sp - " + data['quanlysanpham']['vipham_datamkhoa_dam'])
        logging.info(check_danhsachsanpham_vipham_tensp1)
        logging.info(check_danhsachsanpham_vipham_tensp1 == data['quanlysanpham']['vipham_datamkhoa_dam'])
        driver.find_element(By.XPATH, var.nhaplai).click()

        # Tìm kiếm - Tên sản phẩm
        driver.find_element(By.XPATH, var.danhsachsanpham_vipham_tensanpham_input).send_keys(data['quanlysanpham']['vipham_datamkhoa_son'])
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(1.5)
        check_danhsachsanpham_vipham_tensp1 = driver.find_element(By.XPATH,
                                                                  var.check_danhsachsanpham_vipham_tensp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm - Vi phạm")
        logging.info("check font-end: Tìm kiếm  - Tên sản phẩm  - " + data['quanlysanpham']['vipham_datamkhoa_son'])
        logging.info(check_danhsachsanpham_vipham_tensp1)
        logging.info(check_danhsachsanpham_vipham_tensp1 == data['quanlysanpham']['vipham_datamkhoa_son'])
        driver.find_element(By.XPATH, var.nhaplai).click()
        time.sleep(1)

        # Loại vi phạm
        # driver.find_element(By.XPATH, var.danhsachsanpham_vipham_loaivipham).click()        #Chức năng ko hoạt động

        driver.find_element(By.XPATH, var.capnhat).click()
        time.sleep(1)
        check_danhsachsanpham_vipham_capnhat = driver.find_element(By.XPATH,var.check_danhsachsanpham_vipham_capnhat).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Vi phạm - Cập nhật - Có chuyển tới trang cập nhật sản phẩm không")
        logging.info(check_danhsachsanpham_vipham_capnhat)
        logging.info(check_danhsachsanpham_vipham_capnhat == "Cập nhật sản phẩm")
        time.sleep(1)
        driver.back()
        driver.find_element(By.XPATH, var.roikhoi).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.danhsachsanpham_vipham).click()
        time.sleep(1)
        # driver.find_element(By.XPATH, var.an).click()       #ds sản phẩm - vi pham - ẩn => ko hoạt động
        # time.sleep(1)
        # driver.find_element(By.XPATH, var.xoa).click()      ##ds sản phẩm - vi pham - xóa => ko hoạt động
        # time.sleep(1)

        # Tất cả
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca).click()
        time.sleep(1.5)
        # Tên/SKU sản phẩm
        # SKU sản phẩm
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku).click()
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_sku).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input).send_keys(data['quanlysanpham']['themspmoi_sku'])
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(2)
        check_danhsachsanpham_sp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - Tìm kiếm SKU - " + data['quanlysanpham']['themspmoi_sku'])
        logging.info(check_danhsachsanpham_sp1)
        logging.info(check_danhsachsanpham_sp1 != None)
        driver.find_element(By.XPATH, var.nhaplai).click()
        time.sleep(1)

        # Tên sản phẩm
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku).click()
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_tensp).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input).send_keys(data['quanlysanpham']['tatca_timkiemsp'])
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(2)
        check_danhsachsanpham_sp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - Tìm kiếm sản phẩm - " + data['quanlysanpham']['tatca_timkiemsp'])
        logging.info(check_danhsachsanpham_sp1)
        logging.info(check_danhsachsanpham_sp1 == data['quanlysanpham']['tatca_timkiemsp'])
        driver.find_element(By.XPATH, var.nhaplai).click()
        time.sleep(1)

        # Nghành hàng
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_nghanhhang).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsuanghanhhang_thoitrangnu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsuanghanhhang_ao).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chinhsuanghanhhang_aoong).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.luu).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(2)
        check_danhsachsanpham_sp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - Tìm kiếm Nghành hàng - " + data['quanlysanpham']['tatca_timkiemsp'])
        logging.info(check_danhsachsanpham_sp1)
        logging.info(check_danhsachsanpham_sp1 == data['quanlysanpham']['tatca_timkiemsp'])
        driver.find_element(By.XPATH, var.nhaplai).click()
        time.sleep(1)

        # Sắp xếp
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_sapxep).click()
        time.sleep(0.5)
        # Giá cao tới thấp
        driver.find_element(By.XPATH, var.danhsachsanpham_sapxep_giacaotoithap).click()
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(3)
        check_danhsachsanpham_giacaotoithap1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - Sắp xếp - Giá cao tới thấp - " + data['quanlysanpham']['vipham_datamkhoa_son'])
        logging.info(check_danhsachsanpham_giacaotoithap1)
        logging.info(check_danhsachsanpham_giacaotoithap1 == data['quanlysanpham']['vipham_datamkhoa_son'])
        driver.find_element(By.XPATH, var.nhaplai).click()

        # Giá thấp tới cao
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_sapxep).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham_sapxep_giathaptoicao).click()
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(3)
        check_danhsachsanpham_giathaptoicao1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - Sắp xếp - Giá thấp tới cao - 🌸Kẹp Tóc Hình Trái Tim Màu Hồng Ngọt Ngào Dễ Thương Cho Y2K 2023 YINSAF🌸")
        logging.info(check_danhsachsanpham_giathaptoicao1)
        logging.info(check_danhsachsanpham_giathaptoicao1 == "🌸Kẹp Tóc Hình Trái Tim Màu Hồng Ngọt Ngào Dễ Thương Cho Y2K 2023 YINSAF🌸")
        driver.find_element(By.XPATH, var.nhaplai).click()

        # Số lượng ít
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_sapxep).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham_sapxep_soluongit).click()
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(3)
        check_danhsachsanpham_soluongit1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - Sắp xếp - Số lượng ít - " + data['quanlysanpham']['sanpham_hethang'])
        logging.info(check_danhsachsanpham_soluongit1)
        logging.info(check_danhsachsanpham_soluongit1 == "Váy Trễ Vai Voan 2 Lớp Dáng Xoè Nữ Màu Trắng Dáng Tiểu Thư")
        driver.find_element(By.XPATH, var.nhaplai).click()
        time.sleep(1)

        # Cập nhật cũ nhất
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_sapxep).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham_sapxep_capnhatcunhat).click()
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(3)
        check_danhsachsanpham_capnhatcunhat = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - Sắp xếp - Cập nhật cũ nhất - áo phông nam")
        logging.info(check_danhsachsanpham_capnhatcunhat)
        logging.info(check_danhsachsanpham_capnhatcunhat == "áo phông nam")
        driver.find_element(By.XPATH, var.nhaplai).click()

        # Cập nhật gần nhất
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_sapxep).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham_sapxep_capnhatgannhat).click()
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(3)
        check_danhsachsanpham_capnhatgannhat = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - Sắp xếp - Cập nhật gần nhất - Có hiển thị sản phẩm không?")
        logging.info(check_danhsachsanpham_capnhatgannhat)
        logging.info(check_danhsachsanpham_capnhatgannhat != None)
        driver.find_element(By.XPATH, var.nhaplai).click()

        # Tạo sớm nhất
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_sapxep).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham_sapxep_taosomnhat).click()
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(3)
        check_danhsachsanpham_taosomnhat1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - Sắp xếp - Tạo sớm nhất - Áo polo nam")
        logging.info(check_danhsachsanpham_taosomnhat1)
        logging.info(check_danhsachsanpham_taosomnhat1 == "Áo polo nam")

        # Tạo gần nhất
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_sapxep).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham_sapxep_taogannhat).click()
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(2.5)
        check_danhsachsanpham_taogannhat1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Tất cả - Sắp xếp - Tạo gần nhất - Có hiển thị sản phẩm không?")
        logging.info(check_danhsachsanpham_taogannhat1)
        logging.info(check_danhsachsanpham_taogannhat1 != None)
        driver.find_element(By.XPATH, var.nhaplai).click()
        time.sleep(1)

        # Thêm sản phẩm mới
        driver.find_element(By.XPATH, var.danhsachsanpham_themsanphammoi).click()
        time.sleep(1)
        check_danhsachsanpham_themsanphammoi = driver.find_element(By.XPATH,
                                                                   var.check_danhsachsanpham_themsanphammoi).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Thêm sản phẩm mới - Có chuyển tới trang Thêm sản phẩm mới không?")
        logging.info(check_danhsachsanpham_themsanphammoi)
        logging.info(check_danhsachsanpham_themsanphammoi == "Thêm sản phẩm")
        time.sleep(1)
        driver.back()
        time.sleep(2)

        # Danh sách sản phẩm - Dấu 3 chấm
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku).click()
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_tensp).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input).send_keys(data['quanlysanpham']['dssanpham_dau3cham'])
        driver.find_element(By.XPATH, var.timkiem).click()
        time.sleep(2)
        check_danhsachsanpham_sp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_sp1).text
        if check_danhsachsanpham_sp1 == data['quanlysanpham']['dssanpham_dau3cham']:
            # Ẩn sản phẩm
            driver.find_element(By.XPATH, var.danhsachsanpham_dau3cham).click()
            driver.find_element(By.XPATH, var.danhsachsanpham_dau3cham_an).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.luu).click()
            check_message_ansanpham = driver.find_element(By.XPATH, var.check_message_ansanpham).text
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Dấu 3 chấm - Message Ẩn sản phẩm - Ẩn sản phẩm thành công")
            logging.info(check_message_ansanpham == "Ẩn sản phẩm thành công")
            time.sleep(1)
            check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH,var.check_danhsachsanpham_trangthaisp1_daan).text
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Ẩn sản phẩm - Trạng thái - Đã ẩn")
            logging.info(check_danhsachsanpham_trangthaisp1)
            logging.info(check_danhsachsanpham_trangthaisp1 == "Đã ẩn")
            time.sleep(1)

            driver.get("https://cmc-fe.emso.vn/product/747/about")
            time.sleep(2)
            check_danhsachsanpham_ansanpham = driver.find_element(By.XPATH, var.trangnaykhonghienthi).text
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Dấu 3 chấm - Ẩn - Có ẩn được sản phẩm không")
            logging.info(check_danhsachsanpham_ansanpham == "Trang này không hiển thị")
            driver.back()
            time.sleep(2)

            # Cập nhật
            xoa = driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input)
            xoa.send_keys(Keys.CONTROL, "a")
            driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input).send_keys(data['quanlysanpham']['dssanpham_dau3cham'])
            driver.find_element(By.XPATH, var.timkiem).click()
            time.sleep(2)
            driver.find_element(By.XPATH, var.danhsachsanpham_dau3cham).click()
            driver.find_element(By.XPATH, var.danhsachsanpham_dau3cham_capnhat).click()
            time.sleep(2)
            check_danhsachsanpham_capnhat = driver.find_element(By.XPATH, var.check_danhsachsanpham_capnhat).text
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Cập nhật - Có chuyển tới trang Cập nhật sản phẩm không?")
            logging.info(check_danhsachsanpham_capnhat)
            logging.info(check_danhsachsanpham_capnhat == "Cập nhật sản phẩm")
            driver.back()
            driver.find_element(By.XPATH, var.roikhoi).click()
            time.sleep(2)

            # Hiển sản phẩm
            xoa = driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input)
            xoa.send_keys(Keys.CONTROL, "a")
            driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input).send_keys(data['quanlysanpham']['dssanpham_dau3cham'])
            driver.find_element(By.XPATH, var.timkiem).click()
            time.sleep(2)
            driver.find_element(By.XPATH, var.danhsachsanpham_dau3cham).click()
            driver.find_element(By.XPATH, var.danhsachsanpham_dau3cham_hien).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.luu).click()
            check_message_hiensanpham = driver.find_element(By.XPATH, var.check_message_hiensanpham).text
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Dấu 3 chấm - Message Hiện sản phẩm - Hiện sản phẩm thành công")
            logging.info(check_message_hiensanpham == "Hiện sản phẩm thành công")
            time.sleep(1)
            check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH,var.check_danhsachsanpham_trangthaisp1_danghoatdong).text
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Hiện sản phẩm - Trạng thái - Đang hoạt động")
            logging.info(check_danhsachsanpham_trangthaisp1)
            logging.info(check_danhsachsanpham_trangthaisp1 == "Đang hoạt động")
            time.sleep(1)

            # Sao chép liên kết
            xoa = driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input)
            xoa.send_keys(Keys.CONTROL, "a")
            driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input).send_keys(data['quanlysanpham']['dssanpham_dau3cham'])
            driver.find_element(By.XPATH, var.timkiem).click()
            time.sleep(2)
            driver.find_element(By.XPATH, var.danhsachsanpham_dau3cham).click()
            driver.find_element(By.XPATH, var.danhsachsanpham_dau3cham_saocheplienket).click()
            logging.info(Keys.CONTROL + "v")
            time.sleep(2)
            # check_danhsachsanpham_saocheplienket = driver.find_element(By.XPATH,var.check_danhsachsanpham_saocheplienket).text
            # logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            # logging.info("check font-end: Sao chép liên kết - Có chuyển tới trang liên kết vừa coppy hay không?")
            # logging.info(check_danhsachsanpham_saocheplienket)
            # logging.info(check_danhsachsanpham_saocheplienket == data['quanlysanpham']['dssanpham_dau3cham'])
            # time.sleep(1)
            # driver.back()
            # time.sleep(2)

            # Xem trước
            xoa = driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input)
            xoa.send_keys(Keys.CONTROL, "a")
            driver.find_element(By.XPATH, var.danhsachsanpham_tatca_tensp_sku_input).send_keys(data['quanlysanpham']['dssanpham_dau3cham'])
            driver.find_element(By.XPATH, var.timkiem).click()
            time.sleep(2)
            driver.find_element(By.XPATH, var.danhsachsanpham_dau3cham).click()
            driver.find_element(By.XPATH, var.danhsachsanpham_dau3cham_xemtruoc).click()
            time.sleep(1.5)
            check_danhsachsanpham_xemtruoc = driver.find_element(By.XPATH,var.check_danhsachsanpham_saocheplienket).text
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Xem trước - Có chuyển tới trang Xem trước sản phẩm hay không?")
            logging.info(check_danhsachsanpham_xemtruoc)
            logging.info(check_danhsachsanpham_xemtruoc == data['quanlysanpham']['dssanpham_dau3cham'])
        else:
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Dấu 3 chấm - Không tìm thấy sản phẩm")
            logging.info("False")
        time.sleep(1)


class add_dulieuemso():
    def themsanphammoi1(self):
        driver.implicitly_wait(15)
        login.login4(self, "emsomanagerhd@gmail.com", "khongnhomatkhaucu")
        time.sleep(1.5)
        hang = 135
        while hang<141:
            hang += 1
            tensanpham = readData(var.path_datamarket, 'Sheet1', hang, 1)
            nghanhcha = readData(var.path_datamarket, 'Sheet1', hang, 2)
            data_nghanhcha = "//*[text()='"+nghanhcha+"']"
            conbac1 = readData(var.path_datamarket, 'Sheet1', hang, 3)
            data_conbac1 = "//*[text()='"+conbac1+"']"
            conbac2 = readData(var.path_datamarket, 'Sheet1', hang, 4)
            data_conbac2 = "//*[text()='"+conbac2+"']"
            conbac3 = readData(var.path_datamarket, 'Sheet1', hang, 5)
            data_conbac3 = "//*[text()='"+conbac3+"']"

            mota = readData(var.path_datamarket, 'Sheet1', hang, 6)
            thuonghieu = readData(var.path_datamarket, 'Sheet1', hang, 7)
            data_thuonghieu = "//*[text()='"+thuonghieu+"']"
            gia = readData(var.path_datamarket, 'Sheet1', hang, 8)
            tonkho = readData(var.path_datamarket, 'Sheet1', hang, 9)
            cannang = readData(var.path_datamarket, 'Sheet1', hang, 10)
            anh1 = readData(var.path_datamarket, 'Sheet1', hang, 11)
            anh2 = readData(var.path_datamarket, 'Sheet1', hang, 12)
            anh3 = readData(var.path_datamarket, 'Sheet1', hang, 13)
            sku = readData(var.path_datamarket, 'Sheet1', hang, 14)
            kichthuocdonggoi_r = readData(var.path_datamarket, 'Sheet1', hang, 15)
            kichthuocdonggoi_d = readData(var.path_datamarket, 'Sheet1', hang, 16)
            kichthuocdonggoi_c = readData(var.path_datamarket, 'Sheet1', hang, 17)
            dathangtruoc = readData(var.path_datamarket, 'Sheet1', hang, 18)
            tinhtrang = readData(var.path_datamarket, 'Sheet1', hang, 19)


            driver.get("https://cmc-fe.emso.vn/marketplace/shop/create_product?page_id=108277159419224061")
            time.sleep(2)
            #Thông tin cơ bản
            driver.find_element(By.XPATH, var.themsanpham_tensanpham).send_keys(tensanpham[0:115])
            time.sleep(1)

            #Co nghanh cha con
            driver.find_element(By.XPATH, var.themsanpham_nghanhhang).click()
            time.sleep(1)
            driver.find_element(By.XPATH, data_nghanhcha).click()
            time.sleep(1)
            driver.find_element(By.XPATH, data_conbac1).click()
            time.sleep(1)
            driver.implicitly_wait(1)
            try:
                driver.find_element(By.XPATH, data_conbac2).click()
            except:
                pass
            time.sleep(1)

            try:
                driver.find_element(By.XPATH, data_conbac3).click()
            except:
                pass
            driver.implicitly_wait(15)
            time.sleep(1)
            driver.find_element(By.XPATH, var.luu).click()
            time.sleep(1)

            # Có nghành cha
            driver.find_element(By.XPATH, var.themsanpham_themhinhanh1).send_keys(anh1)
            time.sleep(1)
            driver.find_element(By.XPATH, var.themsanpham_themhinhanh2).send_keys(anh2)
            time.sleep(1)
            actions = ActionChains(driver)
            hover_themmoisp_anh = driver.find_element(By.XPATH, var.hover_themmoisp_anh)
            actions.move_to_element(hover_themmoisp_anh).perform()
            driver.find_element(By.XPATH, "//*[@class='fa-regular fa-xmark']").click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.themsanpham_themhinhanh3).send_keys(anh3)
            time.sleep(1)
            hover_themmoisp_anh4 = driver.find_element(By.XPATH, var.hover_themmoisp_anh4)
            actions.move_to_element(hover_themmoisp_anh4).perform()
            button = driver.find_element(By.XPATH, "//*[@class='fa-regular fa-xmark']")
            driver.execute_script("arguments[0].click();", button)
            time.sleep(1)
            actions.move_to_element(hover_themmoisp_anh4).perform()
            button = driver.find_element(By.XPATH, "//*[@class='fa-regular fa-xmark']")
            driver.execute_script("arguments[0].click();", button)

            # #khong co nghanh cha con
            # driver.find_element(By.XPATH, var.themsanpham_nghanh_goiy1).click()
            # time.sleep(1.5)
            # driver.find_element(By.XPATH, var.themsanpham_themhinhanh1a).send_keys(anh1)
            # time.sleep(1)
            # driver.find_element(By.XPATH, var.themsanpham_themhinhanh2a).send_keys(anh2)
            # time.sleep(1)
            # driver.find_element(By.XPATH, var.themsanpham_themhinhanh3a).send_keys(anh3)
            # time.sleep(1)
            driver.find_element(By.XPATH, var.themsanpham_mota).send_keys(".")
            JS_ADD_TEXT_TO_INPUT = """
              elm = arguments[0], txt = arguments[1];
              elm.value += txt;
              elm.dispatchEvent(new Event('change'));
              """
            elm = driver.find_element(By.XPATH, var.themsanpham_mota)
            driver.execute_script(JS_ADD_TEXT_TO_INPUT, elm, mota[0:2950])
            driver.find_element(By.XPATH, var.themsanpham_mota).send_keys(".")
            time.sleep(2)


            # Thông tin chi tiết
            # driver.execute_script("window.scrollBy(0,400)", "")

            time.sleep(1)
            driver.find_element(By.XPATH, var.themsanpham_thuonghieu).click()
            driver.find_element(By.XPATH, data_thuonghieu).click()

            driver.implicitly_wait(2)
            try:
                driver.find_element(By.XPATH, var.thongtinchitiet_button2).click()
                # driver.find_element(By.XPATH, var.thongtinchitiet_button2_chon1).click()
                button = driver.find_element(By.XPATH, var.thongtinchitiet_button2_chon1)
                driver.execute_script("arguments[0].click();", button)
            except:
                pass

            try:
                driver.find_element(By.XPATH, var.thongtinchitiet_button3).click()
                # driver.find_element(By.XPATH, var.thongtinchitiet_button3_chon1).click()
                button = driver.find_element(By.XPATH, var.thongtinchitiet_button3_chon1)
                driver.execute_script("arguments[0].click();", button)
            except:
                pass
            driver.implicitly_wait(15)
            #Thông tin bán hàng
            driver.find_element(By.XPATH, var.themsanpham_gia1).send_keys(gia)
            driver.find_element(By.XPATH, var.themsanpham_tonkho1).send_keys(tonkho)
            driver.find_element(By.XPATH, var.themsanpham_sku1).send_keys(sku)

            # Vận chuyển
            driver.find_element(By.XPATH, var.themsanpham_cannang).send_keys(cannang)
            driver.find_element(By.XPATH, var.themsanpham_kichthuocdonggoi_r).send_keys(kichthuocdonggoi_r)
            driver.find_element(By.XPATH, var.themsanpham_kichthuocdonggoi_d).send_keys(kichthuocdonggoi_d)
            driver.find_element(By.XPATH, var.themsanpham_kichthuocdonggoi_c).send_keys(kichthuocdonggoi_c)
            time.sleep(1)

            # Thông tin khác
            if dathangtruoc == "Đồng ý":
                driver.find_element(By.XPATH, var.themsanpham_hangdattruoc_dongy).click()
            else:
                driver.find_element(By.XPATH, var.themsanpham_hangdattruoc_khongdongy).click()
            time.sleep(1)
            if tinhtrang == "Đã qua sử dụng":
                driver.find_element(By.XPATH, var.themsanpham_tinhtrang_daquasudung).click()
            else:
                driver.find_element(By.XPATH, var.themsanpham_tinhtrang_moi).click()
            time.sleep(1)

            # # Gửi phê duyệt
            driver.execute_script("window.scrollBy(0,700)", "")
            button = driver.find_element(By.XPATH, var.guipheduyet)
            driver.execute_script("arguments[0].click();", button)
            time.sleep(5)
            print("da tai len san pham hang: ", hang)

            # driver.find_element(By.XPATH, var.quanlydonhang).click()
            # time.sleep(0.5)
            # driver.find_element(By.XPATH, var.roikhoi).click()
            # time.sleep(1)

    def themsanphammoi1_khongconghanhcha(self):
        driver.implicitly_wait(15)
        login.login4(self, "emsomanagerhd@gmail.com", "khongnhomatkhaucu")
        time.sleep(1.5)
        hang = 58
        while hang < 59:
            hang += 1
            tensanpham = readData(var.path_datamarket, 'Sheet1', hang, 1)
            nghanhcha = readData(var.path_datamarket, 'Sheet1', hang, 2)
            data_nghanhcha = "//*[text()='" + nghanhcha + "']"
            conbac1 = readData(var.path_datamarket, 'Sheet1', hang, 3)
            data_conbac1 = "//*[text()='" + conbac1 + "']"
            conbac2 = readData(var.path_datamarket, 'Sheet1', hang, 4)
            data_conbac2 = "//*[text()='" + conbac2 + "']"
            conbac3 = readData(var.path_datamarket, 'Sheet1', hang, 5)
            data_conbac3 = "//*[text()='" + conbac3 + "']"

            mota = readData(var.path_datamarket, 'Sheet1', hang, 6)
            thuonghieu = readData(var.path_datamarket, 'Sheet1', hang, 7)
            data_thuonghieu = "//*[text()='" + thuonghieu + "']"
            gia = readData(var.path_datamarket, 'Sheet1', hang, 8)
            tonkho = readData(var.path_datamarket, 'Sheet1', hang, 9)
            cannang = readData(var.path_datamarket, 'Sheet1', hang, 10)
            anh1 = readData(var.path_datamarket, 'Sheet1', hang, 11)
            anh2 = readData(var.path_datamarket, 'Sheet1', hang, 12)
            anh3 = readData(var.path_datamarket, 'Sheet1', hang, 13)
            sku = readData(var.path_datamarket, 'Sheet1', hang, 14)
            kichthuocdonggoi_r = readData(var.path_datamarket, 'Sheet1', hang, 15)
            kichthuocdonggoi_d = readData(var.path_datamarket, 'Sheet1', hang, 16)
            kichthuocdonggoi_c = readData(var.path_datamarket, 'Sheet1', hang, 17)
            dathangtruoc = readData(var.path_datamarket, 'Sheet1', hang, 18)
            tinhtrang = readData(var.path_datamarket, 'Sheet1', hang, 19)

            driver.get("https://cmc-fe.emso.vn/marketplace/shop/create_product?page_id=108277159419223806")
            time.sleep(2)
            # Thông tin cơ bản
            driver.find_element(By.XPATH, var.themsanpham_tensanpham).send_keys(tensanpham)
            time.sleep(1)
            driver.find_element(By.XPATH, var.themsanpham_nghanh_goiy1).click()
            time.sleep(1)
            driver.implicitly_wait(15)

            # actions = ActionChains(driver)
            # actions.move_to_element(hover_themmoisp_anh4).perform()
            # button = driver.find_element(By.XPATH, "//*[@class='fa-regular fa-xmark']")
            # driver.execute_script("arguments[0].click();", button)

            #khong co nghanh cha con
            driver.find_element(By.XPATH, var.themsanpham_themhinhanh1a).send_keys(anh1)
            time.sleep(1)
            driver.find_element(By.XPATH, var.themsanpham_themhinhanh2a).send_keys(anh2)
            time.sleep(1)
            actions = ActionChains(driver)
            hover_themmoisp_anh3a = driver.find_element(By.XPATH, var.hover_themmoisp_anh3a)
            actions.move_to_element(hover_themmoisp_anh3a).perform()
            button = driver.find_element(By.XPATH, "//*[@class='fa-regular fa-xmark']")
            driver.execute_script("arguments[0].click();", button)
            time.sleep(1)
            driver.find_element(By.XPATH, var.themsanpham_themhinhanh3a).send_keys(anh3)
            time.sleep(1)

            hover_themmoisp_anh4a = driver.find_element(By.XPATH, var.hover_themmoisp_anh4a)
            actions.move_to_element(hover_themmoisp_anh4a).perform()
            button = driver.find_element(By.XPATH, "//*[@class='fa-regular fa-xmark']")
            driver.execute_script("arguments[0].click();", button)
            time.sleep(1)
            hover_themmoisp_anh4a = driver.find_element(By.XPATH, var.hover_themmoisp_anh4a)
            actions.move_to_element(hover_themmoisp_anh4a).perform()
            button = driver.find_element(By.XPATH, "//*[@class='fa-regular fa-xmark']")
            driver.execute_script("arguments[0].click();", button)
            time.sleep(1)

            driver.find_element(By.XPATH, var.themsanpham_mota).send_keys(".")
            JS_ADD_TEXT_TO_INPUT = """
              elm = arguments[0], txt = arguments[1];
              elm.value += txt;
              elm.dispatchEvent(new Event('change'));
              """
            elm = driver.find_element(By.XPATH, var.themsanpham_mota)
            driver.execute_script(JS_ADD_TEXT_TO_INPUT, elm, mota[0:2950])
            driver.find_element(By.XPATH, var.themsanpham_mota).send_keys(".")
            time.sleep(2)

            # Thông tin chi tiết
            # driver.execute_script("window.scrollBy(0,400)", "")

            time.sleep(1)
            driver.find_element(By.XPATH, var.themsanpham_thuonghieu).click()
            driver.find_element(By.XPATH, data_thuonghieu).click()

            driver.implicitly_wait(2)
            try:
                driver.find_element(By.XPATH, var.thongtinchitiet_button2).click()
                # driver.find_element(By.XPATH, var.thongtinchitiet_button2_chon1).click()
                button = driver.find_element(By.XPATH, var.thongtinchitiet_button2_chon1)
                driver.execute_script("arguments[0].click();", button)
            except:
                pass

            try:
                driver.find_element(By.XPATH, var.thongtinchitiet_button3).click()
                # driver.find_element(By.XPATH, var.thongtinchitiet_button3_chon1).click()
                button = driver.find_element(By.XPATH, var.thongtinchitiet_button3_chon1)
                driver.execute_script("arguments[0].click();", button)
            except:
                pass
            driver.implicitly_wait(15)
            # Thông tin bán hàng
            driver.find_element(By.XPATH, var.themsanpham_gia1).send_keys(gia)
            driver.find_element(By.XPATH, var.themsanpham_tonkho1).send_keys(tonkho)
            driver.find_element(By.XPATH, var.themsanpham_sku1).send_keys(sku)

            # Vận chuyển
            driver.find_element(By.XPATH, var.themsanpham_cannang).send_keys(cannang)
            driver.find_element(By.XPATH, var.themsanpham_kichthuocdonggoi_r).send_keys(kichthuocdonggoi_r)
            driver.find_element(By.XPATH, var.themsanpham_kichthuocdonggoi_d).send_keys(kichthuocdonggoi_d)
            driver.find_element(By.XPATH, var.themsanpham_kichthuocdonggoi_c).send_keys(kichthuocdonggoi_c)
            time.sleep(1)

            # Thông tin khác
            if dathangtruoc == "Đồng ý":
                driver.find_element(By.XPATH, var.themsanpham_hangdattruoc_dongy).click()
            else:
                driver.find_element(By.XPATH, var.themsanpham_hangdattruoc_khongdongy).click()
            time.sleep(1)
            if tinhtrang == "Đã qua sử dụng":
                driver.find_element(By.XPATH, var.themsanpham_tinhtrang_daquasudung).click()
            else:
                driver.find_element(By.XPATH, var.themsanpham_tinhtrang_moi).click()
            time.sleep(1)

            # # Gửi phê duyệt
            driver.execute_script("window.scrollBy(0,700)", "")
            button = driver.find_element(By.XPATH, var.guipheduyet)
            driver.execute_script("arguments[0].click();", button)
            time.sleep(5)
            print("da tai len san pham hang: ", hang)

            # driver.find_element(By.XPATH, var.quanlydonhang).click()
            # time.sleep(0.5)
            # driver.find_element(By.XPATH, var.roikhoi).click()
            # time.sleep(1)

    def vietfilesanpham(self):
        hang = 108       #hang = 60 thì sẽ ghi ào hàng 61
        tenfile = int(input("Moi nhap so duoi file: "))
        while hang<315:
            hang += 1
            anh1 = tenfile+1
            anh1 = str(anh1)
            writeData(var.path_datamarket, "Sheet1", hang, 11, "C:/Users\Admin/PycharmProjects/pythonProject/market_anh/" + anh1+".jpg")
            anh1 = int(anh1)

            anh2 = tenfile+2
            anh2 = str(anh2)
            writeData(var.path_datamarket, "Sheet1", hang, 12, "C:/Users\Admin/PycharmProjects/pythonProject/market_anh/" + anh2+".jpg")
            anh2 = int(anh2)

            anh3 = tenfile+3
            anh3 = str(anh3)
            writeData(var.path_datamarket, "Sheet1", hang, 13, "C:/Users\Admin/PycharmProjects/pythonProject/market_anh/" + anh3+".jpg")
            anh3 = int(anh3)
            tenfile = anh3
            print("Đã ghi ào dòng", hang)

    def get_data_lazada(self):
        from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import random
        capa = DesiredCapabilities.CHROME
        capa["pageLoadStrategy"] = "none"
        driver = webdriver.Chrome(desired_capabilities=capa)

        wait = WebDriverWait(driver, 20)
        driver.implicitly_wait(30)

        driver.get("https://www.lazada.vn/?spm=a2o4n.home-vn.header.dhome.7a103bdc3mCnRj")
        driver.maximize_window()
        time.sleep(random.randint(4, 8))
        driver.execute_script("window.scrollBy(0,1700)", "")
        time.sleep(random.randint(5, 10))
        hang = 297
        while hang<1000:
            hang += 1
            a = random.randint(1, 20)
            a = str(a)
            chidanhchoban_sp1 = "//*[@id='J_JFY']/div[2]/div[1]/div/a[" + a + "]"
            try:
                driver.find_element(By.XPATH, chidanhchoban_sp1).click()
            except:
                driver.refresh()
                time.sleep(10)
                driver.find_element(By.XPATH, chidanhchoban_sp1).click()
            a = int(a)
            time.sleep(random.randint(8, 15))
            driver.execute_script("window.stop();")
            try:
                lazada_nghanhcha = driver.find_element(By.XPATH, var.lazada_nghanhcha).text
                lazada_tensp = driver.find_element(By.XPATH, var.lazada_tensp).text
                lazada_gia = driver.find_element(By.XPATH, var.lazada_gia).text
            except NoSuchElementException:
                driver.refresh()
                time.sleep(5)
            lazada_nghanhcha = driver.find_element(By.XPATH, var.lazada_nghanhcha).text
            lazada_tensp = driver.find_element(By.XPATH, var.lazada_tensp).text
            lazada_gia = driver.find_element(By.XPATH, var.lazada_gia).text

            writeData(var.path_datamarket, "Sheet1", hang, 2, lazada_nghanhcha[0:100])
            writeData(var.path_datamarket, "Sheet1", hang, 1, lazada_tensp)
            writeData(var.path_datamarket, "Sheet1", hang, 8, lazada_gia[1::])

            driver.implicitly_wait(1)
            try:
                lazada_anh1a = driver.find_element(By.XPATH, var.lazada_anh1a).get_attribute("src")
                writeData(var.path_datamarket, "Sheet1", hang, 11, lazada_anh1a[0:-6])
            except NoSuchElementException:
                lazada_anh1 = driver.find_element(By.XPATH, var.lazada_anh1).get_attribute("src")
                writeData(var.path_datamarket, "Sheet1", hang, 11, lazada_anh1[0:-6])

            driver.implicitly_wait(2)
            #Lay src ảnh 2
            try:
                actions = ActionChains(driver)
                lazada_hoveranh2 = driver.find_element(By.XPATH, var.lazada_hoveranh2)
                actions.move_to_element(lazada_hoveranh2).perform()
                time.sleep(1)
                lazada_anh2 = driver.find_element(By.XPATH, var.lazada_anh1).get_attribute("src")
                writeData(var.path_datamarket, "Sheet1", hang, 12, lazada_anh2[0:-6])
            except:
                pass
            time.sleep(1)
            #Lay src ảnh 3
            try:
                actions = ActionChains(driver)
                lazada_hoveranh3 = driver.find_element(By.XPATH, var.lazada_hoveranh3)
                actions.move_to_element(lazada_hoveranh3).perform()
                time.sleep(1)
                lazada_anh3 = driver.find_element(By.XPATH, var.lazada_anh1).get_attribute("src")
                writeData(var.path_datamarket, "Sheet1", hang, 13, lazada_anh3[0:-6])
            except:
                pass
            driver.implicitly_wait(15)
            driver.execute_script("window.scrollBy(0,1000)", "")
            #lấy mô tả
            lazada_mota = driver.find_element(By.XPATH, var.lazada_mota).text
            writeData(var.path_datamarket, "Sheet1", hang, 6, lazada_mota)
            time.sleep(3)
            driver.find_element(By.XPATH, var.ladaza_icon).click()
            time.sleep(random.randint(10, 15))
            driver.execute_script("window.stop();")
            driver.execute_script("window.scrollBy(0,2500)", "")
            time.sleep(random.randint(2, 5))
            print("da toi dong,", hang)

    def get_image(self):
        from urllib.request import urlretrieve
        import urllib.request
        driver.implicitly_wait(15)
        hang = 287       #hang = 60 thì sẽ ghi ào hàng 61
        sofile = int(input("Moi nhap so duoi file: "))
        while hang<321:
            hang += 1
            anh1 = readData(var.path_datamarket, 'Sheet1', hang, 11)
            anh2 = readData(var.path_datamarket, 'Sheet1', hang, 12)
            anh3 = readData(var.path_datamarket, 'Sheet1', hang, 13)
            tenanh1 = sofile +1
            tenanh2 = sofile +2
            tenanh3 = sofile +3

            tenanh1 = str(tenanh1)
            tenfile1 = tenanh1+".jpg"
            tenanh2 = str(tenanh2)
            tenfile2 = tenanh2+".jpg"
            tenanh3 = str(tenanh3)
            tenfile3 = tenanh3+".jpg"

            urllib.request.urlretrieve(anh1, tenfile1)
            urllib.request.urlretrieve(anh2, tenfile2)
            urllib.request.urlretrieve(anh3, tenfile3)

            # urlretrieve(anh1, tenfile1)
            # urlretrieve(anh2, tenfile2)
            # urlretrieve(anh3, tenfile3)

            tenanh1 = int(tenanh1)
            tenanh2 = int(tenanh2)
            tenanh3 = int(tenanh3)
            sofile = tenanh3
            print("da tai xong anh hang", hang)







class thuongmai():
    def duyetsanpham_tuchoi(self):
        driver.implicitly_wait(15)
        login.login5(self, "thanghoa1420@gmail.com", "hoathang1420")
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.admin_thuongmai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.thuongmai_duyetsanpham).click()
        time.sleep(2)
        duyetsanpham_tensanpham1 = driver.find_element(By.XPATH, var.duyetsanpham_tensanpham1).text
        check_duyetsanpham_trangthai1 = driver.find_element(By.XPATH, var.check_duyetsanpham_trangthai1).text
        if duyetsanpham_tensanpham1 == data['quanlysanpham']['themspmoi_tensanpham'] and check_duyetsanpham_trangthai1 == "Đang chờ":

            logging.info("Admin - Thương mại - Duyệt sản phẩm - Từ chối sp")
            logging.info("check font-end: Tên thái sản phẩm 1 - " + data['quanlysanpham']['themspmoi_tensanpham'])
            logging.info(duyetsanpham_tensanpham1)
            logging.info(duyetsanpham_tensanpham1 == data['quanlysanpham']['themspmoi_tensanpham'])

            logging.info("Admin - Thương mại - Duyệt sản phẩm - Từ chối sp")
            logging.info("check font-end: Trạng thái sản phẩm 1 - Đang chờ")
            logging.info(check_duyetsanpham_trangthai1)
            logging.info(check_duyetsanpham_trangthai1 == "Đang chờ")

            #Từ chối sản phẩm
            driver.find_element(By.XPATH, var.duyetsanpham_dau3cham).click()
            driver.find_element(By.XPATH, var.duyetsanpham_dau3cham_tuchoi).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.duyetsanpham_dau3cham_tuchoi_lydo).send_keys(data['admin']['tuchoiduyetsp_lydo'])
            driver.find_element(By.XPATH, var.duyetsanpham_dau3cham_tuchoi_tuchoi).click()
            check_message_duyetsanpham_tuchoi = driver.find_element(By.XPATH, var.check_message_duyetsanpham).text
            logging.info("Admin - Thương mại - Duyệt sản phẩm - Từ chối sp")
            logging.info("check font-end: Message từ chối sản phẩm - Dấu 3 chấm - Từ chối")
            logging.info(check_message_duyetsanpham_tuchoi == "Cập nhật thành công")

        else:
            logging.info("Admin - Thương mại - Duyệt sản phẩm - Từ chối sp")
            logging.info("check font-end: Có hiển thị sản phẩm gửi phê duyệt từ người bán không")
            logging.info(duyetsanpham_tensanpham1)
            logging.info(check_duyetsanpham_trangthai1)
            logging.info("False")
        time.sleep(1.5)

        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlysanpham).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham).click()
        time.sleep(1)
        check_danhsachsanpham_tensp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_tensp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Admin Từ chối duyệt sp- Tên sản phẩm 1 - " + data['quanlysanpham']['themspmoi_tensanpham'])
        logging.info(check_danhsachsanpham_tensp1 == data['quanlysanpham']['themspmoi_tensanpham'])

        check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_trangthaisp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Admin Từ chối duyệt sp- Trạng thái sản phẩm 1 - Trả lại")
        logging.info(check_danhsachsanpham_trangthaisp1 == "Trả lại")
        time.sleep(1)

    def duyetsanpham_duyet(self):
        driver.implicitly_wait(15)
        login.login5(self, "thanghoa1420@gmail.com", "hoathang1420")
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.admin_thuongmai).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.thuongmai_duyetsanpham).click()
        time.sleep(1.5)
        duyetsanpham_tensanpham1 = driver.find_element(By.XPATH, var.duyetsanpham_tensanpham1).text
        check_duyetsanpham_trangthai1 = driver.find_element(By.XPATH, var.check_duyetsanpham_trangthai1).text
        if duyetsanpham_tensanpham1 == data['quanlysanpham']['themspmoi_tensanpham'] and check_duyetsanpham_trangthai1 == "Từ chối":

            logging.info("Admin - Thương mại - Duyệt sản phẩm - Duyệt sp")
            logging.info("check font-end: Tên thái sản phẩm 1 - " + data['quanlysanpham']['themspmoi_tensanpham'])
            logging.info(duyetsanpham_tensanpham1 == data['quanlysanpham']['themspmoi_tensanpham'])

            logging.info("Admin - Thương mại - Duyệt sản phẩm - Duyệt sp")
            logging.info("check font-end: Trạng thái sản phẩm 1 - Từ chối")
            logging.info(check_duyetsanpham_trangthai1 == "Từ chối")

            #Duyệt sản phẩm
            driver.find_element(By.XPATH, var.duyetsanpham_dau3cham).click()
            driver.find_element(By.XPATH, var.duyetsanpham_dau3cham_duyet).click()
            time.sleep(0.5)
            driver.find_element(By.XPATH, var.duyetsanpham_dau3cham_duyet_duyet).click()
            check_message_duyetsanpham_tuchoi = driver.find_element(By.XPATH, var.check_message_duyetsanpham).text
            logging.info("Admin - Thương mại - Duyệt sản phẩm - Duyệt sp")
            logging.info("check font-end: Message từ chối sản phẩm - Dấu 3 chấm - Duyêt")
            logging.info(check_message_duyetsanpham_tuchoi == "Cập nhật thành công")

        else:
            logging.info("Admin - Thương mại - Duyệt sản phẩm - Duyệt sp ")
            logging.info("check font-end: Có hiển thị sản phẩm gửi phê duyệt từ người bán không")
            logging.info("False")
        time.sleep(2)

        driver.implicitly_wait(15)
        login.login4(self, "truongvck33@gmail.com", "voncamk22")
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.quanlysanpham).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachsanpham).click()
        time.sleep(1)
        check_danhsachsanpham_tensp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_tensp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Admin Duyệt sp- Tên sản phẩm 1 - " + data['quanlysanpham']['themspmoi_tensanpham'])
        logging.info(check_danhsachsanpham_tensp1 == data['quanlysanpham']['themspmoi_tensanpham'])

        check_danhsachsanpham_trangthaisp1 = driver.find_element(By.XPATH, var.check_danhsachsanpham_trangthaisp1).text
        logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
        logging.info("check font-end: Admin Duyệt sp- Trạng thái sản phẩm 1 - Đang hoạt động")
        logging.info(check_danhsachsanpham_trangthaisp1 == "Đang hoạt động")
        time.sleep(1)
        if check_danhsachsanpham_tensp1 == data['quanlysanpham']['themspmoi_tensanpham'] and check_danhsachsanpham_trangthaisp1 == "Đang hoạt động":
            driver.find_element(By.XPATH, var.danhsachsanpham_sanpham1_dau3cham).click()
            driver.find_element(By.XPATH, var.danhsachsanpham_sanpham1_dau3cham_xoa).click()
            driver.find_element(By.XPATH, var.danhsachsanpham_sanpham1_dau3cham_xoa_xoa).click()
            check_message_danhsachsanpham_xoasp = driver.find_element(By.XPATH,var.check_message_danhsachsanpham_xoasp).text
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Message - Xóa sản phẩm")
            logging.info(check_message_danhsachsanpham_xoasp == "Xóa sản phẩm thành công")

            check_danhsachsanpham_tensp1_daxoa = driver.find_element(By.XPATH, var.check_danhsachsanpham_tensp1).text
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Xóa sản phẩm - Có xóa được sản phẩm không")
            logging.info(check_danhsachsanpham_tensp1 != check_danhsachsanpham_tensp1_daxoa)
        else:
            logging.info("Người bán - Quản lý sản phẩm -  Danh sách sản phẩm")
            logging.info("check font-end: Admin Duyệt sp- Có hiển thị sản phẩm không")
            logging.info("False")





    
class kenhmarketing():
    def magiamgiacuashop(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.kenhmarketting).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.magiamgiacuashop).click()
        time.sleep(1)
        #Tạo Voucher ngay!
        driver.find_element(By.XPATH, var.taovoucherngay).click()
        time.sleep(1)
        #Voucher toàn shop
        driver.find_element(By.XPATH, var.taovoucher_tenchuongtrinh).send_keys(data['kenhmarketing']['tenvoucher'])
        driver.find_element(By.XPATH, var.taovoucher_mavoucher).send_keys(data['kenhmarketing']['mavoucher'])
        #Thời gian sử dụng
        #Voucher - từ ngày
        driver.find_element(By.XPATH, var.taovoucher_icontungay).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_iconchonnam_2025).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_icontungay).click()
        time.sleep(1)
        count = 0
        while (count < 12):
            count = count + 1
            try:
                driver.find_element(By.XPATH, var.taovoucher_iconchuyenthang).click()
                time.sleep(0.2)
                if driver.find_element(By.XPATH, var.check_taovoucher_lich_thang).text == "January 2025":
                    driver.find_element(By.XPATH, var.taovoucher_tungay_ngay1).click()
                    time.sleep(0.5)
                    driver.find_element(By.XPATH, var.taovoucher_tungay_10gio).click()
                    driver.find_element(By.XPATH, var.taovoucher_tungay_15phut).click()
                    time.sleep(2)
                    driver.find_element(By.XPATH, var.xacnhan).click()
                    break
            except:
                pass
        time.sleep(1)

        #Voucher - đến ngày
        driver.find_element(By.XPATH, var.taovoucher_icondenngay).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_iconchonnam_2026).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_icondenngay).click()
        time.sleep(1)
        count = 0
        while (count < 12):
            count = count + 1
            try:
                driver.find_element(By.XPATH, var.taovoucher_iconchuyenthang).click()
                time.sleep(0.2)
                if driver.find_element(By.XPATH, var.check_taovoucher_lich_thang).text == "January 2026":
                    driver.find_element(By.XPATH, var.taovoucher_tungay_ngay10).click()
                    time.sleep(0.5)
                    driver.find_element(By.XPATH, var.taovoucher_tungay_15gio).click()
                    driver.find_element(By.XPATH, var.taovoucher_tungay_30phut).click()
                    time.sleep(2)
                    driver.find_element(By.XPATH, var.xacnhan).click()
                    break
            except:
                pass
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_chophepluuma).click()

        #Thiết lập mã giảm giá
        driver.find_element(By.XPATH, var.taovoucher_loaigiamgia).click()
        driver.find_element(By.XPATH, var.taovoucher_loaigiamgia_theosotien).click()
        driver.find_element(By.XPATH, var.taovoucher_loaigiamgia_theosotien_input).send_keys(data['kenhmarketing']['loaigiamgia_theosotien'])
        driver.find_element(By.XPATH, var.taovoucher_donhangtoithieu).send_keys(data['kenhmarketing']['donhangtoithieu'])

        xoa = driver.find_element(By.XPATH, var.taovoucher_tongluotsudungtoida)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.taovoucher_tongluotsudungtoida).send_keys("2")

        xoa = driver.find_element(By.XPATH, var.taovoucher_luotsudungtoida)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.taovoucher_luotsudungtoida).send_keys("1")
        #Hiển thị mã giảm giá và các sản phẩm áp dụng
        #Hiển thị nhiều nơi
        driver.execute_script("window.scrollBy(0,700)", "")
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.xacnhan)
        driver.execute_script("arguments[0].click();", button)
        check_messsage_taovoucher = driver.find_element(By.XPATH, var.check_messsage_taovoucher).text
        logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
        logging.info("check font-end: Message Tạo voucher - Tạo Voucher thành công!")
        logging.info(check_messsage_taovoucher)
        logging.info(check_messsage_taovoucher == "Tạo Voucher thành công!")
        time.sleep(2)

        #Xóa Voucher
        driver.execute_script("window.scrollBy(0,400)", "")
        check_danhsachmagiamgia1 = driver.find_element(By.XPATH, var.check_danhsachmagiamgia).text
        if check_danhsachmagiamgia1 == data['kenhmarketing']['tenvoucher']:
            logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
            logging.info("check font-end: Tên Voucher vừa tạo - " + data['kenhmarketing']['tenvoucher'])
            logging.info(check_danhsachmagiamgia1)
            logging.info(check_danhsachmagiamgia1 == data['kenhmarketing']['tenvoucher'])
            driver.find_element(By.XPATH, var.danhsachmagiamgia_chonmagiamgia1).click()
            time.sleep(1)
            driver.find_element(By.XPATH, var.danhsachmagiamgia_xoa).click()
            driver.find_element(By.XPATH, var.danhsachmagiamgia_xoa_xoa).click()
            check_messsage_xoavoucher = driver.find_element(By.XPATH, var.check_messsage_xoavoucher).text
            logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
            logging.info("check font-end: Message Xóa voucher - Cập nhật thành công")
            logging.info(check_messsage_xoavoucher)
            logging.info(check_messsage_xoavoucher == "Cập nhật thành công")
            time.sleep(1)

            check_xoavoucher = driver.find_element(By.XPATH, var.check_danhsachmagiamgia).text
            logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
            logging.info("check font-end: Xóa voucher - " + data['kenhmarketing']['tenvoucher'])
            logging.info(check_xoavoucher)
            logging.info(check_xoavoucher != data['kenhmarketing']['tenvoucher'])

        else:
            logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
            logging.info("check font-end: Tên Voucher vừa tạo - " + data['kenhmarketing']['tenvoucher'])
            logging.info("False")

        driver.execute_script("window.scrollBy(0,-400)", "")
        #Voucher tòan shop
        driver.find_element(By.XPATH, var.magiamgiacuashop_vouchertoanshop_taovoucher).click()
        time.sleep(1)
        check_vouchertoanshop_taovoucher = driver.find_element(By.XPATH, var.check_vouchertoanshop_taovoucher).text
        logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
        logging.info("check font-end: Tạo voucher toàn shop - Coa chuyển tới trang tạo voucher không?")
        logging.info(check_vouchertoanshop_taovoucher)
        logging.info(check_vouchertoanshop_taovoucher == "Tạo mã giảm giá mới")
        driver.back()
        time.sleep(1)

        #Voucher sản phẩm
        driver.find_element(By.XPATH, var.magiamgiacuashop_vouchersanpham_taovoucher).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_tenchuongtrinh).send_keys(data['kenhmarketing']['tenvouchersp1'])
        driver.find_element(By.XPATH, var.taovoucher_mavoucher).send_keys(data['kenhmarketing']['mavouchersp1'])
        #Thời gian sử dụng
        #Voucher - từ ngày
        driver.find_element(By.XPATH, var.taovoucher_icontungay).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_iconchonnam_2025).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_icontungay).click()
        time.sleep(1)
        count = 0
        while (count < 12):
            count = count + 1
            try:
                driver.find_element(By.XPATH, var.taovoucher_iconchuyenthang).click()
                time.sleep(0.2)
                if driver.find_element(By.XPATH, var.check_taovoucher_lich_thang).text == "January 2025":
                    driver.find_element(By.XPATH, var.taovoucher_tungay_ngay1).click()
                    time.sleep(0.5)
                    driver.find_element(By.XPATH, var.taovoucher_tungay_10gio).click()
                    driver.find_element(By.XPATH, var.taovoucher_tungay_15phut).click()
                    time.sleep(2)
                    driver.find_element(By.XPATH, var.xacnhan).click()
                    break
            except:
                pass
        time.sleep(1)

        #Voucher - đến ngày
        driver.find_element(By.XPATH, var.taovoucher_icondenngay).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_iconchonnam).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_iconchonnam_2026).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.xacnhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_icondenngay).click()
        time.sleep(1)
        count = 0
        while (count < 12):
            count = count + 1
            try:
                driver.find_element(By.XPATH, var.taovoucher_iconchuyenthang).click()
                time.sleep(0.2)
                if driver.find_element(By.XPATH, var.check_taovoucher_lich_thang).text == "January 2026":
                    driver.find_element(By.XPATH, var.taovoucher_tungay_ngay10).click()
                    time.sleep(0.5)
                    driver.find_element(By.XPATH, var.taovoucher_tungay_15gio).click()
                    driver.find_element(By.XPATH, var.taovoucher_tungay_30phut).click()
                    time.sleep(2)
                    driver.find_element(By.XPATH, var.xacnhan).click()
                    break
            except:
                pass
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_chophepluuma).click()

        #Thiết lập mã giảm giá
        driver.find_element(By.XPATH, var.taovoucher_loaigiamgia).click()
        driver.find_element(By.XPATH, var.taovoucher_loaigiamgia_theophantram).click()
        driver.find_element(By.XPATH, var.taovoucher_loaigiamgia_theosotien_input).send_keys(data['kenhmarketing']['loaigiamgia_theophantram'])
        driver.find_element(By.XPATH, var.taovoucher_donhangtoithieu).send_keys(data['kenhmarketing']['donhangtoithieu'])

        xoa = driver.find_element(By.XPATH, var.taovoucher_tongluotsudungtoida)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.taovoucher_tongluotsudungtoida).send_keys("2")

        xoa = driver.find_element(By.XPATH, var.taovoucher_luotsudungtoida)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.taovoucher_luotsudungtoida).send_keys("1")
        #Hiển thị mã giảm giá và các sản phẩm áp dụng
        #Chia sẻ thông qua mã vocher
        driver.execute_script("window.scrollBy(0,700)", "")
        driver.find_element(By.XPATH, var.taovoucher_chiasethonguqmavoucher).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.themsanpham)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        #Chọn sản phẩm
        driver.find_element(By.XPATH, var.taovoucher_chonsp_tensp_masp).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.taovoucher_chonsp_tensp_masp_masp).click()
        driver.find_element(By.XPATH, var.taovoucher_chonsp_input).send_keys(data['quanlysanpham']['themspmoi_sku'])
        # driver.find_element(By.XPATH, var.taovoucher_chonsp_timkiem).click()    #Không có nút tìm kiếm
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_chonsp_chonsp1).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_chonsp_xacnhan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_sanphamduocapdung_iconxoa).click()
        time.sleep(1)
        driver.implicitly_wait(2)
        try:
            taovoucher_xoasp1 = driver.find_element(By.XPATH, var.taovoucher_sanphamduocapdung_iconxoa).is_displayed()
            logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
            logging.info("check font-end: Tạo voucher sản phẩm - Thêm sản phẩm - Xóa sản phẩm")
            logging.info("Có xóa được sản phẩm không")
            logging.info("False")
        except NoSuchElementException:
            logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
            logging.info("check font-end: Tạo voucher sản phẩm - Thêm sản phẩm - Xóa sản phẩm")
            logging.info("Có xóa được sản phẩm không")
            logging.info("True")
        driver.implicitly_wait(15)
        button = driver.find_element(By.XPATH, var.themsanpham)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        driver.find_element(By.XPATH, var.taovoucher_chonsp_chonsp1).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.taovoucher_chonsp_xacnhan).click()
        time.sleep(1)
        button = driver.find_element(By.XPATH, var.xacnhan)
        driver.execute_script("arguments[0].click();", button)
        check_messsage_taovoucher = driver.find_element(By.XPATH, var.check_messsage_taovoucher).text
        time.sleep(2)
        check_danhsachmagiamgia_loaima1 = driver.find_element(By.XPATH, var.check_danhsachmagiamgia_loaima1).text
        logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
        logging.info("check font-end: Loại Voucher vừa tạo - Mã giảm giá trên sản phẩm")
        logging.info(check_danhsachmagiamgia_loaima1)
        logging.info(check_danhsachmagiamgia_loaima1 == "Mã giảm giá trên sản phẩm")

    def danhsachmagiamgia(self):
        #Danh sách mã giảm giá
        #Đang diễn ra
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/marketplace/shop/vouchers?page_id=108277159419223993&page=1&type=all")
        time.sleep(2)
        driver.execute_script("window.scrollBy(0,900)", "")
        driver.find_element(By.XPATH, var.danhsachmagiamgia_dangdienra).click()
        time.sleep(1.5)
        check_danhsachmagiamgia_trangthai1 = driver.find_element(By.XPATH, var.check_danhsachmagiamgia_trangthai1).text
        logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
        logging.info("check font-end: Danh sách mã giảm giá - Đang diễn ra - Trạng thái - Đang diễn ra")
        logging.info(check_danhsachmagiamgia_trangthai1)
        logging.info(check_danhsachmagiamgia_trangthai1 == "Đang diễn ra")

        #Sắp diễn ra
        driver.find_element(By.XPATH, var.danhsachmagiamgia_sapdienra).click()
        time.sleep(1.5)
        check_danhsachmagiamgia_trangthai1 = driver.find_element(By.XPATH, var.check_danhsachmagiamgia_trangthai1).text
        logging.info("Người bán - Kênh Marketing -sapMã giảm giá của shop")
        logging.info("check font-end: Danh sách mã giảm giá - Sắp diễn ra - Trạng thái - Sắp diễn ra")
        logging.info(check_danhsachmagiamgia_trangthai1)
        logging.info(check_danhsachmagiamgia_trangthai1 == "Sắp diễn ra")

        #Đã kết thúc
        driver.find_element(By.XPATH, var.danhsachmagiamgia_daketthuc).click()
        time.sleep(1.5)
        check_danhsachmagiamgia_trangthai1 = driver.find_element(By.XPATH, var.check_danhsachmagiamgia_trangthai1).text
        logging.info("Người bán - Kênh Marketing -sapMã giảm giá của shop")
        logging.info("check font-end: Danh sách mã giảm giá - Đã kết thúc - Trạng thái - Đã kết thúc")
        logging.info(check_danhsachmagiamgia_trangthai1)
        logging.info(check_danhsachmagiamgia_trangthai1 == "Đã kết thúc")

        #Tất cả
        driver.find_element(By.XPATH, var.danhsachmagiamgia_tatca).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.danhsachmagiamgia_input).send_keys(data['kenhmarketing']['mavouchersp1'])
        driver.find_element(By.XPATH, var.danhsachmagiamgia_timkiem).click()
        time.sleep(1.5)

        check_danhsachmagiamgia_timkiem = driver.find_element(By.XPATH, var.check_danhsachmagiamgia_timkiem).text
        logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
        logging.info("check font-end: Tìm kiếm mã voucher - " + data['kenhmarketing']['mavouchersp1'])
        logging.info(check_danhsachmagiamgia_timkiem)
        logging.info(check_danhsachmagiamgia_timkiem == data['kenhmarketing']['mavouchersp1'])

        driver.find_element(By.XPATH, var.danhsachmagiamgia_timkiem_xoa).click()
        time.sleep(1.5)
        driver.implicitly_wait(2)
        check_danhsachmagiamgia_timkiem2 = driver.find_element(By.XPATH, var.check_danhsachmagiamgia_timkiem2).is_displayed()
        logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
        logging.info("check font-end: Danh sách mã giảm giá - Xóa tìm kiếm")
        logging.info(check_danhsachmagiamgia_timkiem2)

        #Danh sách mã giảm giá - Dấu 3 chấm
        #Dấu 3 chấm - Cập nhật
        xoa = driver.find_element(By.XPATH, var.danhsachmagiamgia_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.danhsachmagiamgia_input).send_keys(data['kenhmarketing']['mavouchersp1'])
        driver.find_element(By.XPATH, var.danhsachmagiamgia_timkiem).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.danhsachmagiamgia_dau3cham).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.danhsachmagiamgia_dau3cham_capnhat).click()
        time.sleep(1.5)
        driver.execute_script("window.scrollBy(0,-1200)", "")
        xoa = driver.find_element(By.XPATH, var.taovoucher_tenchuongtrinh)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.taovoucher_tenchuongtrinh).send_keys(data['kenhmarketing']['mavouchersp_capnhat'])
        driver.find_element(By.XPATH, var.taovoucher_loaigiamgia_theophantram_input).send_keys("20")

        driver.execute_script("window.scrollBy(0,600)", "")
        button = driver.find_element(By.XPATH, var.taovoucher_capnhat)
        driver.execute_script("arguments[0].click();", button)
        # driver.find_element(By.XPATH, var.taovoucher_capnhat).click()

        check_message_danhsachmagiamgia_capnhat = driver.find_element(By.XPATH, var.check_message_danhsachmagiamgia_capnhat).text
        logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
        logging.info("check font-end: Message Cập nhật sản phẩm - Cập nhật Voucher thành công!")
        logging.info(check_message_danhsachmagiamgia_capnhat)
        logging.info(check_message_danhsachmagiamgia_capnhat == "Cập nhật Voucher thành công!")
        time.sleep(1)

        check_capnhatvoucher = driver.find_element(By.XPATH, var.check_danhsachmagiamgia).text
        logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
        logging.info("check font-end: Cập nhật voucher - Tên voucher - " + data['kenhmarketing']['mavouchersp_capnhat'])
        logging.info(check_capnhatvoucher)
        logging.info(check_capnhatvoucher == data['kenhmarketing']['mavouchersp_capnhat'])

        #Dấu 3 chấm - Xóa
        xoa = driver.find_element(By.XPATH, var.danhsachmagiamgia_input)
        xoa.send_keys(Keys.CONTROL, "a")
        driver.find_element(By.XPATH, var.danhsachmagiamgia_input).send_keys(data['kenhmarketing']['mavouchersp1'])
        driver.find_element(By.XPATH, var.danhsachmagiamgia_timkiem).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.danhsachmagiamgia_dau3cham).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.danhsachmagiamgia_dau3cham_xoa).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.danhsachmagiamgia_dau3cham_xoa_xoa).click()

        check_message_danhsachmagiamgia_xoa = driver.find_element(By.XPATH, var.check_message_danhsachmagiamgia_xoa).text
        logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
        logging.info("check font-end: Message Xóa sản phẩm - Cập nhật thành công")
        logging.info(check_message_danhsachmagiamgia_xoa)
        logging.info(check_message_danhsachmagiamgia_xoa == "Cập nhật thành công")
        time.sleep(1)
        driver.find_element(By.XPATH, var.xoa).click()
        time.sleep(1.5)
        check_dsmagiamgia_xoavoucher = driver.find_element(By.XPATH, var.check_danhsachmagiamgia).text
        logging.info("Người bán - Kênh Marketing - Mã giảm giá của shop")
        logging.info("check font-end: Dấu 3 chấm - Xóa voucher - " + data['kenhmarketing']['mavouchersp_capnhat'])
        logging.info(check_dsmagiamgia_xoavoucher)
        logging.info(check_dsmagiamgia_xoavoucher != data['kenhmarketing']['mavouchersp_capnhat'])
        time.sleep(1)



    def chuongtrinhkhuyenmai_tatca(self):
        driver.implicitly_wait(15)
        time.sleep(1.5)
        button = driver.find_element(By.XPATH, var.khonggianthuongmai)
        driver.execute_script("arguments[0].click();", button)
        time.sleep(1)
        driver.find_element(By.XPATH, var.iconshopcuatoi).click()
        driver.find_element(By.XPATH, var.iconshopcuatoi_binhthuan).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.kenhmarketting).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.chuongtrinhkhuyenmai).click()
        time.sleep(1.5)
        #Tất cả - Chương trình Flash Sale cùng EMSO
        ten_chuongtrinh_flashsale_ten1 = driver.find_element(By.XPATH, var.ten_chuongtrinh_flashsale_ten1).text
        ten_chuongtrinh_flashsale_thoigian1 = driver.find_element(By.XPATH, var.ten_chuongtrinh_flashsale_thoigian1).text
        ten_chuongtrinh_flashsale_chitietchuongtrinh1 = driver.find_element(By.XPATH, var.ten_chuongtrinh_flashsale_chitietchuongtrinh1).text
        ten_chuongtrinh_flashsale_gianhangdangky1 = driver.find_element(By.XPATH, var.ten_chuongtrinh_flashsale_gianhangdangky1).text

        logging.info("Người bán - Kênh marketting - Chương trình khuyến mãi -  Flash Sale")
        logging.info("check font-end: Tên chương trình - Có hiển thị không")
        logging.info(ten_chuongtrinh_flashsale_ten1)
        logging.info(ten_chuongtrinh_flashsale_ten1 != None)


        logging.info("Người bán - Kênh marketting - Chương trình khuyến mãi -  Flash Sale")
        logging.info("check font-end: Thời gian - Có hiển thị không")
        logging.info(ten_chuongtrinh_flashsale_thoigian1)
        logging.info(ten_chuongtrinh_flashsale_thoigian1 != None)


        logging.info("Người bán - Kênh marketting - Chương trình khuyến mãi -  Flash Sale")
        logging.info("check font-end: Chi tiết chương trình - Có hiển thị không")
        logging.info(ten_chuongtrinh_flashsale_chitietchuongtrinh1)
        logging.info(ten_chuongtrinh_flashsale_chitietchuongtrinh1 != None)


        logging.info("Người bán - Kênh marketting - Chương trình khuyến mãi -  Flash Sale")
        logging.info("check font-end: Gian hàng đăng ký - Có hiển thị không")
        logging.info(ten_chuongtrinh_flashsale_gianhangdangky1)
        logging.info(ten_chuongtrinh_flashsale_gianhangdangky1 != None)


        #Xem Flash Sale 1
        driver.find_element(By.XPATH, var.chuongtrinh_flashsale_chon1).click()
        time.sleep(2)
        ten_chuongtrinh_flashsale_vaoxemten1 = driver.find_element(By.XPATH, var.ten_chuongtrinh_flashsale_vaoxemten1).text
        logging.info("Người bán - Kênh marketting - Chương trình khuyến mãi -  Flash Sale")
        logging.info("check font-end: Tên Flash Sale khi chọn xem chi tiết")
        logging.info(ten_chuongtrinh_flashsale_ten1 == ten_chuongtrinh_flashsale_vaoxemten1)

        driver.find_element(By.XPATH, var.dieukienthamgia).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.dangkychiendich).click()
        time.sleep(1)

        ten_chuongtrinh_flashsale_dangkychiendich1 = driver.find_element(By.XPATH, var.ten_chuongtrinh_flashsale_dangkychiendich1).text
        logging.info("Người bán - Kênh marketting - Chương trình khuyến mãi -  Flash Sale")
        logging.info("check font-end: Đăng ký chiến djch - Tên chiến dịch 1 có hiển thị không?")
        logging.info(ten_chuongtrinh_flashsale_dangkychiendich1)
        logging.info(ten_chuongtrinh_flashsale_dangkychiendich1 != None)
        driver.back()
        time.sleep(1)
        driver.back()
        time.sleep(1)
        driver.back()
        time.sleep(2)

        #Tất cả - Chương trình Khuyến mãi cùng EMSO
        driver.execute_script("window.scrollBy(0,400)", "")
        ten_chuongtrinh_khuyenmai_ten1 = driver.find_element(By.XPATH, var.ten_chuongtrinh_khuyenmai_ten1).text
        driver.find_element(By.XPATH, var.chuongtrinh_khuyenmai_chon1).click()
        time.sleep(1.5)
        ten_chuongtrinh_khuyenmai_vaoxemten1 = driver.find_element(By.XPATH, var.ten_chuongtrinh_flashsale_vaoxemten1).text
        logging.info("Người bán - Kênh marketting - Chương trình khuyến mãi -  Khuyến mãi")
        logging.info("check font-end: Tên Khuyến mãi khi chọn xem chi tiết")
        logging.info(ten_chuongtrinh_khuyenmai_vaoxemten1 == ten_chuongtrinh_khuyenmai_ten1)

        driver.find_element(By.XPATH, var.dieukienthamgia).click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.dangkychiendich).click()
        time.sleep(1.5)
        driver.back()
        time.sleep(1)
        driver.back()
        time.sleep(1)
        driver.back()
        time.sleep(2)

        #Tất cả - Chương trình Voucher cùng EMSO
        driver.execute_script("window.scrollBy(0,800)", "")
        ten_chuongtrinh_voucher_ten1 = driver.find_element(By.XPATH, var.ten_chuongtrinh_voucher_ten1).text
        driver.find_element(By.XPATH, var.chuongtrinh_voucher_chon1).click()
        time.sleep(1.5)
        ten_chuongtrinh_voucher_vaoxemten1 = driver.find_element(By.XPATH, var.ten_chuongtrinh_voucher_vaoxemten1).text
        logging.info("Người bán - Kênh marketting - Chương trình khuyến mãi -  Voucher")
        logging.info("check font-end: Tên voucher khi chọn xem chi tiết")
        logging.info(ten_chuongtrinh_voucher_ten1 == ten_chuongtrinh_voucher_vaoxemten1)
        time.sleep(1)


    def chuongtrinhkhuyenmai_flashsale(self):
        driver.implicitly_wait(15)
        login.login4(self, "emsomanagerhd@gmail.com", "khongnhomatkhaucu")
        time.sleep(1.5)
        driver.get("https://cmc-fe.emso.vn/marketplace/shop/campaign?page_id=108277159419223806&type=all")
        time.sleep(2)
        driver.find_element(By.XPATH, var.chuongtrinhkhuyenmai_flashsale).click()
        time.sleep(2)

        n = 0
        while (n < 100):
            n = n + 1
            n = str(n)
            danhsach_flashsale = "//*[@class='app']/div/main/div/div[2]/div/div/div[1]/div/div/div[2]/div/div/div/div["+n+"]/div/div"
            danhsach_flashsale_ten = danhsach_flashsale +"/p[1]"
            danhsach_flashsale_ten1 = driver.find_element(By.XPATH, danhsach_flashsale_ten).text
            danhsach_flashsale_trangthaibutton = danhsach_flashsale +"/button"
            danhsach_flashsale_trangthaibutton1 = driver.find_element(By.XPATH, danhsach_flashsale_trangthaibutton).text
            print(danhsach_flashsale_ten1)
            print(danhsach_flashsale_trangthaibutton1)
            n = int(n)
            if danhsach_flashsale_ten1[0:7] == "FLS VT1" and danhsach_flashsale_trangthaibutton1 == "Đăng ký ngay":
                button = driver.find_element(By.XPATH, danhsach_flashsale_trangthaibutton)
                driver.execute_script("arguments[0].click();", button)
                driver.find_element(By.XPATH, var.dangkychiendich).click()
                time.sleep(1)

                danhsachkhuyenmai = driver.find_elements(By.XPATH, var.danhsachkhuyenmai)
                for khuyenmai in danhsachkhuyenmai:
                    khuyenmai1 = khuyenmai.text
                    if khuyenmai1 == "Giảm giá 10% thời trang nữ 2024":
                        kenhmarketing.chuongtrinhkhuyenmai_themsanpham(self, khuyenmai1, var.thoitrangnu, "10", "2", "1")
                    if khuyenmai1 == "Giảm giá 15% thời trang nam 2024":
                        kenhmarketing.chuongtrinhkhuyenmai_themsanpham(self, khuyenmai1, var.thoitrangnam, "10", "3","2")
                    if khuyenmai1 == "Giảm giá thiết bị âm thanh dưới 199k 2024":
                        kenhmarketing.chuongtrinhkhuyenmai_themsanpham(self, khuyenmai1, var.thietbiamthanh, "10", "2","3")
                    print(khuyenmai1)
                break



    def chuongtrinhkhuyenmai_themsanpham(self, chonkhuyenmai, chonnghanhhang, khuyenmai, soluongsanpham, gioihandathang):
        driver.find_element(By.XPATH,"//*[@class='app']/div/main/div/div[2]/div/div/div[1]/div/div/div[2]/div/div[1]/div/nav//*[text()='" + chonkhuyenmai + "']").click()
        time.sleep(1)
        driver.find_element(By.XPATH, var.themsanpham).click()
        time.sleep(2)
        # Chọn sản phẩm
        driver.find_element(By.XPATH, var.themsanpham_nghanhhang1).click()
        time.sleep(1)
        actions = ActionChains(driver)
        themsanpham_hovernghanhhang = driver.find_element(By.XPATH, var.chonnghanhhang + chonnghanhhang)
        actions.move_to_element(themsanpham_hovernghanhhang).perform()
        time.sleep(1)
        driver.find_element(By.XPATH, var.chonnghanhhang + chonnghanhhang).click()
        time.sleep(1.5)
        driver.find_element(By.XPATH, var.themsanpham_chontatcasanpham).click()
        driver.find_element(By.XPATH, var.themsanpham_chonsanpham1).click()
        driver.find_element(By.XPATH, var.themsanpham_xacnhan).click()
        time.sleep(1)
        # Chỉnh sửa sản phẩm
        driver.find_element(By.XPATH, var.themsanpham_chinhsua_khuyenmai).send_keys(khuyenmai)
        driver.find_element(By.XPATH, var.themsanpham_chinhsua_soluongsanpham).send_keys(soluongsanpham)
        driver.find_element(By.XPATH, var.themsanpham_chinhsua_gioihandathang).send_keys(gioihandathang)
        driver.find_element(By.XPATH, var.themsanpham_chinhsua_chontatcasanpham).click()
        time.sleep(0.5)
        driver.find_element(By.XPATH, var.themsanpham_chinhsua_capnhat).click()
        time.sleep(4)
        button = driver.find_element(By.XPATH, var.themsanpham_chinhsua_xacnhan)
        driver.execute_script("arguments[0].click();", button)
        # check_message_chonspkhuyenmai = driver.find_element(By.XPATH, var.check_message_chonspkhuyenmai)
        time.sleep(2.5)



